#-------------------------------------  کتابخانه ها   --------------------
#region
import tkinter as tk
import mysql.connector
from tkinter import ttk
from PIL import Image, ImageTk
from tkinter import filedialog,messagebox,font,scrolledtext
import subprocess
import os
from openpyxl import Workbook
import datetime
from tkinter import filedialog
from docxtpl import DocxTemplate
from tkinter import filedialog
import shutil
import re

def get_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="Mmmm9905",#   entry  در ادرس ها   تبدیل بهtext شود      entry==>text
        #database="state_agency"
    )
#endregion
#---#----#----#----#----#----------  توابع   ----------#----#----#----#-------------
#-------------------------تابع بستن پروژه-----------------
#region
def close_window():#این تابع بعد از اتصال دیتابیس تکمیل مشود 
    response=messagebox.askyesno("تایید خروج","آیا از خارج شدن اطمینان دارید؟")
    if response:
        root.destroy()
    else:
        return
#endregion
# -------------------------------------تابع فراخوانی ادرس با دکمه-----------
#region
def open_file_folder():
    file_path = filedialog.askopenfilename()
    if file_path:
        folder_path = os.path.dirname(file_path)
        subprocess.run(['explorer', '/select,', file_path])
#endregion
#  ------------------------------------------تابع انتخاب فایل عکس------------
#region
def open_file():
    file_path = filedialog.askopenfilename()
    if file_path:
        os.startfile(file_path)
#endregion
#---------------------------تابع خروجی گزارش اکسل-----------
#region
def show_message(label, text):
    label.config(text=text)
    label.update_idletasks()


def export_excel(sql, sheet_title, file_prefix, label):
    try:
        db = get_connection()
        cursor = db.cursor()
        cursor.execute("USE state_agency")
        cursor.execute(sql)

        data = cursor.fetchall()

        if not data:
            show_message(label, "اطلاعاتی برای خروجی وجود ندارد.")
            return
        column_names = [desc[0] for desc in cursor.description]
        workbook = Workbook()
        sheet = workbook.active  # توجه: active نه activess
        sheet.title = sheet_title

        for col_num, title in enumerate(column_names, 1):
            sheet.cell(row=1, column=col_num, value=title)

        for row_num, row in enumerate(data, 2):
            for col_num, value in enumerate(row, 1):
                sheet.cell(row=row_num, column=col_num, value=value)

        file_name = f"{file_prefix}_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

        file_path = filedialog.asksaveasfilename(
            title="ذخیره فایل اکسل",
            initialfile=file_name,
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )

        if file_path:
            workbook.save(file_path)
            show_message(label, "فایل با موفقیت ذخیره شد.")

        cursor.close()
        db.close()

    except Exception as e:
        show_message(label, f"خطا: {str(e)}")
def excel_gozaresh_maskoni():
    value = gozaresh_file_combo_maskoni.get()

    if value == "گزارش فایل اجاره":
        export_excel(
            "SELECT * FROM sabt_ejareh_maskoni",
            "جدول اجاره مسکونی",
            "گزارش اجاره مسکونی",
            error_label_maskoni
        )

    elif value == "گزارش فایل فروش":
        export_excel(
            "SELECT * FROM sabt_forosh_maskoni",
            "جدول فروش مسکونی",
            "گزارش فروش مسکونی",
            error_label_maskoni
        )

    elif value == "گزارش فایل درخواست اجاره":
        export_excel(
            "SELECT * FROM sabt_darkhast_ejareh_maskoni",
            "جدول درخواست اجاره مسکونی",
            "گزارش درخواست اجاره مسکونی",
            error_label_maskoni
        )

    elif value == "گزارش فایل درخواست خرید":
        export_excel(
            "SELECT * FROM sabt_darkhast_kharid_maskoni",
            "جدول درخواست خرید مسکونی",
            "گزارش درخواست خرید مسکونی",
            error_label_maskoni
        )
def excel_gozaresh_edari_tejari():
    value = gozaresh_file_combo_edari_tejari.get()

    if value == "گزارش فایل اجاره":
        export_excel(
            "SELECT * FROM sabt_ejareh_edari_tejari",
            "جدول اجاره اداری_تجاری",
            "گزارش اجاره اداری_تجاری",
            error_label_edari_tejari
        )

    elif value == "گزارش فایل فروش":
        export_excel(
            "SELECT * FROM sabt_forosh_edari_tejari",
            "جدول فروش اداری_تجاری",
            "گزارش فروش اداری_تجاری",
            error_label_edari_tejari
        )

    elif value == "گزارش فایل درخواست اجاره":
        export_excel(
            "SELECT * FROM sabt_darkhast_ejareh_edari_tejari",
            "جدول درخواست اجاره اداری_تجاری",
            "گزارش درخواست اجاره اداری_تجاری",
            error_label_edari_tejari
        )

    elif value == "گزارش فایل درخواست خرید":
        export_excel(
            "SELECT * FROM sabt_darkhast_kharid_edari_tejari",
            "جدول درخواست خرید اداری_تجاری",
            "گزارش درخواست خرید اداری_تجاری",
            error_label_edari_tejari
        )
def excel_gozaresh_bagh_zamin():
    value = gozaresh_file_combo_bagh_zamin.get()

    if value == "گزارش فایل اجاره باغ":
        export_excel(
            "SELECT * FROM ejareh_bagh",
            "جدول اجاره باغ",
            "گزارش اجاره باغ",
            error_label_bagh_zamin
        )

    elif value == "گزارش فایل فروش باغ":
        export_excel(
            "SELECT * FROM forosh_bagh",
            "جدول فروش باغ",
            "گزارش فروش باغ",
            error_label_bagh_zamin
        )

    elif value == "گزارش فایل درخواست اجاره باغ":
        export_excel(
            "SELECT * FROM darkhast_ejareh_bagh",
            "جدول درخواست اجاره باغ",
            "گزارش درخواست باغ",
            error_label_bagh_zamin
        )

    elif value == "گزارش فایل درخواست خرید باغ":
        export_excel(
            "SELECT * FROM darkhast_kharid_bagh",
            "جدول درخواست خرید باغ",
            "گزارش درخواست خرید باغ",
            error_label_bagh_zamin
        )

    elif value == "گزارش فایل اجاره زمین":
        export_excel(
            "SELECT * FROM ejareh_zamin",
            "جدول اجاره زمین",
            "گزارش اجاره زمین",
            error_label_bagh_zamin
        )

    elif value == "گزارش فایل فروش زمین":
        export_excel(
            "SELECT * FROM forosh_zamin",
            "جدول فروش زمین",
            "گزارش فروش زمین",
            error_label_bagh_zamin
        )

    elif value == "گزارش فایل درخواست اجاره زمین":
        export_excel(
            "SELECT * FROM darkhast_ejareh_zamin",
            "جدول درخواست اجاره زمین",
            "گزارش درخواست زمین",
            error_label_bagh_zamin
        )

    elif value == "گزارش فایل درخواست خرید زمین":
        export_excel(
            "SELECT * FROM darkhast_kharid_zamin",
            "جدول درخواست خرید زمین",
            "گزارش درخواست خرید زمین",
            error_label_bagh_zamin
        )
def excel_gozaresh_kargah():
    value = gozaresh_file_combo_kargah.get()

    if value == "گزارش فایل اجاره":
        export_excel(
            "SELECT * FROM sabt_ejareh_kargah",
            "جدول اجاره کارگاه",
            "گزارش اجاره کارگاه",
            error_label_kargah
        )

    elif value == "گزارش فایل فروش":
        export_excel(
            "SELECT * FROM sabt_forosh_kargah",
            "جدول فروش کارگاه",
            "گزارش فروش کارگاه",
            error_label_kargah
        )

    elif value == "گزارش فایل درخواست اجاره":
        export_excel(
            "SELECT * FROM sabt_darkhast_ejareh_kargah",
            "جدول درخواست اجاره کارگاه",
            "گزارش درخواست اجاره کارگاه",
            error_label_kargah
        )

    elif value == "گزارش فایل درخواست خرید":
        export_excel(
            "SELECT * FROM sabt_darkhast_kharid_kargah",
            "جدول درخواست خرید کارگاه",
            "گزارش درخواست خرید کارگاه",
            error_label_kargah
        )
def gharardadeha():
    pass
#endregion
#------------------------------تابع خروجی فایل ورد قرارداد-----------------
#region
def creat_word_gharardad():
    try:
        property_type = type_melk_gharardad_combo.get()
        contract_type = type_gharardad_combo.get()
        party_one = name_shakhs_aval_gharardad_entry.get()
        party_two = name_shakhs_dovom_gharardad_entry.get()
        description = tozih_gharardad_entry.get("1.0", "end-1c")
        # اعتبارسنجی
        if property_type == "":
            messagebox.showwarning("خطا", "نوع ملک را انتخاب کنید")
            return
        if contract_type == "":
            messagebox.showwarning("خطا", "نوع قرارداد را انتخاب کنید")
            return
        if party_one == "":
            messagebox.showwarning("خطا", "نام طرف اول را وارد کنید")
            return
        if party_two == "":
            messagebox.showwarning("خطا", "نام طرف دوم را وارد کنید")
            return
        # اتصال دیتابیس
        db = get_connection()
        cursor = db.cursor()
        cursor.execute("CREATE DATABASE IF NOT EXISTS state_agency")
        cursor.execute("USE state_agency")
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS gharardad(
            id INT AUTO_INCREMENT PRIMARY KEY,
            tracking_code VARCHAR(30),
            property_type VARCHAR(40),
            contract_type VARCHAR(40),
            party_one VARCHAR(40),
            party_two VARCHAR(40),
            description VARCHAR(225)
        )
        """)
        # ثبت اولیه قرارداد
        cursor.execute("""
        INSERT INTO gharardad
        (property_type,contract_type,party_one,party_two,description
        )
        VALUES (%s,%s,%s,%s,%s)
        """,
        (
            property_type,
            contract_type,
            party_one,
            party_two,
            description
        ))

        db.commit()

        # تولید کد رهگیری

        last_id = cursor.lastrowid
        tracking_code = f"GH-{last_id:06d}"
        cursor.execute("""
        UPDATE gharardad
        SET tracking_code=%s
        WHERE id=%s
        """,
        (
            tracking_code,
            last_id
        ))

        db.commit()

        # ساخت فایل ورد

        if contract_type =="خرید و فروش ":
            template_path = "docx files/قرارداد خام خرید و فروش.docx"

        elif contract_type == "اجاره":
            template_path = "docx files/قرارداد خام اجاره.docx"

        elif contract_type == "مشارکت":
            template_path = "docx files/قرارداد خام مشارکت.docx"

        else:
            messagebox.showerror(
                "خطا",
                "نوع قرارداد نامعتبر است"
            )
            db.close()
            return

        # ساخت فایل ورد

        doc = DocxTemplate(template_path)

        doc.render({
            "tracking_code": tracking_code,
            "property_type": property_type,
            "contract_type": contract_type,
            "party_one": party_one,
            "party_two": party_two,
            "description": description
        })

        # انتخاب محل ذخیره

        file_name = f"{tracking_code}_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.docx"

        file_path = filedialog.asksaveasfilename(
            title="ذخیره فایل قرارداد",
            initialfile=file_name,
            defaultextension=".docx",
            filetypes=[("Word Files", "*.docx")]
        )

        if not file_path:
            db.close()
            return

        # ذخیره فایل

        doc.save(file_path)

        # نمایش کد رهگیری

        for widget in code_frame.winfo_children():
            widget.destroy()

        code_label = tk.Label(
            code_frame,
            text=tracking_code,
            fg="#00BFFF",
            bg="#052340",
            font=("Consolas", 11, "bold")
        )

        code_label.place(
            relx=0.5,
            rely=0.5,
            anchor="center"
        )

        db.close()

        messagebox.showinfo(
            "موفق",
            f"قرارداد با موفقیت ایجاد شد\n\nکد رهگیری: {tracking_code}"
        )

    except Exception as e:

        messagebox.showerror(
            "خطا",
            str(e)
        )
#endregion
#---------تابع پاک کردن فرم اصلی----------------
#region
def delete_root():
    melk_mahdode_gheimat_entry.delete(0,tk.END) 
    metraj_entry.delete(0,tk.END)
    metraj_entry.delete(0,tk.END)
    combo_file_type.set("")
    melk_type_combo.set("")
#endregion
#===================================================
#========================================================
# ------------پنجره اصلی-------------------------------
#region
root = tk.Tk()
root.title("Astate Agency")
root.geometry("1200x700")
#تصاویر پروژه
plus=tk.PhotoImage(file="Images/pluse.png")
elvator_pic=tk.PhotoImage(file="Images/elvator.png")
parking_pic=tk.PhotoImage(file="Images/parking.png")
warehouse_pic=tk.PhotoImage(file="Images/anbari.png")
#--------------------- تصاویر صفحه های اجاره -------------------------------

#---------------------- تصاویر صفحه های فروش ------------------------------- 
image_forosh_maskoni=Image.open("Images/forosh_maskoni.jpg")
image_forosh_edari_tejari=Image.open("Images/forosh_edari_tejari.jpg")
image_forosh_bagh_zamin=Image.open("Images/forosh_bagh_zamin.jpg")
image_forosh_karghah=Image.open("Images/forosh_karghah.jpg")
#---------------------- تصاویر صفحه های درخواست -------------------------------
image_darkhast_maskoni=Image.open("Images/darkhast_maskoni.jpg")
image_darkhast_edari_tejari=Image.open("Images/darkhast_edari_tejari.jpg")
image_darkhast_bagh_zamin=Image.open("Images/darkhast_bagh_zamin.jpg")
image_darkhast_kargah=Image.open("Images/darkhast_kargah.jpg")
#---------------------- تصاویر صفحه های گزارش ---------------------------------
image_gozaresh_maskoni=Image.open("Images/gozaresh_maskoni.jpg")
image_gozaresh_edari_tejari=Image.open("Images/gozaresh_edari_tejari.jpg")
image_gozaresh_bagh_zamin=Image.open("Images/gozaresh_bagh_zamin.jpg")
image_gozaresh_kargah=Image.open("Images/gozaresh_kargah.jpg")
#--------------------------تصاویر صفحه قراردادها------------
image_gharardad=tk.PhotoImage(file="Images/gharardad.png")
image_massage=tk.PhotoImage(file="Images/massage.png")
image_melk=tk.PhotoImage(file="Images/melk_image.png")
image_person=tk.PhotoImage(file="Images/person.png")
image_type_gharardad=tk.PhotoImage(file="Images/type_gharardad.png")
image_word=tk.PhotoImage(file="Images/word.png")
# root.attributes("-fullscreen", True) <<<-----  App فول اسکرین شدن
root.configure(bg="#052340")
main_frame=tk.Frame(root)
main_frame.pack(fill="both",expand=True)
#endregion
#------------------------فریم هدر --------------
#region
header=tk.Frame(main_frame,bg="#052340",height=60)
header.pack(fill='x')
title_font=font.Font(family="Shabnam",size=22,weight="bold")
title_label=tk.Label(header,text="آژانس املاک",fg="#00BFFF",bg="#052340",font=title_font)
title_label.pack(pady=20)
#endregion
#-----------قسمت منوبار پروژه------------------------
#---------------------فریم منو----------------------
#region
menu_frame=tk.Frame(main_frame,bg="#ffffff", relief="flat",height=1)
menu_frame.pack(padx=2, pady=2, fill="x")
#endregion
# ------------- لیست کشویی فیلد فایل های ثبتی-----------
#region
def darkhast():
    box_darkhast.deiconify()
    box_darkhast.grab_set()

def gozaresh():
    box_gozaresh.deiconify()
    box_gozaresh.grab_set()

def forosh():
    box_forosh.deiconify()
    box_forosh.grab_set()

def rahn():
    box_rehn_ejareh.deiconify()
    box_rehn_ejareh.grab_set()

def mosharecat():
    box_mosharekat.deiconify()
    box_mosharekat.grab_set()
#endregion
#endregion
#=======================================================
#-----------توابع برگشت صفحات ثبتی به فرم اصلی----------
#region
#-----برگشت از صفحه اجاره مسکونی-------------------------
def back_home_ejareh_maskoni():
    clear_entry_ejareh_maskoni()
    refresh_after_edit()
    clear_errors_labels_ejareh_maskoni()
    root.deiconify()
    ejareh_maskoni_window.withdraw()
    delete_root()
#--------------------------------پاک شدن Entry صفحه اجاره مسکونی--------------------------
def clear_entry_ejareh_maskoni():
    #خالی کردن  باکس های اجاره مسکونی
    sal_sakht_ejareh_maskoni_entry.delete(0,tk.END)
    metraj_ejareh_maskoni_entry.delete(0,tk.END)
    addrres_ejareh_maskoni_entry.delete("1.0",tk.END)
    tabaghe_ejareh_maskoni_entry.delete(0,tk.END)
    vahed_ejareh_maskoni_entry.delete(0,tk.END)
    otagh_ejareh_maskoni_entry.delete(0,tk.END)
    gheimat_ejare_ejare_maskoni_entry.delete(0,tk.END)
    gheimat_pish_ejare_maskoni_entry.delete(0,tk.END)
    name_malek_ejareh_maskoni_entry.delete(0,tk.END)
    shomareh_malek_ejareh_maskoni_entry.delete(0,tk.END)
    #پنجره امکانات
    sarmaesh_ejareh_maskoni_combo.set("")
    garmaesh_ejareh_maskoni_combo.set("")
    kaf_ejareh_maskoni_combo.set("")
    toilet_ejareh_maskoni_combo.set("")
    parking_checkbutton_btn_ejareh_maskoni.deselect()
    asansor_checkbutton_btn_ejareh_maskoni.deselect()
    anbari_checkbutton_btn_ejareh_maskoni.deselect()

def clear_errors_labels_ejareh_maskoni():
    error_lable_sal_sakht_ejareh_maskoni.config(text="")
    error_lable_metraj_ejareh_maskoni.config(text="")
    error_lable_tabaghe_ejareh_maskoni.config(text="")
    error_lable_vahed_ejareh_maskoni.config(text="")
    error_lable_otagh_ejareh_maskoni.config(text="")
    error_lable_gheimat_pish_ejareh_maskoni.config(text="")
    error_lable_gheimat_ejareh_ejareh_maskoni.config(text="")
    error_lable_addrres_ejareh_maskoni.config(text="")
    error_lable_name_malek_ejareh_maskoni.config(text="")
    error_lable_shomareh_malek_ejareh_maskoni.config(text="")
#-----برگشت از صفحه فروش مسکونی-------------------------
def back_home_forosh_maskoni():
    clear_entry_forosh_maskoni()
    refresh_after_edit()
    clear_errors_labels_forosh_maskoni()
    root.deiconify()
    forosh_maskoni_window.withdraw()
    delete_root()
#--------------------------------پاک شدن Entry صفحه فروش مسکونی--------------------------
def clear_entry_forosh_maskoni():
    sal_sakht_forosh_maskoni_entry.delete(0, tk.END)
    metraj_forosh_maskoni_entry.delete(0, tk.END)
    addrres_forosh_maskoni_entry.delete("1.0", tk.END)
    tabaghe_forosh_maskoni_entry.delete(0, tk.END)
    vahed_forosh_maskoni_entry.delete(0, tk.END)
    otagh_forosh_maskoni_entry.delete(0, tk.END)
    gheimat_kol_forosh_maskoni_entry.delete(0, tk.END)
    name_malek_forosh_maskoni_entry.delete(0, tk.END)
    shomareh_malek_forosh_maskoni_entry.delete(0, tk.END)
    
    #پنجره امکانات
    sarmaesh_combo_forosh_maskoni.set("")
    garmaesh_combo_forosh_maskoni.set("")
    kaf_combo_forosh_maskoni.set("")
    toilet_combo_forosh_maskoni.set("")
    parking_ch_btn_forosh_maskoni.deselect()
    asansor_ch_btn_forosh_maskoni.deselect()
    anbari_checkbuton_forosh_maskoni.deselect()

def clear_errors_labels_forosh_maskoni():
    error_lable_sal_sakht_forosh_maskoni.config(text="")
    error_lable_metraj_forosh_maskoni.config(text="")
    error_lable_tabaghe_forosh_maskoni.config(text="")
    error_lable_vahed_forosh_maskoni.config(text="")
    error_lable_otagh_forosh_maskoni.config(text="")
    error_lable_gheimat_kol_forosh_maskoni.config(text="")
    error_lable_addrres_forosh_maskoni.config(text="")
    error_lable_name_malek_forosh_maskoni.config(text="")
    error_lable_shomareh_malek_forosh_maskoni.config(text="")

#------------------------برگشت از صفحه اجاره اداری/تجاری---------------------
def back_home_ejareh_edari_tejari():
    clear_entry_ejareh_edari_tejari()
    refresh_after_edit()
    root.deiconify()
    ejareh_edari_tejari_window.withdraw()
    delete_root()
#--------------------------------پاک شدن Entry صفحه اجاره اداری/تجاری--------------------------
def clear_entry_ejareh_edari_tejari():
    sal_sakht_ejareh_edari_tejari_entry.delete(0,tk.END)
    metraj_melk_ejareh_edari_tejari_entry.delete(0,tk.END)
    addrres_ejareh_edari_tejari_entry.delete("1.0",tk.END)
    tabaghe_ejareh_edari_tejari_entry.delete(0,tk.END)
    vahed_ejareh_edari_tejari_entry.delete(0,tk.END)
    mablagh_ejare_ejareh_edari_tejari_entry.delete(0,tk.END)
    mablagh_pish_ejareh_edari_tejari_entry.delete(0,tk.END)
    name_malek_ejareh_edari_tejari_entry.delete(0,tk.END)
    shomareh_malek_ejareh_edari_tejari_entry.delete(0,tk.END)
    #پنجره امکانات
    ab_va_gaz_combo_emkanat_ejareh_edari_tejari.set("")
    sarmayesh_combo_emkanat_ejareh_edari_tejari.set("")
    garmayesh_combo_emkanat_ejareh_edari_tejari.set("")
    parking_ch_btn_ejareh_edari_tejari.deselect()
    anbari_ch_btn_ejareh_edari_tejari.deselect()
    asansor_ch_btn_ejareh_edari_tejari.deselect()
#---------------------------برگشت از صفحه فروش اداری/تجاری--------------------
def back_home_forosh_edari_tejari():
    clear_entry_forosh_edari_tejari()
    refresh_after_edit()
    root.deiconify()
    forosh_edari_tejari_window.withdraw()
    delete_root()
#--------------------------پاک شدن Entry صفحه فروش اداری و تجاری------------------------
def clear_entry_forosh_edari_tejari():
    sal_sakht_forosh_edari_tejari_entry.delete(0,tk.END)
    addrres_forosh_edari_tejari_entry.delete("1.0",tk.END)
    tabaghe_forosh_edari_tejari_entry.delete(0,tk.END)
    vahed_forosh_edari_tejari_entry.delete(0,tk.END)
    gheimat_kol_forosh_edari_tejari_entry.delete(0,tk.END)
    metraj_forosh_edari_tejari_entry.delete(0,tk.END)
    name_malek_forosh_edari_tejari_entry.delete(0,tk.END)
    shomareh_malek_forosh_edari_tejari_entry.delete(0,tk.END)
    #پنجره امکانات
    aab_va_gaz_combo_forosh_edari_tejari.set("")
    sarmaesh_combo_forosh_edari_tejari.set("")
    garmaesh_combo_forosh_edari_tejari.set("")
    asansor_ch_btn_forosh_edari_tejari.deselect()
    parking_ch_btn_forosh_edari_tejari.deselect()
    anbari_checkbuton_forosh_edari_tejari.deselect()
#----------------------------برگشت از صفحه اجاره باغ / زمین------------------
def back_home_ejareh_bagh_zamin():
    bagh_type_combo.set("باغ")
    frame_down_ejareh_zamin.place_forget()
    frame_down_ejareh_bagh.place(x=10,y=555)
    clear_entry_ejareh_bagh_zamin()
    ejareh_bagh_zamin_window.withdraw()
    root.deiconify()
    delete_root()
#--------------------------پاک شدن Entry صفحه اجاره باغ و زمین------------------------
def clear_entry_ejareh_bagh_zamin():
    metraj_zamin_ejareh_bagh_zamin_entry.delete(0,tk.END)
    bagh_loctaion_entry.delete("1.0",tk.END)
    bagh_gheimat_ejareh_bagh_zamin_entry.delete(0,tk.END)
    bagh_gheimat_har_metr_ejareh_bagh_zamin_entry.delete(0,tk.END)
    bagh_type_combo.set("باغ")
    room_bagh_checkbutton.deselect()
    name_malek_bagh_zamin_entry.delete(0,tk.END)
    number_malek_bagh_zamin_entry.delete(0,tk.END)

    # امکانات فروش باغ و زمین
    type_vila_ejareh_bagh_zamin_combo.set("")
    option_ejareh_bagh_zamin_combo.set("")
    tree_count_entry.delete(0,tk.END)
    metraj_tree_entry.delete(0,tk.END)
    abyari_combo.set("")
    type_tree_combo.set("")
    label_result_add.delete("1.0",tk.END)
    label_result2_add.config(text="")
    metraj_vila_bagh_entry.delete(0,tk.END)
    sal_sakht_vila_bagh_entry.delete(0,tk.END)
    type_vila_forosh_bagh_zamin_combo.set("")
    toilet_bagh_combo.set("")
    hamam_bagh_combo.set("")
    sanad_bagh_combo.set("")
    chah_bagh.deselect()
    estakhr_bagh.deselect()
    bargh_bagh.deselect()
    gaz_bagh.deselect()
    divar_ejareh_bagh_zamin.deselect()
    #تغییر کاربری
    karbari_ejareh_ejareh_bagh_zamin_combo.set("")
    khak_ejareh_ejareh_bagh_zamin_combo.set("")
    ab_ejareh_ejareh_bagh_zamin_combo.set("")
    metraj_vila_bagh_entry.config(state="disabled")
    sal_sakht_vila_bagh_entry.config(state="disabled")
    type_vila_ejareh_bagh_zamin_combo.config(state="disabled")
    toilet_bagh_combo.config(state="disabled")
    hamam_bagh_combo.config(state="disabled")
    sanad_bagh_combo.config(state="disabled")
    option_ejareh_bagh_zamin_combo.config(state="disabled")
    mojavez_sakht_ejareh_bagh_zamin.config(state="disabled")
    mohavate_ejareh_bagh_zamin.config(state="disabled")
    security_room_zamin_ejareh_bagh_zamin.deselect()
    bargh_tak_faz_zamin_ejareh_bagh_zamin.deselect()
    bargh_se_faz_zamin_ejareh_bagh_zamin.deselect()
    anbar_zamin_ejareh_bagh_zamin.deselect()
    fans_zamin_ejareh_bagh_zamin.deselect()
    mojavaz_chah_zamin_ejareh_bagh_zamin.deselect()
#-----------------------------برگشت از صفحه فروش باغ / زمین------------------
def back_home_forosh_bagh_zamin():
    bagh_type_forosh_bagh_zamin_combo.set("باغ")
    frame_down_forosh_zamin.place_forget()
    frame_down_forosh_bagh.place(x=10,y=555)
    clear_entry_forosh_bagh_zamin()
    forosh_bagh_zamin_window.withdraw()
    root.deiconify()    
    delete_root()
#------------------------- پاک شدن Entry صفحه فروش باغ و زمین----------------------------
def clear_entry_forosh_bagh_zamin():
    metraj_zamin_forosh_bagh_zamin_entry.delete(0,tk.END)
    bagh_loctaion_forosh_bagh_zamin_entry.delete("1.0",tk.END)
    gheimat_har_metr_babagh_zamin_forosh_entry.delete(0,tk.END)
    metraj_derakht_forosh_bagh_zamin_entry.delete(0,tk.END)
    tedad_derakht_forosh_bagh_zamin_entry.delete(0,tk.END)
    metraj_vila_forosh_bagh_zamin_entry.delete(0,tk.END)
    sal_sakht_vila_forosh_bagh_zamin_entry.delete(0,tk.END)
    name_malek_forosh_bagh_entry.delete(0,tk.END)
    number_malek_forosh_bagh_entry.delete(0,tk.END)
    gheimat_kol_forosh_bagh_zamin_entry.delete(0,tk.END)

    # امکانات فروش باغ و زمین
    metraj_derakht_forosh_bagh_zamin_entry.delete(0,tk.END)
    tedad_derakht_forosh_bagh_zamin_entry.delete(0,tk.END)
    abyari_forosh_bagh_zamin_combo.set("")
    type_tree_forosh_bagh_zamin_combo.set("")
    label_natige_forosh_bagh_zamin.delete("1.0",tk.END)
    metraj_vila_forosh_bagh_zamin_entry.delete(0,tk.END)
    sal_sakht_vila_forosh_bagh_zamin_entry.delete(0,tk.END)
    type_vila_forosh_bagh_zamin_combo.set("")
    toilet_forosh_bagh_zamin_combo.set("")
    hamam_forosh_bagh_zamin_combo.set("")
    sanad_forosh_bagh_zamin_combo.set("")
    option_forosh_bagh_zamin_combo.set("")
    chah_forosh_bagh_zamin.deselect()
    estakhr_forosh_bagh_zamin.deselect()
    bargh_keshi_forosh_bagh_zamin.deselect()
    gas_keshi_forosh_bagh_zamin.deselect()
    #تغییر کاربری
    karbari_forosh_bagh_zamin_combo.set("")
    khak_forosh_bagh_zamin_combo.set("")
    ab_forosh_bagh_zamin_combo.set("")
    metraj_vila_forosh_bagh_zamin_entry.config(state="disabled")
    sal_sakht_vila_forosh_bagh_zamin_entry.config(state="disabled")
    type_vila_forosh_bagh_zamin_combo.config(state="disabled")
    toilet_forosh_bagh_zamin_combo.config(state="disabled")
    hamam_forosh_bagh_zamin_combo.config(state="disabled")
    sanad_forosh_bagh_zamin_combo.config(state="disabled")
    option_forosh_bagh_zamin_combo.config(state="disabled")
    
    lable_natige_add_forosh_bagh_zamin.config(text="")
    mojavez_sakht_check_btn_forosh_bagh_zamin.config(state="disabled")
    mohavate_sazi_check_btn_forosh_bagh_zamin.config(state="disabled")
    security_zamin_forosh_bagh_zamin.deselect()
    bargh_kesi_zamin_forosh_bagh_zamin.deselect()
    bargh_keshi_zamin_forosh_bagh_zamin2.deselect()
    anbar_zamin_forosh_bagh_zamin.deselect()
    divar_forosh_bagh_zamin.deselect()
    fans_zamin_forosh_bagh_zamin.deselect()
    mojavez_chah_zamin_forosh_bagh_zamin.deselect()
#----------------------- برگشت از صفحه اجاره کارگاه--------------------
def back_home_ejareh_karghah():
    clear_entry_ejareh_karghah()
    refresh_after_edit()
    ejareh_karghah_window.withdraw()
    root.deiconify()
    delete_root()
#------------------------- پاک شدن Entry صفحه اجاره کارگاه ----------------------------
def clear_entry_ejareh_karghah():
    metraj_ejareh_karghah_entry.delete(0,tk.END)
    addrres_ejareh_karghah_entry.delete("1.0",tk.END)
    vadie_ejare_karghah_entry.delete(0,tk.END)
    gheimat_ejare_ejare_karghah_entry.delete(0,tk.END) 
    name_malek_ejareh_karghah_entry.delete(0,tk.END) 
    shomareh_malek_ejareh_karghah_entry.delete(0,tk.END) 
    sal_sakht_ejareh_karghah_entry.delete(0,tk.END)
    #پنجره امکانات
    vaziat_bargh_ejareh_karghah_combo.set("")
    garmaesh_ejareh_karghah_combo.set("")
    vaziat_ab_ejareh_karghah_combo.set("")
    abzaar_ejareh_karghah_combo.set("")
    toilet_ejareh_karghah_combo.set("")
    hamam_ejareh_karghah__combo.set("")
    otagh_ejareh_karghah_combo.set("")

#----------------------- برگشت از صفحه فروش کارگاه--------------------
def back_home_forosh_karghah():
    clear_entry_forosh_kargah()
    refresh_after_edit()
    forosh_karghah_window.withdraw()
    root.deiconify()
    delete_root()
#-------------------------- پاک شدن Entry صفحه فروش کارگاه-----------------------
def clear_entry_forosh_kargah():
    loctaion_forosh_kargah_entry.delete("1.0",tk.END)
    gheimat_kol_forosh_kargah_entry.delete(0,tk.END)
    metraj_forosh_kargah_entry.delete(0,tk.END)
    name_malek_forosh_kargah_entry.delete(0,tk.END)
    shomareh_malek_forosh_kargah_entry.delete(0,tk.END)
    #پنجره امکانات
    sal_sakht_forosh_kargah_entry.delete(0,tk.END)
    vaziat_bargh_forosh_kargah_combo.set("")
    garmayesh_type_forosh_kargah_combo.set("")
    vaziat_ab_forosh_kargah_combo.set("")
    abzar_forosh_kargah_combo.set("")
    toilet_forosh_kargah_combo.set("")
    hamam_forosh_kargah_combo.set("")
    otagh_forosh_kargah_combo.set("")
#----------------------------برگشت از صفحه درخواست مسکونی-------------------
def back_home_darkhast_maskoni():
    clear_entry_darkhast_maskoni()
    refresh_after_edit()
    clear_errors_labels_darkhast_maskoni()
    darkhast_maskoni_window.withdraw()
    root.deiconify()
    delete_root()
#-------------------------- پاک شدن Entry صفحه درخواست مسکونی-----------------------
def clear_entry_darkhast_maskoni():
    sal_sakht_darkhast_maskoni_entry.delete(0,tk.END)
    addrres_darkhast_maskoni_entry.delete("1.0",tk.END)
    metraj_darkhast_maskoni_entry.delete(0,tk.END)
    tabaghe_darkhast_maskoni_entry.delete(0,tk.END)
    vahed_darkhast_maskoni_entry.delete(0,tk.END)
    otagh_darkhast_maskoni_entry.delete(0,tk.END)
    gheimat_kol_darkhast_maskoni_entry.delete(0,tk.END)
    name_moshtari_darkhast_maskoni_entry.delete(0,tk.END)
    shomareh_moshtari_darkhast_maskoni_entry.delete(0,tk.END)
    mablagh_ejare_darkhast_maskoni_entry.delete(0,tk.END)
    gheimat_pish_darkhast_maskoni_entry.delete(0,tk.END)
    #پنجره امکانات
    sarmaesh_combo_darkhast_maskoni.set("")
    garmaesh_combo_darkhast_maskoni.set("")
    kaf_combo_darkhast_maskoni.set("")
    toilet_combo_darkhast_maskoni.set("")
    parking_ch_btn_darkhast_maskoni.deselect()
    asansor_ch_btn_darkhast_maskoni.deselect()
    anbari_checkbuton_darkhast_maskoni.deselect()

def clear_errors_labels_darkhast_maskoni():
    error_lable_sal_sakht_darkhast_maskoni.config(text="")
    error_lable_metraj_darkhast_maskoni.config(text="")
    error_lable_tabaghe_darkhast_maskoni.config(text="")
    error_lable_vahed_darkhast_maskoni.config(text="")
    error_lable_otagh_darkhast_maskoni.config(text="")
    error_lable_gheimat_kol_darkhast_maskoni.config(text="")
    error_lable_addrres_darkhast_maskoni.config(text="")
    error_lable_name_moshtari_darkhast_maskoni.config(text="")
    error_lable_shomareh_moshtari_darkhast_maskoni.config(text="")
    error_lable_mablagh_ejare_darkhast_maskoni.config(text="")
    error_lable_gheimat_pish_darkhast_maskoni.config(text="")
    #---------------------------برگشت از صفحه درخواست اداری/تجاری--------------------
def back_home_darkhast_edari_tejari():
    clear_entry_darkhast_edari_tejari()
    refresh_after_edit()
    darkhast_edari_tejari_window.withdraw()
    root.deiconify()
    delete_root()
#-------------------------- پاک شدن Entry صفحه درخواست اداری/تجاری-----------------------
def clear_entry_darkhast_edari_tejari():
    sal_sakht_darkhast_edari_tejari_entry.delete(0,tk.END)
    addrres_darkhast_edari_tejari_entry.delete("1.0",tk.END)
    tabaghe_darkhast_edari_tejari_entry.delete(0,tk.END)
    vahed_darkhast_edari_tejari_entry.delete(0,tk.END)
    gheimat_kol_darkhast_edari_tejari_entry.delete(0,tk.END)
    metraj_melk_darkhast_edari_tejari_entry.delete(0,tk.END)
    shomareh_moshtari_darkhast_edari_tejari_entry.delete(0,tk.END)
    name_moshtari_darkhast_edari_tejari_entry.delete(0,tk.END)
    mablagh_vadie_darkhast_edari_tejari_entry.delete(0,tk.END)
    mablagh_ejareh_darkhast_edari_tejari_entry.delete(0,tk.END)
    #پنجره امکانات
    aab_va_gaz_combo_emkanat_darkhast_edari_tejari.set("")
    sarmayesh_combo_emkanat_darkhast_edari_tejari.set("")
    garmayesh_combo_emkanat_darkhast_edari_tejari.set("")
    parking_check_btn_darkhast_edari_tejari.deselect()
    asansor_check_btn_darkhast_edari_tejari.deselect()
    anbari_check_btn_darkhast_edari_tejari.deselect()
#-----------------------------برگشت از صفحه درخواست باغ / زمین------------------
def back_home_darkhast_bagh():
    melk_type_darkhast_bagh_zamin_entry.set("درخواست خرید باغ و زمین")
    karbari_darkhast_bagh_zamin_combo.set("باغ")
    gheimat_har_matr_bagh_zamin_darkhast_lable.place(x=490, y=65, anchor="e")
    gheimat_har_metr_bagh_zamin_darkhast_entry.place(x=28, y=55, width=350, height=25)
    gheimat_kol_bagh_zamin_darkhast_lable.place(x=490, y=20, anchor="e")
    gheimat_kol_bagh_zamin_darkhast_entry.place(x=28, y=13, width=350, height=25)
    time_ejareh_bagh_darkhast_zamin_lable.place_forget()
    mablagh_ejareh_mahaneh_darkhast_lable.place_forget()
    mablagh_ejareh_mahaneh_darkhast_entry.place_forget()
    gheimat_ejareh_bagh_darkhast_zamin_lable.place_forget()#ودیعه
    gheimat_ejareh_bagh_darkhast_zamin_entry.place_forget()#
    frame_down_darkhast_bagh.place(x=10,y=555)
    frame_down_darkhast_zamin.place_forget()
    clear_entry_darkhast_bagh_zamin()
    darkhast_bagh_zamin_window.withdraw()
    root.deiconify()  
    delete_root()
#-------------------------- پاک شدن Entry صفحه درخواست باغ/زمین-----------------------
def clear_entry_darkhast_bagh_zamin():  
    metraj_zamin_darkhast_bagh_zamin_entry.delete(0,tk.END)
    bagh_loctaion_darkhast_bagh_zamin_entry.delete("1.0",tk.END)
    gheimat_har_metr_bagh_zamin_darkhast_entry.delete(0,tk.END)
    metraj_derakht_darkhast_bagh_zamin_entry.delete(0,tk.END)
    tedad_derakht_darkhast_bagh_zamin_entry.delete(0,tk.END)
    metraj_vila_darkhast_bagh_zamin_entry.delete(0,tk.END)
    sal_sakht_vila_darkhast_bagh_zamin_entry.delete(0,tk.END)
    shomareh_moshtari_darkhast_bagh_entry.delete(0,tk.END)
    name_moshtari_darkhast_bagh_entry.delete(0,tk.END)
    gheimat_kol_bagh_zamin_darkhast_entry.delete(0,tk.END)
    gheimat_ejareh_bagh_darkhast_zamin_entry.delete(0,tk.END)
    mablagh_ejareh_mahaneh_darkhast_entry.delete(0,tk.END)
    # امکانات فروش باغ و زمین
    metraj_derakht_darkhast_bagh_zamin_entry.delete(0,tk.END)
    tedad_derakht_darkhast_bagh_zamin_entry.delete(0,tk.END)
    abyari_darkhast_bagh_zamin_combo.set("")
    type_tree_darkhast_bagh_zamin_combo.set("")
    lable_natige_add_darkhast_bagh_zamin.config(text="")
    label_natige_darkhast_bagh_zamin.delete("1.0",tk.END)
    metraj_vila_darkhast_bagh_zamin_entry.delete(0,tk.END)
    sal_sakht_vila_darkhast_bagh_zamin_entry.delete(0,tk.END)
    type_vila_darkhast_bagh_zamin_combo.set("")
    toilet_darkhast_bagh_zamin_combo.set("")
    hamam_darkhast_bagh_zamin_combo.set("")
    sanad_darkhast_bagh_zamin_combo.set("")
    option_darkhast_bagh_zamin_combo.set("")
    chah_darkhast_bagh_zamin.deselect()
    estakhr_darkhast_bagh_zamin.deselect()
    bargh_keshi_darkhast_bagh_zamin.deselect()
    gaz_keshi_darkhast_bagh_zamin.deselect()
    #تغییر کاربری
    karbari_darkhast_bagh_zamin_combo.set("")
    khak_darkhast_bagh_zamin_combo.set("")
    ab_darkhast_bagh_zamin_combo.set("")
    metraj_vila_darkhast_bagh_zamin_entry.config(state="disabled")
    sal_sakht_vila_darkhast_bagh_zamin_entry.config(state="disabled")
    type_vila_darkhast_bagh_zamin_combo.config(state="disabled")
    toilet_darkhast_bagh_zamin_combo.config(state="disabled")
    hamam_darkhast_bagh_zamin_combo.config(state="disabled")
    sanad_darkhast_bagh_zamin_combo.config(state="disabled")
    option_darkhast_bagh_zamin_combo.config(state="disabled")
    mojavez_sakht_check_btn_darkhast_bagh_zamin.config(state="disabled")
    mohavate_sazi_check_btn_darkhast_bagh_zamin.config(state="disabled")
    otagh_check_btn_darkhast_bagh_zamin.deselect()
    security_zamin_darkhast_bagh_zamin.deselect()
    bargh_kesi_zamin_darkhast_bagh_zamin.deselect()
    bargh_keshi_zamin_darkhast_bagh_zamin2.deselect()
    anbar_zamin_darkhast_bagh_zamin.deselect()
    fans_zamin_darkhast_bagh_zamin.deselect()
    mojavez_chah_zamin_darkhast_bagh_zamin.deselect()
    divar_darkhast_bagh_zamin.deselect()
#-----------------------------برگشت از صفحه درخواست کارگاه--------------------
def back_home_darkhast_kargah():
    clear_entry_darkhast_kargah()
    refresh_after_edit()
    darkhast_karghah_window.withdraw()   
    root.deiconify()
    delete_root()
#-------------------------- پاک شدن Entry صفحه درخواست کارگاه-----------------------
def clear_entry_darkhast_kargah():
    metraj_darkhast_kargah_entry.delete(0,tk.END)
    loctaion_darkhast_kargah_entry.delete("1.0",tk.END)
    gheimat_kol_darkhast_kargah_entry.delete(0,tk.END)
    shomareh_moshtari_darkhast_kargah_entry.delete(0,tk.END)
    name_moshtari_darkhast_kargah_entry.delete(0,tk.END)
    mablagh_pish_darkhast_kargah_entry.delete(0,tk.END)
    ejareh_mahaneh_darkhast_kargah_entry.delete(0,tk.END)
    #پنجره امکانات
    sal_sakht_darkhast_kargah_entry.delete(0,tk.END)
    vaziat_bargh_darkhast_kargah_combo.set("")
    garmayesh_type_darkhast_kargah_combo.set("")
    vaziat_ab_darkhast_kargah_combo.set("")
    abzar_darkhast_kargah_combo.set("")
    toilet_darkhast_kargah_combo.set("")
    hamam_darkhast_kargah_combo.set("")
    otagh_darkhast_kargah_combo.set("")
#----------------------برگشت از گزارش مسکونی--------
def back_home_gozaresh_maskoni():
    root.deiconify()
    gozaresh_maskoni.withdraw()
    error_label_maskoni.config(text="")
    gozaresh_file_combo_maskoni.set("")
#-----------------------برگشت از صفحه گزارش اداری و تجاری------------------------
def back_home_gozaresh_edari_tejari():
    root.deiconify()
    gozaresh_edari_tejari.withdraw()
    error_label_edari_tejari.config(text="")
    gozaresh_file_combo_edari_tejari.set("")
#--------------------------برگشت از گزارش باغ و زمین----------------------------
def back_home_gozaresh_bagh_zamin():
    root.deiconify()
    gozaresh_bagh_zamin.withdraw()
    error_label_bagh_zamin.config(text="")
    gozaresh_file_combo_bagh_zamin.set("")
#----------------------برگشت از صفحه گزارش کارگاه------------------
def back_home_gozaresh_kargah():
    root.deiconify()
    gozaresh_kargah.withdraw()
    error_label_kargah.config(text="")
    gozaresh_file_combo_kargah.set("")
def back_main_ghararadad():
    root.deiconify()
    gharardad_window.withdraw()
    type_gharardad_combo.set("")
    type_melk_gharardad_combo.set("")
    name_shakhs_aval_gharardad_entry.delete(0,tk.END)
    name_shakhs_dovom_gharardad_entry.delete(0,tk.END)
    tozih_gharardad_entry.delete(0,tk.END)
    code_label.config(text="")
#endregion
#=========================================================
#--------برگشت از امکانات فایل ها به صفحه اصلی ثبتی-------
#region
def  back_to_darkhast_kargah():
     option_file_frame_darkhast_kargah.withdraw()
     option_file_frame_darkhast_kargah.grab_release()
#----------برگشت باکس ها(نوع ملک)-------------
def back_forosh_exit():
    box_forosh.withdraw()
    box_forosh.grab_release()

def back_rehn_ejareh_exit():
    box_rehn_ejareh.withdraw()
    box_rehn_ejareh.grab_release()

def back_darkhast_exit():
    box_darkhast.withdraw()
    box_darkhast.grab_release()

def back_gozaresh_exit():
    box_gozaresh.withdraw()
    box_gozaresh.grab_release()

def back_mosharekat_exit():
    box_mosharekat.withdraw()
    box_mosharekat.grab_release()
#endregion
#============================================
#--------باز و بسته کردن بین باکس ها----------------
#-----بستن باکس و باز کردن صفحه اجاره مسکونی-----------
def ejareh_maskoni_window():
    box_rehn_ejareh.withdraw()
    root.withdraw()
    ejareh_maskoni_window.deiconify()
    box_rehn_ejareh.grab_release()
#-----بستن باکس و باز کردن صفحه اجاره اداری/تجاری-----------
def ejareh_edari_tejari():
    box_rehn_ejareh.withdraw()
    root.withdraw()
    ejareh_edari_tejari_window.deiconify() 
    box_rehn_ejareh.grab_release()
#-----بستن باکس و باز کردن صفحه اجاره باغ/زمین---------
def ejareh_bagh_zamin():
    box_rehn_ejareh.withdraw()
    root.withdraw()
    ejareh_bagh_zamin_window.deiconify()
    box_rehn_ejareh.grab_release()
#-----بستن باکس و باز کردن صفحه اجاره کارگاه---------
def ejareh_karghah():
    box_rehn_ejareh.withdraw()
    root.withdraw()
    ejareh_karghah_window.deiconify() 
    box_rehn_ejareh.grab_release()
#-----بستن باکس و باز کردن صفحه فروش مسکونی-----------
def forosh_maskoni_window():
    box_forosh.withdraw()
    root.withdraw()
    forosh_maskoni_window.deiconify() 
    box_forosh.grab_release()
#-----بستن باکس و باز کردن صفحه فروش اداری/تجاری-----------
def forosh_edari_tejari():
    box_forosh.withdraw()
    root.withdraw()
    forosh_edari_tejari_window.deiconify()
    box_forosh.grab_release()
#-----بستن باکس و باز کردن صفحه فروش باغ/زمین---------
def forosh_bagh_zamin():
    box_forosh.withdraw()
    root.withdraw()
    forosh_bagh_zamin_window.deiconify()
    box_forosh.grab_release()  
#-----بستن باکس و باز کردن صفحه فروش  کارگاه---------
def forosh_karghah():
    box_forosh.withdraw()
    root.withdraw()
    forosh_karghah_window.deiconify() 
    box_forosh.grab_release()
#--------بستن باکس و باز کردن صفحه درخواست مسکونی--------
def darkhast_maskoni_window():
    box_darkhast.withdraw()
    root.withdraw()
    darkhast_maskoni_window.deiconify() 
    box_darkhast.grab_release()
#-----بستن باکس و باز کردن صفحه درخواست اداری/تجاری-----------
def darkhast_edari_tejari():
    box_darkhast.withdraw()
    root.withdraw()
    darkhast_edari_tejari_window.deiconify()
    box_darkhast.grab_release()
#-----بستن باکس و باز کردن صفحه درخواست باغ/زمین---------
def darkhast_bagh_zamin():
    box_darkhast.withdraw()
    root.withdraw()
    darkhast_bagh_zamin_window.deiconify()
    box_darkhast.grab_release()   
#------بستن باکس و باز کردن صفحه درخواست کارگاه-----------
def darkhast_kargah():
    box_darkhast.withdraw()
    root.withdraw()
    darkhast_karghah_window.deiconify()
    box_darkhast.grab_release()
#-------------------باز و بسته کردن صفحه گزارش مسکونی----------
def gozaresh_maskoni():
    box_gozaresh.withdraw()
    root.withdraw()
    gozaresh_maskoni.deiconify()
    box_gozaresh.grab_release()  
#---------------------باز و بسته کردن صفحه گزارش باغ و زمین-------
def gozaresh_bagh_zamin():
    box_gozaresh.withdraw()
    root.withdraw()
    gozaresh_bagh_zamin.deiconify()
    box_gozaresh.grab_release()  
#-----بستن باکس و باز کردن صفحه گزارش اداری/تجاری-----------
def gozaresh_edari_tejari():
    box_gozaresh.withdraw()
    root.withdraw()
    gozaresh_edari_tejari.deiconify()
    box_gozaresh.grab_release()   
#-------------بستن باکس و باز کردن صفحه گزارش کارگاه------------
def gozaresh_kargah():
    box_gozaresh.withdraw()
    root.withdraw()
    gozaresh_kargah.deiconify()
    box_gozaresh.grab_release()
#------------------------باز و بشته کردن پنجره قرادادها-------- 
def  open_ghrardad():
    root.withdraw()
    gharardad_window.deiconify()
#-------------رادیو باتن باکس-------------------
#تابع رادیو باتن باز و بسته کردن صفحات فروش
def sabt_radio_frosh():
    selected2 = forosh_radio_value.get()

    if selected2==0:
            box_forosh.withdraw()
            root.withdraw()
            forosh_maskoni_window.deiconify()
             
            box_forosh.grab_release()

    elif selected2==2:
            edit_btn_ejareh_maskoni.place_forget()
            delete_btn_forosh_edari_tejari.place_forget()
            zakhire_forosh_edari_tejari.place(x=300,y=30)
            back_to_home_forosh_edari_tejari.place(x=400,y=30)
            box_forosh.withdraw()
            root.withdraw()
            forosh_edari_tejari_window.deiconify()
            

            box_forosh.grab_release()
        
    elif selected2==4:
         box_forosh.withdraw()
         root.withdraw()
         forosh_bagh_zamin_window.deiconify()
         box_forosh.grab_release()
        
    elif selected2==6:
        box_forosh.withdraw()
        root.withdraw()
        forosh_karghah_window.deiconify()
        box_forosh.grab_release()

#تابع رادیو باتن باز و بسته کردن صفحات اجاره
def sabt_radio_rehn():
    selected = ejareh_radio_value.get()

    if selected==0:
            box_rehn_ejareh.withdraw()
            root.withdraw()
            ejareh_maskoni_window.deiconify()
            box_rehn_ejareh.grab_release()
        
    elif selected==2:
        box_rehn_ejareh.withdraw()
        root.withdraw()
        ejareh_edari_tejari_window.deiconify() 
        box_rehn_ejareh.grab_release()

    elif selected==4:
            box_rehn_ejareh.withdraw()
            root.withdraw()
            ejareh_bagh_zamin_window.deiconify()
            box_rehn_ejareh.grab_release()

    elif selected==6:
            box_rehn_ejareh.withdraw()
            root.withdraw()
            ejareh_karghah_window.deiconify()
            box_rehn_ejareh.grab_release()

#تابع رادیو باتن باز و بسته کردن صفحات درخواست
def sabt_radio_darkhast():
    selected = darkhast_radio_value.get()

    if selected==0:
            box_darkhast.withdraw()
            root.withdraw()
            darkhast_maskoni_window.deiconify()
            box_darkhast.grab_release()
        
    elif selected==2:
        box_darkhast.withdraw()
        root.withdraw()
        darkhast_edari_tejari_window.deiconify()
        box_darkhast.grab_release()

    elif selected==4:
        box_darkhast.withdraw()
        root.withdraw()
        darkhast_bagh_zamin_window.deiconify()
        box_darkhast.grab_release()

    elif selected==6:
        box_darkhast.withdraw()
        root.withdraw()
        darkhast_karghah_window.deiconify()
        box_darkhast.grab_release()

#تابع رادیو باتن باز و بسته گزارش ها
def sabt_radio_gozaresh():
    selected = gozaresh_radio_value.get()

    if selected==0:
        box_gozaresh.withdraw()
        root.withdraw()
        gozaresh_maskoni.deiconify()
        box_gozaresh.grab_release() 

    elif selected==2:
        box_gozaresh.withdraw()
        root.withdraw()
        gozaresh_edari_tejari.deiconify()
        box_gozaresh.grab_release()

    elif selected==4:
        box_gozaresh.withdraw()
        root.withdraw()
        gozaresh_bagh_zamin.deiconify()
        box_gozaresh.grab_release()

    elif selected==6:
        box_gozaresh.withdraw()
        root.withdraw()
        gozaresh_kargah.deiconify()
        box_gozaresh.grab_release()

#تابع رادیو باتن باز و بسته کردن صفحات مشارکت
def sabt_radio_mosharekat():
    selected = mosharekat_radio_value.get()

    if selected == 0:
        box_mosharekat.withdraw()
        box_mosharekat.grab_release()

        #root.withdraw()
        #mosharecat_window.deiconify()

        #for f in frames:
         #   f.grid_forget()

        #frames[0].grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

        
    elif selected==2:
        box_mosharekat.withdraw()
        root.withdraw()
        #pishforosh_page.deiconify()
        box_mosharekat.grab_release()
#endregion        
#=======================================================
#region
#---------------/جابه جایی کاربری باغ و زمین در قسمت های فروش/درخواست/اجاره-------------
def change_bagh_zamin1(event):
    co=bagh_type_combo.get()
    if co=="باغ":
        frame_down_ejareh_zamin.place_forget()
        frame_down_ejareh_bagh.place(x=10,y=555)
    else:
        frame_down_ejareh_bagh.place_forget()
        frame_down_ejareh_zamin.place(x=10,y=555)
def change_bagh_zamin_forosh_bagh(event):
    co=bagh_type_forosh_bagh_zamin_combo.get()
    frame_down_forosh_zamin.place_forget()
    if co=="باغ":
        frame_down_forosh_bagh.place(x=10,y=555)
    else:
        frame_down_forosh_bagh.place_forget()
        frame_down_forosh_zamin.place(x=10,y=555)

def change_bagh_zamin_darkhast_bagh(event):
    co=bagh_type_darkhast_bagh_zamin_combo.get()
    if co=="باغ":
        frame_down_darkhast_zamin.place_forget()
        frame_down_darkhast_bagh.place(x=10,y=555)
    else:
        frame_down_darkhast_bagh.place_forget()
        frame_down_darkhast_zamin.place(x=10,y=555)
#endregion
#=============================================================  
#region
#---------------------قسمت اضافه کردن اپشن های تفریحی و درختان در قسمت باغ و زمین------------
selected_trees=[]
def add_tree():# برای اضافه کردن درخت به صورت دستی
    t=type_tree_combo.get()
    if t and t not in selected_trees:
        selected_trees.append(t)
        label_result_add.config(state="normal")
        label_result_add.insert(tk.END,t3+",")
        #label_result_add.config(state="disabled")
selected_option=[]
def add_option():
    op=option_ejareh_bagh_zamin_combo.get()
    if op and op not in selected_option:
        selected_option.append(op)
        label_result2_add.config(text=','.join(selected_option))

selected_trees2=[]
def add_tree2():
    t3=type_tree_forosh_bagh_zamin_combo.get()
    if t3 and t3 not in selected_trees2:
        selected_trees2.append(t3)
        label_natige_forosh_bagh_zamin.config(state="normal")
        label_natige_forosh_bagh_zamin.insert(tk.END,t3+",")
        #label_natige_forosh_bagh_zamin.config(state="disabled")
selected_option2=[]
def add_option2():
    op2=option_forosh_bagh_zamin_combo.get()
    if op2 and op2 not in selected_option2:
        selected_option2.append(op2)
        lable_natige_add_forosh_bagh_zamin.config(text=','.join(selected_option2))
selected_trees3=[]
def add_tree3():
    t4=type_tree_darkhast_bagh_zamin_combo.get()
    if t4 and t4 not in selected_trees3:
        selected_trees3.append(t4)
        label_natige_darkhast_bagh_zamin.config(state="normal")
        label_natige_darkhast_bagh_zamin.insert(tk.END,t4+",")
        #label_natige_forosh_bagh_zamin.config(state="disabled")
selected_option3=[]
def add_option3():
    op3=option_darkhast_bagh_zamin_combo.get()
    if op3 and op3 not in selected_option3:
        selected_option3.append(op3)
        lable_natige_add_darkhast_bagh_zamin.config(text=','.join(selected_option3))
#endregion
#====================================================================================
#=====================================================================================
#region
def home_true_false1(): # برای فعال یا غیر فعال کردن ویجت های خونه باغ در اجاره
    if var0.get()==1:
        metraj_vila_bagh_entry.config(state="normal")
        sal_sakht_vila_bagh_entry.config(state="normal")
        type_vila_ejareh_bagh_zamin_combo.config(state="readonly")
        toilet_bagh_combo.config(state="readonly")
        hamam_bagh_combo.config(state="readonly")
        sanad_bagh_combo.config(state="readonly")
        option_ejareh_bagh_zamin_combo.config(state="readonly")
        mojavez_sakht_ejareh_bagh_zamin.config(state="normal")
        mohavate_ejareh_bagh_zamin.config(state="normal")
    else:
        metraj_vila_bagh_entry.config(state="disabled")
        sal_sakht_vila_bagh_entry.config(state="disabled")
        type_vila_ejareh_bagh_zamin_combo.config(state="disabled")
        toilet_bagh_combo.config(state="disabled")
        hamam_bagh_combo.config(state="disabled")
        sanad_bagh_combo.config(state="disabled")
        option_ejareh_bagh_zamin_combo.config(state="disabled")
        mojavez_sakht_ejareh_bagh_zamin.config(state="disabled")
        mohavate_ejareh_bagh_zamin.config(state="disabled")
def home_true_false2(): #برای فعال یا غیر فعال کردن ویجت های خونه باغ در فروش
    if var0_forosh_bagh_zamin.get()==1:
        metraj_vila_forosh_bagh_zamin_entry.config(state="normal")
        sal_sakht_vila_forosh_bagh_zamin_entry.config(state="normal")
        type_vila_forosh_bagh_zamin_combo.config(state="readonly")
        toilet_forosh_bagh_zamin_combo.config(state="readonly")
        hamam_forosh_bagh_zamin_combo.config(state="readonly")
        sanad_forosh_bagh_zamin_combo.config(state="readonly")
        option_forosh_bagh_zamin_combo.config(state="readonly")
        divar_forosh_bagh_zamin.config(state="normal")
        mojavez_sakht_check_btn_forosh_bagh_zamin.config(state="normal")
        mohavate_sazi_check_btn_forosh_bagh_zamin.config(state="normal")
    else:
        metraj_vila_forosh_bagh_zamin_entry.config(state="disabled")
        sal_sakht_vila_forosh_bagh_zamin_entry.config(state="disabled")
        type_vila_forosh_bagh_zamin_combo.config(state="disabled")
        toilet_forosh_bagh_zamin_combo.config(state="disabled")
        hamam_forosh_bagh_zamin_combo.config(state="disabled")
        sanad_forosh_bagh_zamin_combo.config(state="disabled")
        option_forosh_bagh_zamin_combo.config(state="disabled")
        mojavez_sakht_check_btn_forosh_bagh_zamin.config(state="disabled")
        mohavate_sazi_check_btn_forosh_bagh_zamin.config(state="disabled")
def home_true_false3(): #برای فعال یا غیر فعال کردن ویجت های خونه باغ در درخواست
    if var0_darkhast_bagh_zamin.get()==1:
        metraj_vila_darkhast_bagh_zamin_entry.config(state="normal")
        sal_sakht_vila_darkhast_bagh_zamin_entry.config(state="normal")
        type_vila_darkhast_bagh_zamin_combo.config(state="readonly")
        toilet_darkhast_bagh_zamin_combo.config(state="readonly")
        hamam_darkhast_bagh_zamin_combo.config(state="readonly")
        sanad_darkhast_bagh_zamin_combo.config(state="readonly")
        option_darkhast_bagh_zamin_combo.config(state="readonly")
        divar_darkhast_bagh_zamin.config(state="normal")
        mojavez_sakht_check_btn_darkhast_bagh_zamin.config(state="normal")
        mohavate_sazi_check_btn_darkhast_bagh_zamin.config(state="normal")
    else:
        metraj_vila_darkhast_bagh_zamin_entry.config(state="disabled")
        sal_sakht_vila_darkhast_bagh_zamin_entry.config(state="disabled")
        type_vila_darkhast_bagh_zamin_combo.config(state="disabled")
        toilet_darkhast_bagh_zamin_combo.config(state="disabled")
        hamam_darkhast_bagh_zamin_combo.config(state="disabled")
        sanad_darkhast_bagh_zamin_combo.config(state="disabled")
        option_darkhast_bagh_zamin_combo.config(state="disabled")
        mojavez_sakht_check_btn_darkhast_bagh_zamin.config(state="disabled")
        mohavate_sazi_check_btn_darkhast_bagh_zamin.config(state="disabled")
#endregion
#-------------
#---------------------تابع تعویض ویجت درخواست ها-----------------
#region
def change_darkhast_maskoni_type(event=None):

    change_type = melk_type_darkhast_maskoni_entry.get()

    # فیلدهای خرید
    gheimat_kol_darkhast_maskoni_lable.place_forget()
    gheimat_kol_darkhast_maskoni_entry.place_forget()

    # فیلدهای اجاره
    mablagh_ejare_darkhast_maskoni_lable.place_forget()
    mablagh_ejare_darkhast_maskoni_entry.place_forget()
    gheimat_pish_darkhast_maskoni_lable.place_forget()
    gheimat_pish_darkhast_maskoni_entry.place_forget()

    if change_type == "درخواست خرید مسکونی":

        gheimat_kol_darkhast_maskoni_lable.place(x=465, y=30, anchor="e")
        gheimat_kol_darkhast_maskoni_entry.place(x=18, y=20, width=350, height=25)

        addrres_darkhast_maskoni.place(x=465, y=80, anchor="e")
        addrres_darkhast_maskoni_entry.place(x=18, y=70, width=350, height=50)

    elif change_type == "درخواست اجاره مسکونی":

        mablagh_ejare_darkhast_maskoni_lable.place(x=465, y=30, anchor="e")
        mablagh_ejare_darkhast_maskoni_entry.place(x=18, y=20, width=350, height=25)

        gheimat_pish_darkhast_maskoni_lable.place(x=465, y=75, anchor="e")
        gheimat_pish_darkhast_maskoni_entry.place(x=18, y=65, width=350, height=25)

        addrres_darkhast_maskoni.place(x=465, y=120, anchor="e")
        addrres_darkhast_maskoni_entry.place(x=18, y=110, width=350, height=25)

def change_darkhast_edari_tejari_type(event=None):

    change_type = combo_darkhast_edari_tejari_entry.get()

    # فیلد خرید
    gheimat_kol_darkhast_edari_tejari_lable.place_forget()
    gheimat_kol_darkhast_edari_tejari_entry.place_forget()

    # فیلد اجاره
    mablagh_vadie_darkhast_edari_tejari_lable.place_forget()
    mablagh_vadie_darkhast_edari_tejari_entry.place_forget()

    mablagh_ejareh_darkhast_edari_tejari_lable.place_forget()
    mablagh_ejareh_darkhast_edari_tejari_entry.place_forget()

    if change_type == "درخواست خرید اداری و تجاری":

        gheimat_kol_darkhast_edari_tejari_lable.place(x=465, y=30, anchor="e")
        gheimat_kol_darkhast_edari_tejari_entry.place(x=18, y=20, width=350, height=25)

        addrres_darkhast_edari_tejari_lable.place(x=465, y=80, anchor="e")
        addrres_darkhast_edari_tejari_entry.place(x=18, y=70, width=350, height=50)

    elif change_type == "درخواست اجاره اداری و تجاری":

        mablagh_vadie_darkhast_edari_tejari_lable.place(x=465, y=30, anchor="e")
        mablagh_vadie_darkhast_edari_tejari_entry.place(x=18, y=20, width=350, height=25)

        mablagh_ejareh_darkhast_edari_tejari_lable.place(x=465, y=75, anchor="e")
        mablagh_ejareh_darkhast_edari_tejari_entry.place(x=18, y=65, width=350, height=25)

        addrres_darkhast_edari_tejari_lable.place(x=465, y=120, anchor="e")
        addrres_darkhast_edari_tejari_entry.place(x=18, y=110, width=350, height=25)
#endregion

#=================================DataBase========================
#--------------------------- اعتبارسنجی ورودی ها -------------------
#---------------------اعتبارسنجی فروش مسکونی--------------------
def chck_sal_sakht(event=None):
    sal_sakht=sal_sakht_forosh_maskoni_entry.get().strip()

    if sal_sakht.isdigit() and len(sal_sakht) ==4:
        sal_sakht_forosh_maskoni_entry.config(highlightcolor="white",highlightthickness=0)
        error_lable_sal_sakht_forosh_maskoni.config(text="")
        metraj_forosh_maskoni_entry.config(state="normal")
        tabaghe_forosh_maskoni_entry.config(state="normal")
        vahed_forosh_maskoni_entry.config(state="normal")
        otagh_forosh_maskoni_entry.config(state="normal")
    else:
        sal_sakht_forosh_maskoni_entry.config(highlightthickness=2,highlightcolor="red")
        error_lable_sal_sakht_forosh_maskoni.config(text="فیلد(سال ساخت) باید شامل چهار عدد باشد")
        metraj_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")
        tabaghe_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")
        vahed_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")
        otagh_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")

def chck_metraj(event=None):
    metraj=metraj_forosh_maskoni_entry.get().strip()

    if metraj.isdigit():
        metraj_forosh_maskoni_entry.config(highlightcolor="white",highlightthickness=0)
        error_lable_metraj_forosh_maskoni.config(text="")
        sal_sakht_forosh_maskoni_entry.config(state="normal")
        tabaghe_forosh_maskoni_entry.config(state="normal")
        vahed_forosh_maskoni_entry.config(state="normal")
        otagh_forosh_maskoni_entry.config(state="normal")
    else:
        metraj_forosh_maskoni_entry.config(highlightcolor="red",highlightthickness=2)
        error_lable_metraj_forosh_maskoni.config(text=" فیلد(متراژ) باید شامل اعداد باشد")
        sal_sakht_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")
        tabaghe_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")
        vahed_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")
        otagh_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")
    
def chck_tabaghe(event=None):
    tabaghe=tabaghe_forosh_maskoni_entry.get().strip()

    if tabaghe.isdigit():
        tabaghe_forosh_maskoni_entry.config(highlightcolor="white",highlightthickness=0)
        error_lable_tabaghe_forosh_maskoni.config(text="")
        sal_sakht_forosh_maskoni_entry.config(state="normal")
        metraj_forosh_maskoni_entry.config(state="normal")
        vahed_forosh_maskoni_entry.config(state="normal")
        otagh_forosh_maskoni_entry.config(state="normal")

    else:
        tabaghe_forosh_maskoni_entry.config(highlightcolor="red",highlightthickness=2)
        error_lable_tabaghe_forosh_maskoni.config(text=" فیلد(طبقه) باید شامل اعداد باشد")
        sal_sakht_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")
        metraj_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")
        vahed_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")
        otagh_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")

def chck_vahed(event=None):
    vahed=vahed_forosh_maskoni_entry.get().strip()

    if vahed.isdigit():
        vahed_forosh_maskoni_entry.config(highlightcolor="white",highlightthickness=0)
        error_lable_vahed_forosh_maskoni.config(text="")
        sal_sakht_forosh_maskoni_entry.config(state="normal")
        metraj_forosh_maskoni_entry.config(state="normal")
        tabaghe_forosh_maskoni_entry.config(state="normal")
        otagh_forosh_maskoni_entry.config(state="normal")

    else:
        vahed_forosh_maskoni_entry.config(highlightcolor="red",highlightthickness=2)
        error_lable_vahed_forosh_maskoni.config(text=" فیلد(واحد) باید شامل اعداد باشد")
        sal_sakht_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")
        metraj_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")
        tabaghe_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")
        otagh_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")

def chck_otagh(event=None):
    otagh=otagh_forosh_maskoni_entry.get().strip()

    if otagh.isdigit():
        otagh_forosh_maskoni_entry.config(highlightcolor="white",highlightthickness=0)
        error_lable_otagh_forosh_maskoni.config(text="")
        sal_sakht_forosh_maskoni_entry.config(state="normal")
        metraj_forosh_maskoni_entry.config(state="normal")
        tabaghe_forosh_maskoni_entry.config(state="normal")
        vahed_forosh_maskoni_entry.config(state="normal")

    else:
        otagh_forosh_maskoni_entry.config(highlightcolor="red",highlightthickness=2)
        error_lable_otagh_forosh_maskoni.config(text=" فیلد(اتاق) باید شامل اعداد باشد")
        sal_sakht_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")
        metraj_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")
        tabaghe_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")
        vahed_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")

def chck_gheimat_kol(event=None):
    gheimat_kol=gheimat_kol_forosh_maskoni_entry.get().strip()

    if gheimat_kol.isdigit():
        gheimat_kol_forosh_maskoni_entry.config(highlightcolor="white",highlightthickness=0)
        error_lable_gheimat_kol_forosh_maskoni.config(text="")
        addrres_forosh_maskoni_entry.config(state="normal")
        addrres_forosh_maskoni_entry.config(bg="#ffffff",fg="black")

    else:
        gheimat_kol_forosh_maskoni_entry.config(highlightcolor="red",highlightthickness=2)
        error_lable_gheimat_kol_forosh_maskoni.config(text=" فیلد(قیمت) باید شامل اعداد باشد")
        addrres_forosh_maskoni_entry.config(bg="#808080",fg="white")
        addrres_forosh_maskoni_entry.config(state="disabled")

def chck_addrres(event=None):
    addrres = addrres_forosh_maskoni_entry.get("1.0", tk.END).strip()

    if (re.fullmatch(r"[آ-ی0-9۰-۹\s]+", addrres) and re.search(r"[آ-ی]", addrres)):
        addrres_forosh_maskoni_entry.config(highlightcolor="white",highlightthickness=0)
        error_lable_addrres_forosh_maskoni.config(text="")
        gheimat_kol_forosh_maskoni_entry.config(state="normal")
        

    else:
        addrres_forosh_maskoni_entry.config(highlightcolor="red",highlightthickness=2)
        error_lable_addrres_forosh_maskoni.config(text="فیلد (آدرس) باید شامل حروف فارسی باشد")
        gheimat_kol_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")

def chck_name_malek(event=None):
    name_malek = name_malek_forosh_maskoni_entry.get().strip()

    if re.fullmatch(r"[آ-ی\s]+", name_malek):
        name_malek_forosh_maskoni_entry.config(highlightcolor="white",highlightthickness=0)
        error_lable_name_malek_forosh_maskoni.config(text="")
        shomareh_malek_forosh_maskoni_entry.config(state="normal")

    else:
        name_malek_forosh_maskoni_entry.config(highlightcolor="red",highlightthickness=2)
        error_lable_name_malek_forosh_maskoni.config(text="فیلد (نام مالک) باید شامل حروف فارسی باشد")
        shomareh_malek_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")

def chck_shomareh_malek(event=None):
    shomareh_malek=shomareh_malek_forosh_maskoni_entry.get().strip()

    if shomareh_malek.isdigit() and len(shomareh_malek) ==11:
        shomareh_malek_forosh_maskoni_entry.config(highlightcolor="white",highlightthickness=0)
        error_lable_name_malek_forosh_maskoni.config(text="")
        name_malek_forosh_maskoni_entry.config(state="normal")

    else:
        shomareh_malek_forosh_maskoni_entry.config(highlightcolor="red",highlightthickness=2)
        error_lable_name_malek_forosh_maskoni.config(text=" فیلد(شماره مالک) باید شامل 11رقم باشد")
        name_malek_forosh_maskoni_entry.config(state="disabled",disabledbackground="#808080",disabledforeground="white")

#region ============*توابع ثبتی دیتابیس*=============================
#--------------------------------------تابع ثبت فروش---------------------------
#---------------------------forosh_maskoni------------------------------
def sabt_forosh_maskoni():
    db = None
    try:
        db = get_connection()
        cursor = db.cursor()
        
        cursor.execute("CREATE DATABASE IF NOT EXISTS state_agency")
        cursor.execute("USE state_agency")

        # دستور ساده و انگلیسی خالص
        sql_create = """
        CREATE TABLE IF NOT EXISTS sabt_forosh_maskoni (
            id INT AUTO_INCREMENT PRIMARY KEY ,
            type_melk VARCHAR(50) NOT NULL,
            sal_sakht VARCHAR(20),
            address VARCHAR(225),
            tabaghe INT,
            vahed INT,
            otagh INT,
            parking VARCHAR(20),
            asansor VARCHAR(20),
            anbari VARCHAR(20),
            sarmayesh VARCHAR(20),
            garmayesh VARCHAR(20),
            kaf VARCHAR(20),
            toilet VARCHAR(20),
            name_malek VARCHAR(20),
            shomareh_malek TEXT CHECK(length(trim(shomareh_malek)) = 11),
            gheimat_kol DECIMAL(15,2),
            metraj INT

        )
        """
        cursor.execute(sql_create)

        sql_insert = """
        INSERT INTO sabt_forosh_maskoni 
        (type_melk,sal_sakht,address,tabaghe,vahed,otagh,parking,asansor,
        anbari,sarmayesh,garmayesh,kaf,toilet,name_malek,shomareh_malek,gheimat_kol,metraj)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """
        values = (
            melk_type_forosh_maskoni_entry.get(),
            sal_sakht_forosh_maskoni_entry.get(),
            addrres_forosh_maskoni_entry.get("1.0",tk.END),
            tabaghe_forosh_maskoni_entry.get(),
            vahed_forosh_maskoni_entry.get(),
            otagh_forosh_maskoni_entry.get(),
            parking_forosh_maskoni_var.get(),
            asansor_forosh_maskoni_var.get(),
            anbari_forosh_maskoni_var.get(),
            sarmaesh_combo_forosh_maskoni.get(),
            garmaesh_combo_forosh_maskoni.get(),
            kaf_combo_forosh_maskoni.get(),
            toilet_combo_forosh_maskoni.get(),
            name_malek_forosh_maskoni_entry.get(),
            shomareh_malek_forosh_maskoni_entry.get(),
            float(gheimat_kol_forosh_maskoni_entry.get()),
            metraj_forosh_maskoni_entry.get()

        )

        cursor.execute(sql_insert, values)
        last_id = cursor.lastrowid
        user_idcode = f"ID-{last_id}"
        messagebox.showinfo("Success", f"ثبت با کد {user_idcode} انجام شد.")
        error_lable_sal_sakht_forosh_maskoni.config(text="")
        error_lable_metraj_forosh_maskoni.config(text="")
        error_lable_tabaghe_forosh_maskoni.config(text="")
        error_lable_vahed_forosh_maskoni.config(text="")
        error_lable_otagh_forosh_maskoni.config(text="")
        error_lable_gheimat_kol_forosh_maskoni.config(text="")
        error_lable_addrres_forosh_maskoni.config(text="")
        error_lable_name_malek_forosh_maskoni.config(text="")
        error_lable_shomareh_malek_forosh_maskoni.config(text="")
        db.commit()

    except Exception as e:
        messagebox.showerror("Error", f"خطا: {e}")
    finally:
        if db and db.is_connected():
            clear_entry_forosh_maskoni()
            db.close()
#------------------- forosh_edari_tejari database -----------------------------------
def sabt_forosh_edari_tejari():
    db = None
    try:
        db = get_connection()
        cursor = db.cursor()
        cursor.execute("CREATE DATABASE IF NOT EXISTS state_agency")
        cursor.execute("USE state_agency")

        sql_create = """
        CREATE TABLE IF NOT EXISTS sabt_forosh_edari_tejari (
            id INT AUTO_INCREMENT PRIMARY KEY,
            type_melk VARCHAR(50) NOT NULL,
            metraj_melk VARCHAR(20),
            sal_sakht VARCHAR(20),
            address VARCHAR(225),
            tabaghe VARCHAR(10),
            vahed VARCHAR(20),
            parking VARCHAR(20),
            asansor VARCHAR(20),
            anbari VARCHAR(20),
            aab_va_gaz VARCHAR(20),
            system_sarmayesh VARCHAR(20),
            system_garmayesh VARCHAR(20),
            name_malek VARCHAR(20),
            shomareh_malek VARCHAR(11),
            gheimat_kol DECIMAL(15,2)

        )
        """
        cursor.execute(sql_create)

        sql_insert = """
        INSERT INTO sabt_forosh_edari_tejari
        (type_melk,metraj_melk,sal_sakht,address,tabaghe,vahed,parking,
        asansor,anbari,aab_va_gaz,system_sarmayesh,system_garmayesh,name_malek,shomareh_malek,gheimat_kol)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """
        values = (
            melk_type_forosh_edari_tejari_entry.get(),
            metraj_forosh_edari_tejari_entry.get(),
            sal_sakht_forosh_edari_tejari_entry.get(),
            addrres_forosh_edari_tejari_entry.get("1.0",tk.END),
            tabaghe_forosh_edari_tejari_entry.get(),
            vahed_forosh_edari_tejari_entry.get(),
            parking_forosh_edari_tejari_var.get(),
            asansor_forosh_edari_tejari_var.get(),
            anbari_forosh_edari_tejari_var.get(),
            aab_va_gaz_combo_forosh_edari_tejari.get(),
            sarmaesh_combo_forosh_edari_tejari.get(),
            garmaesh_combo_forosh_edari_tejari.get(),
            name_malek_forosh_edari_tejari_entry.get(),
            shomareh_malek_forosh_edari_tejari_entry.get(),
            float(gheimat_kol_forosh_edari_tejari_entry.get()))
        
        cursor.execute(sql_insert, values)
        last_id = cursor.lastrowid
        user_idcode = f"ID-{last_id}"
        messagebox.showinfo("Success", f"ثبت با کد {user_idcode} انجام شد.")
        db.commit()

    except Exception as e:
        messagebox.showerror("Error", f"خطا: {e}")
    finally:
        if db and db.is_connected():
            clear_entry_forosh_edari_tejari()
            db.close()
#---------------------------forosh_bagh/zamin Database-------------------------
#region
#تابع مادر 
selected_option2=[]
selected_trees2=[]
#endregion
def sabt_forosh_bagh_zamin_main():
    db = None
    
    try:
        db = get_connection()
        cursor = db.cursor()
        cursor.execute("CREATE DATABASE IF NOT EXISTS state_agency")
        cursor.execute("USE state_agency")

                # ========== دریافت مقادیر ==========
        karbari = bagh_type_forosh_bagh_zamin_combo.get()
        gheimat_str = gheimat_har_metr_babagh_zamin_forosh_entry.get()
        if not gheimat_str or gheimat_str == "":
            gheimat_str = "0"
        gheimat_value = float(gheimat_str)
        
        if karbari == "باغ": 
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS forosh_bagh(
                id INT AUTO_INCREMENT PRIMARY KEY,
                type_melk VARCHAR(50) NOT NULL,
                metraj VARCHAR(20),
                karbari VARCHAR(20),
                address VARCHAR(255),
                mablagh_metri DECIMAL(15,2),
                name_malek VARCHAR(50),
                shomareh_malek VARCHAR(30),
                gheimat_kol DECIMAL(15,2),
                metraj_derakht VARCHAR(10),
                tedad_derakht VARCHAR(10),
                type_derakht TEXT,
                system_ab VARCHAR(25),
                chah VARCHAR(10),
                estakhr VARCHAR(10),
                divar VARCHAR(10),
                sazeh VARCHAR(10),
                metraj_sazeh VARCHAR(10),
                sal_sakht VARCHAR(10),
                type_sazeh VARCHAR(20),
                emkanat TEXT,
                WC VARCHAR(10),
                hamam VARCHAR(10),
                javaz_sakht VARCHAR(10),
                sanad VARCHAR(20),
                mohavate VARCHAR(10),
                bargh VARCHAR (10),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """) 
            type_derakht_value = ",".join(selected_trees2) if selected_trees2 else ""
            tedad_derakht_value = str(len(selected_trees2)) if selected_trees2 else "0"
            emkanat_value = ",".join(selected_option2) if selected_option2 else ""
            
            sql_bagh = """
                INSERT INTO forosh_bagh(
                    type_melk,metraj,karbari,address,mablagh_metri,name_malek,shomareh_malek,gheimat_kol,
                    metraj_derakht, tedad_derakht, type_derakht,
                    system_ab, chah, estakhr, divar,sazeh, metraj_sazeh,
                    sal_sakht, type_sazeh, emkanat, WC, hamam,
                    javaz_sakht, sanad, mohavate,bargh
                )
                VALUES( %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s
                ,%s,%s,%s,%s,%s,%s,%s,%s)
            """
            
            values_bagh = (
                melk_type_forosh_bagh_zamin_entry.get(),
                metraj_zamin_forosh_bagh_zamin_entry.get(),
                bagh_type_forosh_bagh_zamin_combo.get(),
                bagh_loctaion_forosh_bagh_zamin_entry.get("1.0",tk.END),
                gheimat_har_metr_babagh_zamin_forosh_entry.get(),
                name_malek_forosh_bagh_entry.get(),
                number_malek_forosh_bagh_entry.get(),
                gheimat_kol_forosh_bagh_zamin_entry.get(),
                metraj_derakht_forosh_bagh_zamin_entry.get(),
                tedad_derakht_value,
                type_derakht_value,
                abyari_forosh_bagh_zamin_combo.get(),
                chah_forosh_bagh_var.get(),
                estakhr_forosh_bagh_var.get(),
                divar_forosh_bagh_var.get(),
                var0_forosh_bagh_zamin.get(),
                metraj_vila_forosh_bagh_zamin_entry.get(),
                sal_sakht_vila_forosh_bagh_zamin_entry.get(),
                type_vila_forosh_bagh_zamin_combo.get(),
                emkanat_value,
                toilet_forosh_bagh_zamin_combo.get(),
                hamam_forosh_bagh_zamin_combo.get(),
                mojavez_sakht_forosh_bagh_var.get(),
                sanad_forosh_bagh_zamin_combo.get(),
                mohavate_forosh_bagh_var.get(),
                bargh_keshi_forosh_bagh_var.get()
            )
            cursor.execute(sql_bagh, values_bagh)

            last_id = cursor.lastrowid
            if last_id is None or last_id == 0:
                messagebox.showerror("Error", "خطا: ثبت در جدول  انجام نشد")
                return
        elif karbari == "زمین کشاورزی":
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS forosh_zamin(
                id INT AUTO_INCREMENT PRIMARY KEY,
                type_melk VARCHAR(50) NOT NULL,
                metraj VARCHAR(20),
                karbari VARCHAR(20),
                address VARCHAR(255),
                mablagh_metri DECIMAL(15,2),
                name_malek VARCHAR(50),
                shomareh_malek VARCHAR(30),
                gheimat_kol DECIMAL(15,2),
                karbari_zamin VARCHAR(50),
                type_khak VARCHAR(20),
                manba_ab VARCHAR(20),
                negahbani VARCHAR(20),
                bargh_takfaz VARCHAR(20),
                bargh_sefaz VARCHAR(20),
                anbar VARCHAR(20),
                fans VARCHAR(20),
                chah VARCHAR(20),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP  
            )
        """)

            sql_zamin = """
                INSERT INTO forosh_zamin(
                    type_melk,metraj,karbari,address,mablagh_metri,name_malek,shomareh_malek,gheimat_kol,
                    karbari_zamin,type_khak,
                    manba_ab, negahbani, bargh_takfaz, bargh_sefaz,
                    anbar, fans, chah
                )
                VALUES( %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s,%s,%s,%s,%s,%s,%s)
            """
            
            values_zamin = (
                melk_type_forosh_bagh_zamin_entry.get(),
                metraj_zamin_forosh_bagh_zamin_entry.get(),
                bagh_type_forosh_bagh_zamin_combo.get(),
                bagh_loctaion_forosh_bagh_zamin_entry.get("1.0",tk.END),
                gheimat_har_metr_babagh_zamin_forosh_entry.get(),
                name_malek_forosh_bagh_entry.get(),
                number_malek_forosh_bagh_entry.get(),
                gheimat_kol_forosh_bagh_zamin_entry.get(),
                bagh_type_forosh_bagh_zamin_combo.get(),
                khak_forosh_bagh_zamin_combo.get(),
                ab_forosh_bagh_zamin_combo.get(),
                security_zamin_forosh_zamin_var.get(),
                bargh_kesi_zamin_forosh_zamin_var.get(),
                bargh_kesi_zamin_forosh_zamin2_var.get(),
                anbar_zamin_forosh_zamin_var.get(),
                fans_zamin_forosh_zamin_var.get(),
                javaz_chah_zamin_forosh_zamin_var.get()
            )
            cursor.execute(sql_zamin, values_zamin)
            last_id = cursor.lastrowid
            if last_id is None or last_id == 0:
                messagebox.showerror("Error", "خطا: ثبت در جدول  انجام نشد")
                return
        
        db.commit()
        user_idcode = f"ID-{last_id}"
        messagebox.showinfo("Success", f"ثبت با کد {user_idcode} انجام شد.")      
        
    except Exception as e:
        messagebox.showerror("Error", f"خطا در ثبت داده: {e}")
        
    finally:
        if db and db.is_connected():
            clear_entry_forosh_bagh_zamin()
            db.close()
#---------------------------- forosh_karghah Database ------------------------
def sabt_forosh_kargah():
    db = None
    try:
        db = get_connection()
        cursor = db.cursor()
        cursor.execute("CREATE DATABASE IF NOT EXISTS state_agency")
        
        cursor.execute("USE state_agency")

        sql_create = """
        CREATE TABLE IF NOT EXISTS sabt_forosh_kargah (
            id INT AUTO_INCREMENT PRIMARY KEY,
            karbari_zamin VARCHAR(50) NOT NULL,
            metraj VARCHAR(20),
            address VARCHAR(225),
            sal_sakht VARCHAR(20),
            vaziat_bargh VARCHAR(20),
            garmayesh VARCHAR(20),
            sarmayesh VARCHAR(30),
            vaziat_ab VARCHAR(30),
            abzar VARCHAR(30),
            toilet VARCHAR(20),
            hamam VARCHAR(20),
            otagh VARCHAR(20),
            name_malek VARCHAR(20),
            shomareh_malek VARCHAR(11),
            gheimat_kol DECIMAL(15,2)
        )
        """
        cursor.execute(sql_create)
        sql_insert = """
        INSERT INTO sabt_forosh_kargah
        (karbari_zamin,metraj,address,sal_sakht,
        vaziat_bargh,garmayesh,sarmayesh,vaziat_ab,abzar,toilet,hamam,otagh,name_malek,shomareh_malek,gheimat_kol)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """
        values = (
            karbari_forosh_kargah_entry.get(),
            metraj_forosh_kargah_entry.get(),
            loctaion_forosh_kargah_entry.get("1.0",tk.END),
            sal_sakht_forosh_kargah_entry.get(),
            vaziat_bargh_forosh_kargah_combo.get(),
            garmayesh_type_forosh_kargah_combo.get(),
            sarmayesh_forosh_kargah_combo.get(),
            vaziat_ab_forosh_kargah_combo.get(),
            abzar_forosh_kargah_combo.get(),
            toilet_forosh_kargah_combo.get(),
            hamam_forosh_kargah_combo.get(),
            otagh_forosh_kargah_combo.get(),
            name_malek_forosh_kargah_entry.get(),
            shomareh_malek_forosh_kargah_entry.get(),
            float(gheimat_kol_forosh_kargah_entry.get())
        )
        cursor.execute(sql_insert, values)
        last_id = cursor.lastrowid
        user_idcode = f"ID-{last_id}"
        messagebox.showinfo("Success", f"ثبت با کد {user_idcode} انجام شد.")
        db.commit()

    except Exception as e:
        messagebox.showerror("Error", f"خطا: {e}")
    finally:
        if db and db.is_connected():
            clear_entry_forosh_kargah()
            db.close()
#------------------------------------پایان ثبت فروش-----------------------------
#----------------------------تابع ثبت اجاره----------------------------------
#----------------------- ejareh_maskoni Database -------------------------------
def sabt_ejareh_maskoni():
# اعتبارسنجی 
    sal_sakht= sal_sakht_ejareh_maskoni_entry.get().strip()
    metraj= metraj_ejareh_maskoni_entry.get().strip()
    tabaghe= tabaghe_ejareh_maskoni_entry.get().strip()
    vahed= vahed_ejareh_maskoni_entry.get().strip()
    otagh= otagh_ejareh_maskoni_entry.get().strip()
    gheimat_pish= gheimat_pish_ejare_maskoni_entry.get().strip()
    gheimat_ejare= gheimat_ejare_ejare_maskoni_entry.get().strip()
    addrres = addrres_ejareh_maskoni_entry.get("1.0", "end-1c").strip()
    name_malek= name_malek_ejareh_maskoni_entry.get().strip()
    shomareh_malek= shomareh_malek_ejareh_maskoni_entry.get().strip()

    if len(sal_sakht) != 4 or not sal_sakht.isdigit():
        error_lable_sal_sakht_ejareh_maskoni.config(text="فیلد (سال ساخت) باید 4 رقمی باشد ")
        return
    elif not metraj.isdigit():
        error_lable_metraj_ejareh_maskoni.config(text="فیلد (متراژ) فقط باید شامل اعداد باشد ")
        return
    elif not tabaghe.isdigit():
        error_lable_tabaghe_ejareh_maskoni.config(text="فیلد (طبقه) فقط باید شامل اعداد باشد ")
        return
    elif not vahed.isdigit():
        error_lable_vahed_ejareh_maskoni.config(text="فیلد (واحد) فقط باید شامل اعداد باشد ")
        return
    elif not otagh.isdigit():
        error_lable_otagh_ejareh_maskoni.config(text="فیلد (اتاق) فقط باید شامل اعداد باشد ")
        return
    elif not gheimat_pish.isdigit():
        error_lable_gheimat_pish_ejareh_maskoni.config(text="فیلد (مبلغ پیش) فقط باید شامل اعداد باشد ")
        return
    elif not gheimat_ejare.isdigit():
        error_lable_gheimat_ejareh_ejareh_maskoni.config(text="فیلد (مبلغ اجاره) فقط باید شامل اعداد باشد ")
        return
    elif not re.fullmatch(r"[آ-ی0-9۰-۹\s]+", addrres):
        error_lable_addrres_ejareh_maskoni.config(text="فیلد (آدرس) فقط باید شامل حروف فارسی و اعداد باشد ")
        return
    elif not name_malek or not re.match("^[\u0600-\u06FF\s]+$", name_malek):
        error_lable_addrres_ejareh_maskoni.config(text="")
        error_lable_name_malek_ejareh_maskoni.config(text="فیلد (نام مالک) فقط باید شامل حروف فارسی باشد ")
        return
    elif len(shomareh_malek) != 11 or not shomareh_malek.isdigit():
        error_lable_name_malek_ejareh_maskoni.config(text="")
        error_lable_shomareh_malek_ejareh_maskoni.config(text="فیلد (شماره مالک) باید 11 رقمی باشد ")
        return

    db = None
    try:
        db = get_connection()
        cursor = db.cursor()
        
        cursor.execute("CREATE DATABASE IF NOT EXISTS state_agency")
        cursor.execute("USE state_agency")

        sql_create = """
            CREATE TABLE IF NOT EXISTS sabt_ejareh_maskoni (
            id INT AUTO_INCREMENT PRIMARY KEY,
            type_melk VARCHAR(50) NOT NULL,
            sal_sakht VARCHAR(20),
            address VARCHAR(225),
            tabaghe VARCHAR(10),
            vahed VARCHAR(20),
            otagh INT,
            parking VARCHAR(20),
            asansor VARCHAR(20),
            anbari VARCHAR(20),
            sarmayesh VARCHAR(20),
            garmayesh VARCHAR(20),
            kaf VARCHAR(20),
            toilet VARCHAR(20),
            ejareh VARCHAR(20),
            pish VARCHAR(20),
            name_malek VARCHAR(20),
            shomareh_malek TEXT CHECK(length(trim(shomareh_malek)) = 11),
            metraj VARCHAR(20)
        )
        """

        cursor.execute(sql_create)

        sql_insert = """
        INSERT INTO sabt_ejareh_maskoni 
        (type_melk,sal_sakht,address,tabaghe,vahed,otagh,parking,
        asansor,anbari,sarmayesh,garmayesh,kaf,toilet,ejareh,pish,name_malek,shomareh_malek,metraj)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """

        values = (
            melk_type_ejareh_maskoni_entry.get(),
            sal_sakht_ejareh_maskoni_entry.get(),
            addrres_ejareh_maskoni_entry.get("1.0",tk.END),
            tabaghe_ejareh_maskoni_entry.get(),
            vahed_ejareh_maskoni_entry.get(),
            otagh_ejareh_maskoni_entry.get(),
            parking_ejareh_maskoni_var.get(),
            asansor_ejareh_maskoni_var.get(),
            anbari_ejareh_maskoni_var.get(),
            sarmaesh_ejareh_maskoni_combo.get(),
            garmaesh_ejareh_maskoni_combo.get(),
            kaf_ejareh_maskoni_combo.get(),
            toilet_ejareh_maskoni_combo.get(),
            float(gheimat_ejare_ejare_maskoni_entry.get()),
            float(gheimat_pish_ejare_maskoni_entry.get()),
            name_malek_ejareh_maskoni_entry.get(),
            shomareh_malek_ejareh_maskoni_entry.get(),
            metraj_ejareh_maskoni_entry.get()

        )

        cursor.execute(sql_insert, values)
        last_id = cursor.lastrowid
        user_idcode = f"ID-{last_id}"
        messagebox.showinfo("Success", f"ثبت با کد {user_idcode} انجام شد.")
        error_lable_sal_sakht_ejareh_maskoni.config(text="")
        error_lable_metraj_ejareh_maskoni.config(text="")
        error_lable_tabaghe_ejareh_maskoni.config(text="")
        error_lable_vahed_ejareh_maskoni.config(text="")
        error_lable_otagh_ejareh_maskoni.config(text="")
        error_lable_gheimat_pish_ejareh_maskoni.config(text="")
        error_lable_gheimat_ejareh_ejareh_maskoni.config(text="")
        error_lable_addrres_ejareh_maskoni.config(text="")
        error_lable_name_malek_ejareh_maskoni.config(text="")
        error_lable_shomareh_malek_ejareh_maskoni.config(text="")
        db.commit()

    except Exception as e:
        messagebox.showerror("Error", f"خطا: {e}")
    finally:
        if db and db.is_connected():
            clear_entry_ejareh_maskoni()
            db.close()
#---------------------ejareh_edari/tejari Database------------------------------
def sabt_ejareh_edari_tejari():
    db = None
    try:
        db = get_connection()
        cursor = db.cursor()
        cursor.execute("CREATE DATABASE IF NOT EXISTS state_agency")
        cursor.execute("USE state_agency")

        sql_create = """
        CREATE TABLE IF NOT EXISTS sabt_ejareh_edari_tejari (
            id INT AUTO_INCREMENT PRIMARY KEY,
            type_melk VARCHAR(50) NOT NULL,
            metraj_melk VARCHAR(20),
            sal_sakht VARCHAR(20),
            address VARCHAR(225),
            tabaghe VARCHAR(10),
            vahed VARCHAR(20),
            parking VARCHAR(20),
            asansor VARCHAR(20),
            anbari VARCHAR(20),
            aab_va_gaz VARCHAR(20),
            system_sarmayesh VARCHAR(20),
            system_garmayesh VARCHAR(20),
            gheimat_vadie DECIMAL(15,2),
            gheimat_ejareh DECIMAL(15,2),
            name_malek VARCHAR(20),
            shomareh_malek VARCHAR(11)
        )
        """
        cursor.execute(sql_create)

        sql_insert = """
        INSERT INTO sabt_ejareh_edari_tejari
        (type_melk,metraj_melk,sal_sakht,address,tabaghe,vahed,parking,asansor,
        anbari,aab_va_gaz,system_sarmayesh,system_garmayesh,
        gheimat_vadie,gheimat_ejareh,name_malek,shomareh_malek)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """
        values = (
            melk_type_ejareh_edari_tejari_entry.get(),
            metraj_melk_ejareh_edari_tejari_entry.get(),
            sal_sakht_ejareh_edari_tejari_entry.get(),
            addrres_ejareh_edari_tejari_entry.get("1.0",tk.END),
            tabaghe_ejareh_edari_tejari_entry.get(),
            vahed_ejareh_edari_tejari_entry.get(),
            parking_ejareh_edari_tejari_var.get(),
            asansor_ejareh_edari_tejari_var.get(),
            anbari_ejareh_edari_tejari_var.get(),
            ab_va_gaz_combo_emkanat_ejareh_edari_tejari.get(),
            sarmayesh_combo_emkanat_ejareh_edari_tejari.get(),
            garmayesh_combo_emkanat_ejareh_edari_tejari.get(),
            float(mablagh_pish_ejareh_edari_tejari_entry.get()),
            float(mablagh_ejare_ejareh_edari_tejari_entry.get()),
            name_malek_ejareh_edari_tejari_entry.get(),
            shomareh_malek_ejareh_edari_tejari_entry.get()


        )
        cursor.execute(sql_insert, values)
        last_id = cursor.lastrowid
        user_idcode = f"ID-{last_id}"
        messagebox.showinfo("Success", f"ثبت با کد {user_idcode} انجام شد.")
        db.commit()

    except Exception as e:
        messagebox.showerror("Error", f"خطا: {e}")
    finally:
        if db and db.is_connected():
            clear_entry_ejareh_edari_tejari()
            db.close()
#-----------------ejareh_bagh/zamin Database------------------------------------
selected_option=[]
selected_trees=[]
def sabt_ejareh_bagh_zamin():
    db = None
    try:
        db = get_connection()
        cursor = db.cursor()
        cursor.execute("CREATE DATABASE IF NOT EXISTS state_agency")
        cursor.execute("USE state_agency")
        karbari = bagh_type_combo.get()
        gheimat_str =bagh_gheimat_har_metr_ejareh_bagh_zamin_entry.get()
        if not gheimat_str or gheimat_str == "":
            gheimat_str = "0"
        gheimat_value = float(gheimat_str)
        # ========== ساخت جدول اصلی ==========
        if karbari == "باغ":  
            type_derakht_value = ",".join(selected_trees) if selected_trees else ""
            tedad_derakht_value = str(len(selected_trees)) if selected_trees else "0"
            emkanat_value = ",".join(selected_option) if selected_option else ""

            cursor.execute("""
            CREATE TABLE IF NOT EXISTS ejareh_bagh(
                id INT AUTO_INCREMENT PRIMARY KEY,
                type_melk VARCHAR(50) NOT NULL,
                metraj VARCHAR(20),
                karbari VARCHAR(20),
                address VARCHAR(255),
                mablagh_pish VARCHAR(20),
                mablagh_ejareh DECIMAL(15,2),
                zaman_ejareh VARCHAR(30),
                name_malek VARCHAR(50),
                shomareh_malek VARCHAR(11),
                metraj_derakht VARCHAR(10),
                tedad_derakht VARCHAR(10),
                type_derakht TEXT,
                system_ab VARCHAR(25),
                chah VARCHAR(10),
                estakhr VARCHAR(10),
                divar VARCHAR(10),
                bargh VARCHAR(10),
                sazeh VARCHAR(10),
                metraj_sazeh VARCHAR(10),
                sal_sakht VARCHAR(10),
                type_sazeh VARCHAR(20),
                emkanat TEXT,
                WC VARCHAR(10),
                hamam VARCHAR(10),
                javaz_sakht VARCHAR(10),
                sanad VARCHAR(20),
                mohavate VARCHAR(10),
                gaz VARCHAR(10)
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
            
            sql_bagh = """
                INSERT INTO ejareh_bagh(type_melk, metraj, karbari, address,mablagh_pish,
                    mablagh_ejareh,zaman_ejareh,name_malek,shomareh_malek,
                    metraj_derakht,tedad_derakht, type_derakht,
                    system_ab, chah, estakhr, divar,sazeh, metraj_sazeh,
                    sal_sakht, type_sazeh, emkanat, WC, hamam,
                    javaz_sakht, sanad, mohavate,bargh,gaz
                )
                VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s,%s,
                %s,%s,%s,%s,%s,%s,%s,%s,%s)
            """
            
            values_bagh = (
                melk_type_ejareh_bagh_zamin_entry.get(),
                metraj_zamin_ejareh_bagh_zamin_entry.get(),
                karbari,
                bagh_loctaion_entry.get("1.0",tk.END),
                bagh_gheimat_ejareh_bagh_zamin_entry.get(),
                gheimat_value,
                bagh_time_combo.get(),
                name_malek_bagh_zamin_entry.get(),
                number_malek_bagh_zamin_entry.get(),
                metraj_tree_entry.get(),
                tedad_derakht_value,
                type_derakht_value,
                abyari_combo.get(),
                chah_bagh_ejareh_var.get(),
                estakhr_bagh_var.get(),
                divar_ejareh_bagh_var.get(),
                var0.get(),
                metraj_vila_bagh_entry.get(),
                sal_sakht_vila_bagh_entry.get(),
                type_vila_ejareh_bagh_zamin_combo.get(),
                emkanat_value,
                toilet_bagh_combo.get(),
                hamam_bagh_combo.get(),
                mojavez_sakht_ejareh_bagh_var.get(),
                sanad_bagh_combo.get(),
                mohavate_ejareh_bagh_var.get(),
                bargh_bagh_var.get(),
                gaz_bagh.get()
            )
            cursor.execute(sql_bagh, values_bagh)
            last_id = cursor.lastrowid
            if last_id is None or last_id == 0:
               messagebox.showerror("Error", "خطا: ثبت در جدول اصلی انجام نشد")
               return
            
        elif karbari == "زمین کشاورزی": 
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS ejareh_zamin(
                id INT AUTO_INCREMENT PRIMARY KEY,
                type_melk VARCHAR(50) NOT NULL,
                metraj VARCHAR(20),
                karbari VARCHAR(20),
                address VARCHAR(255),
                mablagh_pish VARCHAR(20),
                mablagh_metri DECIMAL(15,2),
                zaman_ejareh VARCHAR(30),
                name_malek VARCHAR(50),
                shomareh_malek VARCHAR(11),
                karbari_zamin VARCHAR(50),
                type_khak VARCHAR(20),
                manba_ab VARCHAR(20),
                negahbani VARCHAR(20),
                bargh_takfaz VARCHAR(20),
                bargh_sefaz VARCHAR(20),
                anbar VARCHAR(20),
                fans VARCHAR(20),
                chah VARCHAR(20),
               created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """) 
            sql_zamin = """
                INSERT INTO ejareh_zamin(type_melk, metraj, karbari, address,mablagh_pish,
                    mablagh_metri,zaman_ejareh,name_malek,shomareh_malek,
                     karbari_zamin, type_khak,
                    manba_ab, negahbani, bargh_takfaz, bargh_sefaz,
                    anbar, fans, chah
                )
                VALUES(%s, %s,%s, %s, %s, %s, %s, %s, %s, %s,%s,%s,%s,%s,%s,%s,%s,%s)
            """
        # ========== ساخت جدول زمین (بدون hagh_bardasht) ==========

            values_zamin = (
                melk_type_ejareh_bagh_zamin_entry.get(),
                metraj_zamin_ejareh_bagh_zamin_entry.get(),
                karbari,
                bagh_loctaion_entry.get("1.0",tk.END),
                bagh_gheimat_ejareh_bagh_zamin_entry.get(),
                gheimat_value,
                bagh_time_combo.get(),
                name_malek_bagh_zamin_entry.get(),
                number_malek_bagh_zamin_entry.get(),
                karbari_ejareh_ejareh_bagh_zamin_combo.get(),
                khak_ejareh_ejareh_bagh_zamin_combo.get(),
                ab_ejareh_ejareh_bagh_zamin_combo.get(),
                security_zamin_ejareh_var.get(),
                bargh_tak_ejareh_zamin_var.get(),
                bargh_se_faz_ejareh_zamin_var.get(),
                anbar_ejareh_zamin_var.get(),
                fans_ejareh_zamin_var.get(),
                mojavaz_chah_ejareh_zamin_var.get()
            )
            cursor.execute(sql_zamin, values_zamin)
            last_id = cursor.lastrowid
        
            if last_id is None or last_id == 0:
               messagebox.showerror("Error", "خطا: ثبت در جدول اصلی انجام نشد")
               return
            
        db.commit()
        user_idcode = f"ID-{last_id}"
        messagebox.showinfo("Success", f"ثبت با کد {user_idcode} انجام شد")
 
    except Exception as e:
        messagebox.showerror("Error", f"خطا: {e}")
    finally:
        if db and db.is_connected():
            clear_entry_ejareh_bagh_zamin()
            db.close()
#-----------------ejareh_kargah Database----------------------------------------
def sabt_ejareh_kargah():
    db = None
    try:
        db = get_connection()
        cursor = db.cursor()
        cursor.execute("CREATE DATABASE IF NOT EXISTS state_agency")
        
        cursor.execute("USE state_agency")

        sql_create = """
        CREATE TABLE IF NOT EXISTS sabt_ejareh_kargah (
            id INT AUTO_INCREMENT PRIMARY KEY,
            karbari_zamin VARCHAR(50) NOT NULL,
            metraj VARCHAR(20),
            loctaion_and_address VARCHAR(225),
            gheimat_vadie DECIMAL(15,2),
            mablagh_ejareh DECIMAL(15,2),
            time_ejare VARCHAR(20),
            sal_sakht VARCHAR(20),
            vaziat_bargh VARCHAR(30),
            garmayesh VARCHAR(20),
            sarmayesh VARCHAR(30),
            vaziat_ab VARCHAR(30),
            abzar VARCHAR(30),
            toilet VARCHAR(20),
            hamam VARCHAR(20),
            otagh VARCHAR(20),
            name_malek VARCHAR(20),
            shomareh_malek VARCHAR(11)

        )
        """
        cursor.execute(sql_create)
        sql_insert = """
        INSERT INTO sabt_ejareh_kargah
        (karbari_zamin,metraj,loctaion_and_address,
        gheimat_vadie,mablagh_ejareh,time_ejare,sal_sakht,vaziat_bargh,
        garmayesh,sarmayesh,vaziat_ab,abzar,toilet,hamam,otagh,name_malek,shomareh_malek)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """
        values = (
            karbari_zamin_ejareh_karghah_entry.get(),
            metraj_ejareh_karghah_entry.get(),
            addrres_ejareh_karghah_entry.get("1.0",tk.END),
            float(vadie_ejare_karghah_entry.get()),
            float((gheimat_ejare_ejare_karghah_entry).get()),
            time_ejare_ejareh_kargah_combo.get(),
            sal_sakht_ejareh_karghah_entry.get(),
            vaziat_bargh_ejareh_karghah_combo.get(),
            garmaesh_ejareh_karghah_combo.get(),
            sarmaesh_ejareh_karghah_combo.get(),
            vaziat_ab_ejareh_karghah_combo.get(),
            abzaar_ejareh_karghah_combo.get(),
            toilet_ejareh_karghah_combo.get(),
            hamam_ejareh_karghah__combo.get(),
            otagh_ejareh_karghah_combo.get(),
            name_malek_ejareh_karghah_entry.get(),
            shomareh_malek_ejareh_karghah_entry.get()
        )
        cursor.execute(sql_insert, values)
        last_id = cursor.lastrowid
        user_idcode = f"ID-{last_id}"
        messagebox.showinfo("Success", f"ثبت با کد {user_idcode} انجام شد.")
        db.commit()

    except Exception as e:
        messagebox.showerror("Error", f"خطا: {e}")
    finally:
        if db and db.is_connected():
            clear_entry_ejareh_karghah()
            db.close()
#------------------------پایان تابع اجاره-----------------------------------
#endregion
#----------------------تابع ثبت درخواست--------------------------------
#region
#---------------darkhast_maskoni Database------------------------
def sabt_darkhast_maskoni():

    change_type = melk_type_darkhast_maskoni_entry.get()
# اعتبارسنجی 
    skip_save=False
    sal_sakht= sal_sakht_darkhast_maskoni_entry.get().strip()
    metraj= metraj_darkhast_maskoni_entry.get().strip()
    tabaghe= tabaghe_darkhast_maskoni_entry.get().strip()
    vahed= vahed_darkhast_maskoni_entry.get().strip()
    otagh= otagh_darkhast_maskoni_entry.get().strip()
    gheimat_kol= gheimat_kol_darkhast_maskoni_entry.get().strip()
    addrres = addrres_darkhast_maskoni_entry.get("1.0", "end-1c").strip()
    name_moshtari= name_moshtari_darkhast_maskoni_entry.get().strip()
    shomareh_moshtari= shomareh_moshtari_darkhast_maskoni_entry.get().strip()
    mablagh_ejare= mablagh_ejare_darkhast_maskoni_entry.get().strip()
    gheimat_pish= gheimat_pish_darkhast_maskoni_entry.get().strip()
    
    #if len(sal_sakht) != 4 or not sal_sakht.isdigit():
        #error_lable_sal_sakht_darkhast_maskoni.config(text="فیلد (سال ساخت) باید 4 رقمی باشد ")
        #return
    #elif not metraj.isdigit():
        #error_lable_metraj_darkhast_maskoni.config(text="فیلد (متراژ) فقط باید شامل اعداد باشد ")
        #return
    #elif not tabaghe.isdigit():
        #error_lable_tabaghe_darkhast_maskoni.config(text="فیلد (طبقه) فقط باید شامل اعداد باشد ")
        #return
    #elif not vahed.isdigit():
        #error_lable_vahed_darkhast_maskoni.config(text="فیلد (واحد) فقط باید شامل اعداد باشد ")
        #return
    #elif not otagh.isdigit():
        #error_lable_otagh_darkhast_maskoni.config(text="فیلد (اتاق) فقط باید شامل اعداد باشد ")
        #return
    #elif not gheimat_kol.isdigit():
        #error_lable_gheimat_kol_darkhast_maskoni.config(text="فیلد (قیمت کل) فقط باید شامل اعداد باشد ")
        #return
    #elif not re.fullmatch(r"[آ-ی0-9۰-۹\s]+", addrres):
        #error_lable_addrres_darkhast_maskoni.config(text="فیلد (آدرس) فقط باید شامل حروف فارسی و اعداد باشد ")
        #return
    #elif not name_moshtari or not re.match("^[\u0600-\u06FF\s]+$", name_moshtari):
        #error_lable_addrres_darkhast_maskoni.config(text="")
        #error_lable_name_moshtari_darkhast_maskoni.config(text="فیلد (نام مشتری) فقط باید شامل حروف فارسی باشد ")
        #return
    #elif len(shomareh_moshtari) != 11 or not shomareh_moshtari.isdigit():
        #error_lable_name_moshtari_darkhast_maskoni.config(text="")
        #error_lable_shomareh_moshtari_darkhast_maskoni.config(text="فیلد (شماره مشتری) باید 11 رقمی باشد ")
        #return
    #elif not  mablagh_ejare.isdigit():
        #error_lable_mablagh_ejare_darkhast_maskoni.config(text="فیلد (مبلغ اجاره) فقط باید شامل اعداد باشد ")
        #return

    #elif not  gheimat_pish.isdigit():
        #error_lable_gheimat_pish_darkhast_maskoni.config(text="فیلد (مبلغ پیش) فقط باید شامل اعداد باشد ")
        #return
    
    db = None
    try:
        db = get_connection()
        cursor = db.cursor()
        
        cursor.execute("CREATE DATABASE IF NOT EXISTS state_agency")
        cursor.execute("USE state_agency")

        
        
        if change_type=="درخواست خرید مسکونی":
            cursor.execute( """
            CREATE TABLE IF NOT EXISTS sabt_darkhast_kharid_maskoni (
            id INT AUTO_INCREMENT PRIMARY KEY,
            type_melk VARCHAR(50) NOT NULL,
            sal_sakht VARCHAR(20),
            address VARCHAR(225),
            tabaghe VARCHAR(10),
            vahed VARCHAR(20),
            otagh INT,
            parking VARCHAR(20),
            asansor VARCHAR(20),
            anbari VARCHAR(20),
            sarmayesh VARCHAR(20),
            garmayesh VARCHAR(20),
            kaf VARCHAR(20),
            toilet VARCHAR(20),
            gheimat_kol VARCHAR(20),
            name_moshtari VARCHAR(20),
            shomareh_moshtari VARCHAR(11),
            metraj VARCHAR(20)
            )
            """)

            sql_kharid = """
            INSERT INTO sabt_darkhast_kharid_maskoni 
            (type_melk,sal_sakht,address,tabaghe,vahed,otagh,parking,asansor,
            anbari,sarmayesh,garmayesh,kaf,toilet,gheimat_kol,name_moshtari,shomareh_moshtari,metraj)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """

            values_darkhast_kharid_maskoni = (
            melk_type_darkhast_maskoni_entry.get(),
            sal_sakht_darkhast_maskoni_entry.get(),
            addrres_darkhast_maskoni_entry.get("1.0",tk.END),
            tabaghe_darkhast_maskoni_entry.get(),
            vahed_darkhast_maskoni_entry.get(),
            otagh_darkhast_maskoni_entry.get(),
            parking_darkhast_maskoni_var.get(),
            asansor_darkhast_maskoni_var.get(),
            anbari_darkhast_maskoni_var.get(),
            sarmaesh_combo_darkhast_maskoni.get(),
            garmaesh_combo_darkhast_maskoni.get(),
            kaf_combo_darkhast_maskoni.get(),
            toilet_combo_darkhast_maskoni.get(),
            gheimat_kol_darkhast_maskoni_entry.get(),
            name_moshtari_darkhast_maskoni_entry.get(),
            shomareh_moshtari_darkhast_maskoni_entry.get(),
            metraj_darkhast_maskoni_entry.get()     
            )

            cursor.execute(sql_kharid, values_darkhast_kharid_maskoni)
        
            last_id = cursor.lastrowid
            if last_id is None or last_id == 0:
                messagebox.showerror("Error", "خطا: ثبت در جدول  انجام نشد")
                return

        elif change_type=="درخواست اجاره مسکونی":

            cursor.execute( """
            CREATE TABLE IF NOT EXISTS sabt_darkhast_ejareh_maskoni (
            id INT AUTO_INCREMENT PRIMARY KEY,
            type_melk VARCHAR(50) NOT NULL,
            sal_sakht VARCHAR(20),
            address VARCHAR(225),
            tabaghe VARCHAR(10),
            vahed VARCHAR(20),
            otagh INT,
            parking VARCHAR(20),
            asansor VARCHAR(20),
            anbari VARCHAR(20),
            sarmayesh VARCHAR(20),
            garmayesh VARCHAR(20),
            kaf VARCHAR(20),
            toilet VARCHAR(20),
            ejareh VARCHAR(20),
            pish VARCHAR(20),
            name_moshtari VARCHAR(20),
            shomareh_moshtari VARCHAR(11),
            metraj VARCHAR(20)
            )
             """)

            sql_ejareh = """
            INSERT INTO sabt_darkhast_ejareh_maskoni 
            (type_melk,sal_sakht,address,tabaghe,vahed,otagh,parking,
            asansor,anbari,sarmayesh,garmayesh,kaf,toilet,ejareh,pish,name_moshtari,shomareh_moshtari,metraj)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """

            values_darkhast_ejareh_maskoni= (
            melk_type_darkhast_maskoni_entry.get(),
            sal_sakht_darkhast_maskoni_entry.get(),
            addrres_darkhast_maskoni_entry.get("1.0",tk.END),
            tabaghe_darkhast_maskoni_entry.get(),
            vahed_darkhast_maskoni_entry.get(),
            otagh_darkhast_maskoni_entry.get(),
            parking_darkhast_maskoni_var.get(),
            asansor_darkhast_maskoni_var.get(),
            anbari_darkhast_maskoni_var.get(),
            sarmaesh_combo_darkhast_maskoni.get(),
            garmaesh_combo_darkhast_maskoni.get(),
            kaf_combo_darkhast_maskoni.get(),
            toilet_combo_darkhast_maskoni.get(),
            mablagh_ejare_darkhast_maskoni_entry.get(),
            gheimat_pish_darkhast_maskoni_entry.get(),
            name_moshtari_darkhast_maskoni_entry.get(),
            shomareh_moshtari_darkhast_maskoni_entry.get(),
            metraj_darkhast_maskoni_entry.get()
            )

            cursor.execute(sql_ejareh, values_darkhast_ejareh_maskoni)
            last_id = cursor.lastrowid
            if last_id is None or last_id == 0:
                messagebox.showerror("Error", "خطا: ثبت در جدول  انجام نشد")
                return


        db.commit()
        user_idcode = f"ID-{last_id}"
        messagebox.showinfo("Success", f"ثبت با کد {user_idcode} انجام شد.")      
    except Exception as e:
        messagebox.showerror("Error", f"خطا در ثبت داده: {e}")
        
    finally:
        if db and db.is_connected():
            clear_entry_darkhast_maskoni()
            db.close()

#---------------darkhast_edari/tejari Database--------------------
skip_save=False
def sabt_darkhast_edari_tejari(event=None):
    global skip_save
    db = None
    try:
        db = get_connection()
        cursor = db.cursor()
        
        cursor.execute("CREATE DATABASE IF NOT EXISTS state_agency")
        cursor.execute("USE state_agency")

        change_type= combo_darkhast_edari_tejari_entry.get()
        
        #فیلد های اجاره
        mablagh_vadie_darkhast_edari_tejari_lable.place_forget()
        mablagh_vadie_darkhast_edari_tejari_entry.place_forget()
        mablagh_ejareh_darkhast_edari_tejari_lable.place_forget()
        mablagh_ejareh_darkhast_edari_tejari_entry.place_forget()
        gheimat_kol_darkhast_edari_tejari_lable.place_forget()

        if change_type=="درخواست خرید اداری و تجاری":
            gheimat_kol_darkhast_edari_tejari_lable.place(x=465, y=30, anchor="e")
            gheimat_kol_darkhast_edari_tejari_entry.place(x=18, y=20, width=350, height=25)
            addrres_darkhast_edari_tejari_lable.place(x=465, y=80, anchor="e")
            addrres_darkhast_edari_tejari_entry.place(x=18, y=70, width=350, height=50)
        elif change_type=="درخواست اجاره اداری و تجاری":
           mablagh_vadie_darkhast_edari_tejari_lable.place(x=465, y=30, anchor="e")
           mablagh_vadie_darkhast_edari_tejari_entry.place(x=18, y=20, width=350, height=25)
           mablagh_ejareh_darkhast_edari_tejari_lable.place(x=465, y=75 ,anchor="e")
           mablagh_ejareh_darkhast_edari_tejari_entry.place(x=18, y=65, width=350, height=25)
           addrres_darkhast_edari_tejari_lable.place(x=465, y=120, anchor="e")
           addrres_darkhast_edari_tejari_entry.place(x=18, y=110, width=350, height=25)
        if event is not None:#خیلی مهم 
           return
        
        if change_type=="درخواست خرید اداری و تجاری":
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS sabt_darkhast_kharid_edari_tejari (
            id INT AUTO_INCREMENT PRIMARY KEY,
            type_melk VARCHAR(50) NOT NULL,
            metraj_melk VARCHAR(20),
            sal_sakht VARCHAR(20),
            address VARCHAR(225),
            tabaghe VARCHAR(10),
            vahed VARCHAR(20),
            parking VARCHAR(20),
            asansor VARCHAR(20),
            anbari VARCHAR(20),
            aab_va_gaz VARCHAR(20),
            system_sarmayesh VARCHAR(20),
            system_garmayesh VARCHAR(20),
            mablagh_kharid VARCHAR(20),
            name_moshtari VARCHAR(20),
            shomareh_moshtari VARCHAR(11)
            )
            """)

            sql_kharid = """
            INSERT INTO sabt_darkhast_kharid_edari_tejari
            (type_melk,metraj_melk,sal_sakht,address,tabaghe,vahed,parking,
            asansor,anbari,aab_va_gaz,system_sarmayesh,system_garmayesh,mablagh_kharid,name_moshtari,shomareh_moshtari)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """

            values_darkhast_kharid_edari_tejari = (
            combo_darkhast_edari_tejari_entry.get(),
            metraj_melk_darkhast_edari_tejari_entry.get(),
            sal_sakht_darkhast_edari_tejari_entry.get(),
            addrres_darkhast_edari_tejari_entry.get("1.0",tk.END),
            tabaghe_darkhast_edari_tejari_entry.get(),
            vahed_darkhast_edari_tejari_entry.get(),
            parking_darkhast_edari_tejari_var.get(),
            asansor_darkhast_edari_tejari_var.get(),
            anbari_darkhast_edari_tejari_var.get(),
            aab_va_gaz_combo_emkanat_darkhast_edari_tejari.get(),
            sarmayesh_combo_emkanat_darkhast_edari_tejari.get(),
            garmayesh_combo_emkanat_darkhast_edari_tejari.get(),
            gheimat_kol_darkhast_edari_tejari_entry.get(),
            name_moshtari_darkhast_edari_tejari_entry.get(),
            shomareh_moshtari_darkhast_edari_tejari_entry.get()
            )
            
            cursor.execute(sql_kharid,values_darkhast_kharid_edari_tejari)

            last_id = cursor.lastrowid
            if last_id is None or last_id == 0:
                messagebox.showerror("Error", "خطا: ثبت در جدول  انجام نشد")
                return
            
        elif change_type=="درخواست اجاره اداری و تجاری":
            cursor.execute( """
            CREATE TABLE IF NOT EXISTS sabt_darkhast_ejareh_edari_tejari (
            id INT AUTO_INCREMENT PRIMARY KEY,
            type_melk VARCHAR(50) NOT NULL,
            metraj_melk VARCHAR(20),
            sal_sakht VARCHAR(20),
            address VARCHAR(225),
            tabaghe VARCHAR(10),
            vahed VARCHAR(20),
            parking VARCHAR(20),
            asansor VARCHAR(20),
            anbari VARCHAR(20),
            aab_va_gaz VARCHAR(20),
            system_sarmayesh VARCHAR(20),
            system_garmayesh VARCHAR(20),
            mablagh_vadie VARCHAR(20),
            mablagh_ejareh VARCHAR(20),
            name_moshtari VARCHAR(20),
            shomareh_moshtari VARCHAR(11)
            )
            """)

            sql_ejareh = """
            INSERT INTO sabt_darkhast_ejareh_edari_tejari
            (type_melk,metraj_melk,sal_sakht,address,tabaghe,vahed,parking,asansor,
            anbari,aab_va_gaz,system_sarmayesh,system_garmayesh,
            mablagh_vadie,mablagh_ejareh,name_moshtari,shomareh_moshtari)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """
            values_darkhast_ejareh_edari_tejari = (
            combo_darkhast_edari_tejari_entry.get(),
            metraj_melk_darkhast_edari_tejari_entry.get(),
            sal_sakht_darkhast_edari_tejari_entry.get(),
            addrres_darkhast_edari_tejari_entry.get("1.0",tk.END),
            tabaghe_darkhast_edari_tejari_entry.get(),
            vahed_darkhast_edari_tejari_entry.get(),
            parking_darkhast_edari_tejari_var.get(),
            asansor_darkhast_edari_tejari_var.get(),
            anbari_darkhast_edari_tejari_var.get(),
            aab_va_gaz_combo_emkanat_darkhast_edari_tejari.get(),
            sarmayesh_combo_emkanat_darkhast_edari_tejari.get(),
            garmayesh_combo_emkanat_darkhast_edari_tejari.get(),
            mablagh_vadie_darkhast_edari_tejari_entry.get(),
            mablagh_ejareh_darkhast_edari_tejari_entry.get(),
            name_moshtari_darkhast_edari_tejari_entry.get(),
            shomareh_moshtari_darkhast_edari_tejari_entry.get()
            )
            
            cursor.execute(sql_ejareh,values_darkhast_ejareh_edari_tejari)

            last_id = cursor.lastrowid
            if last_id is None or last_id == 0:
                messagebox.showerror("Error", "خطا: ثبت در جدول  انجام نشد")
                return

        db.commit()
        user_idcode = f"ID-{last_id}"
        messagebox.showinfo("Success", f"ثبت با کد {user_idcode} انجام شد.")      
    except Exception as e:
        messagebox.showerror("Error", f"خطا در ثبت داده: {e}")
        
    finally:
        if db and db.is_connected():
            clear_entry_darkhast_edari_tejari()
            db.close()
#--------------darkhast_bagh/zamin Database-----------------------
selected_option3=[]
selected_trees3=[]
skip_save=False
def sabt_darkhast_bagh_zamin(event=None):
    global skip_save
    db = None
    
    try:
        db = get_connection()
        cursor = db.cursor()
        cursor.execute("CREATE DATABASE IF NOT EXISTS state_agency")
        cursor.execute("USE state_agency")
        change_type=melk_type_darkhast_bagh_zamin_entry.get()
        karbari = bagh_type_darkhast_bagh_zamin_combo.get()
        gheimat_str =gheimat_har_metr_bagh_zamin_darkhast_entry.get()
        if not gheimat_str or gheimat_str == "":
            gheimat_str = "0"
        gheimat_value = float(gheimat_str)
        gheimat_har_matr_bagh_zamin_darkhast_lable.place_forget()
        gheimat_har_metr_bagh_zamin_darkhast_entry.place_forget()
        gheimat_kol_bagh_zamin_darkhast_lable.place_forget()
        gheimat_kol_bagh_zamin_darkhast_entry.place_forget()
        time_ejareh_bagh_darkhast_zamin_lable.place_forget()
        bagh_time_darkhast_combo.place_forget()
        mablagh_ejareh_mahaneh_darkhast_lable.place_forget()
        mablagh_ejareh_mahaneh_darkhast_entry.place_forget()
        gheimat_ejareh_bagh_darkhast_zamin_lable.place_forget()#ودیعه
        gheimat_ejareh_bagh_darkhast_zamin_entry.place_forget()#ودیعه
        gheimat_har_matr_bagh_zamin_darkhast_lable.place_forget()
        gheimat_har_metr_bagh_zamin_darkhast_entry.place_forget()
        gheimat_kol_bagh_zamin_darkhast_lable.place_forget()#قیمت کل
        gheimat_kol_bagh_zamin_darkhast_entry.place_forget()#قیمت کل
        time_ejareh_bagh_darkhast_zamin_lable.place_forget()
        bagh_time_darkhast_combo.place_forget()   
        if change_type=="درخواست خرید باغ زمین":
            gheimat_har_matr_bagh_zamin_darkhast_lable.place(x=490, y=65, anchor="e")
            gheimat_har_metr_bagh_zamin_darkhast_entry.place(x=28, y=55, width=350, height=25)
            gheimat_kol_bagh_zamin_darkhast_lable.place(x=490, y=20, anchor="e")
            gheimat_kol_bagh_zamin_darkhast_entry.place(x=28, y=13, width=350, height=25)
        elif change_type=="درخواست اجاره باغ زمین":
           mablagh_ejareh_mahaneh_darkhast_lable.place(x=490, y=65,anchor="e")
           mablagh_ejareh_mahaneh_darkhast_entry.place(x=28, y=55, width=350, height=25)
           gheimat_ejareh_bagh_darkhast_zamin_lable.place(x=490, y=20, anchor="e")#ودیعه
           gheimat_ejareh_bagh_darkhast_zamin_entry.place(x=28, y=13, width=350, height=25)#ودیعه
           time_ejareh_bagh_darkhast_zamin_lable.place(x=490, y=160, anchor="e")
           bagh_time_darkhast_combo.place(x=28, y=150, width=350, height=25)
        if event is not None:#خیلی مهم 
           return
        

        if change_type=="درخواست خرید باغ زمین":
            if karbari=="باغ":
                cursor.execute("""
                CREATE TABLE IF NOT EXISTS darkhast_kharid_bagh(
                id INT AUTO_INCREMENT PRIMARY KEY,
                type_melk VARCHAR(50) NOT NULL,
                metraj VARCHAR(20),
                karbari VARCHAR(20),
                address VARCHAR(255),
                mablagh_metri DECIMAL(15,2),
                gheimat_kol VARCHAR(30),
                name_malek VARCHAR(50),
                shomareh_malek VARCHAR(11),
                metraj_derakht VARCHAR(10),
                tedad_derakht VARCHAR(10),
                type_derakht TEXT,
                system_ab VARCHAR(25),
                chah VARCHAR(10),
                estakhr VARCHAR(10),
                divar VARCHAR(10),
                sazeh VARCHAR(10),
                metraj_sazeh VARCHAR(10),
                sal_sakht VARCHAR(10),
                type_sazeh VARCHAR(20),
                emkanat TEXT,
                WC VARCHAR(10),
                hamam VARCHAR(10),
                javaz_sakht VARCHAR(10),
                sanad VARCHAR(20),
                mohavate VARCHAR(10),
                bargh VARCHAR (10),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                 )
               """) 
                type_derakht_value = ",".join(selected_trees3) if selected_trees3 else ""
                tedad_derakht_value = str(len(selected_trees3)) if selected_trees3 else "0"
                emkanat_value = ",".join(selected_option3) if selected_option3 else ""
            
                sql_bagh = """
                INSERT INTO darkhast_kharid_bagh(
                    type_melk,metraj,karbari,address,mablagh_metri,gheimat_kol,name_malek,shomareh_malek,
                    metraj_derakht, tedad_derakht, type_derakht,
                    system_ab, chah, estakhr, divar,sazeh, metraj_sazeh,
                    sal_sakht, type_sazeh, emkanat, WC, hamam,
                    javaz_sakht, sanad, mohavate,bargh,gaz
                )
                VALUES( %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s,%s
                ,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                 """
                values_kharid=(
                melk_type_darkhast_bagh_zamin_entry.get(),
                metraj_zamin_darkhast_bagh_zamin_entry.get(),
                karbari,
                bagh_loctaion_darkhast_bagh_zamin_entry.get("1.0",tk.END),
                gheimat_value,
                name_moshtari_darkhast_bagh_entry.get(),
                shomareh_moshtari_darkhast_bagh_entry.get(),
                gheimat_kol_bagh_zamin_darkhast_entry.get(),
                metraj_derakht_darkhast_bagh_zamin_entry.get(),
                tedad_derakht_darkhast_bagh_zamin_entry.get(),
                type_tree_darkhast_bagh_zamin_combo.get(),
                abyari_darkhast_bagh_zamin_combo.get(),
                chah_darkhast_bagh_zamin_var.get(),
                estakhr_darkhast_bagh_zamin_var.get(),
                divar_darkhast_bagh_zamin_var.get(),
                var0_darkhast_bagh_zamin.get(),
                metraj_vila_darkhast_bagh_zamin_entry.get(),
                sal_sakht_vila_darkhast_bagh_zamin_entry.get(),
                type_vila_darkhast_bagh_zamin_combo.get(),
                toilet_darkhast_bagh_zamin_combo.get(),
                hamam_darkhast_bagh_zamin_combo.get(),
                sanad_darkhast_bagh_zamin_combo.get(),
                option_darkhast_bagh_zamin_combo.get(),
                mojavez_sakht_darkhast_bagh_zamin_var.get(),
                mohavate_sazi_darkhast_bagh_zamin_var.get(),
                bargh_keshi_darkhast_bagh_zamin_var.get(),
                gaz_keshi_darkhast_bagh_zamin_var.get()
                )
                cursor.execute(sql_bagh, values_kharid)

                last_id = cursor.lastrowid
                if last_id is None or last_id == 0:
                       messagebox.showerror("Error", "خطا: ثبت در جدول  انجام نشد")
                       return
            elif karbari=="زمین کشاورزی":
                cursor.execute("""
                CREATE TABLE IF NOT EXISTS darkhast_kharid_zamin(
                id INT AUTO_INCREMENT PRIMARY KEY,
                type_melk VARCHAR(50) NOT NULL,
                metraj VARCHAR(20),
                karbari VARCHAR(20),
                address VARCHAR(255),
                mablagh_metri DECIMAL(15,2),
                name_moshtari VARCHAR(50),
                shomareh_moshtari VARCHAR(30),
                karbari_zamin VARCHAR(50),
                type_khak VARCHAR(20),
                manba_ab VARCHAR(20),
                negahbani VARCHAR(20),
                bargh_takfaz VARCHAR(20),
                bargh_sefaz VARCHAR(20),
                anbar VARCHAR(20),
                fans VARCHAR(20),
                chah VARCHAR(20),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP  
                )
                """)

                sql_zamin = """
                INSERT INTO darkhast_kharid_zamin(
                    type_melk,metraj,karbari,address,mablagh_metri,name_moshtari,shomareh_moshtari,
                    karbari_zamin,type_khak,
                    manba_ab, negahbani, bargh_takfaz, bargh_sefaz,
                    anbar, fans, chah
                 )
                 VALUES( %s, %s, %s, %s, %s, %s, %s,%s,%s, %s,%s,%s,%s,%s,%s,%s)
                 """
                values_kharid_zamin=(
                   change_type,
                   metraj_zamin_darkhast_bagh_zamin_entry.get(),
                   karbari,
                   bagh_loctaion_darkhast_bagh_zamin_entry.get("1.0",tk.END),
                   gheimat_value,
                   name_moshtari_darkhast_bagh_entry.get(),
                   shomareh_moshtari_darkhast_bagh_entry.get(),

                   karbari_darkhast_bagh_zamin_combo.get(),
                   khak_darkhast_bagh_zamin_combo.get(),
                   ab_darkhast_bagh_zamin_combo.get(),
                   security_zamin_darkhast_bagh_zamin_var.get(),
                   bargh_kesi_zamin_darkhast_bagh_zamin_var.get(),
                   bargh_zamin_darkhast_bagh_zamin2_var.get(),
                   anbar_zamin_darkhast_bagh_zamin_var.get(),
                   fans_zamin_darkhast_bagh_zamin_var.get(),
                   mojavez_chah_zamin_darkhast_bagh_zamin_var.get())
                   
                cursor.execute(sql_zamin,values_kharid_zamin)
                last_id = cursor.lastrowid
                if last_id is None or last_id == 0:
                        messagebox.showerror("Error", "خطا: ثبت در جدول  انجام نشد")
                        return
            
        
        elif change_type=="درخواست اجاره باغ زمین":
            if karbari=="باغ":
                cursor.execute("""
                CREATE TABLE IF NOT EXISTS darkhast_ejareh_bagh(
                id INT AUTO_INCREMENT PRIMARY KEY,
                type_melk VARCHAR(50) NOT NULL,
                metraj VARCHAR(20),
                karbari VARCHAR(20),
                address VARCHAR(255),
                name_moshtari VARCHAR(50),
                shomareh_moshtari VARCHAR(11),
                zaman_ejareh VARCHAR(40),
                mablagh_pish VARCHAR(30),
                mablagh_ejareh VARCHAR(30),
                metraj_derakht VARCHAR(10),
                tedad_derakht VARCHAR(10),
                type_derakht TEXT,
                system_ab VARCHAR(25),
                chah VARCHAR(10),
                estakhr VARCHAR(10),
                divar VARCHAR(10),
                sazeh VARCHAR(10),
                metraj_sazeh VARCHAR(10),
                sal_sakht VARCHAR(10),
                type_sazeh VARCHAR(20),
                emkanat TEXT,
                WC VARCHAR(10),
                hamam VARCHAR(10),
                javaz_sakht VARCHAR(10),
                sanad VARCHAR(20),
                mohavate VARCHAR(10),
                bargh VARCHAR (10),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                 )
               """) 
                type_derakht_value = ",".join(selected_trees3) if selected_trees3 else ""
                tedad_derakht_value = str(len(selected_trees3)) if selected_trees3 else "0"
                emkanat_value = ",".join(selected_option3) if selected_option3 else ""
            
                sql_bagh = """
                INSERT INTO darkhast_ejareh_bagh(
                    type_melk,metraj,karbari,address,name_moshtari,shomareh_moshtari,
                    zaman_ejareh,mablagh_pish,mablagh_ejareh,
                    metraj_derakht, tedad_derakht, type_derakht,
                    system_ab, chah, estakhr, divar,sazeh, metraj_sazeh,
                    sal_sakht, type_sazeh, emkanat, WC, hamam,
                    javaz_sakht, sanad, mohavate,bargh,gaz
                )
                VALUES( %s, %s, %s, %s, %s, %s, %s, %s, %s,%s,%s, %s, %s, %s, %s, %s,%s
                ,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                 """
                values_ejareh=(
                melk_type_darkhast_bagh_zamin_entry.get(),
                metraj_zamin_darkhast_bagh_zamin_entry.get(),
                karbari,
                bagh_loctaion_darkhast_bagh_zamin_entry.get("1.0",tk.END),
                name_moshtari_darkhast_bagh_entry.get(),
                shomareh_moshtari_darkhast_bagh_entry.get(),
                bagh_time_darkhast_combo.get(),
                gheimat_ejareh_bagh_darkhast_zamin_entry.get(),
                mablagh_ejareh_mahaneh_darkhast_entry.get(),
                metraj_derakht_darkhast_bagh_zamin_entry.get(),
                tedad_derakht_darkhast_bagh_zamin_entry.get(),
                type_tree_darkhast_bagh_zamin_combo.get(),
                abyari_darkhast_bagh_zamin_combo.get(),
                chah_darkhast_bagh_zamin_var.get(),
                estakhr_darkhast_bagh_zamin_var.get(),
                divar_darkhast_bagh_zamin_var.get(),
                var0_darkhast_bagh_zamin.get(),
                metraj_vila_darkhast_bagh_zamin_entry.get(),
                sal_sakht_vila_darkhast_bagh_zamin_entry.get(),
                type_vila_darkhast_bagh_zamin_combo.get(),
                toilet_darkhast_bagh_zamin_combo.get(),
                hamam_darkhast_bagh_zamin_combo.get(),
                sanad_darkhast_bagh_zamin_combo.get(),
                option_darkhast_bagh_zamin_combo.get(),
                mojavez_sakht_darkhast_bagh_zamin_var.get(),
                mohavate_sazi_darkhast_bagh_zamin_var.get(),
                bargh_keshi_darkhast_bagh_zamin_var.get(),
                gaz_keshi_darkhast_bagh_zamin_var.get()
                )
                cursor.execute(sql_bagh, values_ejareh)

                last_id = cursor.lastrowid
                if last_id is None or last_id == 0:
                       messagebox.showerror("Error", "خطا: ثبت در جدول")
                       return
            elif karbari=="زمین کشاورزی":
                cursor.execute("""
                CREATE TABLE IF NOT EXISTS darkhast_ejareh_zamin(
                id INT AUTO_INCREMENT PRIMARY KEY,
                type_melk VARCHAR(50) NOT NULL,
                metraj VARCHAR(20),
                karbari VARCHAR(20),
                address VARCHAR(255),
                mablagh_pish VARCHAR(30),
                mablagh_ejareh VARCHAR(30),
                zaman_ejareh VARCHAR(40),
                name_moshtari VARCHAR(50),
                shomareh_moshtari VARCHAR(11),
                karbari_zamin VARCHAR(50),
                type_khak VARCHAR(20),
                manba_ab VARCHAR(20),
                negahbani VARCHAR(20),
                bargh_takfaz VARCHAR(20),
                bargh_sefaz VARCHAR(20),
                anbar VARCHAR(20),
                fans VARCHAR(20),
                chah VARCHAR(20),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP  
                )
              """)

                sql_zamin = """
                INSERT INTO darkhast_ejareh_zamin(
                    type_melk,metraj,karbari,address,mablagh_pish,mablagh_ejareh,zaman_ejareh,name_moshtari,shomareh_moshtari,
                    karbari_zamin,type_khak,
                    manba_ab,negahbani,bargh_takfaz,bargh_sefaz,
                    anbar,fans,chah
                 )
                VALUES( %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """
                values_ejareh_zamin=(
                   change_type,
                   metraj_zamin_darkhast_bagh_zamin_entry.get(),
                   karbari,
                   bagh_loctaion_darkhast_bagh_zamin_entry.get("1.0",tk.END),
                   gheimat_ejareh_bagh_darkhast_zamin_entry.get(),
                   mablagh_ejareh_mahaneh_darkhast_entry.get(),
                   bagh_time_darkhast_combo.get(),
                   name_moshtari_darkhast_bagh_entry.get(),
                   shomareh_moshtari_darkhast_bagh_entry.get(),
                   karbari_darkhast_bagh_zamin_combo.get(),
                   khak_darkhast_bagh_zamin_combo.get(),
                   ab_darkhast_bagh_zamin_combo.get(),
                   security_zamin_darkhast_bagh_zamin_var.get(),
                   bargh_kesi_zamin_darkhast_bagh_zamin_var.get(),
                   bargh_zamin_darkhast_bagh_zamin2_var.get(),
                   anbar_zamin_darkhast_bagh_zamin_var.get(),
                   fans_zamin_darkhast_bagh_zamin_var.get(),
                   mojavez_chah_zamin_darkhast_bagh_zamin_var.get())
            
                cursor.execute(sql_zamin,values_ejareh_zamin)
                last_id = cursor.lastrowid
                if last_id is None or last_id == 0:
                       messagebox.showerror("Error", "خطا: ثبت در جدول  انجام نشد")
                       return    
                        
        db.commit()
        user_idcode = f"ID-{last_id}"
        messagebox.showinfo("Success", f"ثبت با کد {user_idcode} انجام شد.")      
    except Exception as e:
        messagebox.showerror("Error", f"خطا در ثبت داده: {e}")
        
    finally:
        if db and db.is_connected():
            clear_entry_darkhast_bagh_zamin()
            db.close()
                
#------------darkhast_kargah Database--------------------------
skip_save=False
def sabt_darkhast_kargah(event=None):
    global skip_save
    db = None
    try:
        db = get_connection()
        cursor = db.cursor()
        
        cursor.execute("CREATE DATABASE IF NOT EXISTS state_agency")
        cursor.execute("USE state_agency")

        change_type= combo_darkhast_kargah.get()
        #فیلد های خرید
        mablagh_pish_darkhast_kargah_lable.place_forget()
        mablagh_pish_darkhast_kargah_entry.place_forget()
        #فیلد های اجاره
        mablagh_pish_darkhast_kargah_lable.place_forget()
        mablagh_pish_darkhast_kargah_entry.place_forget()
        ejareh_mahaneh_darkhast_kargah_lable.place_forget()
        ejareh_mahaneh_darkhast_kargah_entry.place_forget()


        if change_type=="درخواست خرید کارگاه":
            gheimat_kol_darkhast_kargah_lable.place(x=465, y=30, anchor="e")
            gheimat_kol_darkhast_kargah_entry.place(x=18, y=20, width=350, height=25)
            loctaion_darkhast_kargah.place(x=465, y=80, anchor="e")
            loctaion_darkhast_kargah_entry.place(x=18, y=70, width=350, height=50)

        elif change_type=="درخواست اجاره کارگاه":
            mablagh_pish_darkhast_kargah_lable.place(x=465, y=30, anchor="e")
            mablagh_pish_darkhast_kargah_entry.place(x=18, y=20, width=350, height=25)
            ejareh_mahaneh_darkhast_kargah_lable.place(x=465, y=75, anchor="e")
            ejareh_mahaneh_darkhast_kargah_entry.place(x=18, y=65, width=350, height=25)
            loctaion_darkhast_kargah.place(x=465, y=120, anchor="e")
            loctaion_darkhast_kargah_entry.place(x=18, y=110, width=350, height=25)

        if event is not None:#خیلی مهم 
           return
        
        if change_type=="درخواست خرید کارگاه":
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS sabt_darkhast_kharid_kargah(
            id INT AUTO_INCREMENT PRIMARY KEY,
            type_melk VARCHAR(50) NOT NULL,
            metraj_melk VARCHAR(20),
            address VARCHAR(225),
            name_moshtari VARCHAR(30),
            shomareh_moshtari INT, 
            gheimat_kol VARCHAR(20),              
            sal_sakht VARCHAR(20),
            vaziat_bargh VARCHAR(20),
            garmayesh VARCHAR(20),
            sarmayesh VARCHAR(30),
            vaziat_ab VARCHAR(100),
            abzar VARCHAR(100),
            toilet VARCHAR(20),
            hamam VARCHAR(20),
            otagh VARCHAR(20)
            )
            """)

            sql_kharid = """
            INSERT INTO sabt_darkhast_kharid_kargah
            (type_melk,metraj_melk,address,name_moshtari,shomareh_moshtari,gheimat_kol,sal_sakht,
            vaziat_bargh,garmayesh,sarmayesh,vaziat_ab,abzar,toilet,hamam,otagh)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """

            values_darkhast_kharid_kargah = (
            combo_darkhast_kargah.get(),
            metraj_darkhast_kargah_entry.get(),
            loctaion_darkhast_kargah_entry.get("1.0",tk.END),
            name_moshtari_darkhast_kargah_entry.get(),
            shomareh_moshtari_darkhast_kargah_entry.get(),
            gheimat_kol_darkhast_kargah_entry.get(),
            sal_sakht_darkhast_kargah_entry.get(),
            vaziat_bargh_darkhast_kargah_combo.get(),
            garmayesh_type_darkhast_kargah_combo.get(),
            sarmayesh_darkhast_kargah_combo.get(),
            vaziat_ab_darkhast_kargah_combo.get(),
            abzar_darkhast_kargah_combo.get(),
            toilet_darkhast_kargah_combo.get(),
            hamam_darkhast_kargah_combo.get(),
            otagh_darkhast_kargah_combo.get()
            )
            
            cursor.execute(sql_kharid,values_darkhast_kharid_kargah)

            last_id = cursor.lastrowid
            if last_id is None or last_id == 0:
                messagebox.showerror("Error", "خطا: ثبت در جدول  انجام نشد")
                return
            
        elif change_type=="درخواست اجاره کارگاه":
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS sabt_darkhast_ejareh_kargah(
            id INT AUTO_INCREMENT PRIMARY KEY,
            type_melk VARCHAR(50) NOT NULL,
            metraj_melk VARCHAR(20),
            address VARCHAR(225),
            mablagh_pish VARCHAR(20),
            name_moshtari VARCHAR(30),
            shomareh_moshtari INT,
            ejareh_mahaneh VARCHAR(20),
            sal_sakht VARCHAR(20),
            vaziat_bargh VARCHAR(20),
            garmayesh VARCHAR(20),
            sarmayesh VARCHAR(30),
            vaziat_ab VARCHAR(100),
            abzar VARCHAR(100),
            toilet VARCHAR(20),
            hamam VARCHAR(20),
            otagh VARCHAR(20)
            )
            """)

            sql_ejareh = """
            INSERT INTO sabt_darkhast_ejareh_kargah
            (type_melk,metraj_melk,address,mablagh_pish,name_moshtari,shomareh_moshtari,
            ejareh_mahaneh,sal_sakht,vaziat_bargh,garmayesh,sarmayesh,vaziat_ab,
            abzar,toilet,hamam,otagh)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """
            values_darkhast_ejareh_kargah = (
            combo_darkhast_kargah.get(),
            metraj_darkhast_kargah_entry.get(),
            loctaion_darkhast_kargah_entry.get("1.0",tk.END),
            mablagh_pish_darkhast_kargah_entry.get(),
            name_moshtari_darkhast_kargah_entry.get(),
            shomareh_moshtari_darkhast_kargah_entry.get(),
            ejareh_mahaneh_darkhast_kargah_entry.get(),
            sal_sakht_darkhast_kargah_entry.get(),
            vaziat_bargh_darkhast_kargah_combo.get(),
            garmayesh_type_darkhast_kargah_combo.get(),
            vaziat_ab_darkhast_kargah_combo.get(),
            abzar_darkhast_kargah_combo.get(),
            toilet_darkhast_kargah_combo.get(),
            hamam_darkhast_kargah_combo.get(),
            otagh_darkhast_kargah_combo.get()
            )
            
            cursor.execute(sql_ejareh,values_darkhast_ejareh_kargah)

            last_id = cursor.lastrowid
            if last_id is None or last_id == 0:
                messagebox.showerror("Error", "خطا: ثبت در جدول  انجام نشد")
                return

        db.commit()
        user_idcode = f"ID-{last_id}"
        messagebox.showinfo("Success", f"ثبت با کد {user_idcode} انجام شد.")      
    except Exception as e:
        messagebox.showerror("Error", f"خطا در ثبت داده: {e}")
        
    finally:
        if db and db.is_connected():
            clear_entry_darkhast_kargah()
            db.close()
#--------------------پایان تابع ثبت درخواست----------------
#endregion
#----------------------------سیو تصاویر--------------------------
#===================سیو تصاویر در پنجره فروش اداری و تجاری================
#region
def open_file_forosh_edari_tejari():

    global selected_images_forosh_edari_tejari
    global photo_refs_forosh_edari_tejari

    file_paths = filedialog.askopenfilenames(
        title="انتخاب تصاویر ملک",
        filetypes=[
            ("Image Files", "*.png *.jpg *.jpeg *.bmp *.webp")
        ]
    )

    # حداکثر 4 تصویر
    file_paths = list(file_paths[:4])

    # ساخت پوشه روی دسکتاپ
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    save_folder = os.path.join(desktop_path, "Forosh_Edari_Tejari")

    os.makedirs(save_folder, exist_ok=True)

    new_paths = []

    # کپی تصاویر به پوشه پروژه
    for path in file_paths:

        filename = os.path.basename(path)

        destination = os.path.join(save_folder, filename)

        shutil.copy2(path, destination)

        new_paths.append(destination)

    # ذخیره مسیرهای جدید
    selected_images_forosh_edari_tejari = new_paths

    # پاک کردن تصاویر قبلی
    for widget in image_frame_forosh_edari_tejari.winfo_children():
        widget.destroy()

    photo_refs_forosh_edari_tejari.clear()

    # نمایش تصاویر
    for i, path in enumerate(selected_images_forosh_edari_tejari):

        try:

            img = Image.open(path)

            img.thumbnail((150, 110))

            photo = ImageTk.PhotoImage(img)

            photo_refs_forosh_edari_tejari.append(photo)

            lbl = tk.Label(image_frame_forosh_edari_tejari,image=photo,bg="white",bd=1,relief="solid")

            lbl.grid(row=i // 2,column=i % 2,padx=5,pady=5)

        except Exception as e:
            print("خطا در بارگذاری تصویر:", e)
#endregion
#------------------------توابع سرچ--------------------
#regoin
def search():
    for item in tree.get_children():
        tree.delete(item)
    file = combo_file_type.get()
    melk = melk_type_combo.get()

    db = get_connection()
    cursor = db.cursor()
    cursor.execute("CREATE DATABASE IF NOT EXISTS state_agency") 
    cursor.execute("USE state_agency")

    try:
        if file == "فروش":

            if melk == "مسکونی":
                cursor.execute("""
                    SELECT id,"sabt_forosh_maskoni",address, gheimat_kol, type_melk, name_malek ,'فروش' AS noe_file
                    FROM sabt_forosh_maskoni
                    WHERE address LIKE %s
                """, (f"%{address_entry.get().strip()}%",))

            elif melk == "اداری_تجاری":
                cursor.execute("""
                    SELECT id,"sabt_forosh_edari_tejari",address, gheimat_kol, type_melk, name_malek,'فروش' AS noe_file
                    FROM sabt_forosh_edari_tejari
                    WHERE address LIKE %s
                """, (f"%{address_entry.get().strip()}%",))

            elif melk == "زمین":
                cursor.execute("""
                    SELECT id,"forosh_zamin",address, gheimat_kol, type_melk, name_malek,'فروش' AS noe_file
                    FROM forosh_zamin
                    WHERE address LIKE %s
                """, (f"%{address_entry.get().strip()}%",))

            elif melk == "باغ":
                cursor.execute("""
                    SELECT id,"forosh_bagh",address, gheimat_kol, type_melk, name_malek,'فروش' AS noe_file
                    FROM forosh_bagh
                    WHERE address LIKE %s
                """, (f"%{address_entry.get().strip()}%",))

            elif melk == "کارگاه":
                cursor.execute("""
                    SELECT id,"sabt_forosh_kargah",address, gheimat_kol,karbari_zamin, name_malek,'فروش' AS noe_file
                    FROM sabt_forosh_kargah
                    WHERE address LIKE %s
                """, (f"%{address_entry.get().strip()}%",))
        elif file=="رهن_اجاره":
            if melk=="مسکونی":
                    cursor.execute("""
                    SELECT id,"sabt_ejareh_maskoni",address, pish, type_melk, name_malek
                    ,'اجاره' AS noe_file
                    FROM sabt_ejareh_maskoni
                    WHERE address LIKE %s
                """, (f"%{address_entry.get().strip()}%",))
            elif melk == "اداری_تجاری":
                    cursor.execute("""
                    SELECT id,"sabt_ejareh_edari_tejari",address, gheimat_ejareh, type_melk, name_malek ,'اجاره' AS noe_file
                    FROM sabt_ejareh_edari_tejari
                    WHERE address LIKE %s
                """, (f"%{address_entry.get().strip()}%",))
            elif melk == "زمین":
                    cursor.execute("""
                    SELECT id,"ejareh_zamin",address, gheimat_kol, type_melk, name_malek ,'اجاره' AS noe_file
                    FROM ejareh_zamin
                    WHERE address LIKE %s
                """, (f"%{address_entry.get().strip()}%",))
            elif melk == "باغ":
                    cursor.execute("""
                    SELECT id,"ejareh_bagh",address, gheimat_kol, type_melk, name_malek ,'اجاره' AS noe_file
                    FROM ejareh_bagh
                    WHERE address LIKE %s
                """, (f"%{address_entry.get().strip()}%",))
            elif melk == "کارگاه":
                    cursor.execute("""
                    SELECT id,"sabt_forosh_kargah",address, gheimat_kol, type_melk, name_malek ,'اجاره' AS noe_file
                    FROM sabt_forosh_kargah
                    WHERE address LIKE %s
                """, (f"%{address_entry.get().strip()}%",))
        elif file=="درخواست ها":
            if melk=="مسکونی":
                queries = []
                params = []
                cursor.execute("SHOW TABLES LIKE 'sabt_darkhast_kharid_maskoni'")
                if cursor.fetchone():
                    queries.append("""
                    SELECT id,'sabt_darkhast_kharid_maskoni',address,gheimat_kol,type_melk,name_moshtari,'درخواست خرید' AS noe_file
                    FROM sabt_darkhast_kharid_maskoni
                    WHERE address LIKE %s
                    """)
                    params.append(f"%{address_entry.get().strip()}%")

                cursor.execute("SHOW TABLES LIKE 'sabt_darkhast_ejareh_maskoni'")
                if cursor.fetchone():
                    queries.append("""
                    SELECT id,'sabt_darkhast_ejareh_maskoni',address,pish,type_melk,name_moshtari,'درخواست اجاره' AS noe_file
                    FROM sabt_darkhast_ejareh_maskoni
                    WHERE address LIKE %s
                    """)
                    params.append(f"%{address_entry.get().strip()}%")

                if queries:
                    sql = "\nUNION ALL\n".join(queries)
                    cursor.execute(sql, tuple(params))
            elif melk == "اداری_تجاری":

                queries = []
                params = []
                cursor.execute("SHOW TABLES LIKE 'sabt_darkhast_kharid_edari_tejari'")
                if cursor.fetchone():
                    queries.append("""SELECT
                        id,'sabt_darkhast_kharid_edari_tejari',address,mablagh_kharid,type_melk,name_moshtari,
                        'درخواست خرید' AS noe_file
                         FROM sabt_darkhast_kharid_edari_tejari
                        WHERE address LIKE %s
                        """)
                    params.append(f"%{address_entry.get().strip()}%")

                cursor.execute("SHOW TABLES LIKE 'sabt_darkhast_ejareh_edari_tejari'")
                if cursor.fetchone():
                    queries.append("""SELECT
                        id,
                        'sabt_darkhast_ejareh_edari_tejari',
                        address,
                        mablagh_vadie,
                        type_melk,
                        name_moshtari,
                        'درخواست اجاره' AS noe_file
                    FROM sabt_darkhast_ejareh_edari_tejari
                    WHERE address LIKE %s
                    """)
                    params.append(f"%{address_entry.get().strip()}%")

                if queries:
                    sql = "\nUNION ALL\n".join(queries)
                    cursor.execute(sql, tuple(params))
            elif melk=="باغ":
                cursor.execute("""
                SELECT id,"darkhast_kharid_bagh",address, gheimat_kol, type_melk, name_malk, 'درخواست خرید' AS noe_file
                FROM darkhast_kharid_bagh
                WHERE address LIKE %s
                UNION ALL
                SELECT id,"darkhast_ejareh_bagh",address, gheimat_kol, type_melk, name_malk, 'درخواست اجاره' AS noe_file
                FROM darkhast_ejareh_bagh
                WHERE address LIKE %s
                """,
                (
                    f"%{address_entry.get().strip()}%",
                    f"%{address_entry.get().strip()}%"
                ))
            elif melk=="زمین":
                cursor.execute("""
                SELECT id,"darkhast_kharid_zamin",address, gheimat_kol, type_melk, name_malk, 'درخواست خرید' AS noe_file
                FROM darkhast_kharid_zamin
                WHERE address LIKE %s
                UNION ALL
                SELECT id,"darkhast_ejareh_zamin",address, gheimat_kol, type_melk, name_malk, 'درخواست اجاره' AS noe_file
                FROM darkhast_ejareh_zamin
                WHERE address LIKE %s
                """,
                (
                    f"%{address_entry.get().strip()}%",
                    f"%{address_entry.get().strip()}%"
                ))
            elif melk=="کارگاه":
                cursor.execute("""
                SELECT id,"darkhast_kharid_kargah",address, gheimat_kol, type_melk, name_malk, 'درخواست خرید' AS noe_file
                FROM darkhast_kharid_kargah
                WHERE address LIKE %s
                UNION ALL
                SELECT id,"darkhast_ejareh_kargah",address, gheimat_kol, type_melk, name_malk,'درخواست اجاره' AS noe_file
                FROM darkhast_ejareh_kargah
                WHERE address LIKE %s
                """,
                (
                    f"%{address_entry.get().strip()}%",
                    f"%{address_entry.get().strip()}%"
                ))

        else:
            messagebox.showerror("خطا", "نوع ملک نامعتبر است")
            return

        results = cursor.fetchall()

        # پاک کردن نتایج قبلی
        for item in tree.get_children():
            tree.delete(item)

        if results:
        # پاک کردن نتایج قبلی
            for item in tree.get_children():
                tree.delete(item)
            for row in results:
                tree.insert("",tk.END,values=row)

        else:
            messagebox.showinfo("یافت نشد", "هیچ موردی پیدا نشد")
    except Exception as e:
        messagebox.showerror("Error", f"خطا: {e}")

    finally:
        cursor.close()
        db.close()
#endregion
#-----------------------توابع ویرایش---------------------------
#region
#---------------تابع وارد کردن جزییات به باکس سوم----------
selected_id = None
selected_table = None

def show_details(event):
    global selected_id, selected_table

    item = tree.focus()

    if item == "":
        return

    values = tree.item(item)["values"]

    selected_id = int(values[0])
    selected_table = values[1]

    db = get_connection()
    cursor = db.cursor()
    cursor.execute("USE state_agency")

    cursor.execute(f"SELECT * FROM {selected_table} WHERE id=%s", (selected_id,))

    data = cursor.fetchone()
#--------------------مسکونی --------------------
    if data is None:
        return

    if selected_table == "sabt_forosh_maskoni":
        entry_malek_phone_number.delete(0, tk.END)
        entry_malek_phone_number.insert(0, data[15])

        metraj_lable_right_entry.delete(0, tk.END)
        metraj_lable_right_entry.insert(0, data[17])
        options = []

        options.append(f"پارکینگ : {'دارد' if data[7] == '1' else 'ندارد'}")
        options.append(f"آسانسور : {'دارد' if data[8] == '1' else 'ندارد'}")
        options.append(f"انباری : {'دارد' if data[9] == '1' else 'ندارد'}")

        options.append(f"سرمایش : {data[10] if data[10] else 'ندارد'}")
        options.append(f"گرمایش : {data[11] if data[11] else 'ندارد'}")
        options.append(f"کف : {data[12] if data[12] else 'ثبت نشده'}")
        options.append(f"سرویس : {data[13] if data[13] else 'ثبت نشده'}")
        
        options_text_entry.delete("1.0", tk.END)
        options_text_entry.insert("1.0", "\n".join(options))
        options_text_entry.tag_add("right", "1.0", "end")
        

    elif selected_table == "sabt_ejareh_maskoni":
        entry_malek_phone_number.delete(0, tk.END)
        entry_malek_phone_number.insert(0, data[17])

        metraj_lable_right_entry.delete(0, tk.END)
        metraj_lable_right_entry.insert(0, data[18])
        options = []

        options.append(f"پارکینگ : {'دارد' if data[7] == '1' else 'ندارد'}")
        options.append(f"آسانسور : {'دارد' if data[8] == '1' else 'ندارد'}")
        options.append(f"انباری : {'دارد' if data[9] == '1' else 'ندارد'}")

        options.append(f"سرمایش : {data[10] if data[10] else 'ندارد'}")
        options.append(f"گرمایش : {data[11] if data[11] else 'ندارد'}")
        options.append(f"کف : {data[12] if data[12] else 'ثبت نشده'}")
        options.append(f"سرویس : {data[13] if data[13] else 'ثبت نشده'}")
        
        options_text_entry.delete("1.0", tk.END)
        options_text_entry.insert("1.0", "\n".join(options))
        options_text_entry.tag_add("right", "1.0", "end")
    elif selected_table == "sabt_darkhast_kharid_maskoni":

        entry_malek_phone_number.delete(0, tk.END)
        entry_malek_phone_number.insert(0, data[16])

        metraj_lable_right_entry.delete(0, tk.END)
        metraj_lable_right_entry.insert(0, data[17])

        options = []

        options.append(f"پارکینگ : {'دارد' if data[7] == '1' else 'ندارد'}")
        options.append(f"آسانسور : {'دارد' if data[8] == '1' else 'ندارد'}")
        options.append(f"انباری : {'دارد' if data[9] == '1' else 'ندارد'}")

        options.append(f"سرمایش : {data[10] if data[10] else 'ندارد'}")
        options.append(f"گرمایش : {data[11] if data[11] else 'ندارد'}")
        options.append(f"کف : {data[12] if data[12] else 'ثبت نشده'}")
        options.append(f"سرویس : {data[13] if data[13] else 'ثبت نشده'}")

        options_text_entry.delete("1.0", tk.END)
        options_text_entry.insert("1.0", "\n".join(options))
        options_text_entry.tag_add("right", "1.0", "end")
    elif selected_table == "sabt_darkhast_ejareh_maskoni":

        entry_malek_phone_number.delete(0, tk.END)
        entry_malek_phone_number.insert(0, data[17])

        metraj_lable_right_entry.delete(0, tk.END)
        metraj_lable_right_entry.insert(0, data[18])

        options = []

        options.append(f"پارکینگ : {'دارد' if data[7] == '1' else 'ندارد'}")
        options.append(f"آسانسور : {'دارد' if data[8] == '1' else 'ندارد'}")
        options.append(f"انباری : {'دارد' if data[9] == '1' else 'ندارد'}")

        options.append(f"سرمایش : {data[10] if data[10] else 'ندارد'}")
        options.append(f"گرمایش : {data[11] if data[11] else 'ندارد'}")
        options.append(f"کف : {data[12] if data[12] else 'ثبت نشده'}")
        options.append(f"سرویس : {data[13] if data[13] else 'ثبت نشده'}")

        options_text_entry.delete("1.0", tk.END)
        options_text_entry.insert("1.0", "\n".join(options))
        options_text_entry.tag_add("right", "1.0", "end")
#--------------------------اداری تجاری-----------------------------
    elif selected_table == "sabt_forosh_edari_tejari":

        entry_malek_phone_number.delete(0, tk.END)
        entry_malek_phone_number.insert(0, data[14])

        metraj_lable_right_entry.delete(0, tk.END)
        metraj_lable_right_entry.insert(0, data[2])

        options = []

        options.append(f"پارکینگ : {'دارد' if data[7]=='1' else 'ندارد'}")
        options.append(f"آسانسور : {'دارد' if data[8]=='1' else 'ندارد'}")
        options.append(f"انباری : {'دارد' if data[9]=='1' else 'ندارد'}")
        options.append(f"آب و گاز : {data[10] if data[10] else 'ثبت نشده'}")
        options.append(f"سرمایش : {data[11] if data[11] else 'ثبت نشده'}")
        options.append(f"گرمایش : {data[12] if data[12] else 'ثبت نشده'}")

        options_text_entry.delete("1.0", tk.END)
        options_text_entry.insert("1.0", "\n".join(options))
        options_text_entry.tag_add("right", "1.0", "end")
    elif selected_table == "sabt_ejareh_edari_tejari":

        entry_malek_phone_number.delete(0, tk.END)
        entry_malek_phone_number.insert(0, data[16])

        metraj_lable_right_entry.delete(0, tk.END)
        metraj_lable_right_entry.insert(0, data[2])

        options = []

        options.append(f"پارکینگ : {'دارد' if data[7]=='1' else 'ندارد'}")
        options.append(f"آسانسور : {'دارد' if data[8]=='1' else 'ندارد'}")
        options.append(f"انباری : {'دارد' if data[9]=='1' else 'ندارد'}")
        options.append(f"آب و گاز : {data[10] if data[10] else 'ثبت نشده'}")
        options.append(f"سرمایش : {data[11] if data[11] else 'ثبت نشده'}")
        options.append(f"گرمایش : {data[12] if data[12] else 'ثبت نشده'}")
        options.append(f"ودیعه : {data[13]}")
        options.append(f"اجاره : {data[14]}")

        options_text_entry.delete("1.0", tk.END)
        options_text_entry.insert("1.0", "\n".join(options))
        options_text_entry.tag_add("right", "1.0", "end")

    elif selected_table == "sabt_darkhast_kharid_edari_tejari":
        entry_malek_phone_number.delete(0, tk.END)
        entry_malek_phone_number.insert(0, data[15])

        metraj_lable_right_entry.delete(0, tk.END)
        metraj_lable_right_entry.insert(0, data[2])

        options = []

        options.append(f"پارکینگ : {'دارد' if data[7]=='1' else 'ندارد'}")
        options.append(f"آسانسور : {'دارد' if data[8]=='1' else 'ندارد'}")
        options.append(f"انباری : {'دارد' if data[9]=='1' else 'ندارد'}")

        options.append(f"آب و گاز : {data[10] if data[10] else 'ندارد'}")
        options.append(f"سرمایش : {data[11] if data[11] else 'ندارد'}")
        options.append(f"گرمایش : {data[12] if data[12] else 'ندارد'}")

        options_text_entry.delete("1.0", tk.END)
        options_text_entry.insert("1.0", "\n".join(options))
        options_text_entry.tag_add("right", "1.0", "end")
    
    elif selected_table == "sabt_darkhast_ejareh_edari_tejari":
        entry_malek_phone_number.delete(0, tk.END)
        entry_malek_phone_number.insert(0, data[16])

        metraj_lable_right_entry.delete(0, tk.END)
        metraj_lable_right_entry.insert(0, data[2])

        options = []

        options.append(f"پارکینگ : {'دارد' if data[7]=='1' else 'ندارد'}")
        options.append(f"آسانسور : {'دارد' if data[8]=='1' else 'ندارد'}")
        options.append(f"انباری : {'دارد' if data[9]=='1' else 'ندارد'}")

        options.append(f"آب و گاز : {data[10] if data[10] else 'ندارد'}")
        options.append(f"سرمایش : {data[11] if data[11] else 'ندارد'}")
        options.append(f"گرمایش : {data[12] if data[12] else 'ندارد'}")

        options_text_entry.delete("1.0", tk.END)
        options_text_entry.insert("1.0", "\n".join(options))
        options_text_entry.tag_add("right", "1.0", "end")
#------------------------کارگاه------------------------
    elif selected_table == "sabt_forosh_kargah":

        entry_malek_phone_number.delete(0, tk.END)
        entry_malek_phone_number.insert(0, data[14])

        metraj_lable_right_entry.delete(0, tk.END)
        metraj_lable_right_entry.insert(0, data[2])

        options = []

        options.append(f"کاربری : {data[1] if data[1] else 'ثبت نشده'}")
        options.append(f"برق : {data[5] if data[5] else 'ثبت نشده'}")
        options.append(f"گرمایش : {data[6] if data[6] else 'ثبت نشده'}")
        options.append(f"سرمایش : {data[7] if data[7] else 'ثبت نشده'}")
        options.append(f"آب : {data[8] if data[8] else 'ثبت نشده'}")
        options.append(f"ابزار : {data[9] if data[9] else 'ثبت نشده'}")
        options.append(f"سرویس : {data[10] if data[10] else 'ثبت نشده'}")
        options.append(f"حمام : {data[11] if data[11] else 'ثبت نشده'}")
        options.append(f"اتاق : {data[12] if data[12] else 'ثبت نشده'}")

        options_text_entry.delete("1.0", tk.END)
        options_text_entry.insert("1.0", "\n".join(options))
        options_text_entry.tag_configure("right", justify="right")
        options_text_entry.tag_add("right", "1.0", "end")






    cursor.close()
    db.close()

#---------------------تابع باز کردن ادیت----------
def open_edit():
#--------------------------مسکونی------------------  
    if selected_table == "sabt_forosh_maskoni":
        zakhire_forosh_maskoni.place_forget()
        edit_btn_forosh_maskoni.place(x=300,y=30)
        delete_btn_forosh_maskoni.place(x=200,y=30)
        root.withdraw()
        forosh_maskoni_window.deiconify()
        db = get_connection()
        cursor = db.cursor()
        cursor.execute("USE state_agency")

        cursor.execute(
            "SELECT * FROM sabt_forosh_maskoni WHERE id=%s",
            (selected_id,)
        )

        data = cursor.fetchone()
        print(data)

        sal_sakht_forosh_maskoni_entry.delete(0, tk.END)
        sal_sakht_forosh_maskoni_entry.insert(0, data[2])

        addrres_forosh_maskoni_entry.delete("1.0", tk.END)
        addrres_forosh_maskoni_entry.insert("1.0", data[3])

        tabaghe_forosh_maskoni_entry.delete(0, tk.END)
        tabaghe_forosh_maskoni_entry.insert(0, data[4])

        vahed_forosh_maskoni_entry.delete(0, tk.END)
        vahed_forosh_maskoni_entry.insert(0, data[5])

        otagh_forosh_maskoni_entry.delete(0, tk.END)
        otagh_forosh_maskoni_entry.insert(0, data[6])

        parking_forosh_maskoni_var.set(data[7])
        asansor_forosh_maskoni_var.set(data[8])
        anbari_forosh_maskoni_var.set(data[9])

        sarmaesh_combo_forosh_maskoni.set(data[10])
        garmaesh_combo_forosh_maskoni.set(data[11])
        kaf_combo_forosh_maskoni.set(data[12])
        toilet_combo_forosh_maskoni.set(data[13])

        name_malek_forosh_maskoni_entry.delete(0, tk.END)
        name_malek_forosh_maskoni_entry.insert(0, data[14])

        shomareh_malek_forosh_maskoni_entry.delete(0, tk.END)
        shomareh_malek_forosh_maskoni_entry.insert(0, data[15])

        gheimat_kol_forosh_maskoni_entry.delete(0, tk.END)
        gheimat_kol_forosh_maskoni_entry.insert(0, data[16])

        metraj_forosh_maskoni_entry.delete(0, tk.END)
        metraj_forosh_maskoni_entry.insert(0, data[17])

        
        cursor.close()
        db.close()

    elif selected_table == "sabt_ejareh_maskoni":
        save_button_ejareh_maskooni.place_forget()
        edit_btn_ejareh_maskoni.place(x=300,y=30)
        delete_btn_ejareh_maskoni.place(x=200,y=30)
        root.withdraw()
        ejareh_maskoni_window.deiconify()
        db = get_connection()
        cursor = db.cursor()
        cursor.execute("USE state_agency")

        cursor.execute(
            "SELECT * FROM sabt_ejareh_maskoni WHERE id=%s",
            (selected_id,)
        )

        data = cursor.fetchone()
        print(data)

        sal_sakht_ejareh_maskoni_entry.delete(0, tk.END)
        sal_sakht_ejareh_maskoni_entry.insert(0, data[2])

        addrres_ejareh_maskoni_entry.delete("1.0", tk.END)
        addrres_ejareh_maskoni_entry.insert("1.0", data[3])

        tabaghe_ejareh_maskoni_entry.delete(0, tk.END)
        tabaghe_ejareh_maskoni_entry.insert(0, data[4])

        vahed_ejareh_maskoni_entry.delete(0, tk.END)
        vahed_ejareh_maskoni_entry.insert(0, data[5])

        otagh_ejareh_maskoni_entry.delete(0, tk.END)
        otagh_ejareh_maskoni_entry.insert(0, data[6])

        parking_ejareh_maskoni_var.set(data[7])
        asansor_ejareh_maskoni_var.set(data[8])
        anbari_ejareh_maskoni_var.set(data[9])

        sarmaesh_ejareh_maskoni_combo.set(data[10])
        garmaesh_ejareh_maskoni_combo.set(data[11])
        kaf_ejareh_maskoni_combo.set(data[12])
        toilet_ejareh_maskoni_combo.set(data[13])

        gheimat_pish_ejare_maskoni_entry.delete(0, tk.END)
        gheimat_pish_ejare_maskoni_entry.insert(0, data[14])

        gheimat_ejare_ejare_maskoni_entry.delete(0, tk.END)
        gheimat_ejare_ejare_maskoni_entry.insert(0, data[15])


        name_malek_ejareh_maskoni_entry.delete(0, tk.END)
        name_malek_ejareh_maskoni_entry.insert(0, data[16])

        shomareh_malek_ejareh_maskoni_entry.delete(0, tk.END)
        shomareh_malek_ejareh_maskoni_entry.insert(0, data[17])

        metraj_ejareh_maskoni_entry.delete(0, tk.END)
        metraj_ejareh_maskoni_entry.insert(0, data[18])


        cursor.close()
        db.close()
    elif selected_table == "sabt_darkhast_kharid_maskoni":
        edit_btn_darkhast_maskoni.place(x=300, y=30)
        delete_btn_darkhast_maskoni.place(x=200, y=30)
        zakhire_darkhast_maskoni.place_forget()

        root.withdraw()
        darkhast_maskoni_window.deiconify()



        db = get_connection()
        cursor = db.cursor()
        cursor.execute("USE state_agency")

        cursor.execute(
            "SELECT * FROM sabt_darkhast_kharid_maskoni WHERE id=%s",
            (selected_id,)
        )

        data = cursor.fetchone()

        melk_type_darkhast_maskoni_entry.set(data[1])
        change_darkhast_maskoni_type()

        sal_sakht_darkhast_maskoni_entry.delete(0, tk.END)
        sal_sakht_darkhast_maskoni_entry.insert(0, data[2])

        addrres_darkhast_maskoni_entry.delete("1.0", tk.END)
        addrres_darkhast_maskoni_entry.insert("1.0", data[3])

        tabaghe_darkhast_maskoni_entry.delete(0, tk.END)
        tabaghe_darkhast_maskoni_entry.insert(0, data[4])

        vahed_darkhast_maskoni_entry.delete(0, tk.END)
        vahed_darkhast_maskoni_entry.insert(0, data[5])

        otagh_darkhast_maskoni_entry.delete(0, tk.END)
        otagh_darkhast_maskoni_entry.insert(0, data[6])

        parking_darkhast_maskoni_var.set(data[7])
        asansor_darkhast_maskoni_var.set(data[8])
        anbari_darkhast_maskoni_var.set(data[9])

        sarmaesh_combo_darkhast_maskoni.set(data[10])
        garmaesh_combo_darkhast_maskoni.set(data[11])
        kaf_combo_darkhast_maskoni.set(data[12])
        toilet_combo_darkhast_maskoni.set(data[13])

        gheimat_kol_darkhast_maskoni_entry.delete(0, tk.END)
        gheimat_kol_darkhast_maskoni_entry.insert(0, data[14])

        name_moshtari_darkhast_maskoni_entry.delete(0, tk.END)
        name_moshtari_darkhast_maskoni_entry.insert(0, data[15])

        shomareh_moshtari_darkhast_maskoni_entry.delete(0, tk.END)
        shomareh_moshtari_darkhast_maskoni_entry.insert(0, data[16])

        metraj_darkhast_maskoni_entry.delete(0, tk.END)
        metraj_darkhast_maskoni_entry.insert(0, data[17])

        cursor.close()
        db.close()
    elif selected_table == "sabt_darkhast_ejareh_maskoni":
        edit_btn_darkhast_maskoni.place(x=300, y=30)
        delete_btn_darkhast_maskoni.place(x=200, y=30)
        zakhire_darkhast_maskoni.place_forget()

        root.withdraw()
        darkhast_maskoni_window.deiconify()

        db = get_connection()
        cursor = db.cursor()
        cursor.execute("USE state_agency")

        cursor.execute(
            "SELECT * FROM sabt_darkhast_ejareh_maskoni WHERE id=%s",
            (selected_id,)
        )

        data = cursor.fetchone()

        melk_type_darkhast_maskoni_entry.set(data[1])
        change_darkhast_maskoni_type()

        sal_sakht_darkhast_maskoni_entry.delete(0, tk.END)
        sal_sakht_darkhast_maskoni_entry.insert(0, data[2])

        addrres_darkhast_maskoni_entry.delete("1.0", tk.END)
        addrres_darkhast_maskoni_entry.insert("1.0", data[3])

        tabaghe_darkhast_maskoni_entry.delete(0, tk.END)
        tabaghe_darkhast_maskoni_entry.insert(0, data[4])

        vahed_darkhast_maskoni_entry.delete(0, tk.END)
        vahed_darkhast_maskoni_entry.insert(0, data[5])

        otagh_darkhast_maskoni_entry.delete(0, tk.END)
        otagh_darkhast_maskoni_entry.insert(0, data[6])

        parking_darkhast_maskoni_var.set(data[7])
        asansor_darkhast_maskoni_var.set(data[8])
        anbari_darkhast_maskoni_var.set(data[9])

        sarmaesh_combo_darkhast_maskoni.set(data[10])
        garmaesh_combo_darkhast_maskoni.set(data[11])
        kaf_combo_darkhast_maskoni.set(data[12])
        toilet_combo_darkhast_maskoni.set(data[13])

        mablagh_ejare_darkhast_maskoni_entry.delete(0, tk.END)
        mablagh_ejare_darkhast_maskoni_entry.insert(0, data[14])

        gheimat_pish_darkhast_maskoni_entry.delete(0, tk.END)
        gheimat_pish_darkhast_maskoni_entry.insert(0, data[15])

        name_moshtari_darkhast_maskoni_entry.delete(0, tk.END)
        name_moshtari_darkhast_maskoni_entry.insert(0, data[16])

        shomareh_moshtari_darkhast_maskoni_entry.delete(0, tk.END)
        shomareh_moshtari_darkhast_maskoni_entry.insert(0, data[17])

        metraj_darkhast_maskoni_entry.delete(0, tk.END)
        metraj_darkhast_maskoni_entry.insert(0, data[18])

        cursor.close()
        db.close()
#-------------------------------اداری تجاری---------------------
    elif selected_table == "sabt_forosh_edari_tejari":
        edit_btn_forosh_edari_tejari.place(x=300,y=30)
        delete_btn_forosh_edari_tejari.place(x=200,y=30)
        zakhire_forosh_edari_tejari.place_forget()
        root.withdraw()
        forosh_edari_tejari_window.deiconify()

        db=get_connection()
        cursor=db.cursor()
        cursor.execute("USE state_agency")

        cursor.execute(
            "SELECT * FROM sabt_forosh_edari_tejari WHERE id=%s",
            (selected_id,)
        )

        data=cursor.fetchone()

    
        metraj_forosh_edari_tejari_entry.delete(0,tk.END)
        metraj_forosh_edari_tejari_entry.insert(0,data[2])

        sal_sakht_forosh_edari_tejari_entry.delete(0,tk.END)
        sal_sakht_forosh_edari_tejari_entry.insert(0,data[3])

        addrres_forosh_edari_tejari_entry.delete("1.0",tk.END)
        addrres_forosh_edari_tejari_entry.insert("1.0",data[4])

        tabaghe_forosh_edari_tejari_entry.delete(0,tk.END)
        tabaghe_forosh_edari_tejari_entry.insert(0,data[5])

        vahed_forosh_edari_tejari_entry.delete(0,tk.END)
        vahed_forosh_edari_tejari_entry.insert(0,data[6])

        parking_forosh_edari_tejari_var.set(data[7])
        asansor_forosh_edari_tejari_var.set(data[8])
        anbari_forosh_edari_tejari_var.set(data[9])

        aab_va_gaz_combo_forosh_edari_tejari.set(data[10])
        sarmaesh_combo_forosh_edari_tejari.set(data[11])
        garmaesh_combo_forosh_edari_tejari.set(data[12])

        name_malek_forosh_edari_tejari_entry.delete(0,tk.END)
        name_malek_forosh_edari_tejari_entry.insert(0,data[13])

        shomareh_malek_forosh_edari_tejari_entry.delete(0,tk.END)
        shomareh_malek_forosh_edari_tejari_entry.insert(0,data[14])

        gheimat_kol_forosh_edari_tejari_entry.delete(0,tk.END)
        gheimat_kol_forosh_edari_tejari_entry.insert(0,data[15])

        cursor.close()
        db.close()
    elif selected_table == "sabt_ejareh_edari_tejari":

        root.withdraw()
        ejareh_edari_tejari_window.deiconify()

        edit_btn_ejareh_edari_tejari.place(x=300,y=30)
        delete_btn_ejareh_edari_tejari.place(x=200,y=30)
        zakhire_ejareh_edari_tejari.place_forget()
        db=get_connection()
        cursor=db.cursor()
        cursor.execute("USE state_agency")

        cursor.execute(
            "SELECT * FROM sabt_ejareh_edari_tejari WHERE id=%s",
            (selected_id,)
        )
        data=cursor.fetchone()

        metraj_melk_ejareh_edari_tejari_entry.delete(0,tk.END)
        metraj_melk_ejareh_edari_tejari_entry.insert(0,data[2])

        sal_sakht_ejareh_edari_tejari_entry.delete(0,tk.END)
        sal_sakht_ejareh_edari_tejari_entry.insert(0,data[3])

        addrres_ejareh_edari_tejari_entry.delete("1.0",tk.END)
        addrres_ejareh_edari_tejari_entry.insert("1.0",data[4])

        tabaghe_ejareh_edari_tejari_entry.delete(0,tk.END)
        tabaghe_ejareh_edari_tejari_entry.insert(0,data[5])

        vahed_ejareh_edari_tejari_entry.delete(0,tk.END)
        vahed_ejareh_edari_tejari_entry.insert(0,data[6])

        parking_ejareh_edari_tejari_var.set(data[7])
        asansor_ejareh_edari_tejari_var.set(data[8])
        anbari_ejareh_edari_tejari_var.set(data[9])

        ab_va_gaz_combo_emkanat_ejareh_edari_tejari.set(data[10])
        sarmayesh_combo_emkanat_ejareh_edari_tejari.set(data[11])
        garmayesh_combo_emkanat_ejareh_edari_tejari.set(data[12])

        mablagh_pish_ejareh_edari_tejari_entry.delete(0,tk.END)
        mablagh_pish_ejareh_edari_tejari_entry.insert(0,data[13])

        mablagh_ejare_ejareh_edari_tejari_entry.delete(0,tk.END)
        mablagh_ejare_ejareh_edari_tejari_entry.insert(0,data[14])

        name_malek_ejareh_edari_tejari_entry.delete(0,tk.END)
        name_malek_ejareh_edari_tejari_entry.insert(0,data[15])

        shomareh_malek_ejareh_edari_tejari_entry.delete(0,tk.END)
        shomareh_malek_ejareh_edari_tejari_entry.insert(0,data[16])

        cursor.close()
        db.close()
    elif selected_table == "sabt_darkhast_kharid_edari_tejari":

        root.withdraw()
        darkhast_edari_tejari_window.deiconify()

        edit_btn_darkhast_edari_tejari.place(x=300, y=30)
        delete_btn_darkhast_edari_tejari.place(x=200, y=30)
        zakhire_darkhast_edari_tejari.place_forget()

        db = get_connection()
        cursor = db.cursor()
        cursor.execute("USE state_agency")

        cursor.execute(
            "SELECT * FROM sabt_darkhast_kharid_edari_tejari WHERE id=%s",
            (selected_id,)
        )

        data = cursor.fetchone()

        combo_darkhast_edari_tejari_entry.set(data[1])
        change_darkhast_edari_tejari_type(event=None)

        metraj_melk_darkhast_edari_tejari_entry.delete(0, tk.END)
        metraj_melk_darkhast_edari_tejari_entry.insert(0, data[2])

        sal_sakht_darkhast_edari_tejari_entry.delete(0, tk.END)
        sal_sakht_darkhast_edari_tejari_entry.insert(0, data[3])

        addrres_darkhast_edari_tejari_entry.delete("1.0", tk.END)
        addrres_darkhast_edari_tejari_entry.insert("1.0", data[4])

        tabaghe_darkhast_edari_tejari_entry.delete(0, tk.END)
        tabaghe_darkhast_edari_tejari_entry.insert(0, data[5])

        vahed_darkhast_edari_tejari_entry.delete(0, tk.END)
        vahed_darkhast_edari_tejari_entry.insert(0, data[6])

        parking_darkhast_edari_tejari_var.set(data[7])
        asansor_darkhast_edari_tejari_var.set(data[8])
        anbari_darkhast_edari_tejari_var.set(data[9])

        aab_va_gaz_combo_emkanat_darkhast_edari_tejari.set(data[10])
        sarmayesh_combo_emkanat_darkhast_edari_tejari.set(data[11])
        garmayesh_combo_emkanat_darkhast_edari_tejari.set(data[12])

        gheimat_kol_darkhast_edari_tejari_entry.delete(0, tk.END)
        gheimat_kol_darkhast_edari_tejari_entry.insert(0, data[13])

        name_moshtari_darkhast_edari_tejari_entry.delete(0, tk.END)
        name_moshtari_darkhast_edari_tejari_entry.insert(0, data[14])

        shomareh_moshtari_darkhast_edari_tejari_entry.delete(0, tk.END)
        shomareh_moshtari_darkhast_edari_tejari_entry.insert(0, data[15])

        cursor.close()
        db.close()
    elif selected_table == "sabt_darkhast_ejareh_edari_tejari":

        root.withdraw()
        darkhast_edari_tejari_window.deiconify()

        edit_btn_darkhast_edari_tejari.place(x=300, y=30)
        delete_btn_darkhast_edari_tejari.place(x=200, y=30)
        zakhire_darkhast_edari_tejari.place_forget()

        db = get_connection()
        cursor = db.cursor()
        cursor.execute("USE state_agency")

        cursor.execute(
            "SELECT * FROM sabt_darkhast_ejareh_edari_tejari WHERE id=%s",
            (selected_id,)
        )

        data = cursor.fetchone()

        combo_darkhast_edari_tejari_entry.set(data[1])
        change_darkhast_edari_tejari_type(event=None)

        metraj_melk_darkhast_edari_tejari_entry.delete(0, tk.END)
        metraj_melk_darkhast_edari_tejari_entry.insert(0, data[2])

        sal_sakht_darkhast_edari_tejari_entry.delete(0, tk.END)
        sal_sakht_darkhast_edari_tejari_entry.insert(0, data[3])

        addrres_darkhast_edari_tejari_entry.delete("1.0", tk.END)
        addrres_darkhast_edari_tejari_entry.insert("1.0", data[4])

        tabaghe_darkhast_edari_tejari_entry.delete(0, tk.END)
        tabaghe_darkhast_edari_tejari_entry.insert(0, data[5])

        vahed_darkhast_edari_tejari_entry.delete(0, tk.END)
        vahed_darkhast_edari_tejari_entry.insert(0, data[6])

        parking_darkhast_edari_tejari_var.set(data[7])
        asansor_darkhast_edari_tejari_var.set(data[8])
        anbari_darkhast_edari_tejari_var.set(data[9])

        aab_va_gaz_combo_emkanat_darkhast_edari_tejari.set(data[10])
        sarmayesh_combo_emkanat_darkhast_edari_tejari.set(data[11])
        garmayesh_combo_emkanat_darkhast_edari_tejari.set(data[12])

        mablagh_vadie_darkhast_edari_tejari_entry.delete(0, tk.END)
        mablagh_vadie_darkhast_edari_tejari_entry.insert(0, data[13])

        mablagh_ejareh_darkhast_edari_tejari_entry.delete(0, tk.END)
        mablagh_ejareh_darkhast_edari_tejari_entry.insert(0, data[14])

        name_moshtari_darkhast_edari_tejari_entry.delete(0, tk.END)
        name_moshtari_darkhast_edari_tejari_entry.insert(0, data[15])

        shomareh_moshtari_darkhast_edari_tejari_entry.delete(0, tk.END)
        shomareh_moshtari_darkhast_edari_tejari_entry.insert(0, data[16])

        cursor.close()
        db.close()

#----------------------------کارگاه---------------------
    elif selected_table == "sabt_forosh_kargah":

        root.withdraw()
        forosh_karghah_window.deiconify()

        edit_btn_forosh_kargah.place(x=300, y=30)
        delete_btn_forosh_kargah.place(x=200, y=30)
        zakhire_forosh_kargah.place_forget()

        db = get_connection()
        cursor = db.cursor()
        cursor.execute("USE state_agency")

        cursor.execute(
            "SELECT * FROM sabt_forosh_kargah WHERE id=%s",
            (selected_id,)
        )

        data = cursor.fetchone()

        metraj_forosh_kargah_entry.delete(0, tk.END)
        metraj_forosh_kargah_entry.insert(0, data[2])

        loctaion_forosh_kargah_entry.delete("1.0", tk.END)
        loctaion_forosh_kargah_entry.insert("1.0", data[3])

        sal_sakht_forosh_kargah_entry.delete(0, tk.END)
        sal_sakht_forosh_kargah_entry.insert(0, data[4])

        vaziat_bargh_forosh_kargah_combo.set(data[5])
        garmayesh_type_forosh_kargah_combo.set(data[6])
        sarmayesh_forosh_kargah_combo.set(data[7])
        vaziat_ab_forosh_kargah_combo.set(data[8])
        abzar_forosh_kargah_combo.set(data[9])
        toilet_forosh_kargah_combo.set(data[10])
        hamam_forosh_kargah_combo.set(data[11])
        otagh_forosh_kargah_combo.set(data[12])

        name_malek_forosh_kargah_entry.delete(0, tk.END)
        name_malek_forosh_kargah_entry.insert(0, data[13])

        shomareh_malek_forosh_kargah_entry.delete(0, tk.END)
        shomareh_malek_forosh_kargah_entry.insert(0, data[14])

        gheimat_kol_forosh_kargah_entry.delete(0, tk.END)
        gheimat_kol_forosh_kargah_entry.insert(0, data[15])

        cursor.close()
        db.close()
#--------------------------------------توابع ویرایش صفحات---------------
#------------مسکونی------------
def update_forosh_maskoni():
    db = get_connection()
    cursor = db.cursor()
    cursor.execute("USE state_agency")

    sql = """
    UPDATE sabt_forosh_maskoni SET type_melk=%s,sal_sakht=%s,address=%s,tabaghe=%s,vahed=%s,otagh=%s,parking=%s,
        asansor=%s,anbari=%s,sarmayesh=%s,garmayesh=%s,kaf=%s,toilet=%s,name_malek=%s,shomareh_malek=%s,gheimat_kol=%s,metraj=%s
    WHERE id=%s
    """

    values = (
        melk_type_forosh_maskoni_entry.get(),
        sal_sakht_forosh_maskoni_entry.get(),
        addrres_forosh_maskoni_entry.get("1.0", tk.END),
        tabaghe_forosh_maskoni_entry.get(),
        vahed_forosh_maskoni_entry.get(),
        otagh_forosh_maskoni_entry.get(),
        parking_forosh_maskoni_var.get(),
        asansor_forosh_maskoni_var.get(),
        anbari_forosh_maskoni_var.get(),
        sarmaesh_combo_forosh_maskoni.get(),
        garmaesh_combo_forosh_maskoni.get(),
        kaf_combo_forosh_maskoni.get(),
        toilet_combo_forosh_maskoni.get(),
        name_malek_forosh_maskoni_entry.get(),
        shomareh_malek_forosh_maskoni_entry.get(),
        float(gheimat_kol_forosh_maskoni_entry.get()),
        metraj_forosh_maskoni_entry.get(),
        selected_id
    )

    cursor.execute(sql, values)
    db.commit()

    cursor.close()
    db.close()
    refresh_after_edit()

    messagebox.showinfo("موفق", "اطلاعات با موفقیت ویرایش شد.")
def update_ejareh_maskoni():
    db = get_connection()
    cursor = db.cursor()
    cursor.execute("USE state_agency")

    sql = """
    UPDATE sabt_ejareh_maskoni SET type_melk=%s,sal_sakht=%s,address=%s,tabaghe=%s,vahed=%s,otagh=%s,parking=%s,
        asansor=%s,anbari=%s,sarmayesh=%s,garmayesh=%s,kaf=%s,toilet=%s,name_malek=%s,shomareh_malek=%s,pish=%s,ejareh=%s
    WHERE id=%s
    """

    values = (
        melk_type_ejareh_maskoni_entry.get(),
        sal_sakht_ejareh_maskoni_entry.get(),
        addrres_ejareh_maskoni_entry.get("1.0", tk.END),
        tabaghe_ejareh_maskoni_entry.get(),
        vahed_ejareh_maskoni_entry.get(),
        otagh_ejareh_maskoni_entry.get(),
        parking_ejareh_maskoni_var.get(),
        asansor_ejareh_maskoni_var.get(),
        anbari_ejareh_maskoni_var.get(),
        sarmaesh_ejareh_maskoni_combo.get(),
        garmaesh_ejareh_maskoni_combo.get(),
        kaf_ejareh_maskoni_combo.get(),
        toilet_ejareh_maskoni_combo.get(),
        name_malek_ejareh_maskoni_entry.get(),
        shomareh_malek_ejareh_maskoni_entry.get(),
        float(gheimat_pish_ejare_maskoni_entry.get()),
        gheimat_ejare_ejare_maskoni_entry.get(),
        selected_id
    )

    cursor.execute(sql, values)
    db.commit()

    cursor.close()
    db.close()
    refresh_after_edit()

    messagebox.showinfo("موفق", "اطلاعات با موفقیت ویرایش شد.")
def update_darkhast_maskoni():
    db = get_connection()
    cursor = db.cursor()
    cursor.execute("USE state_agency")

    if melk_type_darkhast_maskoni_entry.get() == "درخواست خرید مسکونی":

        sql = """
        UPDATE sabt_darkhast_kharid_maskoni SET
        type_melk=%s,sal_sakht=%s,address=%s,tabaghe=%s,vahed=%s,
        otagh=%s,parking=%s,asansor=%s,anbari=%s,sarmayesh=%s,
        garmayesh=%s,kaf=%s,toilet=%s,gheimat_kol=%s,
        name_moshtari=%s,shomareh_moshtari=%s,metraj=%s
        WHERE id=%s
        """

        values = (
            melk_type_darkhast_maskoni_entry.get(),
            sal_sakht_darkhast_maskoni_entry.get(),
            addrres_darkhast_maskoni_entry.get("1.0", tk.END),
            tabaghe_darkhast_maskoni_entry.get(),
            vahed_darkhast_maskoni_entry.get(),
            otagh_darkhast_maskoni_entry.get(),
            parking_darkhast_maskoni_var.get(),
            asansor_darkhast_maskoni_var.get(),
            anbari_darkhast_maskoni_var.get(),
            sarmaesh_combo_darkhast_maskoni.get(),
            garmaesh_combo_darkhast_maskoni.get(),
            kaf_combo_darkhast_maskoni.get(),
            toilet_combo_darkhast_maskoni.get(),
            gheimat_kol_darkhast_maskoni_entry.get(),
            name_moshtari_darkhast_maskoni_entry.get(),
            shomareh_moshtari_darkhast_maskoni_entry.get(),
            metraj_darkhast_maskoni_entry.get(),
            selected_id
        )

    else:

        sql = """
        UPDATE sabt_darkhast_ejareh_maskoni SET
        type_melk=%s,sal_sakht=%s,address=%s,tabaghe=%s,vahed=%s,
        otagh=%s,parking=%s,asansor=%s,anbari=%s,sarmayesh=%s,
        garmayesh=%s,kaf=%s,toilet=%s,ejareh=%s,pish=%s,
        name_moshtari=%s,shomareh_moshtari=%s,metraj=%s
        WHERE id=%s
        """

        values = (
            melk_type_darkhast_maskoni_entry.get(),
            sal_sakht_darkhast_maskoni_entry.get(),
            addrres_darkhast_maskoni_entry.get("1.0", tk.END),
            tabaghe_darkhast_maskoni_entry.get(),
            vahed_darkhast_maskoni_entry.get(),
            otagh_darkhast_maskoni_entry.get(),
            parking_darkhast_maskoni_var.get(),
            asansor_darkhast_maskoni_var.get(),
            anbari_darkhast_maskoni_var.get(),
            sarmaesh_combo_darkhast_maskoni.get(),
            garmaesh_combo_darkhast_maskoni.get(),
            kaf_combo_darkhast_maskoni.get(),
            toilet_combo_darkhast_maskoni.get(),
            mablagh_ejare_darkhast_maskoni_entry.get(),
            gheimat_pish_darkhast_maskoni_entry.get(),
            name_moshtari_darkhast_maskoni_entry.get(),
            shomareh_moshtari_darkhast_maskoni_entry.get(),
            metraj_darkhast_maskoni_entry.get(),
            selected_id
        )

    cursor.execute(sql, values)
    db.commit()

    cursor.close()
    db.close()

    refresh_after_edit()
    messagebox.showinfo("موفق", "اطلاعات با موفقیت ویرایش شد.")
#----------------------اداری تجاری-----------------
def update_forosh_edari_tejari():

    db=get_connection()
    cursor=db.cursor()
    cursor.execute("USE state_agency")
    sql = """
    UPDATE sabt_forosh_edari_tejari
    SET
    type_melk=%s,
    metraj_melk=%s,
    sal_sakht=%s,
    address=%s,
    tabaghe=%s,
    vahed=%s,
    parking=%s,
    asansor=%s,
    anbari=%s,
    aab_va_gaz=%s,
    system_sarmayesh=%s,
    system_garmayesh=%s,
    name_malek=%s,
    shomareh_malek=%s,
    gheimat_kol=%s
    WHERE id=%s
    """

    values=(
        melk_type_forosh_edari_tejari_entry.get(),
        metraj_forosh_edari_tejari_entry.get(),
        sal_sakht_forosh_edari_tejari_entry.get(),
        addrres_forosh_edari_tejari_entry.get("1.0",tk.END),
        tabaghe_forosh_edari_tejari_entry.get(),
        vahed_forosh_edari_tejari_entry.get(),
        parking_forosh_edari_tejari_var.get(),
        asansor_forosh_edari_tejari_var.get(),
        anbari_forosh_edari_tejari_var.get(),
        aab_va_gaz_combo_forosh_edari_tejari.get(),
        sarmaesh_combo_forosh_edari_tejari.get(),
        garmaesh_combo_forosh_edari_tejari.get(),
        name_malek_forosh_edari_tejari_entry.get(),
        shomareh_malek_forosh_edari_tejari_entry.get(),
        float(gheimat_kol_forosh_edari_tejari_entry.get()),
        selected_id
    )

    cursor.execute(sql,values)
    db.commit()

    cursor.close()
    db.close()

    refresh_after_edit()

    messagebox.showinfo("موفق","اطلاعات با موفقیت ویرایش شد.")
def update_ejareh_edari_tejari():

    db=get_connection()
    cursor=db.cursor()
    cursor.execute("USE state_agency")

    sql="""
    UPDATE sabt_ejareh_edari_tejari
    SET
    type_melk=%s,
    metraj_melk=%s,
    sal_sakht=%s,
    address=%s,
    tabaghe=%s,
    vahed=%s,
    parking=%s,
    asansor=%s,
    anbari=%s,
    aab_va_gaz=%s,
    system_sarmayesh=%s,
    system_garmayesh=%s,
    gheimat_vadie=%s,
    gheimat_ejareh=%s,
    name_malek=%s,
    shomareh_malek=%s
    WHERE id=%s
    """

    values=(
        melk_type_ejareh_edari_tejari_entry.get(),
        metraj_melk_ejareh_edari_tejari_entry.get(),
        sal_sakht_ejareh_edari_tejari_entry.get(),
        addrres_ejareh_edari_tejari_entry.get("1.0",tk.END),
        tabaghe_ejareh_edari_tejari_entry.get(),
        vahed_ejareh_edari_tejari_entry.get(),
        parking_ejareh_edari_tejari_var.get(),
        asansor_ejareh_edari_tejari_var.get(),
        anbari_ejareh_edari_tejari_var.get(),
        ab_va_gaz_combo_emkanat_ejareh_edari_tejari.get(),
        sarmayesh_combo_emkanat_ejareh_edari_tejari.get(),
        garmayesh_combo_emkanat_ejareh_edari_tejari.get(),
        float(mablagh_pish_ejareh_edari_tejari_entry.get()),
        float(mablagh_ejare_ejareh_edari_tejari_entry.get()),
        name_malek_ejareh_edari_tejari_entry.get(),
        shomareh_malek_ejareh_edari_tejari_entry.get(),
        selected_id
    )

    cursor.execute(sql,values)
    db.commit()

    cursor.close()
    db.close()

    refresh_after_edit()

    messagebox.showinfo("موفق","اطلاعات با موفقیت ویرایش شد.")
def update_darkhast_edari_tejari():

    db = get_connection()
    cursor = db.cursor()
    cursor.execute("USE state_agency")
    if combo_darkhast_edari_tejari_entry.get() =="درخواست خرید اداری و تجاری":
        sql = """
        UPDATE sabt_darkhast_kharid_edari_tejari
        SET
        type_melk=%s,
        metraj_melk=%s,
        sal_sakht=%s,
        address=%s,
        tabaghe=%s,
        vahed=%s,
        parking=%s,
        asansor=%s,
        anbari=%s,
        aab_va_gaz=%s,
        system_sarmayesh=%s,
        system_garmayesh=%s,
        mablagh_kharid=%s,
        name_moshtari=%s,
        shomareh_moshtari=%s
        WHERE id=%s
        """

        values = (
            combo_darkhast_edari_tejari_entry.get(),
            metraj_melk_darkhast_edari_tejari_entry.get(),
            sal_sakht_darkhast_edari_tejari_entry.get(),
            addrres_darkhast_edari_tejari_entry.get("1.0", tk.END),
            tabaghe_darkhast_edari_tejari_entry.get(),
            vahed_darkhast_edari_tejari_entry.get(),
            parking_darkhast_edari_tejari_var.get(),
            asansor_darkhast_edari_tejari_var.get(),
            anbari_darkhast_edari_tejari_var.get(),
            aab_va_gaz_combo_emkanat_darkhast_edari_tejari.get(),
            sarmayesh_combo_emkanat_darkhast_edari_tejari.get(),
            garmayesh_combo_emkanat_darkhast_edari_tejari.get(),
            gheimat_kol_darkhast_edari_tejari_entry.get(),
            name_moshtari_darkhast_edari_tejari_entry.get(),
            shomareh_moshtari_darkhast_edari_tejari_entry.get(),
            selected_id
        )
    else:
        sql = """
        UPDATE sabt_darkhast_ejareh_edari_tejari
        SET
        type_melk=%s,
        metraj_melk=%s,
        sal_sakht=%s,
        address=%s,
        tabaghe=%s,
        vahed=%s,
        parking=%s,
        asansor=%s,
        anbari=%s,
        aab_va_gaz=%s,
        system_sarmayesh=%s,
        system_garmayesh=%s,
        mablagh_vadie=%s,
        mablagh_ejareh=%s,
        name_moshtari=%s,
        shomareh_moshtari=%s
        WHERE id=%s
        """

        values = (
            combo_darkhast_edari_tejari_entry.get(),
            metraj_melk_darkhast_edari_tejari_entry.get(),
            sal_sakht_darkhast_edari_tejari_entry.get(),
            addrres_darkhast_edari_tejari_entry.get("1.0", tk.END),
            tabaghe_darkhast_edari_tejari_entry.get(),
            vahed_darkhast_edari_tejari_entry.get(),
            parking_darkhast_edari_tejari_var.get(),
            asansor_darkhast_edari_tejari_var.get(),
            anbari_darkhast_edari_tejari_var.get(),
            aab_va_gaz_combo_emkanat_darkhast_edari_tejari.get(),
            sarmayesh_combo_emkanat_darkhast_edari_tejari.get(),
            garmayesh_combo_emkanat_darkhast_edari_tejari.get(),
            mablagh_vadie_darkhast_edari_tejari_entry.get(),
            mablagh_ejareh_darkhast_edari_tejari_entry.get(),
            name_moshtari_darkhast_edari_tejari_entry.get(),
            shomareh_moshtari_darkhast_edari_tejari_entry.get(),
            selected_id
        )

    cursor.execute(sql, values)
    db.commit()

    cursor.close()
    db.close()

    refresh_after_edit()

    messagebox.showinfo("موفق", "اطلاعات با موفقیت ویرایش شد.")

#---------------------------------کارگاه----------------------
def update_forosh_kargah():

    db = get_connection()
    cursor = db.cursor()
    cursor.execute("USE state_agency")

    sql = """
    UPDATE sabt_forosh_kargah
    SET
        karbari_zamin=%s,
        metraj=%s,
        address=%s,
        sal_sakht=%s,
        vaziat_bargh=%s,
        garmayesh=%s,
        sarmayesh=%s,
        vaziat_ab=%s,
        abzar=%s,
        toilet=%s,
        hamam=%s,
        otagh=%s,
        name_malek=%s,
        shomareh_malek=%s,
        gheimat_kol=%s
    WHERE id=%s
    """

    values = (
        karbari_forosh_kargah_entry.get(),
        metraj_forosh_kargah_entry.get(),
        loctaion_forosh_kargah_entry.get("1.0", tk.END),
        sal_sakht_forosh_kargah_entry.get(),
        vaziat_bargh_forosh_kargah_combo.get(),
        garmayesh_type_forosh_kargah_combo.get(),
        sarmayesh_forosh_kargah_combo.get(),
        vaziat_ab_forosh_kargah_combo.get(),
        abzar_forosh_kargah_combo.get(),
        toilet_forosh_kargah_combo.get(),
        hamam_forosh_kargah_combo.get(),
        otagh_forosh_kargah_combo.get(),
        name_malek_forosh_kargah_entry.get(),
        shomareh_malek_forosh_kargah_entry.get(),
        float(gheimat_kol_forosh_kargah_entry.get()),
        selected_id
    )

    cursor.execute(sql, values)
    db.commit()

    cursor.close()
    db.close()

    refresh_after_edit()

    messagebox.showinfo("موفق", "اطلاعات با موفقیت ویرایش شد.")
#endregion
#-------------------------توابع حذف--------------------------
#region
def delete_forosh_maskoni():
    if not messagebox.askyesno("تأیید", "آیا از حذف این فایل مطمئن هستید؟"):
        return

    db = get_connection()
    cursor = db.cursor()
    cursor.execute("USE state_agency")

    cursor.execute(
        "DELETE FROM sabt_forosh_maskoni WHERE id=%s",
        (selected_id,)
    )
    db.commit()

    cursor.close()
    db.close()

    messagebox.showinfo("موفق", "فایل حذف شد.")

    refresh_after_edit()
    forosh_maskoni_window.withdraw()
    root.deiconify()

def delete_ejareh_maskoni():
    if not messagebox.askyesno("تأیید", "آیا از حذف این فایل مطمئن هستید؟"):
        return

    db = get_connection()
    cursor = db.cursor()
    cursor.execute("USE state_agency")

    cursor.execute(
        "DELETE FROM sabt_ejareh_maskoni WHERE id=%s",
        (selected_id,)
    )
    db.commit()

    cursor.close()
    db.close()

    messagebox.showinfo("موفق", "فایل حذف شد.")

    refresh_after_edit()
    ejareh_maskoni_window.withdraw()
    root.deiconify()
def delete_darkhast_maskoni():

    if not messagebox.askyesno("تأیید", "آیا از حذف این فایل مطمئن هستید؟"):
        return

    db = get_connection()
    cursor = db.cursor()
    cursor.execute("USE state_agency")

    if selected_table == "sabt_darkhast_kharid_maskoni":
        cursor.execute(
            "DELETE FROM sabt_darkhast_kharid_maskoni WHERE id=%s",
            (selected_id,)
        )

    elif selected_table == "sabt_darkhast_ejareh_maskoni":
        cursor.execute(
            "DELETE FROM sabt_darkhast_ejareh_maskoni WHERE id=%s",
            (selected_id,)
        )

    db.commit()

    cursor.close()
    db.close()

    messagebox.showinfo("موفق", "فایل حذف شد.")

    refresh_after_edit()
#-----------------------اداری تجاری--------------
def delete_forosh_edari_tejari():

    if not messagebox.askyesno("تأیید","آیا از حذف این فایل مطمئن هستید؟"):
        return

    db=get_connection()
    cursor=db.cursor()
    cursor.execute("USE state_agency")

    cursor.execute(
        "DELETE FROM sabt_forosh_edari_tejari WHERE id=%s",
        (selected_id,)
    )

    db.commit()

    cursor.close()
    db.close()

    messagebox.showinfo("موفق","فایل حذف شد.")

    refresh_after_edit()
def delete_ejareh_edari_tejari():

    if not messagebox.askyesno("تأیید","آیا از حذف این فایل مطمئن هستید؟"):
        return

    db=get_connection()
    cursor=db.cursor()
    cursor.execute("USE state_agency")

    cursor.execute(
        "DELETE FROM sabt_ejareh_edari_tejari WHERE id=%s",
        (selected_id,)
    )

    db.commit()

    cursor.close()
    db.close()

    messagebox.showinfo("موفق","فایل حذف شد.")

    refresh_after_edit()
def delete_darkhast_edari_tejari():

    if not messagebox.askyesno("تأیید", "آیا از حذف این فایل مطمئن هستید؟"):
        return

    db = get_connection()
    cursor = db.cursor()
    cursor.execute("USE state_agency")

    if selected_table == "sabt_darkhast_kharid_edari_tejari":

        cursor.execute(
            "DELETE FROM sabt_darkhast_kharid_edari_tejari WHERE id=%s",
            (selected_id,)
        )
    elif selected_table == "sabt_darkhast_ejareh_edari_tejari":
        cursor.execute(
            "DELETE FROM sabt_darkhast_ejareh_edari_tejari WHERE id=%s",
            (selected_id,)
        )

    db.commit()

    cursor.close()
    db.close()

    messagebox.showinfo("موفق", "فایل حذف شد.")

    refresh_after_edit()
#-----------------------------کارگاه---------------
def delete_forosh_kargah():

    if not messagebox.askyesno("تأیید", "آیا از حذف این فایل مطمئن هستید؟"):
        return

    db = get_connection()
    cursor = db.cursor()
    cursor.execute("USE state_agency")

    cursor.execute(
        "DELETE FROM sabt_forosh_kargah WHERE id=%s",
        (selected_id,)
    )

    db.commit()

    cursor.close()
    db.close()

    messagebox.showinfo("موفق", "فایل با موفقیت حذف شد.")

    refresh_after_edit()
#------------------رفرش----------------
def refresh_after_edit():
    clear_entry_forosh_maskoni()#پاک کردن صفحات
    clear_entry_ejareh_maskoni()
    clear_entry_darkhast_maskoni()
    clear_entry_forosh_edari_tejari()
    clear_entry_darkhast_edari_tejari()
    clear_entry_ejareh_edari_tejari()
    clear_entry_forosh_kargah()
    clear_entry_ejareh_karghah()
    clear_entry_darkhast_kargah()
    clear_entry_forosh_bagh_zamin()
    clear_entry_ejareh_bagh_zamin()
    clear_entry_darkhast_bagh_zamin()

    edit_btn_forosh_edari_tejari.place_forget()#تنظیم دکمه ها
    delete_btn_forosh_edari_tejari.place_forget()
    zakhire_forosh_edari_tejari.place(x=200, y=30)

    edit_btn_ejareh_edari_tejari.place_forget()
    delete_btn_ejareh_edari_tejari.place_forget()
    zakhire_ejareh_edari_tejari.place(x=300,y=20)

    edit_btn_darkhast_edari_tejari.place_forget()
    delete_btn_darkhast_edari_tejari.place_forget()
    zakhire_darkhast_edari_tejari.place(x=300,y=30)
    #------------------مسکونی---------
    zakhire_forosh_maskoni.place(x=300,y=30)
    edit_btn_forosh_maskoni.place_forget()
    delete_btn_forosh_maskoni.place_forget()

    save_button_ejareh_maskooni.place(x=300,y=30)
    edit_btn_ejareh_maskoni.place_forget()
    delete_btn_ejareh_maskoni.place_forget()

    edit_btn_darkhast_maskoni.place_forget()
    delete_btn_darkhast_maskoni.place_forget()
    zakhire_darkhast_maskoni.place(x=300,y=30)
    #---------------کارگاه-----------------
    zakhire_forosh_kargah.place(x=300,y=30)
    edit_btn_forosh_kargah.place_forget()
    delete_btn_forosh_kargah.place_forget()


    for item in tree.get_children():
        tree.delete(item)

    combo_file_type.set("فروش")#تنظیم صفحه اصلی
    melk_type_combo.set("مسکونی")

    melk_mahdode_gheimat_entry.delete(0, tk.END)
    mahdode_ta_entry.delete(0, tk.END)
    metraj_entry.delete(0, tk.END)
    address_entry.delete(0, tk.END)

    entry_malek_phone_number.delete(0, tk.END)
    metraj_lable_right_entry.delete(0, tk.END)

    options_text_entry.config(state="normal")
    options_text_entry.delete("1.0", tk.END)

    forosh_maskoni_window.withdraw()#بستن پنجره ها
    ejareh_maskoni_window.withdraw()
    darkhast_maskoni_window.withdraw()
    forosh_edari_tejari_window.withdraw()
    ejareh_edari_tejari_window.withdraw()
    darkhast_edari_tejari_window.withdraw()
    forosh_karghah_window.withdraw()
    ejareh_karghah_window.withdraw()
    darkhast_karghah_window.withdraw()
    root.deiconify()
#endregion
#---#----#----#----#----#----------  گرافیک   ----------#----#----#----#-----#-----------
# ---------دکمه فایل با منوی کشویی ------------------
#region 
menubar = tk.Menu(root, font=("Shabnam", 10))
# ایجاد منوی "ثبت فایل ها"
file_menu_sabt_file = tk.Menu(menubar, tearoff=0, font=("Shabnam", 10))
file_menu_sabt_file.add_command(label="فروش", command=forosh)
file_menu_sabt_file.add_command(label="رهن/اجاره", command=rahn)
file_menu_sabt_file.add_command(label="مشارکت", command=mosharecat)

# اضافه کردن منوی "ثبت فایل ها" به منوبار
menubar.add_cascade(label="ثبت فایل ها", menu=file_menu_sabt_file)

# اتصال منوبار به پنجره
root.config(menu=menubar)
#endregion
# ---------------اضافه کردن فیلد قرارداد ---------------
#region
file_menu_gharardad= tk.Menu(menubar, tearoff=0, font=("Shabnam", 10))
file_menu_gharardad.add_command(label="ثبت قراردادها", command=open_ghrardad)
# اضافه کردن منوی قرار دادها به منوبار
menubar.add_cascade(label="قراردادها", menu=file_menu_gharardad)
#endregion
#----------------دکمه گزارش ها با منوی کشویی ------------------------
#region
# ----------لیست کشویی فیلد گزارش ها-----------------
file_menu_gozaresh = tk.Menu(menubar, tearoff=0, font=("Shabnam", 10))
file_menu_gozaresh.add_command(label="خروجی اکسل", command=gozaresh)
file_menu_gozaresh.add_command(label="قراردادها", command=None)
# اضافه کردن منوی گزارش ها به منوبار
menubar.add_cascade(label="گزارش ها", menu=file_menu_gozaresh)
#endregion
# ----------------------اضافه کردن فیلد درخواست ها-------
#region
file_menu_darkhast= tk.Menu(menubar, tearoff=0, font=("Shabnam", 10))
file_menu_darkhast.add_command(label="درخواست خرید و اجاره", command=darkhast)

# اضافه کردن منوی گزارش ها به منوبار
menubar.add_cascade(label="درخواست ها", menu=file_menu_darkhast)
#endregion
#-------------------------اضافه کردن فیلد کاربران---------
#region
file_menu_karbaran= tk.Menu(menubar, tearoff=0, font=("Shabnam", 10))
file_menu_karbaran.add_command(label="", command=None)
file_menu_karbaran.add_command(label="", command=None)

# اضافه کردن منوی گزارش ها به منوبار
menubar.add_cascade(label="کاربران", menu=file_menu_karbaran)
#endregion
#========================================================
# -----------باکس سمت چپ - جستجو در فایل های ملک----------
#region
contant_frame=tk.Frame(root)
contant_frame.pack(fill="both",expand=True)
frame_jostojo_melk_left= tk.LabelFrame(contant_frame, text="جستجوی ملک", width=200, bg="#052340",fg="#00BFFF", font=("Shabnam", 16))
frame_jostojo_melk_left.pack(side="left", fill="both",expand=True, padx=6, pady=15)

box_jostojo_malk1= tk.Frame(frame_jostojo_melk_left,bg="#052340")
box_jostojo_malk1.pack(padx=6, pady=15)

file_type = tk.Label(box_jostojo_malk1,text="نوع فایل",bg="#052340", fg="#FFFFFF",font=("Shabnam", 13))
file_type.pack(padx=15,pady=10, side="right")
combo_file_type= ttk.Combobox(box_jostojo_malk1)
combo_file_type["values"] = ("رهن_اجاره","فروش","درخواست ها",)
combo_file_type["state"]=["readonly"]
combo_file_type.pack(padx=10, pady=10) 

box_jostojo_malk2= tk.Frame(frame_jostojo_melk_left,bg="#052340")
box_jostojo_malk2.pack(padx=6, pady=15)

melk_type_lable = tk.Label(box_jostojo_malk2,text="نوع ملک",bg="#052340", fg="#FFFFFF",font=("Shabnam", 13))
melk_type_lable.pack(padx=15,pady=10, side="right")
melk_type_combo= ttk.Combobox(box_jostojo_malk2)
melk_type_combo["values"] = ("مسکونی","اداری_تجاری","زمین","کارگاه","باغ")
melk_type_combo["state"]=["readonly"]
melk_type_combo.pack(padx=10, pady=10) 

box_jostojo_malk3 = tk.Frame(frame_jostojo_melk_left,bg="#052340")
box_jostojo_malk3.pack(padx=6, pady=5)

melk_mahdode_gheimat_lable = tk.Label(box_jostojo_malk3,text="محدوده قیمت",bg="#052340", fg="#FFFFFF",font=("Shabnam", 13))
melk_mahdode_gheimat_lable.pack(padx=5, pady=1)

melk_mahdode_gheimat_entry = tk.Entry(box_jostojo_malk3,bg="#FFFFFF", fg="#000000",font=("Shabnam", 13))
melk_mahdode_gheimat_entry.pack(padx=20,pady=10)


mahdode_ta_lable = tk.Label(box_jostojo_malk3,text="تا",bg="#052340", fg="#FFFFFF",font=("Shabnam", 13))
mahdode_ta_lable.pack(padx=5, pady=1)

mahdode_ta_entry = tk.Entry(box_jostojo_malk3,bg="#FFFFFF", fg="#000000",font=("Shabnam", 13))
mahdode_ta_entry.pack(padx=20,pady=10)


metraj_lable = tk.Label(box_jostojo_malk3,text=" متراژ",bg="#052340", fg="#FFFFFF",font=("Shabnam", 13))
metraj_lable.pack(padx=5, pady=1)
metraj_entry = tk.Entry(box_jostojo_malk3,bg="#FFFFFF", fg="#000000",font=("Shabnam", 13))
metraj_entry.pack(padx=20,pady=10)

address_lable = tk.Label(box_jostojo_malk3,text=" منطقه / آدرس",bg="#052340", fg="#FFFFFF",font=("Shabnam", 13))
address_lable.pack(padx=5, pady=1)
address_entry = tk.Entry(box_jostojo_malk3,bg="#FFFFFF", fg="#000000",font=("Shabnam", 13))
address_entry.pack(padx=20,pady=10)

# ---------------------دکمه جستجوی ملک-----------------
search_btn = tk.Button(frame_jostojo_melk_left, text="    جستجو    " , bg="#00BFFF", fg="#000000", font=("Shabnam", 13), command=search)
search_btn.pack(pady=10)
#=====================================================
#endregion
# ---------------------------باکس وسط - نمایش جستجوی --------------
#region
frame_list_amlack_centre = tk.LabelFrame(contant_frame, text="لیست املاک", bg="#052340",fg="#00BFFF", font=("Shabnam", 13))
frame_list_amlack_centre.pack(side="left", fill="both", expand=True, padx=4, pady=15)

columns = (
    "id","table_name","آدرس","قیمت","نوع ملک","نام مالک","نوع فایل")
tree = ttk.Treeview(frame_list_amlack_centre,columns=columns,show="headings")

for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=120)

# مخفی کردن ستون‌های id و table_name
tree.column("id", width=0, stretch=False)
tree.heading("id", text="")

tree.column("table_name", width=0, stretch=False)
tree.heading("table_name", text="")

tree.pack(fill="both", expand=True)

tree.bind("<Double-1>", show_details)

lable_list_amlack_centre = tk.Label(frame_list_amlack_centre,text="تمامی حقوق قانونی این نرم افزار متعلق به گروه نیواد است",bg="#052340", fg="white",font=("Shabnam", 10))
lable_list_amlack_centre.pack(padx=10)
#===================================================
#endregion
# --------------------باکس سمت راست - نمایش جزئیات فایل های موجود املاک---------------
#region
frame_joziat_amlack = tk.LabelFrame(contant_frame, text="جزئیات ملک", width=200, bg="#052340",fg="#00BFFF", font=("Shabnam", 13))
frame_joziat_amlack.pack(side="right", fill="both",expand=True, padx=6, pady=15)


malek_phone_number = tk.Label(frame_joziat_amlack,text="شماره مالک",bg="#052340", fg="#F7F7FA",font=("Shabnam", 13))
malek_phone_number.pack(padx=6,pady=4)

entry_malek_phone_number = tk.Entry(frame_joziat_amlack,bg="#FFFFFF", fg="#000000",font=("Shabnam", 13))
entry_malek_phone_number.pack(padx=20,pady=4)

metraj_lable_right= tk.Label(frame_joziat_amlack,text="متراژ ",bg="#052340", fg="#F7F7FA",font=("Shabnam", 13))
metraj_lable_right.pack(padx=6,pady=4)

metraj_lable_right_entry = tk.Entry(frame_joziat_amlack,bg="#FFFFFF", fg="#000000",font=("Shabnam", 13))
metraj_lable_right_entry.pack(padx=20,pady=4)

options_text = tk.Label(frame_joziat_amlack,text="امکانات فایل",bg="#052340", fg="#F7F7FA",font=("Shabnam", 13))
options_text.pack(padx=6,pady=4)

options_text_entry = tk.Text(frame_joziat_amlack,bg="#FFFFFF", fg="#000000",font=("Shabnam", 13),height=10)
options_text_entry .pack(padx=20,pady=4)
options_text_entry.tag_configure("right", justify="right")

edit_btn=tk.Button(frame_joziat_amlack,text="ویرایش",command=open_edit,bg="#00BFFF", fg="#000000", font=("Shabnam", 13) )
edit_btn.pack(padx=20,pady=4)
#=======================================================
#endregion
#-------------------------باکس های نوع ثبتی فایل ها----------------------
#-------------------نوع انتخاب ثبتی فایل برای پنجره های رهن و اجاره--------------
#region
box_rehn_ejareh=tk.Toplevel(root)
box_rehn_ejareh.title("انتخاب نوع ملک رهن و اجاره")
box_rehn_ejareh.geometry("500x270")
box_rehn_ejareh.withdraw()
box_rehn_ejareh.configure(bg="#052340")

ejareh_radio_value = tk.IntVar(value=0)

ejareh_maskoni_radio = tk.Radiobutton(box_rehn_ejareh,text="ثبت فایل مسکونی",background="#052340",fg="#00BFFF",variable=ejareh_radio_value,value=0,font=("Shabnam", 11))
ejareh_maskoni_radio.place(x=330, y=50)

ejareh_edari_radio = tk.Radiobutton( box_rehn_ejareh,text="ثبت فایل اداری/تجاری",
    bg="#052340",fg="#00BFFF",
    variable=ejareh_radio_value,
    value=2,
    font=("Shabnam", 11))
ejareh_edari_radio.place(x=330, y=90)

ejareh_bagh_radio = tk.Radiobutton(box_rehn_ejareh,text="ثبت فایل باغ/زمین",bg="#052340",fg="#00BFFF",
    variable=ejareh_radio_value,
    value=4,
    font=("Shabnam", 11))
ejareh_bagh_radio.place(x=330, y=130)

ejareh_kargah_radio = tk.Radiobutton(box_rehn_ejareh,text="ثبت فایل کارگاه",bg="#052340",fg="#00BFFF",variable=ejareh_radio_value,value=6,
    font=("Shabnam", 11))
ejareh_kargah_radio.place(x=330, y=170)

back_to_home_box_ejareh=tk.Button(box_rehn_ejareh,text="بازگشت",bg="#00BFFF",fg="#000000",width=12,height=2,command=back_rehn_ejareh_exit)
back_to_home_box_ejareh.place(x=190,y=210)


zakhire_radio_box_ejareh=tk.Button(box_rehn_ejareh,text="ادامه",bg="#00BFFF",fg="#000000",width=12,height=2,command=sabt_radio_rehn)
zakhire_radio_box_ejareh.place(x=50,y=210)

box_rehn_ejareh.protocol("WM_DELETE_WINDOW", lambda: None)
box_rehn_ejareh.resizable(False, False) 

#=====================================================================
#endregion
#----------------------------------نوع انتخاب ثبتی فایل برای پنجره های فروش-----------------
#region
box_forosh=tk.Toplevel(root)
box_forosh.title("انتخاب نوع ملک فروش")
box_forosh.geometry("500x270")
box_forosh.withdraw()
box_forosh.configure(bg="#052340")

# یک متغیر مشترک برای همه رادیوباتن‌ها
forosh_radio_value = tk.IntVar(value=0)  # مقدار پیش‌فرض -1 یعنی هیچکدام انتخاب نشده

forosh_maskoni_radio = tk.Radiobutton(box_forosh, value=0, text="ثبت فایل مسکونی", background="#052340",fg="#00BFFF", variable=forosh_radio_value, font=("Shabnam",11))
forosh_maskoni_radio.place(x=330,y=50)

forosh_edari_radio = tk.Radiobutton(box_forosh, value=2, text="ثبت فایل اداری/تجاری",bg="#052340",fg="#00BFFF", variable=forosh_radio_value, font=("Shabnam",11))
forosh_edari_radio.place(x=330,y=90)

forosh_bagh_radio = tk.Radiobutton(box_forosh, value=4, text="ثبت فایل باغ/زمین",bg="#052340",fg="#00BFFF", variable=forosh_radio_value, font=("Shabnam",11))
forosh_bagh_radio.place(x=330,y=130)

forosh_kargah_radio = tk.Radiobutton(box_forosh, value=6, text="ثبت فایل کارگاه",bg="#052340",fg="#00BFFF", variable=forosh_radio_value, font=("Shabnam",11))
forosh_kargah_radio.place(x=330,y=170)


back_to_home_box_forosh=tk.Button(box_forosh,text="بازگشت",bg="#00BFFF",fg="#000000",width=12,height=2,command=back_forosh_exit)
back_to_home_box_forosh.place(x=190,y=210)

zakhire_radio_box_forosh=tk.Button(box_forosh,text="ادامه",bg="#00BFFF",fg="#000000",width=12,height=2,command=sabt_radio_frosh)
zakhire_radio_box_forosh.place(x=50,y=210)

box_forosh.protocol("WM_DELETE_WINDOW", lambda: None)
box_forosh.resizable(False, False) 




#======================================================================
#endregion
#----------------------------------نوع انتخاب ثبتی فایل برای پنجره های درخواستی-----------------
#region
box_darkhast=tk.Toplevel(root)
box_darkhast.title("انتخاب نوع ملک درخواستی")
box_darkhast.geometry("500x270")
box_darkhast.withdraw()
box_darkhast.configure(bg="#052340")

# یک متغیر مشترک برای همه رادیوباتن‌ها
darkhast_radio_value = tk.IntVar(value=0)  # مقدار پیش‌فرض -1 یعنی هیچکدام انتخاب نشده

darkhast_maskoni_radio = tk.Radiobutton(box_darkhast, value=0, text="ثبت درخواست مسکونی", background="#052340",fg="#00BFFF", variable=darkhast_radio_value, font=("Shabnam",11))
darkhast_maskoni_radio.place(x=295,y=50)

darkhast_edari_radio = tk.Radiobutton(box_darkhast, value=2, text="ثبت درخواست اداری/تجاری",bg="#052340",fg="#00BFFF", variable=darkhast_radio_value, font=("Shabnam",11))
darkhast_edari_radio.place(x=295,y=90)

darkhast_bagh_radio = tk.Radiobutton(box_darkhast, value=4, text="ثبت درخواست باغ/زمین",bg="#052340",fg="#00BFFF", variable=darkhast_radio_value, font=("Shabnam",11))
darkhast_bagh_radio.place(x=295,y=130)

darkhast_kargah_radio = tk.Radiobutton(box_darkhast, value=6, text="ثبت درخواست کارگاه",bg="#052340",fg="#00BFFF", variable=darkhast_radio_value, font=("Shabnam",11))
darkhast_kargah_radio.place(x=295,y=170)


back_to_home_box_darkhast=tk.Button(box_darkhast,text="بازگشت",bg="#00BFFF",fg="#000000",width=12,height=2,command=back_darkhast_exit)
back_to_home_box_darkhast.place(x=190,y=210)

zakhire_radio_box_darkhast=tk.Button(box_darkhast,text="ادامه",bg="#00BFFF",fg="#000000",width=12,height=2,command=sabt_radio_darkhast)
zakhire_radio_box_darkhast.place(x=50,y=210)

box_darkhast.protocol("WM_DELETE_WINDOW", lambda: None)
box_darkhast.resizable(False, False)
#endregion 
#---------------------------------نوع انتخاب ثبتی فایل برای  پنجره خروجی گزارش------------
#region
box_gozaresh=tk.Toplevel(root)
box_gozaresh.title("انتخاب نوع ملکی گزارش")
box_gozaresh.geometry("500x270")
box_gozaresh.withdraw()
box_gozaresh.configure(bg="#052340")

# یک متغیر مشترک برای همه رادیوباتن‌ها
gozaresh_radio_value = tk.IntVar(value=0)  # مقدار پیش‌فرض -1 یعنی هیچکدام انتخاب نشده

gozaresh_maskoni_radio = tk.Radiobutton(box_gozaresh, value=0, text="گزارش مسکونی", background="#052340",fg="#00BFFF", variable=gozaresh_radio_value, font=("Shabnam",11))
gozaresh_maskoni_radio.place(x=295,y=50)

gozaresh_edari_radio = tk.Radiobutton(box_gozaresh, value=2, text="گزارش اداری/تجاری",bg="#052340",fg="#00BFFF", variable=gozaresh_radio_value, font=("Shabnam",11))
gozaresh_edari_radio.place(x=295,y=90)

gozaresh_bagh_radio = tk.Radiobutton(box_gozaresh, value=4, text="گزارش باغ/زمین",bg="#052340",fg="#00BFFF", variable=gozaresh_radio_value, font=("Shabnam",11))
gozaresh_bagh_radio.place(x=295,y=130)

gozaresh_kargah_radio = tk.Radiobutton(box_gozaresh, value=6, text="گزارش کارگاه",bg="#052340",fg="#00BFFF", variable=gozaresh_radio_value, font=("Shabnam",11))
gozaresh_kargah_radio.place(x=295,y=170)


back_to_home_box_gozaresh=tk.Button(box_gozaresh,text="بازگشت",bg="#00BFFF",fg="#000000",width=12,height=2,command=back_gozaresh_exit)
back_to_home_box_gozaresh.place(x=190,y=210)

zakhire_radio_box_gozaresh=tk.Button(box_gozaresh,text="ادامه",bg="#00BFFF",fg="#000000",width=12,height=2,command=sabt_radio_gozaresh)
zakhire_radio_box_gozaresh.place(x=50,y=210)

box_gozaresh.protocol("WM_DELETE_WINDOW", lambda: None)
box_gozaresh.resizable(False, False)
#endregion
#----------------------------------نوع انتخاب ثبتی فایل برای پنجره های مشارکت-----------------
#region
box_mosharekat=tk.Toplevel(root)
box_mosharekat.title("انتخاب نوع ملک مشارکت")
box_mosharekat.geometry("350x300")
box_mosharekat.withdraw()
box_mosharekat.configure(bg="#052340")

# یک متغیر مشترک برای همه رادیوباتن‌ها
mosharekat_radio_value = tk.IntVar(value=0)  # مقدار پیش‌فرض -1 یعنی هیچکدام انتخاب نشده

mosharekat_sakht_radio = tk.Radiobutton(box_mosharekat, value=0, text="مشارکت در ساخت", background="#052340",fg="#00BFFF", variable=mosharekat_radio_value, font=("Shabnam",11))
mosharekat_sakht_radio.place(x=200,y=50)

mosharekat_pishforosh_radio = tk.Radiobutton(box_mosharekat, value=2, text="پیش فروش",
bg="#052340",fg="#00BFFF", variable=mosharekat_radio_value, font=("Shabnam",11))
mosharekat_pishforosh_radio.place(x=200,y=90)

back_to_home_box_mosharekat=tk.Button(box_mosharekat,text="بازگشت",bg="#00BFFF",fg="#000000",width=10,height=2,command=back_mosharekat_exit)
back_to_home_box_mosharekat.place(x=190,y=210)

zakhire_radio_box_mosharekat=tk.Button(box_mosharekat,text="ادامه",bg="#00BFFF",fg="#000000",width=10,height=2,command=sabt_radio_mosharekat)
zakhire_radio_box_mosharekat.place(x=50,y=210)

box_mosharekat.protocol("WM_DELETE_WINDOW", lambda: None)
box_mosharekat.resizable(False, False)
#endregion
#-----------------پنجره های ثبتی بخش رهن و اجاره----------------------------
#--------------------------پنجره اجاره مسکونی----------------
#region
ejareh_maskoni_window = tk.Toplevel(root)
ejareh_maskoni_window.title("رهن و اجاره مسکونی")
ejareh_maskoni_window.geometry("1200x700")
ejareh_maskoni_window.configure(bg="#052340")
ejareh_maskoni_window.withdraw()

#------------------کادر اجاره مسکونی------------#

frame_up_right_ejareh_maskoni = tk.Frame(ejareh_maskoni_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=500,height=300)
frame_up_right_ejareh_maskoni.configure(bg="#052340")
frame_up_right_ejareh_maskoni.place(x=670,y=90)

frame_up_left_ejareh_maskoni= tk.Frame(ejareh_maskoni_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=300)
frame_up_left_ejareh_maskoni.configure(bg="#052340")
frame_up_left_ejareh_maskoni.place(x=10,y=90)

frame_midde_right_ejareh_maskoni= tk.Frame(ejareh_maskoni_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=500,height=150)
frame_midde_right_ejareh_maskoni.configure(bg="#052340")
frame_midde_right_ejareh_maskoni.place(x=670,y=410)

frame_midde_left_ejareh_maskoni= tk.Frame(ejareh_maskoni_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=150)
frame_midde_left_ejareh_maskoni.configure(bg="#052340")
frame_midde_left_ejareh_maskoni.place(x=10,y=410)

frame_down_ejareh_maskoni= tk.Frame(ejareh_maskoni_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=1160,height=100)
frame_down_ejareh_maskoni.configure(bg="#052340")
frame_down_ejareh_maskoni.place(x=10,y=580)


title_label_up1_ejareh_maskoni = tk.Label(ejareh_maskoni_window,text="رهن و اجاره مسکونی",bg="#052340",fg="#00BFFF",font=("Shabnam", 16))
title_label_up1_ejareh_maskoni.place(x=555, y=17)   

title_label_up2_ejareh_maskoni = tk.Label(ejareh_maskoni_window,text="ثبت اطلاعات اجاره مسکونی",bg="#052340",fg="#ffffff",font=("Shabnam", 11))
title_label_up2_ejareh_maskoni.place(x=555, y=45)

label_up_right_ejareh_maskoni=tk.Label(ejareh_maskoni_window,text="اطلاعات ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_right_ejareh_maskoni.place(x=800,y=73)

label_up_left_ejareh_maskoni=tk.Label(ejareh_maskoni_window,text="تصویر ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_left_ejareh_maskoni.place(x=100,y=73)

label_midde_right_ejareh_maskoni=tk.Label(ejareh_maskoni_window,text="اطلاعات معامله",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_right_ejareh_maskoni.place(x=800,y=395)

label_midde_left_ejareh_maskoni=tk.Label(ejareh_maskoni_window,text="اطلاعات مالک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_left_ejareh_maskoni.place(x=100,y=395)

label_down_ejareh_maskoni=tk.Label(ejareh_maskoni_window,text="امکانات",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_down_ejareh_maskoni.place(x=100,y=565)


#-------------------------------فریم بالا سمت راست----------------------------

melk_type_ejareh_maskoni_lable = tk.Label(frame_up_right_ejareh_maskoni, text="نوع ملک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
melk_type_ejareh_maskoni_lable.place(x=465, y=40, anchor="e")

melk_type_ejareh_maskoni_entry = tk.Entry(frame_up_right_ejareh_maskoni, bg="#06294B", fg="#ffffff", font=("Shabnam", 10), justify="center")
melk_type_ejareh_maskoni_entry.insert(0, "اجاره مسکونی")
melk_type_ejareh_maskoni_entry.config(state="readonly",readonlybackground="#06294B",fg="#ffffff")
melk_type_ejareh_maskoni_entry.place(x=18, y=30, width=350, height=25)


sal_sakht_ejareh_maskoni_lable = tk.Label(frame_up_right_ejareh_maskoni, text="سال ساخت", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
sal_sakht_ejareh_maskoni_lable.place(x=465, y=80, anchor="e")

sal_sakht_ejareh_maskoni_entry = tk.Entry(frame_up_right_ejareh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
sal_sakht_ejareh_maskoni_entry.place(x=18, y=70, width=350, height=25)

metraj_ejareh_maskoni_lable = tk.Label(frame_up_right_ejareh_maskoni, text="متراژ", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
metraj_ejareh_maskoni_lable.place(x=465, y=120,anchor="e")

metraj_ejareh_maskoni_entry = tk.Entry(frame_up_right_ejareh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
metraj_ejareh_maskoni_entry.place(x=18, y=110, width=350, height=25)

tabaghe_ejare_maskoni_lable = tk.Label(frame_up_right_ejareh_maskoni, text="طبقه", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
tabaghe_ejare_maskoni_lable.place(x=465, y=160, anchor="e")

tabaghe_ejareh_maskoni_entry = tk.Entry(frame_up_right_ejareh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
tabaghe_ejareh_maskoni_entry.place(x=18, y=150, width=350, height=25)

vahed_ejareh_maskoni_lable = tk.Label(frame_up_right_ejareh_maskoni, text="واحد", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
vahed_ejareh_maskoni_lable.place(x=465, y=200, anchor="e")

vahed_ejareh_maskoni_entry = tk.Entry(frame_up_right_ejareh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
vahed_ejareh_maskoni_entry.place(x=18, y=190, width=350, height=25)

otagh_ejare_maskoni_lable = tk.Label(frame_up_right_ejareh_maskoni, text="اتاق", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
otagh_ejare_maskoni_lable.place(x=465, y=240, anchor="e")

otagh_ejareh_maskoni_entry = tk.Entry(frame_up_right_ejareh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
otagh_ejareh_maskoni_entry.place(x=18 ,y=230, width=350, height=25)

#--------------------فریم چپ بالا---------------------------
photo_lbl2_ejare_maskoni = tk.Label(frame_up_left_ejareh_maskoni, text="[تصویر ملک]", bg="#FFFFFF", width=79, height=15,relief="solid")
photo_lbl2_ejare_maskoni.place(x=40, y=10)

add_img_btn_ejare_maskoni = tk.Button(frame_up_left_ejareh_maskoni, text="افزودن تصویر", bg="#00BFFF", fg="#ffffff",command=open_file, height=2,width=13)
add_img_btn_ejare_maskoni.place(x=240, y=250)
#--------------------------فریم راست وسط---------------------
gheimat_pish_ejare_maskoni_lable = tk.Label(frame_midde_right_ejareh_maskoni, text="مبلغ پیش", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
gheimat_pish_ejare_maskoni_lable.place(x=465, y=30, anchor="e")

gheimat_pish_ejare_maskoni_entry = tk.Entry(frame_midde_right_ejareh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
gheimat_pish_ejare_maskoni_entry.place(x=18, y=20, width=350, height=25)

gheimat_ejare_ejare_maskoni_lable = tk.Label(frame_midde_right_ejareh_maskoni, text="مبلغ اجاره", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
gheimat_ejare_ejare_maskoni_lable.place(x=465, y=80, anchor="e")

gheimat_ejare_ejare_maskoni_entry = tk.Entry(frame_midde_right_ejareh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
gheimat_ejare_ejare_maskoni_entry.place(x=18, y=70, width=350, height=25)

addrres_ejareh_maskoni_lable = tk.Label(frame_midde_right_ejareh_maskoni, text="آدرس", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
addrres_ejareh_maskoni_lable.place(x=465, y=125, anchor="e")

addrres_ejareh_maskoni_entry = tk.Text(frame_midde_right_ejareh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
addrres_ejareh_maskoni_entry.place(x=18, y=115, width=350, height=25)
#------------------------------------فریم چپ وسط-----------------------
name_malek_ejareh_maskoni_lable = tk.Label(frame_midde_left_ejareh_maskoni,text="نام مالک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
name_malek_ejareh_maskoni_lable.place(x=600, y=30,anchor="e")

name_malek_ejareh_maskoni_entry = tk.Entry(frame_midde_left_ejareh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
name_malek_ejareh_maskoni_entry.place(x=30, y=20, width=350, height=25)

shomareh_malek_ejareh_maskoni_lable = tk.Label(frame_midde_left_ejareh_maskoni, text="شماره مالک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
shomareh_malek_ejareh_maskoni_lable.place(x=600, y=80,anchor="e")

shomareh_malek_ejareh_maskoni_entry = tk.Entry(frame_midde_left_ejareh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
shomareh_malek_ejareh_maskoni_entry.place(x=30, y=70, width=350, height=25)
#---------------------------------فریم پایین--------------------------------
parking_ejareh_maskoni_var=tk.IntVar(value=0)
anbari_ejareh_maskoni_var=tk.IntVar(value=0)
asansor_ejareh_maskoni_var=tk.IntVar(value=0)


parking_checkbutton_btn_ejareh_maskoni = tk.Checkbutton(frame_down_ejareh_maskoni, image=parking_pic,variable=parking_ejareh_maskoni_var,bg="#052340")
parking_checkbutton_btn_ejareh_maskoni.place(x=1050, y=10)
parking_ch_btn_ejareh_maskoni_label=tk.Label(frame_down_ejareh_maskoni,text="پارکینگ", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
parking_ch_btn_ejareh_maskoni_label.place(x=1055,y=60)

asansor_checkbutton_btn_ejareh_maskoni = tk.Checkbutton(frame_down_ejareh_maskoni, image=elvator_pic,variable=asansor_ejareh_maskoni_var, bg="#052340")
asansor_checkbutton_btn_ejareh_maskoni.place(x=950, y=10)
asansor_ch_btn_ejareh_maskoni_label=tk.Label(frame_down_ejareh_maskoni,text="آسانسور", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
asansor_ch_btn_ejareh_maskoni_label.place(x=955,y=60)


anbari_checkbutton_btn_ejareh_maskoni = tk.Checkbutton(frame_down_ejareh_maskoni, image=warehouse_pic,variable=anbari_ejareh_maskoni_var,bg="#052340")
anbari_checkbutton_btn_ejareh_maskoni.place(x=850, y=10)
anbari_checkbuton_ejareh_maskoni_label=tk.Label(frame_down_ejareh_maskoni,text="انباری", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
anbari_checkbuton_ejareh_maskoni_label.place(x=855,y=60)

sarmaesh_ejareh_maskoni = tk.Label(frame_down_ejareh_maskoni, text="سرمایش", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
sarmaesh_ejareh_maskoni.place(x=650, y=15)
sarmaesh_ejareh_maskoni_combo = ttk.Combobox(frame_down_ejareh_maskoni)
sarmaesh_ejareh_maskoni_combo["values"] = ("ندارد", "پنکه سقفی", "کولر آبی", "کولر گازی ", "آبی/گازی")
sarmaesh_ejareh_maskoni_combo["state"] = "readonly"
sarmaesh_ejareh_maskoni_combo.configure(justify="center")
sarmaesh_ejareh_maskoni_combo.place(x=475, y=15)

garmaesh_ejareh_maskoni = tk.Label(frame_down_ejareh_maskoni, text="گرمایش", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
garmaesh_ejareh_maskoni.place(x=650, y=45)
garmaesh_ejareh_maskoni_combo = ttk.Combobox(frame_down_ejareh_maskoni)
garmaesh_ejareh_maskoni_combo["values"] = ("ندارد", "بخاری", " شوفاژ", "گرمایش از کف ")
garmaesh_ejareh_maskoni_combo["state"] = "readonly"
garmaesh_ejareh_maskoni_combo.configure(justify="center")
garmaesh_ejareh_maskoni_combo.place(x=475, y=45)

kaf_ejareh_maskoni = tk.Label(frame_down_ejareh_maskoni, text="کف", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
kaf_ejareh_maskoni.place(x=350, y=15)
kaf_ejareh_maskoni_combo = ttk.Combobox(frame_down_ejareh_maskoni)
kaf_ejareh_maskoni_combo["values"] = ("سرامیک", "موزاییک", "پارکت")
kaf_ejareh_maskoni_combo["state"] = "readonly"
kaf_ejareh_maskoni_combo.configure(justify="center")
kaf_ejareh_maskoni_combo.place(x=150, y=15)


toilet_ejareh_maskoni = tk.Label(frame_down_ejareh_maskoni, text="سرویس بهداشتی", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
toilet_ejareh_maskoni.place(x=330, y=45)
toilet_ejareh_maskoni_combo = ttk.Combobox(frame_down_ejareh_maskoni)
toilet_ejareh_maskoni_combo["values"] = ("ایرانی", "فرنگی", "هردو")
toilet_ejareh_maskoni_combo["state"] = "readonly"
toilet_ejareh_maskoni_combo.configure(justify="center")
toilet_ejareh_maskoni_combo.place(x=150, y=45)


back_to_home_ejareh_maskoni=tk.Button(ejareh_maskoni_window,text="بازگشت",bg="#052340", fg="#ffffff",width=10,height=1,command=back_home_ejareh_maskoni)
back_to_home_ejareh_maskoni.place(x=400,y=30)

save_button_ejareh_maskooni=tk.Button(ejareh_maskoni_window,text="ذخیره",bg="#00BFFF", fg="#ffffff",width=10,height=1,command=sabt_ejareh_maskoni)
save_button_ejareh_maskooni.place(x=300,y=30)

delete_btn_ejareh_maskoni=tk.Button(ejareh_maskoni_window,text="حذف",command=delete_ejareh_maskoni,bg="#8B0000",fg="#ffffff",height=1,width=10 )
delete_btn_ejareh_maskoni.place_forget()

edit_btn_ejareh_maskoni=tk.Button(ejareh_maskoni_window,text="ثبت ویرایش",command=update_ejareh_maskoni,bg="#00BFFF", fg="#ffffff",width=10,height=1,)
edit_btn_ejareh_maskoni.place_forget()

error_lable_sal_sakht_ejareh_maskoni= tk.Label(ejareh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_sal_sakht_ejareh_maskoni.place(x=900 , y=20)

error_lable_metraj_ejareh_maskoni= tk.Label(ejareh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_metraj_ejareh_maskoni.place(x=900 , y=20)

error_lable_tabaghe_ejareh_maskoni= tk.Label(ejareh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_tabaghe_ejareh_maskoni.place(x=900 , y=20)

error_lable_vahed_ejareh_maskoni= tk.Label(ejareh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_vahed_ejareh_maskoni.place(x=900 , y=20)

error_lable_otagh_ejareh_maskoni= tk.Label(ejareh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_otagh_ejareh_maskoni.place(x=900 , y=20)

error_lable_gheimat_pish_ejareh_maskoni= tk.Label(ejareh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_gheimat_pish_ejareh_maskoni.place(x=900 , y=20)

error_lable_gheimat_ejareh_ejareh_maskoni= tk.Label(ejareh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_gheimat_ejareh_ejareh_maskoni.place(x=900 , y=20)

error_lable_addrres_ejareh_maskoni= tk.Label(ejareh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_addrres_ejareh_maskoni.place(x=850 , y=20)

error_lable_name_malek_ejareh_maskoni= tk.Label(ejareh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_name_malek_ejareh_maskoni.place(x=850 , y=20)

error_lable_shomareh_malek_ejareh_maskoni= tk.Label(ejareh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_shomareh_malek_ejareh_maskoni.place(x=850 , y=20)

ejareh_maskoni_window.protocol("WM_DELETE_WINDOW", lambda: None)
ejareh_maskoni_window.resizable(False, False)
#endregion
#------------------------پنجره اجاره اداری/تجاری--------------------
#region
ejareh_edari_tejari_window = tk.Toplevel(root)
ejareh_edari_tejari_window.title("اجاره اداری و تجاری")
ejareh_edari_tejari_window.configure(bg="#052340")
ejareh_edari_tejari_window.geometry("1200x700")
ejareh_edari_tejari_window.withdraw()

#----------------------کادر اجاره اداری و تجاری------------------#

frame_up_right_ejareh_edari_tejari= tk.Frame(ejareh_edari_tejari_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=500,height=300)
frame_up_right_ejareh_edari_tejari.configure(bg="#052340")
frame_up_right_ejareh_edari_tejari.place(x=670,y=90)

frame_up_left_ejareh_edari_tejari= tk.Frame(ejareh_edari_tejari_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=300)
frame_up_left_ejareh_edari_tejari.configure(bg="#052340")
frame_up_left_ejareh_edari_tejari.place(x=10,y=90)

frame_midde_right_ejareh_edari_tejari= tk.Frame(ejareh_edari_tejari_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=500,height=150)
frame_midde_right_ejareh_edari_tejari.configure(bg="#052340")
frame_midde_right_ejareh_edari_tejari.place(x=670,y=410)

frame_midde_left_ejareh_edari_tejari= tk.Frame(ejareh_edari_tejari_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=150)
frame_midde_left_ejareh_edari_tejari.configure(bg="#052340")
frame_midde_left_ejareh_edari_tejari.place(x=10,y=410)

frame_down_ejareh_edari_tejari= tk.Frame(ejareh_edari_tejari_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=1160,height=100)
frame_down_ejareh_edari_tejari.configure(bg="#052340")
frame_down_ejareh_edari_tejari.place(x=10,y=580)

title_label_up1 = tk.Label(ejareh_edari_tejari_window,text="اجاره اداری و تجاری",bg="#052340",fg="#00BFFF",font=("Shabnam", 16))
title_label_up1.place(x=570, y=17)

title_label_up2 = tk.Label(ejareh_edari_tejari_window,text="ثبت اطلاعات اجاره اداری و تجاری",bg="#052340",fg="#ffffff",font=("Shabnam", 11))
title_label_up2.place(x=555, y=45)

label_up_right_ejareh_edari_tejari=tk.Label(ejareh_edari_tejari_window,text="اطلاعات ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_right_ejareh_edari_tejari.place(x=800,y=73)

label_up_left_ejareh_edari_tejari=tk.Label(ejareh_edari_tejari_window,text="تصویر ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_left_ejareh_edari_tejari.place(x=100,y=73)

label_midde_right_ejareh_edari_tejari=tk.Label(ejareh_edari_tejari_window,text="اطلاعات معامله",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_right_ejareh_edari_tejari.place(x=800,y=395)

label_midde_left_ejareh_edari_tejari=tk.Label(ejareh_edari_tejari_window,text="اطلاعات مالک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_left_ejareh_edari_tejari.place(x=100,y=395)

label_down_ejareh_edari_tejari=tk.Label(ejareh_edari_tejari_window,text="امکانات",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_down_ejareh_edari_tejari.place(x=100,y=565)

#-------------------------------------فریم بالا سمت راست------------------------------

melk_type_ejareh_edari_tejari_lable=tk.Label(frame_up_right_ejareh_edari_tejari, text="نوع ملک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
melk_type_ejareh_edari_tejari_lable.place(x=465,y=40, anchor="e")

melk_type_ejareh_edari_tejari_entry=tk.Entry(frame_up_right_ejareh_edari_tejari, bg="#06294B", fg="#FFFFFF", font=("Shabnam", 10), justify="center")
melk_type_ejareh_edari_tejari_entry.insert(0,"اجاره اداری و تجاری")
melk_type_ejareh_edari_tejari_entry.config(state="readonly",readonlybackground="#06294B",fg="#FFFFFF")
melk_type_ejareh_edari_tejari_entry.place(x=18, y=30, width=350, height=25)

sal_sakht_ejareh_edari_tejari_lable=tk.Label(frame_up_right_ejareh_edari_tejari, text="سال ساخت", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
sal_sakht_ejareh_edari_tejari_lable.place(x=465, y=80, anchor="e")

sal_sakht_ejareh_edari_tejari_entry=tk.Entry(frame_up_right_ejareh_edari_tejari, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
sal_sakht_ejareh_edari_tejari_entry.place(x=18, y=70, width=350, height=25)

metraj_melk_ejareh_edari_tejari_lable=tk.Label(frame_up_right_ejareh_edari_tejari, text=" متراژ ", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
metraj_melk_ejareh_edari_tejari_lable.place(x=465, y=120, anchor="e")

metraj_melk_ejareh_edari_tejari_entry=tk.Entry(frame_up_right_ejareh_edari_tejari, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
metraj_melk_ejareh_edari_tejari_entry.place(x=18, y=110, width=350, height=25)

tabaghe_ejareh_edari_tejari_lable=tk.Label(frame_up_right_ejareh_edari_tejari, text="طبقه", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
tabaghe_ejareh_edari_tejari_lable.place(x=465, y=160, anchor="e")

tabaghe_ejareh_edari_tejari_entry=tk.Entry(frame_up_right_ejareh_edari_tejari, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
tabaghe_ejareh_edari_tejari_entry.place(x=18, y=150, width=350, height=25)

vahed_ejareh_edari_tejari_lable=tk.Label(frame_up_right_ejareh_edari_tejari, text="واحد", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
vahed_ejareh_edari_tejari_lable.place(x=465, y=200, anchor="e")

vahed_ejareh_edari_tejari_entry=tk.Entry(frame_up_right_ejareh_edari_tejari, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
vahed_ejareh_edari_tejari_entry.place(x=18, y=190, width=350, height=25)

mablagh_pish_ejareh_edari_tejari_lable=tk.Label(frame_up_right_ejareh_edari_tejari, text="مبلغ و ودیعه", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
mablagh_pish_ejareh_edari_tejari_lable.place(x=465, y=240, anchor="e")

mablagh_pish_ejareh_edari_tejari_entry=tk.Entry(frame_up_right_ejareh_edari_tejari, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
mablagh_pish_ejareh_edari_tejari_entry.place(x=18, y=230, width=350, height=25)

#----------------------------فریم بالا سمت چپ----------------------------------

photo_lbl2_ejareh_edari_tejari =tk.Label(frame_up_left_ejareh_edari_tejari, text="[تصویر ملک]", bg="#ffffff", width=79, height=15)
photo_lbl2_ejareh_edari_tejari.place(x=40 ,y=10)

add_img_btn_ejareh_edari_tejari =tk.Button(frame_up_left_ejareh_edari_tejari, text="افزودن تصویر", bg="#00BFFF", fg="#ffffff",command=open_file,height=2,width=13)
add_img_btn_ejareh_edari_tejari.place(x=240, y=250)

#----------------------------------فریم وسط سمت راست--------------------------

mablagh_ejare_ejareh_edari_tejari_lable=tk.Label(frame_midde_right_ejareh_edari_tejari, text=" مبلغ اجاره ", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
mablagh_ejare_ejareh_edari_tejari_lable.place(x=465, y=30, anchor="e")

mablagh_ejare_ejareh_edari_tejari_entry=tk.Entry(frame_midde_right_ejareh_edari_tejari, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
mablagh_ejare_ejareh_edari_tejari_entry.place(x=18, y=20, width=350, height=25)

addrres_ejareh_edari_tejari_lable=tk.Label(frame_midde_right_ejareh_edari_tejari, text="آدرس", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
addrres_ejareh_edari_tejari_lable.place(x=465, y=80, anchor="e")

addrres_ejareh_edari_tejari_entry=tk.Text(frame_midde_right_ejareh_edari_tejari, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
addrres_ejareh_edari_tejari_entry.place(x=18, y=70, width=350, height=50)

#-------------------------------------فریم وسط سمت چپ---------------------------

name_malek_ejareh_edari_tejari_lable=tk.Label(frame_midde_left_ejareh_edari_tejari, text="نام مالک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
name_malek_ejareh_edari_tejari_lable.place(x=600, y=30,anchor="e")

name_malek_ejareh_edari_tejari_entry=tk.Entry(frame_midde_left_ejareh_edari_tejari, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
name_malek_ejareh_edari_tejari_entry.place(x=30, y=20, width=350, height=25)

shomareh_malek_ejareh_edari_tejari_lable= tk.Label(frame_midde_left_ejareh_edari_tejari, text="شماره مالک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
shomareh_malek_ejareh_edari_tejari_lable.place(x=600, y=80,anchor="e")

shomareh_malek_ejareh_edari_tejari_entry=tk.Entry(frame_midde_left_ejareh_edari_tejari, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
shomareh_malek_ejareh_edari_tejari_entry.place(x=30, y=70, width=350, height=25)

#-----------------------------------------فریم پایین--------------------------------

parking_ejareh_edari_tejari_var=tk.IntVar(value=0)
anbari_ejareh_edari_tejari_var=tk.IntVar(value=0)
asansor_ejareh_edari_tejari_var=tk.IntVar(value=0)

parking_ch_btn_ejareh_edari_tejari=tk.Checkbutton(frame_down_ejareh_edari_tejari,variable=parking_ejareh_edari_tejari_var,image=parking_pic, bg="#052340")
parking_ch_btn_ejareh_edari_tejari.place(x=1050, y=10)
parking_btn_ejareh_edari_tejari=tk.Label(frame_down_ejareh_edari_tejari,text="پارکینگ", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
parking_btn_ejareh_edari_tejari.place(x=1055,y=60)

asansor_ch_btn_ejareh_edari_tejari=tk.Checkbutton(frame_down_ejareh_edari_tejari,variable=asansor_ejareh_edari_tejari_var,image=elvator_pic,background="#052340")
asansor_ch_btn_ejareh_edari_tejari.place(x=950, y=10)
asansor_btn_ejareh_edari_tejari=tk.Label(frame_down_ejareh_edari_tejari,text="آسانسور", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
asansor_btn_ejareh_edari_tejari.place(x=955,y=60)

anbari_ch_btn_ejareh_edari_tejari=tk.Checkbutton(frame_down_ejareh_edari_tejari,variable=anbari_ejareh_edari_tejari_var,image=warehouse_pic,background="#052340")
anbari_ch_btn_ejareh_edari_tejari.place(x=850, y=10)
anbari_btn_ejareh_edari_tejari=tk.Label(frame_down_ejareh_edari_tejari,text="انباری", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
anbari_btn_ejareh_edari_tejari.place(x=855,y=60)

ab_va_gaz_emkanat_ejareh_edari_tejari=tk.Label(frame_down_ejareh_edari_tejari, text="وضعیت آب و گاز", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
ab_va_gaz_emkanat_ejareh_edari_tejari.place(x=650, y=15)

ab_va_gaz_combo_emkanat_ejareh_edari_tejari=ttk.Combobox(frame_down_ejareh_edari_tejari)
ab_va_gaz_combo_emkanat_ejareh_edari_tejari["state"]=["readonly"]
ab_va_gaz_combo_emkanat_ejareh_edari_tejari.configure(justify="center")
ab_va_gaz_combo_emkanat_ejareh_edari_tejari["values"] = ("آب و گاز دارد","فقط گاز دارد","فقط آب دارد ")
ab_va_gaz_combo_emkanat_ejareh_edari_tejari.place(x=475, y=15)

sarmayesh_emkanat_ejareh_edari_tejari=tk.Label(frame_down_ejareh_edari_tejari, text="وضعیت سرمایش", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
sarmayesh_emkanat_ejareh_edari_tejari.place(x=350, y=45)

sarmayesh_combo_emkanat_ejareh_edari_tejari=ttk.Combobox(frame_down_ejareh_edari_tejari)
sarmayesh_combo_emkanat_ejareh_edari_tejari["state"]=["readonly"]
sarmayesh_combo_emkanat_ejareh_edari_tejari.configure(justify="center")
sarmayesh_combo_emkanat_ejareh_edari_tejari["values"] = ("پنکه سقفی","کولر آبی","کولر گازی ","ندارد")
sarmayesh_combo_emkanat_ejareh_edari_tejari.place(x=150, y=45)

garmayesh_emkanat_ejareh_edari_tejari=tk.Label(frame_down_ejareh_edari_tejari, text="وضعیت گرمایش", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
garmayesh_emkanat_ejareh_edari_tejari.place(x=350, y=15)

garmayesh_combo_emkanat_ejareh_edari_tejari=ttk.Combobox(frame_down_ejareh_edari_tejari)
garmayesh_combo_emkanat_ejareh_edari_tejari["values"] = ("بخاری","شوفاژ","ندارد")
garmayesh_combo_emkanat_ejareh_edari_tejari["state"]=["readonly"]
garmayesh_combo_emkanat_ejareh_edari_tejari.place(x=150, y=15)

back_to_home_ejareh_edari_tejari=tk.Button(ejareh_edari_tejari_window,text="بازگشت",bg="#052340",fg="#ffffff",width=10,height=1,command=back_home_ejareh_edari_tejari)
back_to_home_ejareh_edari_tejari.place(x=400,y=30)

zakhire_ejareh_edari_tejari=tk.Button(ejareh_edari_tejari_window,text="ذخیره",bg="#00BFFF",fg="#ffffff",width=10,height=1,command=sabt_ejareh_edari_tejari)
zakhire_ejareh_edari_tejari.place(x=300,y=30)

delete_btn_ejareh_edari_tejari=tk.Button(ejareh_edari_tejari_window,text="حذف",command=delete_ejareh_edari_tejari,bg="#8B0000",fg="#ffffff",height=1,width=10 )
delete_btn_ejareh_edari_tejari.place_forget()

edit_btn_ejareh_edari_tejari=tk.Button(ejareh_edari_tejari_window,text="ثبت ویرایش",command=update_ejareh_edari_tejari,bg="#00BFFF", fg="#ffffff",width=10,height=1,)
edit_btn_ejareh_edari_tejari.place_forget()

ejareh_edari_tejari_window.protocol("WM_DELETE_WINDOW", lambda: None)
ejareh_edari_tejari_window.resizable(False, False)
#endregion
#-------------------پنجره اجاره باغ/زمین------------------------
#region
ejareh_bagh_zamin_window = tk.Toplevel(root)
ejareh_bagh_zamin_window.title(" اجاره باغ و زمین")
ejareh_bagh_zamin_window.geometry("1200x700")
ejareh_bagh_zamin_window.configure(bg="#052340")
ejareh_bagh_zamin_window.withdraw()

#---------------------کادر اجاره باغ و زمین---------------------#
frame_up_right_ejareh_bagh_zamin= tk.Frame(ejareh_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=520,height=300)
frame_up_right_ejareh_bagh_zamin.configure(bg="#052340")
frame_up_right_ejareh_bagh_zamin.place(x=670,y=75)


frame_up_left_ejareh_bagh_zamin= tk.Frame(ejareh_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=300)
frame_up_left_ejareh_bagh_zamin.configure(bg="#052340")
frame_up_left_ejareh_bagh_zamin.place(x=10,y=75)

frame_midde_right_ejareh_bagh_zamin= tk.Frame(ejareh_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=520,height=150)
frame_midde_right_ejareh_bagh_zamin.configure(bg="#052340")
frame_midde_right_ejareh_bagh_zamin.place(x=670,y=390)

frame_midde_left_ejareh_bagh_zamin= tk.Frame(ejareh_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=150)
frame_midde_left_ejareh_bagh_zamin.configure(bg="#052340")
frame_midde_left_ejareh_bagh_zamin.place(x=10,y=390)

frame_down_ejareh_bagh= tk.Frame(ejareh_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=1180,height=130)
frame_down_ejareh_bagh.configure(bg="#052340")
frame_down_ejareh_bagh.place(x=10,y=555)

frame_down_ejareh_zamin=tk.Frame(ejareh_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=1180,height=130)
frame_down_ejareh_zamin.configure(bg="#052340")
frame_down_ejareh_zamin.place_forget()

title_label_ejareh_bagh_zamin_up1 = tk.Label(ejareh_bagh_zamin_window,text="اجاره باغ و زمین",bg="#052340",fg="#00BFFF",font=("Shabnam", 16))
title_label_ejareh_bagh_zamin_up1.place(x=600, y=10)

title_label_ejareh_bagh_zamin_up2 = tk.Label(ejareh_bagh_zamin_window,text="ثبت اطلاعات اجاره باغ و زمین",bg="#052340",fg="#ffffff",font=("Shabnam", 11))
title_label_ejareh_bagh_zamin_up2.place(x=585, y=40)

label_up_right_ejareh_bagh_zamin=tk.Label(ejareh_bagh_zamin_window,text="اطلاعات ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_right_ejareh_bagh_zamin.place(x=760,y=60)

label_up_left_ejareh_bagh_zamin=tk.Label(ejareh_bagh_zamin_window,text="تصویر ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_left_ejareh_bagh_zamin.place(x=100,y=60)

label_midde_right_ejareh_bagh_zamin=tk.Label(ejareh_bagh_zamin_window,text="اطلاعات معامله",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_right_ejareh_bagh_zamin.place(x=800,y=375)

label_midde_left_ejareh_bagh_zamin=tk.Label(ejareh_bagh_zamin_window,text="اطلاعات مالک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_left_ejareh_bagh_zamin.place(x=100,y=375)

label_down_ejareh_bagh_zamin=tk.Label(ejareh_bagh_zamin_window,text="امکانات",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_down_ejareh_bagh_zamin.place(x=100,y=545)

#------------------------------------فریم بالا سمت راست--------------------------------------
melk_type_ejareh_bagh_zamin_lable=tk.Label(frame_up_right_ejareh_bagh_zamin, text="نوع ملک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
melk_type_ejareh_bagh_zamin_lable.place(x=490, y=40, anchor="e")

melk_type_ejareh_bagh_zamin_entry=tk.Entry(frame_up_right_ejareh_bagh_zamin, bg="#06294B", fg="#FFFFFF", font=("Shabnam", 10), justify="center")
melk_type_ejareh_bagh_zamin_entry.place(x=28, y=30, width=350, height=25)
melk_type_ejareh_bagh_zamin_entry.insert(0,"اجاره باغ و زمین")
melk_type_ejareh_bagh_zamin_entry.config(state="readonly",readonlybackground="#06294B",fg="#ffffff")

metraj_zamin_ejareh_bagh_zamin_lable=tk.Label(frame_up_right_ejareh_bagh_zamin, text="متراژ", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
metraj_zamin_ejareh_bagh_zamin_lable.place(x=490, y=80, anchor="e")

metraj_zamin_ejareh_bagh_zamin_entry=tk.Entry(frame_up_right_ejareh_bagh_zamin, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
metraj_zamin_ejareh_bagh_zamin_entry.place(x=28, y=70, width=350, height=25)

bagh_type_lable=tk.Label(frame_up_right_ejareh_bagh_zamin, text="کاربری زمین", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
bagh_type_lable.place(x=490, y=120, anchor="e")

bagh_type_combo=ttk.Combobox(frame_up_right_ejareh_bagh_zamin,state="readonly")
bagh_type_combo["values"]=("باغ","زمین کشاورزی")
bagh_type_combo.set("باغ")
bagh_type_combo.configure(justify="center")
bagh_type_combo.place(x=28, y=110, width=350, height=25)
bagh_type_combo.bind("<<ComboboxSelected>>",change_bagh_zamin1)

time_bagh_ejareh_bagh_zamin_lable=tk.Label(frame_up_right_ejareh_bagh_zamin,text="مدت اجاره",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
time_bagh_ejareh_bagh_zamin_lable.place(x=490, y=180, anchor="e")

bagh_time_combo=ttk.Combobox(frame_up_right_ejareh_bagh_zamin,state="readonly")
bagh_time_combo["values"]=("بلندمدت","کوتاه مدت","فصلی","سالانه")
bagh_time_combo.set("فصلی")
bagh_time_combo.configure(justify="center")
bagh_time_combo.place(x=28, y=170, width=350, height=25)
#-------------------------------------------فریم بالا سمت چپ-------------------
photo_lbl2_ejareh_bagh_zamin = tk.Label(frame_up_left_ejareh_bagh_zamin, text="[تصویر ملک]", bg="#FFFFFF", width=79, height=15)
photo_lbl2_ejareh_bagh_zamin.place(x=40, y=10)

add_img_btn_ejareh_bagh_zamin = tk.Button(frame_up_left_ejareh_bagh_zamin, text="افزودن تصویر", bg="#00BFFF", fg="black",command=open_file,height=2,width=13)
add_img_btn_ejareh_bagh_zamin.place(x=240, y=250)
#-----------------------------فریم وسط سمت راست-------------------------------

bagh_gheimat_ejareh_bagh_zamin_lable=tk.Label(frame_midde_right_ejareh_bagh_zamin, text="ودیعه", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
bagh_gheimat_ejareh_bagh_zamin_lable.place(x=490, y=20, anchor="e")

bagh_gheimat_ejareh_bagh_zamin_entry=tk.Entry(frame_midde_right_ejareh_bagh_zamin, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
bagh_gheimat_ejareh_bagh_zamin_entry.place(x=28, y=13, width=350, height=25)

bagh_gheimat_har_metr_ejareh_bagh_zamin_lable=tk.Label(frame_midde_right_ejareh_bagh_zamin, text="اجاره ماهانه", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
bagh_gheimat_har_metr_ejareh_bagh_zamin_lable.place(x=490, y=65, anchor="e")

bagh_gheimat_har_metr_ejareh_bagh_zamin_entry=tk.Entry(frame_midde_right_ejareh_bagh_zamin, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
bagh_gheimat_har_metr_ejareh_bagh_zamin_entry.place(x=28, y=55, width=350, height=25)

bagh_loctaion_lable=tk.Label(frame_midde_right_ejareh_bagh_zamin, text=" منطقه و آدرس ", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
bagh_loctaion_lable.place(x=490, y=110, anchor="e")

bagh_loctaion_entry=tk.Text(frame_midde_right_ejareh_bagh_zamin, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
bagh_loctaion_entry.place(x=28, y=95, width=350, height=40)

#-------------------------------------فریم وسط سمت چپ---------------------------

name_malek_bagh_zamin_lable=tk.Label(frame_midde_left_ejareh_bagh_zamin, text="نام مالک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
name_malek_bagh_zamin_lable.place(x=600, y=30, anchor="e")

name_malek_bagh_zamin_entry=tk.Entry(frame_midde_left_ejareh_bagh_zamin, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
name_malek_bagh_zamin_entry.place(x=30, y=20, width=350, height=25)

number_malek_bagh_zamin_lable=tk.Label(frame_midde_left_ejareh_bagh_zamin, text="شماره مالک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
number_malek_bagh_zamin_lable.place(x=600, y=80 ,anchor="e")

number_malek_bagh_zamin_entry=tk.Entry(frame_midde_left_ejareh_bagh_zamin, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
number_malek_bagh_zamin_entry.place(x=30, y=70, width=350, height=25)
#-----------------------------------فریم پایین------------------------------------------

abyari=tk.Label(frame_down_ejareh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="نوع آبیاری")
abyari.place(x=1080, y=20)

abyari_combo=ttk.Combobox(frame_down_ejareh_bagh,width=12)
abyari_combo["values"]=("سطحی","بارانی","قطره ای","تحت فشار")
abyari_combo.set("سطحی")
abyari_combo.configure(justify="center")
abyari_combo["state"]=["readonly"]
abyari_combo.place(x=970, y=20)

metraj_tree=tk.Label(frame_down_ejareh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="متراژ درخت کاری")
metraj_tree.place(x=1080, y=60)

metraj_tree_entry=tk.Entry(frame_down_ejareh_bagh,width=10,bg="#ffffff",fg="#000000")
metraj_tree_entry.place(x=1000, y=60)

tree_count=tk.Label(frame_down_ejareh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="تعداد درخت")
tree_count.place(x=1080, y=100)

tree_count_entry=tk.Entry(frame_down_ejareh_bagh,width=10,bg="#ffffff",fg="#000000")
tree_count_entry.place(x=1000, y=100)

type_tree=tk.Label(frame_down_ejareh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="نوع درخت")
type_tree.place(x=870, y=20)

type_tree_combo=ttk.Combobox(frame_down_ejareh_bagh,width=12)
type_tree_combo["values"]=(" ","پسته","بادام","گردو","شلیل","هلو","سیب","انگور","انجیر","زردالو","گیلاس","آلبالو")
type_tree_combo["state"]=["readonly"]
type_tree_combo.configure(justify="center")
type_tree_combo.set("گردو")
type_tree_combo.place(x=770, y=20)

add_tree_button=tk.Button(frame_down_ejareh_bagh,text="افزودن درخت",font=("Shabnam",9),command=add_tree,bg="#00BFFF",width=10,height=1)
add_tree_button.place(x=870, y=60)

label_result_add=tk.Text(frame_down_ejareh_bagh,width=11,wrap="word",height=4,font=("Shabnam",9))
label_result_add.place(x=770,y=50)

chah_bagh_ejareh_var= tk.IntVar(value=0)
chah_bagh=tk.Checkbutton(frame_down_ejareh_bagh,variable=chah_bagh_ejareh_var,text="چاه",font=("Shabnam",9),background="#052340",fg="#00BFFF")
chah_bagh.place(x=685, y=20)

estakhr_bagh_var=tk.IntVar(value=0)
estakhr_bagh=tk.Checkbutton(frame_down_ejareh_bagh,variable=estakhr_bagh_var,text="استخر",font=("Shabnam",9),background="#052340",fg="#00BFFF")
estakhr_bagh.place(x=600, y=20)

bargh_bagh_var=tk.IntVar(value=0)
bargh_bagh=tk.Checkbutton(frame_down_ejareh_bagh,variable=bargh_bagh_var,text="برق کشی",font=("Shabnam",9),background="#052340",fg="#00BFFF")
bargh_bagh.place(x=685, y=80)

gaz_bagh_var=tk.IntVar(value=0)
gaz_bagh=tk.Checkbutton(frame_down_ejareh_bagh,variable=gaz_bagh_var,text="گاز کشی",font=("Shabnam",9),background="#052340",fg="#00BFFF")
gaz_bagh.place(x=685,y=50)

divar_ejareh_bagh_var=tk.IntVar(value=0)
divar_ejareh_bagh_zamin=tk.Checkbutton(frame_down_ejareh_bagh,variable=divar_ejareh_bagh_var,text="دیوار کشی",background="#052340",fg="#00BFFF",font=("Shabnam",9))
divar_ejareh_bagh_zamin.place(x=600, y=50)

var0=tk.IntVar(value=0)#چک باتن پیش فرض تیک نخورده باشه

room_bagh_checkbutton=tk.Checkbutton(frame_down_ejareh_bagh,variable=var0,image=warehouse_pic,background="#052340",text="ساختمان",command=home_true_false1)
room_bagh_checkbutton.place(x=600, y=70)
otagh_check_btn_ejareh_bagh_zamin_label=tk.Label(frame_down_ejareh_bagh,text="ویلا", bg="#052340", fg="#ffffff", font=("Shabnam", 7), width=7)
otagh_check_btn_ejareh_bagh_zamin_label.place(x=620,y=110)

type_vila_ejareh_bagh_zamin=tk.Label(frame_down_ejareh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="نوع سازه")
type_vila_ejareh_bagh_zamin.place(x=490, y=20)

type_vila_ejareh_bagh_zamin_combo=ttk.Combobox(frame_down_ejareh_bagh,state="disabled",width=12)
type_vila_ejareh_bagh_zamin_combo["values"]=("آجری","بلوکی","کانکس","چوبی")
type_vila_ejareh_bagh_zamin_combo.set("آجری")
type_vila_ejareh_bagh_zamin_combo.configure(justify="center")
type_vila_ejareh_bagh_zamin_combo.place(x=380, y=20)

toilet_bagh=tk.Label(frame_down_ejareh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="سرویس بهداشتی")
toilet_bagh.place(x=490,y=60)

toilet_bagh_combo=ttk.Combobox(frame_down_ejareh_bagh,state="disabled",width=12)
toilet_bagh_combo["values"]=("ندارد","فرنگی","ایرانی","هردو")
toilet_bagh_combo.set("ندارد")
toilet_bagh_combo.configure(justify="center")
toilet_bagh_combo.place(x=380, y=60)

hamam_bagh=tk.Label(frame_down_ejareh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="حمام")
hamam_bagh.place(x=490, y=95)

hamam_bagh_combo=ttk.Combobox(frame_down_ejareh_bagh,state="disabled",width=12)
hamam_bagh_combo["values"]=("ندارد","دارد")
hamam_bagh_combo.set("ندارد")
hamam_bagh_combo.configure(justify="center")
hamam_bagh_combo.place(x=380, y=100)

sanad_bagh=tk.Label(frame_down_ejareh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="سند")
sanad_bagh.place(x=300, y=60)

sanad_bagh_combo=ttk.Combobox(frame_down_ejareh_bagh,width=12)
sanad_bagh_combo["values"]=("ندارد","تک برگ","قولنامه ای","مشاع")
sanad_bagh_combo.set("ندارد")
sanad_bagh_combo.configure(justify="center")
sanad_bagh_combo.place(x=180, y=60)

option_ejareh_bagh_zamin=tk.Label(frame_down_ejareh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="امکانات تفریحی")
option_ejareh_bagh_zamin.place(x=310, y=20)

option_ejareh_bagh_zamin_combo=ttk.Combobox(frame_down_ejareh_bagh,width=12)
option_ejareh_bagh_zamin_combo["values"]=("استخر","جکوزی","باربیکیو")
option_ejareh_bagh_zamin_combo.set("استخر")
option_ejareh_bagh_zamin_combo.configure(justify="center")
option_ejareh_bagh_zamin_combo.place(x=180, y=20)

add_option_button=tk.Button(frame_down_ejareh_bagh,text="افزودن امکانات",command=add_option,bg="#00BFFF",font=("Shabnam",9),width=10,height=1)
add_option_button.place(x=85, y=20)

label_result2_add=tk.Label(frame_down_ejareh_bagh,text="",width=10)
label_result2_add.place(x=5, y=20)

metraj_vila_bagh_lable=tk.Label(frame_down_ejareh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="متراژ سازه")
metraj_vila_bagh_lable.place(x=290, y=100)

metraj_vila_bagh_entry=tk.Entry(frame_down_ejareh_bagh,width=10,bg="#ffffff",fg="#000000",state="disabled")
metraj_vila_bagh_entry.place(x=230, y=100)

sal_sakht_vila_bagh_lable=tk.Label(frame_down_ejareh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="سال ساخت")
sal_sakht_vila_bagh_lable.place(x=80, y=60)

sal_sakht_vila_bagh_entry=tk.Entry(frame_down_ejareh_bagh,width=10,bg="#ffffff",fg="#000000",state="disabled")
sal_sakht_vila_bagh_entry.place(x=10, y=60)


mojavez_sakht_ejareh_bagh_var=tk.IntVar(value=0)
mojavez_sakht_ejareh_bagh_zamin=tk.Checkbutton(frame_down_ejareh_bagh,variable=mojavez_sakht_ejareh_bagh_var,text="مجوز ساختن",background="#052340",fg="#00BFFF",font=("Shabnam",9),state="disabled")
mojavez_sakht_ejareh_bagh_zamin.place(x=10, y=95)

mohavate_ejareh_bagh_var=tk.IntVar(value=0)
mohavate_ejareh_bagh_zamin=tk.Checkbutton(frame_down_ejareh_bagh,variable=mohavate_ejareh_bagh_var,text="محوطه سازی",background="#052340",fg="#00BFFF",font=("Shabnam",9),state="disabled")
mohavate_ejareh_bagh_zamin.place(x=120, y=95)

#-------------------تعویض کاربری به زمین در قسمت اجاره باغ/زمین--------------

karbari_ejareh_ejareh_bagh_zamin=tk.Label(frame_down_ejareh_zamin,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="نوع کاربری")
karbari_ejareh_ejareh_bagh_zamin.place(x=1000, y=20)

karbari_ejareh_ejareh_bagh_zamin_combo=ttk.Combobox(frame_down_ejareh_zamin)
karbari_ejareh_ejareh_bagh_zamin_combo["values"]=(" ","زراعی","باغی","گلخانه ای","دامداری ","مرغداری","دامداری و مرغداری","آیش")     
karbari_ejareh_ejareh_bagh_zamin_combo["state"]=["readonly"]                        
karbari_ejareh_ejareh_bagh_zamin_combo.set(" ")
karbari_ejareh_ejareh_bagh_zamin_combo.configure(justify="center")
karbari_ejareh_ejareh_bagh_zamin_combo.place(x=800, y=20)

khak_ejareh_ejareh_bagh_zamin=tk.Label(frame_down_ejareh_zamin,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="نوع خاک")
khak_ejareh_ejareh_bagh_zamin.place(x=1000, y=60)

khak_ejareh_ejareh_bagh_zamin_combo=ttk.Combobox(frame_down_ejareh_zamin)
khak_ejareh_ejareh_bagh_zamin_combo["values"]=(" ","رسی","شنی","لومی","رسی_شنی","شنی_لومی","رسی_لومی")
khak_ejareh_ejareh_bagh_zamin_combo["state"]=["readonly"]                              
khak_ejareh_ejareh_bagh_zamin_combo.set(" ")
khak_ejareh_ejareh_bagh_zamin_combo.configure(justify="center")
khak_ejareh_ejareh_bagh_zamin_combo.place(x=800, y=60)

ab_ejareh_ejareh_bagh_zamin=tk.Label(frame_down_ejareh_zamin,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="منبع آب")
ab_ejareh_ejareh_bagh_zamin.place(x=1000, y=100)

ab_ejareh_ejareh_bagh_zamin_combo=ttk.Combobox(frame_down_ejareh_zamin)
ab_ejareh_ejareh_bagh_zamin_combo["values"]=(" ","چاه","قنات","رودخانه","کانال آبیاری","چشمه","آب لوله کشی کشاورزی","تانکر","استخر") 
ab_ejareh_ejareh_bagh_zamin_combo["state"]=["readonly"] 
ab_ejareh_ejareh_bagh_zamin_combo.configure(justify="center")                   
ab_ejareh_ejareh_bagh_zamin_combo.set(" ")
ab_ejareh_ejareh_bagh_zamin_combo.place(x=800, y=100)

security_zamin_ejareh_var=tk.IntVar(value=0)
security_room_zamin_ejareh_bagh_zamin=tk.Checkbutton(frame_down_ejareh_zamin,variable=security_zamin_ejareh_var,text="اتاق نگهبان",background="#052340",fg="#00BFFF",font=("Shabnam",9))
security_room_zamin_ejareh_bagh_zamin.place(x=300, y=20)

bargh_tak_ejareh_zamin_var=tk.IntVar(value=0)
bargh_tak_faz_zamin_ejareh_bagh_zamin=tk.Checkbutton(frame_down_ejareh_zamin,variable=bargh_tak_ejareh_zamin_var,text="برق تک فاز",background="#052340",fg="#00BFFF",font=("Shabnam",9))
bargh_tak_faz_zamin_ejareh_bagh_zamin.place(x=500, y=20)

bargh_se_faz_ejareh_zamin_var=tk.IntVar(value=0)
bargh_se_faz_zamin_ejareh_bagh_zamin=tk.Checkbutton(frame_down_ejareh_zamin,variable=bargh_se_faz_ejareh_zamin_var,text="برق سه فاز",background="#052340",fg="#00BFFF",font=("Shabnam",9))
bargh_se_faz_zamin_ejareh_bagh_zamin.place(x=400, y=20)

anbar_ejareh_zamin_var=tk.IntVar(value=0)
anbar_zamin_ejareh_bagh_zamin=tk.Checkbutton(frame_down_ejareh_zamin,variable=anbar_ejareh_zamin_var,text="انبار/سوله",background="#052340",fg="#00BFFF",font=("Shabnam",9))
anbar_zamin_ejareh_bagh_zamin.place(x=500, y=60)

fans_ejareh_zamin_var=tk.IntVar(value=0)
fans_zamin_ejareh_bagh_zamin=tk.Checkbutton(frame_down_ejareh_zamin,variable=fans_ejareh_zamin_var,text="فنس/دیوار",background="#052340",fg="#00BFFF",font=("Shabnam",9))
fans_zamin_ejareh_bagh_zamin.place(x=400, y=60)

mojavaz_chah_ejareh_zamin_var=tk.IntVar(value=0)
mojavaz_chah_zamin_ejareh_bagh_zamin=tk.Checkbutton(frame_down_ejareh_zamin,variable=mojavaz_chah_ejareh_zamin_var,text="اجازه حفر چاه",background="#052340",fg="#00BFFF",font=("Shabnam",9))
mojavaz_chah_zamin_ejareh_bagh_zamin.place(x=300, y=60)


back_to_home_ejareh_bagh_zamin=tk.Button(ejareh_bagh_zamin_window,text="بازگشت",bg="#052340",fg="#ffffff",width=10,height=1,command=back_home_ejareh_bagh_zamin)
back_to_home_ejareh_bagh_zamin.place(x=300,y=30)

zakhire_ejareh_bagh_zamin=tk.Button(ejareh_bagh_zamin_window,text="ذخیره",bg="#00BFFF",fg="#ffffff",width=10,height=1,command=sabt_ejareh_bagh_zamin)
zakhire_ejareh_bagh_zamin.place(x=200,y=30)

ejareh_bagh_zamin_window.protocol("WM_DELETE_WINDOW", lambda: None)
ejareh_bagh_zamin_window.resizable(False, False)

#endregion
#-------------------پنجره اجاره کارگاه------------------------
#region
ejareh_karghah_window = tk.Toplevel(root)
ejareh_karghah_window.title(" اجاره کارگاه")
ejareh_karghah_window.geometry("1200x700")
ejareh_karghah_window.configure(bg="#052340")
ejareh_karghah_window.withdraw()

frame_up_right_ejareh_karghah = tk.Frame(ejareh_karghah_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=510,height=300)
frame_up_right_ejareh_karghah.configure(bg="#052340")
frame_up_right_ejareh_karghah.place(x=670,y=90)

frame_up_left_ejareh_karghah= tk.Frame(ejareh_karghah_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=300)
frame_up_left_ejareh_karghah.configure(bg="#052340")
frame_up_left_ejareh_karghah.place(x=10,y=90)

frame_midde_right_ejareh_karghah= tk.Frame(ejareh_karghah_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=510,height=150)
frame_midde_right_ejareh_karghah.configure(bg="#052340")
frame_midde_right_ejareh_karghah.place(x=670,y=410)

frame_midde_left_ejareh_karghah= tk.Frame(ejareh_karghah_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=150)
frame_midde_left_ejareh_karghah.configure(bg="#052340")
frame_midde_left_ejareh_karghah.place(x=10,y=410)

frame_down_ejareh_karghah= tk.Frame(ejareh_karghah_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=1170,height=85)
frame_down_ejareh_karghah.configure(bg="#052340")
frame_down_ejareh_karghah.place(x=10,y=580)

title_label_up1_ejareh_karghah= tk.Label(ejareh_karghah_window,text="رهن و اجاره کارگاه",bg="#052340",fg="#00BFFF",font=("Shabnam", 16))
title_label_up1_ejareh_karghah.place(x=570, y=17)   

title_label_up2_ejareh_karghah = tk.Label(ejareh_karghah_window,text="ثبت اطلاعات اجاره کارگاه",bg="#052340",fg="#ffffff",font=("Shabnam", 11))
title_label_up2_ejareh_karghah.place(x=565, y=45)

label_up_right_ejareh_karghah=tk.Label(ejareh_karghah_window,text="اطلاعات ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_right_ejareh_karghah.place(x=800,y=73)

label_up_left_ejareh_karghah=tk.Label(ejareh_karghah_window,text="تصویر ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_left_ejareh_karghah.place(x=100,y=73)

label_midde_right_ejareh_karghah=tk.Label(ejareh_karghah_window,text="اطلاعات معامله",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_right_ejareh_karghah.place(x=800,y=395)

label_midde_left_ejareh_karghah=tk.Label(ejareh_karghah_window,text="اطلاعات مالک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_left_ejareh_karghah.place(x=100,y=395)

label_down_ejareh_karghah=tk.Label(ejareh_karghah_window,text="امکانات",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_down_ejareh_karghah.place(x=100,y=565)


#-------------------------------فریم بالا سمت راست----------------------------

sal_sakht_ejareh_karghah_lable = tk.Label(frame_up_right_ejareh_karghah, text="سال ساخت", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
sal_sakht_ejareh_karghah_lable.place(x=465, y=80, anchor="e")

sal_sakht_ejareh_karghah_entry = tk.Entry(frame_up_right_ejareh_karghah, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
sal_sakht_ejareh_karghah_entry.place(x=18, y=70, width=350, height=25)

karbari_zamin_ejareh_karghah_lable = tk.Label(frame_up_right_ejareh_karghah, text="کاربری زمین", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
karbari_zamin_ejareh_karghah_lable.place(x=465, y=40, anchor="e")

karbari_zamin_ejareh_karghah_entry = tk.Entry(frame_up_right_ejareh_karghah, bg="#06294B", fg="#FFFFFF", font=("Shabnam", 10), justify="center")
karbari_zamin_ejareh_karghah_entry.insert(0, "اجاره کارگاه")
karbari_zamin_ejareh_karghah_entry.config(state="readonly",readonlybackground="#06294B",fg="#FFFFFF")
karbari_zamin_ejareh_karghah_entry.place(x=18, y=30, width=350, height=25)

metraj_ejareh_karghah_lable = tk.Label(frame_up_right_ejareh_karghah, text="متراژ", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
metraj_ejareh_karghah_lable.place(x=465, y=120,anchor="e")

metraj_ejareh_karghah_entry = tk.Entry(frame_up_right_ejareh_karghah, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
metraj_ejareh_karghah_entry.place(x=18, y=110, width=350, height=25)

time_ejate_ejareh_kargah_lable=tk.Label(frame_up_right_ejareh_karghah, text="مدت اجاره", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
time_ejate_ejareh_kargah_lable.place(x=465, y=160, anchor="e")

time_ejare_ejareh_kargah_combo=ttk.Combobox(frame_up_right_ejareh_karghah,state="readonly")
time_ejare_ejareh_kargah_combo["values"]=("بلندمدت","کوتاه مدت","فصلی","سالانه")
time_ejare_ejareh_kargah_combo.set("فصلی")
time_ejare_ejareh_kargah_combo.configure(justify="center")
time_ejare_ejareh_kargah_combo.place(x=18, y=152, width=350, height=25)


#--------------------فریم چپ بالا---------------------------
photo_lbl2_ejareh_karghah = tk.Label(frame_up_left_ejareh_karghah, text="[تصویر ملک]", bg="#FFFFFF", width=79, height=15,relief="solid")
photo_lbl2_ejareh_karghah.place(x=40, y=10)

add_img_btn_ejareh_karghah = tk.Button(frame_up_left_ejareh_karghah, text="افزودن تصویر", bg="#00BFFF", fg="#ffffff",command=open_file, height=2,width=13)
add_img_btn_ejareh_karghah.place(x=240, y=250)

#--------------------------فریم راست وسط---------------------
vadie_ejare_karghah_lable = tk.Label(frame_midde_right_ejareh_karghah, text="مبلغ پیش", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
vadie_ejare_karghah_lable.place(x=465, y=30, anchor="e")

vadie_ejare_karghah_entry = tk.Entry(frame_midde_right_ejareh_karghah, bg="#ffffff", fg="#ffffff", font=("Shabnam", 10))
vadie_ejare_karghah_entry.place(x=18, y=20, width=350, height=25)

gheimat_ejare_ejare_karghah_lable = tk.Label(frame_midde_right_ejareh_karghah, text="مبلغ اجاره", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
gheimat_ejare_ejare_karghah_lable.place(x=465, y=70, anchor="e")

gheimat_ejare_ejare_karghah_entry = tk.Entry(frame_midde_right_ejareh_karghah, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
gheimat_ejare_ejare_karghah_entry.place(x=18, y=55, width=350, height=25)

addrres_ejareh_karghah_lable = tk.Label(frame_midde_right_ejareh_karghah, text="آدرس", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=12)
addrres_ejareh_karghah_lable.place(x=465, y=110, anchor="e")
addrres_ejareh_karghah_entry = tk.Text(frame_midde_right_ejareh_karghah, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
addrres_ejareh_karghah_entry.place(x=18, y=90, width=350, height=50)

#------------------------------------فریم چپ وسط-----------------------
name_malek_ejareh_karghah_lable = tk.Label(frame_midde_left_ejareh_karghah,text="نام مالک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
name_malek_ejareh_karghah_lable.place(x=600, y=30,anchor="e")

name_malek_ejareh_karghah_entry = tk.Entry(frame_midde_left_ejareh_karghah, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
name_malek_ejareh_karghah_entry.place(x=30, y=20, width=350, height=25)

shomareh_malek_ejareh_karghah_lable = tk.Label(frame_midde_left_ejareh_karghah, text="شماره مالک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
shomareh_malek_ejareh_karghah_lable.place(x=600, y=80,anchor="e")

shomareh_malek_ejareh_karghah_entry = tk.Entry(frame_midde_left_ejareh_karghah, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
shomareh_malek_ejareh_karghah_entry.place(x=30, y=70, width=350, height=25)

#---------------------------------فریم پایین--------------------------------
sarmaesh_ejareh_karghah = tk.Label(frame_down_ejareh_karghah, text="سیستم سرمایش", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
sarmaesh_ejareh_karghah.place(x=485, y=15)
sarmaesh_ejareh_karghah_combo = ttk.Combobox(frame_down_ejareh_karghah)
sarmaesh_ejareh_karghah_combo["values"] = ("ندارد", "پنکه سقفی", "کولر ابی", "کولر گازی ", "ابی/گازی")
sarmaesh_ejareh_karghah_combo["state"] = "readonly"
sarmaesh_ejareh_karghah_combo.configure(justify="center")
sarmaesh_ejareh_karghah_combo.place(x=325, y=15)

garmaesh_ejareh_karghah = tk.Label(frame_down_ejareh_karghah, text="سیستم گرمایش", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
garmaesh_ejareh_karghah.place(x=485, y=45)
garmaesh_ejareh_karghah_combo = ttk.Combobox(frame_down_ejareh_karghah)
garmaesh_ejareh_karghah_combo["values"] = ("ندارد", "بخاری", " شوفاژ", "گرمایش از کف ")
garmaesh_ejareh_karghah_combo["state"] = "readonly"
garmaesh_ejareh_karghah_combo.configure(justify="center")
garmaesh_ejareh_karghah_combo.place(x=325, y=45)

toilet_ejareh_karghah = tk.Label(frame_down_ejareh_karghah, text="سرویس بهداشتی", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
toilet_ejareh_karghah.place(x=190, y=45)
toilet_ejareh_karghah_combo = ttk.Combobox(frame_down_ejareh_karghah)
toilet_ejareh_karghah_combo["values"] = ("ایرانی", "فرنگی", "هردو")
toilet_ejareh_karghah_combo["state"] = "readonly"
toilet_ejareh_karghah_combo.configure(justify="center")
toilet_ejareh_karghah_combo.place(x=20, y=45)

vaziat_bagh_ejareh_karghah=tk.Label(frame_down_ejareh_karghah,bg="#052340",fg="#ffffff",font=("Shabnam", 11),width=15,text="وضعیت برق")
vaziat_bagh_ejareh_karghah.place(x=190, y=15)

vaziat_bargh_ejareh_karghah_combo=ttk.Combobox(frame_down_ejareh_karghah)
vaziat_bargh_ejareh_karghah_combo["values"]=("","برق شهری","سه فاز","تک فاز")
vaziat_bargh_ejareh_karghah_combo.set("")
vaziat_bargh_ejareh_karghah_combo["state"]=["readonly"]
vaziat_bargh_ejareh_karghah_combo.configure(justify="center")
vaziat_bargh_ejareh_karghah_combo.place(x=20, y=15)

vaziat_ab_ejareh_karghah=tk.Label(frame_down_ejareh_karghah,bg="#052340",fg="#ffffff",width=13,text=" وضعیت آب",font=("Shabnam", 9))
vaziat_ab_ejareh_karghah.place(x=1064, y=13)

vaziat_ab_ejareh_karghah_combo=ttk.Combobox(frame_down_ejareh_karghah,width=24)
vaziat_ab_ejareh_karghah_combo["values"]=(""," آب  لوله کشی (بدون فشار) " ," آب لوله کشی (با موتور فشار) ","دارای منبع(با موتور فشار)","دارای منبع(بدون فشار)")
vaziat_ab_ejareh_karghah_combo.set("")
vaziat_ab_ejareh_karghah_combo["state"]=["readonly"]
vaziat_ab_ejareh_karghah_combo.configure(justify="center")
vaziat_ab_ejareh_karghah_combo.place(x=910, y=13)

abzar_ejareh_karghah=tk.Label(frame_down_ejareh_karghah,bg="#052340",fg="#ffffff",font=("Shabnam", 11),width=15,text=" ابزار صنعتی ")
abzar_ejareh_karghah.place(x=745, y=45)

abzaar_ejareh_karghah_combo=ttk.Combobox(frame_down_ejareh_karghah,width=15)
abzaar_ejareh_karghah_combo["values"]=("","(کارگاه خالی) بدون دستگاه ","دارای دستگاه")
abzaar_ejareh_karghah_combo.set("")
abzaar_ejareh_karghah_combo["state"]=["readonly"]
abzaar_ejareh_karghah_combo.configure(justify="center")
abzaar_ejareh_karghah_combo.place(x=620, y=45)

hamam_ejareh_karghah=tk.Label(frame_down_ejareh_karghah,bg="#052340",fg="#ffffff",font=("Shabnam", 9),width=13,text="حمام")
hamam_ejareh_karghah.place(x=1064, y=43)

hamam_ejareh_karghah__combo=ttk.Combobox(frame_down_ejareh_karghah,width=24)
hamam_ejareh_karghah__combo["values"]=("","ندارد","دارد")
hamam_ejareh_karghah__combo.set("")
hamam_ejareh_karghah__combo["state"]=["readonly"]
hamam_ejareh_karghah__combo.configure(justify="center")
hamam_ejareh_karghah__combo.place(x=910, y=45)

otagh_ejareh_karghah=tk.Label(frame_down_ejareh_karghah,bg="#052340",fg="#ffffff",font=("Shabnam", 11),width=17,text="اتاق رختکن")
otagh_ejareh_karghah.place(x=740, y=15)

otagh_ejareh_karghah_combo=ttk.Combobox(frame_down_ejareh_karghah,width=15)
otagh_ejareh_karghah_combo["values"]=("","ندارد","دارد")
otagh_ejareh_karghah_combo.set("")
otagh_ejareh_karghah_combo["state"]=["readonly"]
otagh_ejareh_karghah_combo.configure(justify="center")
otagh_ejareh_karghah_combo.place(x=620, y=15)

back_to_home_ejareh_karghah=tk.Button(ejareh_karghah_window,text="بازگشت",bg="#052340", fg="#ffffff",width=10,height=1,command=back_home_ejareh_karghah)
back_to_home_ejareh_karghah.place(x=300,y=30)

save_button_ejareh_mkarghah=tk.Button(ejareh_karghah_window,text="ذخیره",bg="#00BFFF", fg="#ffffff",width=10,height=1,command=sabt_ejareh_kargah)
save_button_ejareh_mkarghah.place(x=200,y=30)

ejareh_karghah_window.protocol("WM_DELETE_WINDOW", lambda: None)
ejareh_karghah_window.resizable(False, False)
#endregion
#---------------------------پنجره های ثبتی بخش فروش--------------------
#-------------------پنجره فروش مسکونی----------------------
#region
forosh_maskoni_window = tk.Toplevel(root)
forosh_maskoni_window.title("فروش مسکونی")
forosh_maskoni_window.configure(bg="#052340")
forosh_maskoni_window.geometry("1200x700")
forosh_maskoni_window.withdraw()

#------------------کادر فروش مسکونی-----------------------------#
frame_up_right_forosh_maskoni= tk.Frame(forosh_maskoni_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=500,height=300)
frame_up_right_forosh_maskoni.configure(bg="#052340")
frame_up_right_forosh_maskoni.place(x=670,y=90)


frame_up_left_forosh_maskoni= tk.Frame(forosh_maskoni_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=300)
frame_up_left_forosh_maskoni.configure(bg="#052340")
frame_up_left_forosh_maskoni.place(x=10,y=90)

frame_midde_right_forosh_maskoni= tk.Frame(forosh_maskoni_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=500,height=150)
frame_midde_right_forosh_maskoni.configure(bg="#052340")
frame_midde_right_forosh_maskoni.place(x=670,y=410)

frame_midde_left_forosh_maskoni= tk.Frame(forosh_maskoni_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=150)
frame_midde_left_forosh_maskoni.configure(bg="#052340")
frame_midde_left_forosh_maskoni.place(x=10,y=410)

frame_down_forosh_maskoni= tk.Frame(forosh_maskoni_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=1160,height=100)
frame_down_forosh_maskoni.configure(bg="#052340")
frame_down_forosh_maskoni.place(x=10,y=580)

title_label_up1 = tk.Label(forosh_maskoni_window,text="فروش مسکونی",bg="#052340",fg="#00BFFF",font=("Shabnam", 16))
title_label_up1.place(x=570, y=17)

title_label_up2 = tk.Label(forosh_maskoni_window,text="ثبت اطلاعات فروش مسکونی",bg="#052340",fg="#ffffff",font=("Shabnam", 11))
title_label_up2.place(x=555, y=45)

label_up_right_forosh_maskoni=tk.Label(forosh_maskoni_window,text="اطلاعات ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_right_forosh_maskoni.place(x=800,y=73)

label_up_left_forosh_maskoni=tk.Label(forosh_maskoni_window,text="تصویر ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_left_forosh_maskoni.place(x=100,y=73)

label_midde_right_forosh_maskoni=tk.Label(forosh_maskoni_window,text="اطلاعات معامله",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_right_forosh_maskoni.place(x=800,y=395)

label_midde_left_forosh_maskoni=tk.Label(forosh_maskoni_window,text="اطلاعات مالک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_left_forosh_maskoni.place(x=100,y=395)

label_down_forosh_maskoni=tk.Label(forosh_maskoni_window,text="امکانات",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_down_forosh_maskoni.place(x=100,y=565)
#-------------------------------------فریم بالا سمت راست------------------------------
melk_type_forosh_maskoni_lable = tk.Label(frame_up_right_forosh_maskoni, text="نوع ملک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
melk_type_forosh_maskoni_lable.place(x=465,y=40, anchor="e")

melk_type_forosh_maskoni_entry=tk.Entry(frame_up_right_forosh_maskoni, bg="#06294B", fg="#ffffff", font=("Shabnam", 10), justify="center")
melk_type_forosh_maskoni_entry.insert(0,"فروش مسکونی")
melk_type_forosh_maskoni_entry.config(state="readonly",readonlybackground="#06294B",fg="#ffffff")
melk_type_forosh_maskoni_entry.place(x=18, y=30, width=350, height=25)

sal_sakht_forosh_maskoni=tk.Label(frame_up_right_forosh_maskoni, text="سال ساخت", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
sal_sakht_forosh_maskoni.place(x=465, y=80, anchor="e")

sal_sakht_forosh_maskoni_entry=tk.Entry(frame_up_right_forosh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
sal_sakht_forosh_maskoni_entry.place(x=18, y=70, width=350, height=25)
sal_sakht_forosh_maskoni_entry.bind("<KeyRelease>",chck_sal_sakht)

metraj_forosh_maskoni=tk.Label(frame_up_right_forosh_maskoni, text=" متراژ ", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
metraj_forosh_maskoni.place(x=465, y=120, anchor="e")

metraj_forosh_maskoni_entry=tk.Entry(frame_up_right_forosh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
metraj_forosh_maskoni_entry.place(x=18, y=110, width=350, height=25)
metraj_forosh_maskoni_entry.bind("<KeyRelease>",chck_metraj)

tabaghe_forosh_maskoni= tk.Label(frame_up_right_forosh_maskoni, text="طبقه", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
tabaghe_forosh_maskoni.place(x=465, y=160, anchor="e")

tabaghe_forosh_maskoni_entry=tk.Entry(frame_up_right_forosh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
tabaghe_forosh_maskoni_entry.place(x=18, y=150, width=350, height=25)
tabaghe_forosh_maskoni_entry.bind("<KeyRelease>",chck_tabaghe)

vahed_forosh_maskoni=tk.Label(frame_up_right_forosh_maskoni, text="واحد", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
vahed_forosh_maskoni.place(x=465, y=200, anchor="e")

vahed_forosh_maskoni_entry=tk.Entry(frame_up_right_forosh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
vahed_forosh_maskoni_entry.place(x=18, y=190, width=350, height=25)
vahed_forosh_maskoni_entry.bind("<KeyRelease>",chck_vahed)

otagh_forosh_maskoni= tk.Label(frame_up_right_forosh_maskoni, text="اتاق", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
otagh_forosh_maskoni.place(x=465, y=240, anchor="e")

otagh_forosh_maskoni_entry=tk.Entry(frame_up_right_forosh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
otagh_forosh_maskoni_entry.place(x=18, y=230, width=350, height=25)
otagh_forosh_maskoni_entry.bind("<KeyRelease>",chck_otagh)

#----------------------------فریم بالا سمت چپ----------------------------------
photo_lbl2_forosh_maskoni = tk.Label(frame_up_left_forosh_maskoni, text="[تصویر ملک]", bg="#ffffff", width=79, height=15)
photo_lbl2_forosh_maskoni.place(x=40 ,y=10)

add_img_btn_forosh_maskoni = tk.Button(frame_up_left_forosh_maskoni, text="افزودن تصویر", bg="#00BFFF", fg="#ffffff",command=open_file,height=2,width=13)
add_img_btn_forosh_maskoni.place(x=240, y=250)
#----------------------------------فریم وسط سمت راست--------------------------
gheimat_kol_forosh_maskoni=tk.Label(frame_midde_right_forosh_maskoni, text=" قیمت کل ", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
gheimat_kol_forosh_maskoni.place(x=465, y=30, anchor="e")

gheimat_kol_forosh_maskoni_entry=tk.Entry(frame_midde_right_forosh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
gheimat_kol_forosh_maskoni_entry.place(x=18, y=20, width=350, height=25)
gheimat_kol_forosh_maskoni_entry.bind("<KeyRelease>",chck_gheimat_kol)

addrres_forosh_maskoni=tk.Label(frame_midde_right_forosh_maskoni, text="آدرس", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
addrres_forosh_maskoni.place(x=465, y=80, anchor="e")

addrres_forosh_maskoni_entry=tk.Text(frame_midde_right_forosh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
addrres_forosh_maskoni_entry.place(x=18, y=70, width=350, height=50)
addrres_forosh_maskoni_entry.bind("<KeyRelease>",chck_addrres)
#-------------------------------------فریم وسط سمت چپ---------------------------
name_malek_forosh_maskoni_lable = tk.Label(frame_midde_left_forosh_maskoni, text="نام مالک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
name_malek_forosh_maskoni_lable.place(x=600, y=30,anchor="e")

name_malek_forosh_maskoni_entry = tk.Entry(frame_midde_left_forosh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
name_malek_forosh_maskoni_entry.place(x=30, y=20, width=350, height=25)
name_malek_forosh_maskoni_entry.bind("<KeyRelease>",chck_name_malek)

shomareh_malek_forosh_maskoni_lable = tk.Label(frame_midde_left_forosh_maskoni, text="شماره مالک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
shomareh_malek_forosh_maskoni_lable.place(x=600, y=80,anchor="e")

shomareh_malek_forosh_maskoni_entry = tk.Entry(frame_midde_left_forosh_maskoni, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
shomareh_malek_forosh_maskoni_entry.place(x=30, y=70, width=350, height=25)
shomareh_malek_forosh_maskoni_entry.bind("<KeyRelease>",chck_shomareh_malek)

#-----------------------------------------فریم پایین--------------------------------

parking_forosh_maskoni_var=tk.IntVar(value=0)
anbari_forosh_maskoni_var=tk.IntVar(value=0)
asansor_forosh_maskoni_var=tk.IntVar(value=0)

parking_ch_btn_forosh_maskoni=tk.Checkbutton(frame_down_forosh_maskoni,variable=parking_forosh_maskoni_var,image=parking_pic, bg="#052340")
parking_ch_btn_forosh_maskoni.place(x=1050, y=10)
parking_ch_btn_forosh_maskoni_label=tk.Label(frame_down_forosh_maskoni,text="پارکینگ", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
parking_ch_btn_forosh_maskoni_label.place(x=1055,y=60)

asansor_ch_btn_forosh_maskoni=tk.Checkbutton(frame_down_forosh_maskoni,variable=asansor_forosh_maskoni_var,image=elvator_pic,background="#052340")
asansor_ch_btn_forosh_maskoni.place(x=950, y=10)
asansor_ch_btn_forosh_maskoni_label=tk.Label(frame_down_forosh_maskoni,text="آسانسور", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
asansor_ch_btn_forosh_maskoni_label.place(x=955,y=60)

anbari_checkbuton_forosh_maskoni=tk.Checkbutton(frame_down_forosh_maskoni,variable=anbari_forosh_maskoni_var,image=warehouse_pic,background="#052340")
anbari_checkbuton_forosh_maskoni.place(x=850, y=10)
anbari_checkbuton_forosh_maskoni_label=tk.Label(frame_down_forosh_maskoni,text="انباری", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
anbari_checkbuton_forosh_maskoni_label.place(x=855,y=60)

sarmaesh_forosh_maskoni=tk.Label(frame_down_forosh_maskoni, text="سرمایش", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
sarmaesh_forosh_maskoni.place(x=650, y=15)


sarmaesh_combo_forosh_maskoni=ttk.Combobox(frame_down_forosh_maskoni)
sarmaesh_combo_forosh_maskoni["state"]=["readonly"]
sarmaesh_combo_forosh_maskoni.configure(justify="center")
sarmaesh_combo_forosh_maskoni["values"] = ("ندارد","پنکه سقفی","کولر آبی","کولر گازی ","آبی/گازی")
sarmaesh_combo_forosh_maskoni.place(x=475, y=15)

garmaesh_forosh_maskoni=tk.Label(frame_down_forosh_maskoni, text="گرمایش", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
garmaesh_forosh_maskoni.place(x=650, y=45)


garmaesh_combo_forosh_maskoni=ttk.Combobox(frame_down_forosh_maskoni)
garmaesh_combo_forosh_maskoni["values"] = ("ندارد","بخاری"," شوفاژ","گرمایش از کف ")
garmaesh_combo_forosh_maskoni["state"]=["readonly"]
garmaesh_combo_forosh_maskoni.place(x=475, y=45)

kaf_forosh_maskoni= tk.Label(frame_down_forosh_maskoni, text="کف", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
kaf_forosh_maskoni.place(x=350, y=15)

kaf_combo_forosh_maskoni=ttk.Combobox(frame_down_forosh_maskoni)
kaf_combo_forosh_maskoni["state"]=["readonly"]
kaf_combo_forosh_maskoni["values"] = ("سرامیک","موزاییک","پارکت")
kaf_combo_forosh_maskoni.place(x=150, y=15)

toilet_forosh_maskoni=tk.Label(frame_down_forosh_maskoni,text="سرویس بهداشتی",background="#052340",fg="#ffffff",font=("Shabnam",11))
toilet_forosh_maskoni.place(x=330, y=45)

toilet_combo_forosh_maskoni=ttk.Combobox(frame_down_forosh_maskoni)
toilet_combo_forosh_maskoni["state"]=["readonly"]
toilet_combo_forosh_maskoni["values"] = ("ایرانی","فرنگی","هردو")
toilet_combo_forosh_maskoni.place(x=150, y=45)

back_to_home_forosh_maskoni=tk.Button(forosh_maskoni_window,text="بازگشت",bg="#052340", fg="#ffffff",width=10,height=1,command=back_home_forosh_maskoni)
back_to_home_forosh_maskoni.place(x=400,y=30)

zakhire_forosh_maskoni=tk.Button(forosh_maskoni_window,text="ذخیره",bg="#00BFFF", fg="#ffffff",width=10,height=1,command=sabt_forosh_maskoni)
zakhire_forosh_maskoni.place(x=300,y=30)

delete_btn_forosh_maskoni=tk.Button(forosh_maskoni_window,text="حذف",command=delete_forosh_maskoni,bg="#8B0000",fg="#ffffff",height=1,width=10 )
delete_btn_forosh_maskoni.place_forget()

edit_btn_forosh_maskoni=tk.Button(forosh_maskoni_window,text="ثبت ویرایش",command=update_forosh_maskoni,bg="#00BFFF", fg="#ffffff",width=10,height=1,)
edit_btn_forosh_maskoni.place_forget()
#------------------------------------ارور لیبل های فروش مسکونی--------------------
error_lable_sal_sakht_forosh_maskoni= tk.Label(forosh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_sal_sakht_forosh_maskoni.place(x=900 , y=20)

error_lable_metraj_forosh_maskoni= tk.Label(forosh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_metraj_forosh_maskoni.place(x=900 , y=20)

error_lable_tabaghe_forosh_maskoni= tk.Label(forosh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_tabaghe_forosh_maskoni.place(x=900 , y=20)

error_lable_vahed_forosh_maskoni= tk.Label(forosh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_vahed_forosh_maskoni.place(x=900 , y=20)

error_lable_otagh_forosh_maskoni= tk.Label(forosh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_otagh_forosh_maskoni.place(x=900 , y=20)

error_lable_gheimat_kol_forosh_maskoni= tk.Label(forosh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_gheimat_kol_forosh_maskoni.place(x=900 , y=20)

error_lable_addrres_forosh_maskoni= tk.Label(forosh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_addrres_forosh_maskoni.place(x=835 , y=20)

error_lable_name_malek_forosh_maskoni= tk.Label(forosh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_name_malek_forosh_maskoni.place(x=835 , y=20)

error_lable_shomareh_malek_forosh_maskoni= tk.Label(forosh_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_shomareh_malek_forosh_maskoni.place(x=900 , y=20)

forosh_maskoni_window.protocol("WM_DELETE_WINDOW", lambda: None)
forosh_maskoni_window.resizable(False, False)
#endregion
#-----------------پنجره فروش اداری/تجاری-------------------
#region
forosh_edari_tejari_window = tk.Toplevel(root)
forosh_edari_tejari_window.title("فروش اداری و تجاری")
forosh_edari_tejari_window.configure(bg="#052340")
forosh_edari_tejari_window.geometry("1200x700")
forosh_edari_tejari_window.withdraw()

#----------------------کادر فروش اداری و تجاری------------------#

frame_up_right_forosh_edari_tejari= tk.Frame(forosh_edari_tejari_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=500,height=300)
frame_up_right_forosh_edari_tejari.configure(bg="#052340")
frame_up_right_forosh_edari_tejari.place(x=670,y=90)

frame_up_left_forosh_edari_tejari= tk.Frame(forosh_edari_tejari_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=300)
frame_up_left_forosh_edari_tejari.configure(bg="#052340")
frame_up_left_forosh_edari_tejari.place(x=10,y=90)

frame_midde_right_forosh_edari_tejari= tk.Frame(forosh_edari_tejari_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=500,height=150)
frame_midde_right_forosh_edari_tejari.configure(bg="#052340")
frame_midde_right_forosh_edari_tejari.place(x=670,y=410)

frame_midde_left_forosh_edari_tejari= tk.Frame(forosh_edari_tejari_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=150)
frame_midde_left_forosh_edari_tejari.configure(bg="#052340")
frame_midde_left_forosh_edari_tejari.place(x=10,y=410)

frame_down_forosh_edari_tejari= tk.Frame(forosh_edari_tejari_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=1160,height=100)
frame_down_forosh_edari_tejari.configure(bg="#052340")
frame_down_forosh_edari_tejari.place(x=10,y=580)

title_label_up1 = tk.Label(forosh_edari_tejari_window,text="فروش اداری و تجاری",bg="#052340",fg="#00BFFF",font=("Shabnam", 16))
title_label_up1.place(x=570, y=17)

title_label_up2 = tk.Label(forosh_edari_tejari_window,text="ثبت اطلاعات فروش اداری و تجاری",bg="#052340",fg="#ffffff",font=("Shabnam", 11))
title_label_up2.place(x=555, y=45)

label_up_right_forosh_edari_tejari=tk.Label(forosh_edari_tejari_window,text="اطلاعات ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_right_forosh_edari_tejari.place(x=800,y=73)

label_up_left_forosh_edari_tejari=tk.Label(forosh_edari_tejari_window,text="تصویر ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_left_forosh_edari_tejari.place(x=100,y=73)

label_midde_right_forosh_edari_tejari=tk.Label(forosh_edari_tejari_window,text="اطلاعات معامله",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_right_forosh_edari_tejari.place(x=800,y=395)

label_midde_left_forosh_edari_tejari=tk.Label(forosh_edari_tejari_window,text="اطلاعات مالک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_left_forosh_edari_tejari.place(x=100,y=395)

label_down_forosh_edari_tejari=tk.Label(forosh_edari_tejari_window,text="امکانات",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_down_forosh_edari_tejari.place(x=100,y=565)

#-------------------------------------فریم بالا سمت راست------------------------------

melk_type_forosh_edari_tejari_lable = tk.Label(frame_up_right_forosh_edari_tejari, text="نوع ملک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
melk_type_forosh_edari_tejari_lable.place(x=465,y=40, anchor="e")

melk_type_forosh_edari_tejari_entry=tk.Entry(frame_up_right_forosh_edari_tejari, bg="#06294B", fg="#FFFFFF", font=("Shabnam", 10), justify="center")
melk_type_forosh_edari_tejari_entry.insert(0,"فروش اداری و تجاری")
melk_type_forosh_edari_tejari_entry.config(state="readonly",readonlybackground="#06294B",fg="#FFFFFF")
melk_type_forosh_edari_tejari_entry.place(x=18, y=30, width=350, height=25)

sal_sakht_forosh_edari_tejari=tk.Label(frame_up_right_forosh_edari_tejari, text="سال ساخت", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
sal_sakht_forosh_edari_tejari.place(x=465, y=80, anchor="e")

sal_sakht_forosh_edari_tejari_entry=tk.Entry(frame_up_right_forosh_edari_tejari, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
sal_sakht_forosh_edari_tejari_entry.place(x=18, y=70, width=350, height=25)

metraj_forosh_edari_tejari=tk.Label(frame_up_right_forosh_edari_tejari, text=" متراژ ", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
metraj_forosh_edari_tejari.place(x=465, y=120, anchor="e")

metraj_forosh_edari_tejari_entry=tk.Entry(frame_up_right_forosh_edari_tejari, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
metraj_forosh_edari_tejari_entry.place(x=18, y=110, width=350, height=25)

tabaghe_forosh_edari_tejari= tk.Label(frame_up_right_forosh_edari_tejari, text="طبقه", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
tabaghe_forosh_edari_tejari.place(x=465, y=160, anchor="e")

tabaghe_forosh_edari_tejari_entry=tk.Entry(frame_up_right_forosh_edari_tejari, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
tabaghe_forosh_edari_tejari_entry.place(x=18, y=150, width=350, height=25)

vahed_forosh_edari_tejari=tk.Label(frame_up_right_forosh_edari_tejari, text="واحد", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
vahed_forosh_edari_tejari.place(x=465, y=200, anchor="e")

vahed_forosh_edari_tejari_entry=tk.Entry(frame_up_right_forosh_edari_tejari, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
vahed_forosh_edari_tejari_entry.place(x=18, y=190, width=350, height=25)

#otagh_forosh_edari_tejari= tk.Label(frame_up_right_forosh_edari_tejari, text="اتاق", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
#otagh_forosh_edari_tejari.place(x=465, y=240, anchor="e")

#otagh_forosh_edari_tejari_entry=tk.Entry(frame_up_right_forosh_edari_tejari, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
#otagh_forosh_edari_tejari_entry.place(x=18, y=230, width=350, height=25)

#----------------------------فریم بالا سمت چپ----------------------------------

photo_lbl2_forosh_edari_tejari = tk.Label(frame_up_left_forosh_edari_tejari, text="[تصویر ملک]", bg="#ffffff", width=79, height=15)
photo_lbl2_forosh_edari_tejari.place(x=40 ,y=10)

add_img_btn_forosh_edari_tejari = tk.Button(frame_up_left_forosh_edari_tejari, text="افزودن تصویر", bg="#00BFFF", fg="#ffffff",command=open_file,height=2,width=13)
add_img_btn_forosh_edari_tejari.place(x=240, y=250)

#----------------------------------فریم وسط سمت راست--------------------------

gheimat_kol_forosh_edari_tejari=tk.Label(frame_midde_right_forosh_edari_tejari, text=" قیمت کل ", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
gheimat_kol_forosh_edari_tejari.place(x=465, y=30, anchor="e")

gheimat_kol_forosh_edari_tejari_entry=tk.Entry(frame_midde_right_forosh_edari_tejari, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
gheimat_kol_forosh_edari_tejari_entry.place(x=18, y=20, width=350, height=25)

addrres_forosh_edari_tejari=tk.Label(frame_midde_right_forosh_edari_tejari, text="آدرس", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
addrres_forosh_edari_tejari.place(x=465, y=80, anchor="e")

addrres_forosh_edari_tejari_entry=tk.Text(frame_midde_right_forosh_edari_tejari, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
addrres_forosh_edari_tejari_entry.place(x=18, y=70, width=350, height=50)

#-------------------------------------فریم وسط سمت چپ---------------------------

name_malek_forosh_edari_tejari_lable = tk.Label(frame_midde_left_forosh_edari_tejari, text="نام مالک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
name_malek_forosh_edari_tejari_lable.place(x=600, y=30,anchor="e")

name_malek_forosh_edari_tejari_entry = tk.Entry(frame_midde_left_forosh_edari_tejari, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
name_malek_forosh_edari_tejari_entry.place(x=30, y=20, width=350, height=25)

shomareh_malek_forosh_edari_tejari_lable = tk.Label(frame_midde_left_forosh_edari_tejari, text="شماره مالک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
shomareh_malek_forosh_edari_tejari_lable.place(x=600, y=80,anchor="e")

shomareh_malek_forosh_edari_tejari_entry = tk.Entry(frame_midde_left_forosh_edari_tejari, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
shomareh_malek_forosh_edari_tejari_entry.place(x=30, y=70, width=350, height=25)

#-----------------------------------------فریم پایین--------------------------------

parking_forosh_edari_tejari_var=tk.IntVar(value=0)
anbari_forosh_edari_tejari_var=tk.IntVar(value=0)
asansor_forosh_edari_tejari_var=tk.IntVar(value=0)

parking_ch_btn_forosh_edari_tejari=tk.Checkbutton(frame_down_forosh_edari_tejari,variable=parking_forosh_edari_tejari_var,image=parking_pic, bg="#052340")
parking_ch_btn_forosh_edari_tejari.place(x=1050, y=10)
parking_ch_btn_forosh_edari_tejari_label=tk.Label(frame_down_forosh_edari_tejari,text="پارکینگ", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
parking_ch_btn_forosh_edari_tejari_label.place(x=1055,y=60)

asansor_ch_btn_forosh_edari_tejari=tk.Checkbutton(frame_down_forosh_edari_tejari,variable=asansor_forosh_edari_tejari_var,image=elvator_pic,background="#052340")
asansor_ch_btn_forosh_edari_tejari.place(x=950, y=10)
asansor_ch_btn_forosh_edari_tejari_label=tk.Label(frame_down_forosh_edari_tejari,text="آسانسور", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
asansor_ch_btn_forosh_edari_tejari_label.place(x=955,y=60)

anbari_checkbuton_forosh_edari_tejari=tk.Checkbutton(frame_down_forosh_edari_tejari,variable=anbari_forosh_edari_tejari_var,image=warehouse_pic,background="#052340")
anbari_checkbuton_forosh_edari_tejari.place(x=850, y=10)
anbari_checkbuton_forosh_edari_tejari_label=tk.Label(frame_down_forosh_edari_tejari,text="انباری", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
anbari_checkbuton_forosh_edari_tejari_label.place(x=855,y=60)

aab_va_gaz_forosh_edari_tejari=tk.Label(frame_down_forosh_edari_tejari, text="وضعیت آب و گاز", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
aab_va_gaz_forosh_edari_tejari.place(x=650, y=15)

aab_va_gaz_combo_forosh_edari_tejari=ttk.Combobox(frame_down_forosh_edari_tejari)
aab_va_gaz_combo_forosh_edari_tejari["state"]=["readonly"]
aab_va_gaz_combo_forosh_edari_tejari.configure(justify="center")
aab_va_gaz_combo_forosh_edari_tejari["values"] = ("آب و گاز دارد","فقط گاز دارد","فقط آب دارد ")
aab_va_gaz_combo_forosh_edari_tejari.place(x=475, y=15)

sarmaesh_forosh_edari_tejari=tk.Label(frame_down_forosh_edari_tejari, text="وضعیت سرمایش", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
sarmaesh_forosh_edari_tejari.place(x=350, y=45)

sarmaesh_combo_forosh_edari_tejari=ttk.Combobox(frame_down_forosh_edari_tejari)
sarmaesh_combo_forosh_edari_tejari["state"]=["readonly"]
sarmaesh_combo_forosh_edari_tejari.configure(justify="center")
sarmaesh_combo_forosh_edari_tejari["values"] = ("پنکه سقفی","کولر آبی","کولر گازی ","ندارد")
sarmaesh_combo_forosh_edari_tejari.place(x=150, y=45)

garmaesh_forosh_edari_tejari=tk.Label(frame_down_forosh_edari_tejari, text="وضعیت گرمایش", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
garmaesh_forosh_edari_tejari.place(x=350, y=15)

garmaesh_combo_forosh_edari_tejari=ttk.Combobox(frame_down_forosh_edari_tejari)
garmaesh_combo_forosh_edari_tejari["values"] = ("بخاری","شوفاژ","ندارد")
garmaesh_combo_forosh_edari_tejari["state"]=["readonly"]
garmaesh_combo_forosh_edari_tejari.place(x=150, y=15)

back_to_home_forosh_edari_tejari=tk.Button(forosh_edari_tejari_window,text="بازگشت",bg="#052340", fg="#ffffff",width=10,height=1,command=back_home_forosh_edari_tejari)
back_to_home_forosh_edari_tejari.place(x=400,y=30)

zakhire_forosh_edari_tejari=tk.Button(forosh_edari_tejari_window,text="ذخیره",bg="#00BFFF", fg="#ffffff",width=10,height=1,command=sabt_forosh_edari_tejari)
zakhire_forosh_edari_tejari.place(x=300,y=30)

delete_btn_forosh_edari_tejari=tk.Button(forosh_edari_tejari_window,text="حذف",command=delete_forosh_edari_tejari,bg="#8B0000",fg="#ffffff",height=1,width=10 )
delete_btn_forosh_edari_tejari.place_forget()

edit_btn_forosh_edari_tejari=tk.Button(forosh_edari_tejari_window,text="ثبت ویرایش",command=update_forosh_edari_tejari,bg="#00BFFF", fg="#ffffff",width=10,height=1,)
edit_btn_forosh_edari_tejari.place_forget()

forosh_edari_tejari_window.protocol("WM_DELETE_WINDOW", lambda: None)
forosh_edari_tejari_window.resizable(False, False)
#endregion
#--------------------پنجره فروش باغ/زمین-----------------------
#region
forosh_bagh_zamin_window = tk.Toplevel(root)
forosh_bagh_zamin_window.title("فروش باغ و زمین")
forosh_bagh_zamin_window.geometry("1200x700")
forosh_bagh_zamin_window.configure(bg="#052340")
forosh_bagh_zamin_window.withdraw()

#------------------کادر فروش باغ و زمین-----------------------------#
frame_up_right_forosh_bagh_zamin= tk.Frame(forosh_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=520,height=300)
frame_up_right_forosh_bagh_zamin.configure(bg="#052340")
frame_up_right_forosh_bagh_zamin.place(x=670,y=75)


frame_up_left_forosh_bagh_zamin= tk.Frame(forosh_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=300)
frame_up_left_forosh_bagh_zamin.configure(bg="#052340")
frame_up_left_forosh_bagh_zamin.place(x=10,y=75)

frame_midde_right_forosh_bagh_zamin= tk.Frame(forosh_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=520,height=150)
frame_midde_right_forosh_bagh_zamin.configure(bg="#052340")
frame_midde_right_forosh_bagh_zamin.place(x=670,y=390)

frame_midde_left_forosh_bagh_zamin= tk.Frame(forosh_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=150)
frame_midde_left_forosh_bagh_zamin.configure(bg="#052340")
frame_midde_left_forosh_bagh_zamin.place(x=10,y=390)

frame_down_forosh_bagh= tk.Frame(forosh_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=1180,height=130)
frame_down_forosh_bagh.configure(bg="#052340")
frame_down_forosh_bagh.place(x=10,y=555)

frame_down_forosh_zamin=tk.Frame(forosh_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=1180,height=130)
frame_down_forosh_zamin.configure(bg="#052340")
frame_down_forosh_zamin.place_forget()

title_label_forosh_bagh_zamin_up1 = tk.Label(forosh_bagh_zamin_window,text="فروش باغ و زمین",bg="#052340",fg="#00BFFF",font=("Shabnam", 16))
title_label_forosh_bagh_zamin_up1.place(x=600, y=10)

title_label_forosh_bagh_zamin_up2 = tk.Label(forosh_bagh_zamin_window,text="ثبت اطلاعات فروش باغ و زمین",bg="#052340",fg="#ffffff",font=("Shabnam", 11))
title_label_forosh_bagh_zamin_up2.place(x=585, y=40)

label_up_right_forosh_bagh_zamin=tk.Label(forosh_bagh_zamin_window,text="اطلاعات ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_right_forosh_bagh_zamin.place(x=760,y=60)

label_up_left_forosh_bagh_zamin=tk.Label(forosh_bagh_zamin_window,text="تصویر ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_left_forosh_bagh_zamin.place(x=100,y=60)

label_midde_right_forosh_bagh_zamin=tk.Label(forosh_bagh_zamin_window,text="اطلاعات معامله",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_right_forosh_bagh_zamin.place(x=800,y=375)

label_midde_left_forosh_bagh_zamin=tk.Label(forosh_bagh_zamin_window,text="اطلاعات مالک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_left_forosh_bagh_zamin.place(x=100,y=375)

label_down_forosh_bagh_zamin=tk.Label(forosh_bagh_zamin_window,text="امکانات",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_down_forosh_bagh_zamin.place(x=100,y=545)

#-----------------------فریم بالا سمت راست---------------------------
melk_type_forosh_bagh_zamin_lable=tk.Label(frame_up_right_forosh_bagh_zamin,text="نوع ملک",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
melk_type_forosh_bagh_zamin_lable.place(x=490, y=40, anchor="e")

melk_type_forosh_bagh_zamin_entry=tk.Entry(frame_up_right_forosh_bagh_zamin,bg="#06294B", fg="#ffffff",font=("Shabnam", 10),justify="center")
melk_type_forosh_bagh_zamin_entry.place(x=28, y=30, width=350, height=25)
melk_type_forosh_bagh_zamin_entry.insert(0,"فروش باغ و زمین")
melk_type_forosh_bagh_zamin_entry.config(state="readonly",readonlybackground="#06294B",fg="#ffffff")

metraj_zamin_forosh_bagh_zamin_lable=tk.Label(frame_up_right_forosh_bagh_zamin,text="متراژ",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
metraj_zamin_forosh_bagh_zamin_lable.place(x=490, y=80, anchor="e")

metraj_zamin_forosh_bagh_zamin_entry=tk.Entry(frame_up_right_forosh_bagh_zamin,bg="#FFFFFF", fg="#000000",font=("Shabnam", 10),textvariable="متر مربع")
metraj_zamin_forosh_bagh_zamin_entry.place(x=28, y=70, width=350, height=25)

bagh_type_forosh_bagh_zamin_lable=tk.Label(frame_up_right_forosh_bagh_zamin,text="کاربری زمین",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
bagh_type_forosh_bagh_zamin_lable.place(x=490, y=120, anchor="e")

bagh_type_forosh_bagh_zamin_combo=ttk.Combobox(frame_up_right_forosh_bagh_zamin,state="readonly")
bagh_type_forosh_bagh_zamin_combo["values"]=("باغ","زمین کشاورزی")
bagh_type_forosh_bagh_zamin_combo.set("باغ")
bagh_type_forosh_bagh_zamin_combo["state"]=["readonly"]
bagh_type_forosh_bagh_zamin_combo.place(x=28, y=110, width=350, height=25)
bagh_type_forosh_bagh_zamin_combo.bind("<<ComboboxSelected>>",change_bagh_zamin_forosh_bagh)
#------------------------------------فریم بالا سمت چپ-------------------------------
photo_forosh_bagh_zamin_lable= tk.Label(frame_up_left_forosh_bagh_zamin, text="[تصویر ملک]", bg="#ffffff", width=79, height=15)
photo_forosh_bagh_zamin_lable.place(x=40, y=10)
add_img_btn_forosh_bagh_zamin = tk.Button(frame_up_left_forosh_bagh_zamin, text="افزودن تصویر", bg="#00BFFF", fg="black",command=open_file,height=2,width=13)
add_img_btn_forosh_bagh_zamin.place(x=240, y=250)
#-------------------------فریم وسط سمت راست-----------------------------------------
gheimat_har_matr_babagh_zamin_forosh_bagh_zamin_lable=tk.Label(frame_midde_right_forosh_bagh_zamin,text='قیمت هر متر',bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
gheimat_har_matr_babagh_zamin_forosh_bagh_zamin_lable.place(x=490, y=20, anchor="e")

gheimat_har_metr_babagh_zamin_forosh_entry=tk.Entry(frame_midde_right_forosh_bagh_zamin,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
gheimat_har_metr_babagh_zamin_forosh_entry.place(x=28, y=13, width=350, height=25)

gheimat_kol_forosh_bagh_zamin_lable=tk.Label(frame_midde_right_forosh_bagh_zamin,text='قیمت کل',bg="#052340",fg="#ffffff",font=("Shabnam",12),width=10)
gheimat_kol_forosh_bagh_zamin_lable.place(x=490, y=65, anchor="e")

gheimat_kol_forosh_bagh_zamin_entry=tk.Entry(frame_midde_right_forosh_bagh_zamin,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
gheimat_kol_forosh_bagh_zamin_entry.place(x=28,y=55, width=350, height=25)

bagh_loctaion_forosh_bagh_zamin_lable=tk.Label(frame_midde_right_forosh_bagh_zamin,text="منطقه و ادرس ",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
bagh_loctaion_forosh_bagh_zamin_lable.place(x=490, y=110, anchor="e")

bagh_loctaion_forosh_bagh_zamin_entry=tk.Text(frame_midde_right_forosh_bagh_zamin,bg="#FFFFFF", fg="#000000",font=("Shabnam", 10))
bagh_loctaion_forosh_bagh_zamin_entry.place(x=28, y=90, width=350, height=25)
#----------------------------------------------فریم وسط سمت چپ---------------------------
name_malek_forosh_bagh_lable=tk.Label(frame_midde_left_forosh_bagh_zamin,text="نام مالک",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
name_malek_forosh_bagh_lable.place(x=600,y=30, anchor="e")

name_malek_forosh_bagh_entry=tk.Entry(frame_midde_left_forosh_bagh_zamin,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
name_malek_forosh_bagh_entry.place(x=30, y=20, width=350, height=25)

number_malek_forosh_bagh_lable=tk.Label(frame_midde_left_forosh_bagh_zamin,text="شماره مالک",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
number_malek_forosh_bagh_lable.place(x=600,y=80, anchor="e")

number_malek_forosh_bagh_entry=tk.Entry(frame_midde_left_forosh_bagh_zamin,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
number_malek_forosh_bagh_entry.place(x=30, y=70, width=350, height=25)
#-------------------------------------------فریم پایین-------------------------------------

abyari_forosh_bagh_zamin_lable=tk.Label(frame_down_forosh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="نوع آبیاری")
abyari_forosh_bagh_zamin_lable.place(x=1080,y=20)

abyari_forosh_bagh_zamin_combo=ttk.Combobox(frame_down_forosh_bagh,width=12)
abyari_forosh_bagh_zamin_combo["values"]=("سطحی","بارانی","قطره ای","تحت فشار")
abyari_forosh_bagh_zamin_combo["state"]=["readonly"]
abyari_forosh_bagh_zamin_combo.set("سطحی")
abyari_forosh_bagh_zamin_combo.place(x=970, y=20)

metraj_derakht_forosh_bagh_zamin_lable=tk.Label(frame_down_forosh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",10),width=13,text="متراژ درخت کاری")
metraj_derakht_forosh_bagh_zamin_lable.place(x=1080, y=60)

metraj_derakht_forosh_bagh_zamin_entry=tk.Entry(frame_down_forosh_bagh,width=10,bg="#ffffff",fg="#000000")
metraj_derakht_forosh_bagh_zamin_entry.place(x=1000, y=60)

tedad_derakht_forosh_bagh_zamin_lable=tk.Label(frame_down_forosh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",10),width=10,text="تعداد درخت")
tedad_derakht_forosh_bagh_zamin_lable.place(x=1080, y=100)

tedad_derakht_forosh_bagh_zamin_entry=tk.Entry(frame_down_forosh_bagh,width=10,bg="#ffffff",fg="#000000")
tedad_derakht_forosh_bagh_zamin_entry.place(x=1000, y=100)


type_tree_forosh_bagh_zamin_lable=tk.Label(frame_down_forosh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="نوع درخت")
type_tree_forosh_bagh_zamin_lable.place(x=870, y=20)

type_tree_forosh_bagh_zamin_combo=ttk.Combobox(frame_down_forosh_bagh,width=12,font=("Shabnam",9))
type_tree_forosh_bagh_zamin_combo["values"]=(" ","پسته","بادام","گردو","شلیل","هلو","سیب","انگور","انجیر","زردالو","گیلاس","آلبالو")
type_tree_forosh_bagh_zamin_combo["state"]=["readonly"]
type_tree_forosh_bagh_zamin_combo.configure(justify="center")
type_tree_forosh_bagh_zamin_combo.set(" ")
type_tree_forosh_bagh_zamin_combo.place(x=770, y=20)

type_tree_forosh_btn=tk.Button(frame_down_forosh_bagh,text="افزودن درخت",command=add_tree2,bg="#00BFFF",font=("Shabnam",9),width=10)
type_tree_forosh_btn.place(x=870, y=60)

label_natige_forosh_bagh_zamin=tk.Text(frame_down_forosh_bagh,width=11,height=4,wrap="word",font=("Shabnam",9))
label_natige_forosh_bagh_zamin.place(x=770, y=50)
#------------------------------------------------
chah_forosh_bagh_var=tk.IntVar(value=0)
chah_forosh_bagh_zamin=tk.Checkbutton(frame_down_forosh_bagh,variable=chah_forosh_bagh_var,text="چاه",background="#052340",fg="#00BFFF",font=("Shabnam",9))
chah_forosh_bagh_zamin.place(x=685, y=20)

estakhr_forosh_bagh_var=tk.IntVar(value=0)
estakhr_forosh_bagh_zamin=tk.Checkbutton(frame_down_forosh_bagh,variable=estakhr_forosh_bagh_var,text="استخر",background="#052340",fg="#00BFFF",font=("Shabnam",9))
estakhr_forosh_bagh_zamin.place(x=600, y=20)

bargh_keshi_forosh_bagh_var=tk.IntVar(value=0)
bargh_keshi_forosh_bagh_zamin=tk.Checkbutton(frame_down_forosh_bagh,variable=bargh_keshi_forosh_bagh_var,text="برق کشی",background="#052340",fg="#00BFFF",font=("Shabnam",9))
bargh_keshi_forosh_bagh_zamin.place(x=685, y=80)

gas_keshi_forosh_bagh_var=tk.IntVar(value=0)
gas_keshi_forosh_bagh_zamin=tk.Checkbutton(frame_down_forosh_bagh,variable=gas_keshi_forosh_bagh_var,text="گاز کشی",background="#052340",fg="#00BFFF",font=("Shabnam",9))
gas_keshi_forosh_bagh_zamin.place(x=685, y=50)

divar_forosh_bagh_var=tk.IntVar(value=0)
divar_forosh_bagh_zamin=tk.Checkbutton(frame_down_forosh_bagh,variable=divar_forosh_bagh_var,text="دیوار کشی",background="#052340",fg="#00BFFF",font=("Shabnam",9))
divar_forosh_bagh_zamin.place(x=600, y=50)
#-----------------------------------------------

var0_forosh_bagh_zamin=tk.IntVar(value=0)#چک باتن پیش فرض تیک نخورده باشه

otagh_check_btn_forosh_bagh_zamin=tk.Checkbutton(frame_down_forosh_bagh,variable=var0_forosh_bagh_zamin,image=warehouse_pic,background="#052340",text="ساختمان",command=home_true_false2)
otagh_check_btn_forosh_bagh_zamin.place(x=600, y=70)
otagh_check_btn_forosh_bagh_zamin_label=tk.Label(frame_down_forosh_bagh,text="ویلا", bg="#052340", fg="#ffffff", font=("Shabnam", 6), width=7)
otagh_check_btn_forosh_bagh_zamin_label.place(x=620,y=110)

type_vila_forosh_bagh_zamin=tk.Label(frame_down_forosh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="نوع سازه")
type_vila_forosh_bagh_zamin.place(x=490, y=20)

type_vila_forosh_bagh_zamin_combo=ttk.Combobox(frame_down_forosh_bagh,width=12)
type_vila_forosh_bagh_zamin_combo["values"]=("آجری","بلوکی","کانکس","چوبی")
type_vila_forosh_bagh_zamin_combo.set("آجری")
type_vila_forosh_bagh_zamin_combo["state"]=["disabled"]
type_vila_forosh_bagh_zamin_combo.configure(justify="center")
type_vila_forosh_bagh_zamin_combo.place(x=380, y=20)

toilet_forosh_bagh_zamin_lable=tk.Label(frame_down_forosh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="سرویس بهداشتی")
toilet_forosh_bagh_zamin_lable.place(x=490, y=60)

toilet_forosh_bagh_zamin_combo=ttk.Combobox(frame_down_forosh_bagh,width=12)
toilet_forosh_bagh_zamin_combo["values"]=("ندارد","فرنگی","ایرانی","هردو")
toilet_forosh_bagh_zamin_combo.set("ندارد")
toilet_forosh_bagh_zamin_combo.configure(justify="center")
toilet_forosh_bagh_zamin_combo["state"]=["disabled"]
toilet_forosh_bagh_zamin_combo.place(x=380, y=60)

hamam_forosh_bagh_zamin=tk.Label(frame_down_forosh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="حمام")
hamam_forosh_bagh_zamin.place(x=490, y=95)

hamam_forosh_bagh_zamin_combo=ttk.Combobox(frame_down_forosh_bagh,width=12)
hamam_forosh_bagh_zamin_combo["values"]=("ندارد","دارد")
hamam_forosh_bagh_zamin_combo.set("ندارد")
hamam_forosh_bagh_zamin_combo.configure(justify="center")
hamam_forosh_bagh_zamin_combo["state"]=["disabled"]
hamam_forosh_bagh_zamin_combo.place(x=380, y=100)

sanad_forosh_bagh_zamin_lable=tk.Label(frame_down_forosh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="سند")
sanad_forosh_bagh_zamin_lable.place(x=300, y=60)

sanad_forosh_bagh_zamin_combo=ttk.Combobox(frame_down_forosh_bagh,width=12)
sanad_forosh_bagh_zamin_combo["values"]=("ندارد","تک برگ","قولنامه ای","مشاع")
sanad_forosh_bagh_zamin_combo.set("ندارد")
sanad_forosh_bagh_zamin_combo.configure(justify="center")
sanad_forosh_bagh_zamin_combo["state"]=["disabled"]
sanad_forosh_bagh_zamin_combo.place(x=180, y=60)

option_forosh_bagh_zamin=tk.Label(frame_down_forosh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="امکانات تفریحی")
option_forosh_bagh_zamin.place(x=310, y=20)

option_forosh_bagh_zamin_combo=ttk.Combobox(frame_down_forosh_bagh,width=12)
option_forosh_bagh_zamin_combo["values"]=("استخر","جکوزی","باربیکیو")
option_forosh_bagh_zamin_combo.set("استخر")
option_forosh_bagh_zamin_combo.configure(justify="center")
option_forosh_bagh_zamin_combo["state"]=["disabled"]
option_forosh_bagh_zamin_combo.place(x=180, y=20)

add_option_button_forosh_bagh_zamin=tk.Button(frame_down_forosh_bagh,text="افزودن امکانات",command=add_option2,bg="#00BFFF",font=("Shabnam",9),width=10)
add_option_button_forosh_bagh_zamin.place(x=85, y=20)

lable_natige_add_forosh_bagh_zamin=tk.Label(frame_down_forosh_bagh,text="",width=10)
lable_natige_add_forosh_bagh_zamin.place(x=5, y=20)

metraj_vila_forosh_bagh_zamin=tk.Label(frame_down_forosh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="متراژ سازه")
metraj_vila_forosh_bagh_zamin.place(x=300, y=95)

metraj_vila_forosh_bagh_zamin_entry=tk.Entry(frame_down_forosh_bagh,width=10,bg="#ffffff",fg="#000000",state="disabled")
metraj_vila_forosh_bagh_zamin_entry.place(x=240, y=100)

sal_sakht_vila_forosh_bagh_zamin_lable=tk.Label(frame_down_forosh_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="سال ساخت")
sal_sakht_vila_forosh_bagh_zamin_lable.place(x=80, y=60)

sal_sakht_vila_forosh_bagh_zamin_entry=tk.Entry(frame_down_forosh_bagh,width=10,bg="#ffffff",fg="#000000",state="disabled")
sal_sakht_vila_forosh_bagh_zamin_entry.place(x=10, y=60)

mojavez_sakht_forosh_bagh_var=tk.IntVar(value=0)
mojavez_sakht_check_btn_forosh_bagh_zamin=tk.Checkbutton(frame_down_forosh_bagh,variable=mojavez_sakht_forosh_bagh_var,text="مجوز ساختن",background="#052340",fg="#00BFFF",font=("Shabnam",9),state="disabled")
mojavez_sakht_check_btn_forosh_bagh_zamin.place(x=10, y=95)

mohavate_forosh_bagh_var=tk.IntVar(value=0)
mohavate_sazi_check_btn_forosh_bagh_zamin=tk.Checkbutton(frame_down_forosh_bagh,variable=mohavate_forosh_bagh_var,text="محوطه سازی",background="#052340",fg="#00BFFF",font=("Shabnam",9),state="disabled")
mohavate_sazi_check_btn_forosh_bagh_zamin.place(x=120, y=95)
#-------------------------تعویض کاربری به زمین در قسمت فروش باغ/زمین-------------



karbari_forosh_bagh_zamin_lable=tk.Label(frame_down_forosh_zamin,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="نوع کاربری")
karbari_forosh_bagh_zamin_lable.place(x=1000,y=20)

karbari_forosh_bagh_zamin_combo=ttk.Combobox(frame_down_forosh_zamin)
karbari_forosh_bagh_zamin_combo["values"]=(" ","زراعی","باغی","گلخانه ای","دامداری ","مرغداری","دامداری و مرغداری","آیش")        
karbari_forosh_bagh_zamin_combo["state"]=["readonly"]                        
karbari_forosh_bagh_zamin_combo.set(" ")
karbari_forosh_bagh_zamin_combo.configure(justify="center")
karbari_forosh_bagh_zamin_combo.place(x=800,y=20)

khak_forosh_bagh_zamin=tk.Label(frame_down_forosh_zamin,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="نوع خاک")
khak_forosh_bagh_zamin.place(x=1000,y=60)

khak_forosh_bagh_zamin_combo=ttk.Combobox(frame_down_forosh_zamin)
khak_forosh_bagh_zamin_combo["values"]=(" ","رسی","شنی","لومی","رسی_شنی","شنی_لومی","رسی_لومی")       
khak_forosh_bagh_zamin_combo["state"]=["readonly"]                         
khak_forosh_bagh_zamin_combo.set(" ")
khak_forosh_bagh_zamin_combo.configure(justify="center")
khak_forosh_bagh_zamin_combo.place(x=800,y=60)

ab_forosh_bagh_zamin=tk.Label(frame_down_forosh_zamin,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="منبع آب")
ab_forosh_bagh_zamin.place(x=1000,y=100)

ab_forosh_bagh_zamin_combo=ttk.Combobox(frame_down_forosh_zamin)
ab_forosh_bagh_zamin_combo["values"]=(" ","چاه","قنات","رودخانه","کانال آبیاری","چشمه","آب لوله کشی کشاورزی","تانکر","استخر")    
ab_forosh_bagh_zamin_combo["state"]=["readonly"]                            
ab_forosh_bagh_zamin_combo.set(" ")
ab_forosh_bagh_zamin_combo.configure(justify="center")
ab_forosh_bagh_zamin_combo.place(x=800,y=100)

security_zamin_forosh_zamin_var=tk.IntVar(value=0)
security_zamin_forosh_bagh_zamin=tk.Checkbutton(frame_down_forosh_zamin,variable=security_zamin_forosh_zamin_var,text="اتاق نگهبان",background="#052340",fg="#00BFFF",font=("Shabnam",9))
security_zamin_forosh_bagh_zamin.place(x=500,y=20)

bargh_kesi_zamin_forosh_zamin_var=tk.IntVar(value=0)
bargh_kesi_zamin_forosh_bagh_zamin=tk.Checkbutton(frame_down_forosh_zamin,variable=bargh_kesi_zamin_forosh_zamin_var,text="برق تک فاز",background="#052340",fg="#00BFFF",font=("Shabnam",9))
bargh_kesi_zamin_forosh_bagh_zamin.place(x=400,y=20)

bargh_kesi_zamin_forosh_zamin2_var=tk.IntVar(value=0)
bargh_keshi_zamin_forosh_bagh_zamin2=tk.Checkbutton(frame_down_forosh_zamin,variable=bargh_kesi_zamin_forosh_zamin2_var,text="برق سه فاز",background="#052340",fg="#00BFFF",font=("Shabnam",9))
bargh_keshi_zamin_forosh_bagh_zamin2.place(x=300,y=20)

anbar_zamin_forosh_zamin_var=tk.IntVar(value=0)
anbar_zamin_forosh_bagh_zamin=tk.Checkbutton(frame_down_forosh_zamin,variable=anbar_zamin_forosh_zamin_var,text="انبار/سوله",background="#052340",fg="#00BFFF",font=("Shabnam",9))
anbar_zamin_forosh_bagh_zamin.place(x=500,y=60)

fans_zamin_forosh_zamin_var=tk.IntVar(value=0)
fans_zamin_forosh_bagh_zamin=tk.Checkbutton(frame_down_forosh_zamin,variable=fans_zamin_forosh_zamin_var,text="فنس/دیوار",background="#052340",fg="#00BFFF",font=("Shabnam",9))
fans_zamin_forosh_bagh_zamin.place(x=400,y=60)

javaz_chah_zamin_forosh_zamin_var=tk.IntVar(value=0)
mojavez_chah_zamin_forosh_bagh_zamin=tk.Checkbutton(frame_down_forosh_zamin,variable=javaz_chah_zamin_forosh_zamin_var,text="اجازه حفر چاه",background="#052340",fg="#00BFFF",font=("Shabnam",9))
mojavez_chah_zamin_forosh_bagh_zamin.place(x=300,y=60)

back_to_home_forosh_bagh_zamin=tk.Button(forosh_bagh_zamin_window,text="بازگشت",bg="#052340",fg="#ffffff",width=10,height=1,command=back_home_forosh_bagh_zamin)
back_to_home_forosh_bagh_zamin.place(x=300,y=30)

zakhire_forosh_bagh_zamin=tk.Button(forosh_bagh_zamin_window,text="ذخیره",bg="#00BFFF", fg="#ffffff",width=10,height=1,command=sabt_forosh_bagh_zamin_main)
zakhire_forosh_bagh_zamin.place(x=200,y=30)

forosh_bagh_zamin_window.protocol("WM_DELETE_WINDOW", lambda: None)
forosh_bagh_zamin_window.resizable(False, False)
#endregion
#-------------------پنجره فروش کارگاه------------------------
#region
forosh_karghah_window = tk.Toplevel(root)
forosh_karghah_window.title(" فروش کارگاه")
forosh_karghah_window.geometry("1200x700")
forosh_karghah_window.configure(bg="#052340")
forosh_karghah_window.withdraw()

frame_up_right_forosh_karghah = tk.Frame(forosh_karghah_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=510,height=300)
frame_up_right_forosh_karghah.configure(bg="#052340")
frame_up_right_forosh_karghah.place(x=670,y=90)

frame_up_left_forosh_karghah= tk.Frame(forosh_karghah_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=300)
frame_up_left_forosh_karghah.configure(bg="#052340")
frame_up_left_forosh_karghah.place(x=40,y=90)

frame_midde_right_forosh_karghah= tk.Frame(forosh_karghah_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=510,height=150)
frame_midde_right_forosh_karghah.configure(bg="#052340")
frame_midde_right_forosh_karghah.place(x=670,y=410)

frame_midde_left_forosh_karghah= tk.Frame(forosh_karghah_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=150)
frame_midde_left_forosh_karghah.configure(bg="#052340")
frame_midde_left_forosh_karghah.place(x=10,y=410)

frame_down_forosh_karghah= tk.Frame(forosh_karghah_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=1170,height=85)
frame_down_forosh_karghah.configure(bg="#052340")
frame_down_forosh_karghah.place(x=10,y=580)

title_label_up1_forosh_karghah= tk.Label(forosh_karghah_window,text="فروش کارگاه",bg="#052340",fg="#00BFFF",font=("Shabnam", 16))
title_label_up1_forosh_karghah.place(x=570, y=17)  

title_label_up2_forosh_karghah = tk.Label(forosh_karghah_window,text="ثبت اطلاعات فروش کارگاه",bg="#052340",fg="#ffffff",font=("Shabnam", 11))
title_label_up2_forosh_karghah.place(x=555, y=45)

label_up_right_forosh_karghah=tk.Label(forosh_karghah_window,text="اطلاعات ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_right_forosh_karghah.place(x=800,y=73)

label_up_left_forosh_karghah=tk.Label(forosh_karghah_window,text="تصویر ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_left_forosh_karghah.place(x=100,y=73)

label_midde_right_forosh_karghah=tk.Label(forosh_karghah_window,text="اطلاعات معامله",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_right_forosh_karghah.place(x=800,y=395)

label_midde_left_forosh_karghah=tk.Label(forosh_karghah_window,text="اطلاعات مالک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_left_forosh_karghah.place(x=100,y=395)

label_down_forosh_karghah=tk.Label(forosh_karghah_window,text="امکانات",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_down_forosh_karghah.place(x=100,y=565)

#-------------------------------فریم بالا سمت راست----------------------------
sal_sakht_forosh_kargah_lable = tk.Label(frame_up_right_forosh_karghah, text="سال ساخت", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
sal_sakht_forosh_kargah_lable.place(x=465, y=80, anchor="e")

sal_sakht_forosh_kargah_entry = tk.Entry(frame_up_right_forosh_karghah, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
sal_sakht_forosh_kargah_entry.place(x=18, y=70, width=350, height=25)

karbari_forosh_kargah = tk.Label(frame_up_right_forosh_karghah, text="کاربری زمین", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
karbari_forosh_kargah.place(x=465, y=40, anchor="e")

karbari_forosh_kargah_entry = tk.Entry(frame_up_right_forosh_karghah, bg="#06294B", fg="#ffffff", font=("Shabnam", 10), justify="center")
karbari_forosh_kargah_entry.insert(0, "فروش کارگاه")
karbari_forosh_kargah_entry.config(state="readonly",readonlybackground="#06294B",fg="#ffffff")
karbari_forosh_kargah_entry.place(x=18, y=30, width=350, height=25)

metraj_forosh_kargah = tk.Label(frame_up_right_forosh_karghah, text="متراژ", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
metraj_forosh_kargah.place(x=465, y=120,anchor="e")

metraj_forosh_kargah_entry = tk.Entry(frame_up_right_forosh_karghah, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
metraj_forosh_kargah_entry.place(x=18, y=110, width=350, height=25)

#--------------------فریم چپ بالا---------------------------
photo_lbl2_forosh_kargah = tk.Label(frame_up_left_forosh_karghah, text="[تصویر ملک]", bg="#FFFFFF", width=79, height=15,relief="solid")
photo_lbl2_forosh_kargah.place(x=40, y=10)

add_img_btn_forosh_kargah = tk.Button(frame_up_left_forosh_karghah, text="افزودن تصویر", bg="#00BFFF", fg="#ffffff",command=open_file, height=2,width=13)
add_img_btn_forosh_kargah.place(x=240, y=250)

#--------------------------فریم راست وسط---------------------
loctaion_forosh_kargah = tk.Label(frame_midde_right_forosh_karghah, text="آدرس", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=12)
loctaion_forosh_kargah.place(x=465, y=80, anchor="e")

loctaion_forosh_kargah_entry = tk.Text(frame_midde_right_forosh_karghah, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
loctaion_forosh_kargah_entry.place(x=18, y=70, width=350, height=50)

gheimat_kol_forosh_kargah_lable=tk.Label(frame_midde_right_forosh_karghah,text="قیمت کل ",bg="#052340",fg="#ffffff",font=("Shabnam", 12),width=9)
gheimat_kol_forosh_kargah_lable.place(x=465, y=30, anchor="e")

gheimat_kol_forosh_kargah_entry=tk.Entry(frame_midde_right_forosh_karghah,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
gheimat_kol_forosh_kargah_entry.place(x=18, y=20, width=350, height=25)

#------------------------------------فریم چپ وسط-----------------------
name_malek_forosh_kargah_lable = tk.Label(frame_midde_left_forosh_karghah,text="نام مالک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
name_malek_forosh_kargah_lable.place(x=600, y=30,anchor="e")

name_malek_forosh_kargah_entry = tk.Entry(frame_midde_left_forosh_karghah, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
name_malek_forosh_kargah_entry.place(x=30, y=20, width=350, height=25)

shomareh_malek_forosh_kargah_lable = tk.Label(frame_midde_left_forosh_karghah, text="شماره مالک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
shomareh_malek_forosh_kargah_lable.place(x=600, y=80,anchor="e")

shomareh_malek_forosh_kargah_entry = tk.Entry(frame_midde_left_forosh_karghah, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
shomareh_malek_forosh_kargah_entry.place(x=30, y=70, width=350, height=25)

#---------------------------------فریم پایین--------------------------------
sarmayesh_forosh_kargah = tk.Label(frame_down_forosh_karghah, text="سیستم سرمایش", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
sarmayesh_forosh_kargah.place(x=485, y=15)
sarmayesh_forosh_kargah_combo = ttk.Combobox(frame_down_forosh_karghah)
sarmayesh_forosh_kargah_combo["values"] = ("ندارد", "پنکه سقفی", "کولر ابی", "کولر گازی ", "ابی/گازی")
sarmayesh_forosh_kargah_combo["state"] = "readonly"
sarmayesh_forosh_kargah_combo.configure(justify="center")
sarmayesh_forosh_kargah_combo.place(x=325, y=15)

garmayesh_type_forosh_kargah = tk.Label(frame_down_forosh_karghah, text="سیستم گرمایش", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
garmayesh_type_forosh_kargah.place(x=485, y=45)
garmayesh_type_forosh_kargah_combo = ttk.Combobox(frame_down_forosh_karghah)
garmayesh_type_forosh_kargah_combo["values"] = ("ندارد", "بخاری", " شوفاژ", "گرمایش از کف ")
garmayesh_type_forosh_kargah_combo["state"] = "readonly"
garmayesh_type_forosh_kargah_combo.configure(justify="center")
garmayesh_type_forosh_kargah_combo.place(x=325, y=45)

toilet_forosh_kargah = tk.Label(frame_down_forosh_karghah, text="سرویس بهداشتی", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
toilet_forosh_kargah.place(x=190, y=45)
toilet_forosh_kargah_combo = ttk.Combobox(frame_down_forosh_karghah)
toilet_forosh_kargah_combo["values"] = ("ایرانی", "فرنگی", "هردو")
toilet_forosh_kargah_combo["state"] = "readonly"
toilet_forosh_kargah_combo.configure(justify="center")
toilet_forosh_kargah_combo.place(x=20, y=45)

vaziat_bagh_forosh_kargah=tk.Label(frame_down_forosh_karghah,bg="#052340",fg="#ffffff",font=("Shabnam", 11),width=15,text="وضعیت برق")
vaziat_bagh_forosh_kargah.place(x=190, y=15)

vaziat_bargh_forosh_kargah_combo=ttk.Combobox(frame_down_forosh_karghah)
vaziat_bargh_forosh_kargah_combo["values"]=("","برق شهری","سه فاز","تک فاز")
vaziat_bargh_forosh_kargah_combo.set("")
vaziat_bargh_forosh_kargah_combo["state"]=["readonly"]
vaziat_bargh_forosh_kargah_combo.place(x=20, y=15)


vaziat_ab_forosh_kargah=tk.Label(frame_down_forosh_karghah,bg="#052340",fg="#ffffff",width=13,text=" وضعیت آب",font=("Shabnam", 9))
vaziat_ab_forosh_kargah.place(x=1064, y=13)

vaziat_ab_forosh_kargah_combo=ttk.Combobox(frame_down_forosh_karghah,width=23)
vaziat_ab_forosh_kargah_combo["values"]=(""," آب  لوله کشی (بدون فشار) " ," آب لوله کشی (با موتور فشار) ","دارای منبع(با موتور فشار)","دارای منبع(بدون فشار)")
vaziat_ab_forosh_kargah_combo.set("")
vaziat_ab_forosh_kargah_combo["state"]=["readonly"]
vaziat_ab_forosh_kargah_combo.place(x=910, y=13)

abzar_forosh_kargah=tk.Label(frame_down_forosh_karghah,bg="#052340",fg="#ffffff",font=("Shabnam", 11),width=15,text=" ابزار صنعتی ")
abzar_forosh_kargah.place(x=745, y=45)

abzar_forosh_kargah_combo=ttk.Combobox(frame_down_forosh_karghah,width=15)
abzar_forosh_kargah_combo["values"]=("","(کارگاه خالی) بدون دستگاه ","دارای دستگاه")
abzar_forosh_kargah_combo.set("")
abzar_forosh_kargah_combo["state"]=["readonly"]
abzar_forosh_kargah_combo.place(x=620, y=45)

hamam_forosh_kargah=tk.Label(frame_down_forosh_karghah,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="حمام")
hamam_forosh_kargah.place(x=1064, y=43)

hamam_forosh_kargah_combo=ttk.Combobox(frame_down_forosh_karghah,width=23)
hamam_forosh_kargah_combo["values"]=("","ندارد","دارد")
hamam_forosh_kargah_combo.set("")
hamam_forosh_kargah_combo["state"]=["readonly"]
hamam_forosh_kargah_combo.place(x=910, y=45)

otagh_forosh_kargah=tk.Label(frame_down_forosh_karghah,bg="#052340",fg="#ffffff",font=("Shabnam", 11),width=17,text="اتاق رختکن")
otagh_forosh_kargah.place(x=740, y=15)

otagh_forosh_kargah_combo=ttk.Combobox(frame_down_forosh_karghah,width=15)
otagh_forosh_kargah_combo["values"]=("","ندارد","دارد")
otagh_forosh_kargah_combo.set("")
otagh_forosh_kargah_combo["state"]=["readonly"]
otagh_forosh_kargah_combo.place(x=620, y=15)

back_to_home_forosh_kargah=tk.Button(forosh_karghah_window,text="بازگشت",bg="#052340", fg="#ffffff",width=10,height=1,command=back_home_forosh_karghah)
back_to_home_forosh_kargah.place(x=400,y=30)

zakhire_forosh_kargah=tk.Button(forosh_karghah_window,text="ذخیره",bg="#00BFFF", fg="#ffffff",width=10,height=1,command=sabt_forosh_kargah)
zakhire_forosh_kargah.place(x=300,y=30)

delete_btn_forosh_kargah=tk.Button(forosh_karghah_window,text="حذف",command=delete_forosh_kargah,bg="#8B0000",fg="#ffffff",height=1,width=10 )
delete_btn_forosh_kargah.place_forget()

edit_btn_forosh_kargah=tk.Button(forosh_karghah_window,text="ثبت ویرایش",command=update_forosh_kargah,bg="#00BFFF", fg="#ffffff",width=10,height=1,)
edit_btn_forosh_kargah.place_forget()

forosh_karghah_window.protocol("WM_DELETE_WINDOW", lambda: None)
forosh_karghah_window.resizable(False, False)
#endregion
#--------------------پنجره های ثبتی بخش درخواست-----------------------
#----------------------پنجره درخواست مسکونی--------------------------
#region
darkhast_maskoni_window = tk.Toplevel(root)
darkhast_maskoni_window.title("درخواست مسکونی")
darkhast_maskoni_window.geometry("1200x700")
darkhast_maskoni_window.configure(bg="#052340")
darkhast_maskoni_window.withdraw()

#------------------کادر درخواست مسکونی-----------------------------#
frame_up_right_darkhast_maskoni = tk.Frame(darkhast_maskoni_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=500,height=300)
frame_up_right_darkhast_maskoni.configure(bg="#052340")
frame_up_right_darkhast_maskoni.place(x=670,y=90)

frame_up_left_darkhast_maskoni= tk.Frame(darkhast_maskoni_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=300)
frame_up_left_darkhast_maskoni.configure(bg="#052340")
frame_up_left_darkhast_maskoni.place(x=10,y=90)

frame_midde_right_darkhast_maskoni= tk.Frame(darkhast_maskoni_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=500,height=150)
frame_midde_right_darkhast_maskoni.configure(bg="#052340")
frame_midde_right_darkhast_maskoni.place(x=670,y=410)

frame_midde_left_darkhast_maskoni= tk.Frame(darkhast_maskoni_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=150)
frame_midde_left_darkhast_maskoni.configure(bg="#052340")
frame_midde_left_darkhast_maskoni.place(x=10,y=410)

frame_down_darkhast_maskoni= tk.Frame(darkhast_maskoni_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=1160,height=100)
frame_down_darkhast_maskoni.configure(bg="#052340")
frame_down_darkhast_maskoni.place(x=10,y=580)


title_label_up1_darkhast_maskoni = tk.Label(darkhast_maskoni_window,text="درخواست خرید و اجاره مسکونی",bg="#052340",fg="#00BFFF",font=("Shabnam", 16))
title_label_up1_darkhast_maskoni.place(x=570, y=17)   

title_label_up2_darkhast_maskoni = tk.Label(darkhast_maskoni_window,text="ثبت درخواست های خرید و اجاره مسکونی",bg="#052340",fg="#ffffff",font=("Shabnam", 11))
title_label_up2_darkhast_maskoni.place(x=555, y=45)

label_up_right_darkhast_maskoni=tk.Label(darkhast_maskoni_window,text="اطلاعات ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_right_darkhast_maskoni.place(x=800,y=73)

label_up_left_darkhast_maskoni=tk.Label(darkhast_maskoni_window,text="تصویر ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_left_darkhast_maskoni.place(x=100,y=73)

label_midde_right_darkhast_maskoni=tk.Label(darkhast_maskoni_window,text="اطلاعات معامله",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_right_darkhast_maskoni.place(x=800,y=395)

label_midde_left_darkhast_maskoni=tk.Label(darkhast_maskoni_window,text="اطلاعات مالک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_left_darkhast_maskoni.place(x=100,y=395)

label_down_darkhast_maskoni=tk.Label(darkhast_maskoni_window,text="امکانات",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_down_darkhast_maskoni.place(x=100,y=565)
 
#------------------فریم بالا سمت راست-----------------------------#
melk_type_darkhast_maskoni=tk.Label(frame_up_right_darkhast_maskoni,text="نوع ملک",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
melk_type_darkhast_maskoni.place(x=465, y=40, anchor="e")

melk_type_darkhast_maskoni_entry=ttk.Combobox(frame_up_right_darkhast_maskoni,state="readonly")
melk_type_darkhast_maskoni_entry["values"] = ("درخواست خرید مسکونی","درخواست اجاره مسکونی")
melk_type_darkhast_maskoni_entry.set("درخواست خرید مسکونی")
melk_type_darkhast_maskoni_entry.configure(justify="center")
melk_type_darkhast_maskoni_entry.bind("<<ComboboxSelected>>", change_darkhast_maskoni_type)
melk_type_darkhast_maskoni_entry.place(x=18, y=30, width=350, height=25)

sal_sakht_darkhast_maskoni=tk.Label(frame_up_right_darkhast_maskoni,text="سال ساخت",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
sal_sakht_darkhast_maskoni.place(x=465, y=80, anchor="e")

sal_sakht_darkhast_maskoni_entry=tk.Entry(frame_up_right_darkhast_maskoni,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
sal_sakht_darkhast_maskoni_entry.place(x=18, y=70, width=350, height=25)

metraj_darkhast_maskoni_lable=tk.Label(frame_up_right_darkhast_maskoni,text="متراژ",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
metraj_darkhast_maskoni_lable.place(x=465, y=120, anchor="e")

metraj_darkhast_maskoni_entry=tk.Entry(frame_up_right_darkhast_maskoni,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
metraj_darkhast_maskoni_entry.place(x=18, y=110, width=350, height=25)

tabaghe_darkhast_maskoni=tk.Label(frame_up_right_darkhast_maskoni,text="طبقه",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
tabaghe_darkhast_maskoni.place(x=465, y=160, anchor="e")

tabaghe_darkhast_maskoni_entry=tk.Entry(frame_up_right_darkhast_maskoni,bg="#ffffff", fg="#000000",font=("Shabnam", 10),)
tabaghe_darkhast_maskoni_entry.place(x=18, y=150, width=350, height=25)

vahed_darkhast_maskoni=tk.Label(frame_up_right_darkhast_maskoni,text="واحد",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
vahed_darkhast_maskoni.place(x=465, y=200, anchor="e")

vahed_darkhast_maskoni_entry=tk.Entry(frame_up_right_darkhast_maskoni,bg="#ffffff", fg="#000000",font=("Shabnam", 10),)
vahed_darkhast_maskoni_entry.place(x=18, y=190, width=350, height=25)

otagh_darkhast_maskoni=tk.Label(frame_up_right_darkhast_maskoni,text="اتاق",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
otagh_darkhast_maskoni.place(x=465, y=240, anchor="e")

otagh_darkhast_maskoni_entry=tk.Entry(frame_up_right_darkhast_maskoni,bg="#ffffff", fg="#000000",font=("Shabnam", 10),)
otagh_darkhast_maskoni_entry.place(x=18, y=230, width=350, height=25)
#-------------------------------فریم بالا سمت چپ-------------------------------------------
photo_lbl2_darkhast_maskoni = tk.Label(frame_up_left_darkhast_maskoni, text="[تصویر ملک]", bg="#ffffff",width=79, height=15,relief="solid")
photo_lbl2_darkhast_maskoni.place(x=40, y=10)

add_img_btn_darkhast_maskoni = tk.Button(frame_up_left_ejareh_maskoni, text="افزودن تصویر", bg="#00BFFF", fg="black",command=open_file,height=2,width=13)
add_img_btn_darkhast_maskoni.place(x=240, y=250)
#---------------------------فریم وسط سمت راست----------------------------------------

gheimat_kol_darkhast_maskoni_lable=tk.Label(frame_midde_right_darkhast_maskoni,text="قیمت کل",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
gheimat_kol_darkhast_maskoni_lable.place(x=465, y=30, anchor="e")

gheimat_kol_darkhast_maskoni_entry=tk.Entry(frame_midde_right_darkhast_maskoni,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
gheimat_kol_darkhast_maskoni_entry.place(x=18, y=20, width=350, height=25)

gheimat_pish_darkhast_maskoni_lable = tk.Label(frame_midde_right_darkhast_maskoni, text="مبلغ پیش", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
gheimat_pish_darkhast_maskoni_lable.place_forget()

gheimat_pish_darkhast_maskoni_entry = tk.Entry(frame_midde_right_darkhast_maskoni, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
gheimat_pish_darkhast_maskoni_entry.place_forget()

mablagh_ejare_darkhast_maskoni_lable = tk.Label(frame_midde_right_darkhast_maskoni, text="مبلغ اجاره", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
mablagh_ejare_darkhast_maskoni_lable.place_forget()

mablagh_ejare_darkhast_maskoni_entry = tk.Entry(frame_midde_right_darkhast_maskoni, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
mablagh_ejare_darkhast_maskoni_entry.place_forget()

addrres_darkhast_maskoni=tk.Label(frame_midde_right_darkhast_maskoni,text="آدرس",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
addrres_darkhast_maskoni.place(x=465, y=80, anchor="e")

addrres_darkhast_maskoni_entry=tk.Text(frame_midde_right_darkhast_maskoni,bg="#ffffff", fg="#000000",font=("Shabnam", 10),)
addrres_darkhast_maskoni_entry.place(x=18, y=70, width=350, height=50)

#---------------------------فریم سمت چپ وسط-----------------------------------
name_moshtari_darkhast_maskoni_lable = tk.Label(frame_midde_left_darkhast_maskoni, text="نام مشتری", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
name_moshtari_darkhast_maskoni_lable.place(x=600, y=30,anchor="e")

name_moshtari_darkhast_maskoni_entry = tk.Entry(frame_midde_left_darkhast_maskoni, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
name_moshtari_darkhast_maskoni_entry.place(x=30, y=20, width=350, height=25)

shomareh_moshtari_darkhast_maskoni_lable = tk.Label(frame_midde_left_darkhast_maskoni, text="شماره مشتری", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
shomareh_moshtari_darkhast_maskoni_lable.place(x=600, y=80,anchor="e")

shomareh_moshtari_darkhast_maskoni_entry = tk.Entry(frame_midde_left_darkhast_maskoni, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
shomareh_moshtari_darkhast_maskoni_entry.place(x=30, y=70, width=350, height=25)

#-------------------------فریم پایین----------------------------------

parking_darkhast_maskoni_var=tk.IntVar(value=0)
anbari_darkhast_maskoni_var=tk.IntVar(value=0)
asansor_darkhast_maskoni_var=tk.IntVar(value=0)

parking_ch_btn_darkhast_maskoni=tk.Checkbutton(frame_down_darkhast_maskoni,variable=parking_darkhast_maskoni_var,image=parking_pic,background="#052340")
parking_ch_btn_darkhast_maskoni.place(x=1050, y=10)
parking_ch_btn_darkhast_maskoni_label=tk.Label(frame_down_darkhast_maskoni,text="پارکینگ", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
parking_ch_btn_darkhast_maskoni_label.place(x=1055,y=60)

asansor_ch_btn_darkhast_maskoni=tk.Checkbutton(frame_down_darkhast_maskoni,variable=asansor_darkhast_maskoni_var,image=elvator_pic,background="#052340")
asansor_ch_btn_darkhast_maskoni.place(x=950, y=10)
asansor_ch_btn_darkhast_maskoni_label=tk.Label(frame_down_darkhast_maskoni,text="اسانسور", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
asansor_ch_btn_darkhast_maskoni_label.place(x=955,y=60)

anbari_checkbuton_darkhast_maskoni=tk.Checkbutton(frame_down_darkhast_maskoni,variable=anbari_darkhast_maskoni_var,image=warehouse_pic,background="#052340")
anbari_checkbuton_darkhast_maskoni.place(x=850, y=10)
anbari_checkbuton_darkhast_maskoni_label=tk.Label(frame_down_darkhast_maskoni,text="انباری", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
anbari_checkbuton_darkhast_maskoni_label.place(x=855,y=60)


sarmaesh_darkhast_maskoni=tk.Label(frame_down_darkhast_maskoni,text="سرمایش",background="#052340",fg="#ffffff",font=("Shabnam",11))
sarmaesh_darkhast_maskoni.place(x=650,y=15)

sarmaesh_combo_darkhast_maskoni=ttk.Combobox(frame_down_darkhast_maskoni)
sarmaesh_combo_darkhast_maskoni["values"] = ("ندارد","پنکه سقفی","کولر ابی","کولر گازی ","ابی/گازی")
sarmaesh_combo_darkhast_maskoni["state"]=["readonly"]
sarmaesh_combo_darkhast_maskoni.configure(justify="center")
sarmaesh_combo_darkhast_maskoni.place(x=475,y=15)

garmaesh_darkhast_maskoni=tk.Label(frame_down_darkhast_maskoni,text="گرمایش",background="#052340",fg="#ffffff",font=("Shabnam",11))
garmaesh_darkhast_maskoni.place(x=650, y=45)

garmaesh_combo_darkhast_maskoni=ttk.Combobox(frame_down_darkhast_maskoni)
garmaesh_combo_darkhast_maskoni["values"] = ("ندارد","بخاری"," شوفاژ","گرمایش از کف ")
garmaesh_combo_darkhast_maskoni["state"]=["readonly"]
garmaesh_combo_darkhast_maskoni.configure(justify="center")
garmaesh_combo_darkhast_maskoni.place(x=475, y=45)

kaf_darkhast_maskoni=tk.Label(frame_down_darkhast_maskoni,text="کف",background="#052340",fg="#ffffff",font=("Shabnam",11))
kaf_darkhast_maskoni.place(x=350, y=15)

kaf_combo_darkhast_maskoni=ttk.Combobox(frame_down_darkhast_maskoni)
kaf_combo_darkhast_maskoni["state"]=["readonly"]
kaf_combo_darkhast_maskoni.configure(justify="center")
kaf_combo_darkhast_maskoni["values"] = ("سرامیک","موزاییک","پارکت")
kaf_combo_darkhast_maskoni.place(x=150, y=15)

toilet_darkhast_maskoni=tk.Label(frame_down_darkhast_maskoni,text="سرویس بهداشتی",background="#052340",fg="#ffffff",font=("Shabnam",11))
toilet_darkhast_maskoni.place(x=330, y=45)

toilet_combo_darkhast_maskoni=ttk.Combobox(frame_down_darkhast_maskoni)
toilet_combo_darkhast_maskoni["state"]=["readonly"]
toilet_combo_darkhast_maskoni.configure(justify="center")
toilet_combo_darkhast_maskoni["values"] = ("ایرانی","فرنگی","هردو")
toilet_combo_darkhast_maskoni.place(x=150, y=45)

back_to_home_darkhast_maskoni=tk.Button(darkhast_maskoni_window,text="بازگشت",bg="#052340", fg="#ffffff",width=10,height=1,command=back_home_darkhast_maskoni)
back_to_home_darkhast_maskoni.place(x=400,y=30)

delete_btn_darkhast_maskoni=tk.Button(darkhast_maskoni_window,text="حذف",command=delete_darkhast_maskoni,bg="#8B0000",fg="#ffffff",height=1,width=10 )
delete_btn_darkhast_maskoni.place_forget()

edit_btn_darkhast_maskoni=tk.Button(darkhast_maskoni_window,text="ثبت ویرایش",command=update_darkhast_maskoni,bg="#00BFFF", fg="#ffffff",width=10,height=1,)
edit_btn_darkhast_maskoni.place_forget()


zakhire_darkhast_maskoni=tk.Button(darkhast_maskoni_window,text="ذخیره",bg="#00BFFF",fg="#ffffff",width=10,height=1,command=sabt_darkhast_maskoni)
zakhire_darkhast_maskoni.place(x=300,y=30)

error_lable_sal_sakht_darkhast_maskoni= tk.Label(darkhast_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_sal_sakht_darkhast_maskoni.place(x=900 , y=20)

error_lable_metraj_darkhast_maskoni= tk.Label(darkhast_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_metraj_darkhast_maskoni.place(x=900 , y=20)

error_lable_tabaghe_darkhast_maskoni= tk.Label(darkhast_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_tabaghe_darkhast_maskoni.place(x=900 , y=20)

error_lable_vahed_darkhast_maskoni= tk.Label(darkhast_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_vahed_darkhast_maskoni.place(x=900 , y=20)

error_lable_otagh_darkhast_maskoni= tk.Label(darkhast_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_otagh_darkhast_maskoni.place(x=900 , y=20)

error_lable_gheimat_kol_darkhast_maskoni= tk.Label(darkhast_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_gheimat_kol_darkhast_maskoni.place(x=900 , y=20)

error_lable_addrres_darkhast_maskoni= tk.Label(darkhast_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_addrres_darkhast_maskoni.place(x=835 , y=20)

error_lable_name_moshtari_darkhast_maskoni= tk.Label(darkhast_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_name_moshtari_darkhast_maskoni.place(x=835 , y=20)

error_lable_shomareh_moshtari_darkhast_maskoni= tk.Label(darkhast_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_shomareh_moshtari_darkhast_maskoni.place(x=900 , y=20)

error_lable_mablagh_ejare_darkhast_maskoni= tk.Label(darkhast_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_mablagh_ejare_darkhast_maskoni.place(x=900 , y=20)

error_lable_gheimat_pish_darkhast_maskoni= tk.Label(darkhast_maskoni_window, text="",fg="red",bg="#052340",font=("Shabnam",11))
error_lable_gheimat_pish_darkhast_maskoni.place(x=900 , y=20)

darkhast_maskoni_window.protocol("WM_DELETE_WINDOW", lambda: None)
darkhast_maskoni_window.resizable(False, False)
#endregion
#-----------------پنجره درخواست اداری/تجاری-------------------
#region
darkhast_edari_tejari_window = tk.Toplevel(root)
darkhast_edari_tejari_window.title("درخواست اداری و تجاری")
darkhast_edari_tejari_window.configure(bg="#052340")
darkhast_edari_tejari_window.geometry("1200x700")
darkhast_edari_tejari_window.withdraw()

#----------------------کادر درخواست اداری و تجاری------------------#

frame_up_right_darkhast_edari_tejari= tk.Frame(darkhast_edari_tejari_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=500,height=300)
frame_up_right_darkhast_edari_tejari.configure(bg="#052340")
frame_up_right_darkhast_edari_tejari.place(x=670,y=90)

frame_up_left_darkhast_edari_tejari= tk.Frame(darkhast_edari_tejari_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=300)
frame_up_left_darkhast_edari_tejari.configure(bg="#052340")
frame_up_left_darkhast_edari_tejari.place(x=10,y=90)

frame_midde_right_darkhast_edari_tejari= tk.Frame(darkhast_edari_tejari_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=500,height=150)
frame_midde_right_darkhast_edari_tejari.configure(bg="#052340")
frame_midde_right_darkhast_edari_tejari.place(x=670,y=410)

frame_midde_left_darkhast_edari_tejari= tk.Frame(darkhast_edari_tejari_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=150)
frame_midde_left_darkhast_edari_tejari.configure(bg="#052340")
frame_midde_left_darkhast_edari_tejari.place(x=10,y=410)

frame_down_darkhast_edari_tejari= tk.Frame(darkhast_edari_tejari_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=1160,height=100)
frame_down_darkhast_edari_tejari.configure(bg="#052340")
frame_down_darkhast_edari_tejari.place(x=10,y=580)

title_label_up1 = tk.Label(darkhast_edari_tejari_window,text="درخواست اداری و تجاری",bg="#052340",fg="#00BFFF",font=("Shabnam", 16))
title_label_up1.place(x=550, y=17)

title_label_up2 = tk.Label(darkhast_edari_tejari_window,text="ثبت اطلاعات درخواست اداری و تجاری",bg="#052340",fg="#ffffff",font=("Shabnam", 11))
title_label_up2.place(x=535, y=45)

label_up_right_darkhast_edari_tejari=tk.Label(darkhast_edari_tejari_window,text="اطلاعات ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_right_darkhast_edari_tejari.place(x=800,y=73)

label_up_left_darkhast_edari_tejari=tk.Label(darkhast_edari_tejari_window,text="تصویر ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_left_darkhast_edari_tejari.place(x=100,y=73)

label_midde_right_darkhast_edari_tejari=tk.Label(darkhast_edari_tejari_window,text="اطلاعات معامله",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_right_darkhast_edari_tejari.place(x=800,y=395)

label_midde_left_darkhast_edari_tejari=tk.Label(darkhast_edari_tejari_window,text="اطلاعات مالک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_left_darkhast_edari_tejari.place(x=100,y=395)

label_down_darkhast_edari_tejari=tk.Label(darkhast_edari_tejari_window,text="امکانات",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_down_darkhast_edari_tejari.place(x=100,y=565)

#-------------------------------------فریم بالا سمت راست------------------------------
melk_type_darkhast_edari_tejari_lable=tk.Label(frame_up_right_darkhast_edari_tejari, text="نوع ملک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
melk_type_darkhast_edari_tejari_lable.place(x=465,y=40, anchor="e")

combo_darkhast_edari_tejari_entry=ttk.Combobox(frame_up_right_darkhast_edari_tejari,state="readonly")
combo_darkhast_edari_tejari_entry["values"] = ("درخواست اجاره اداری و تجاری","درخواست خرید اداری و تجاری")
combo_darkhast_edari_tejari_entry.set("درخواست خرید اداری و تجاری")
combo_darkhast_edari_tejari_entry.configure(justify="center")
combo_darkhast_edari_tejari_entry.bind("<<ComboboxSelected>>",change_darkhast_edari_tejari_type)
combo_darkhast_edari_tejari_entry.place(x=18, y=30, width=350, height=25)

sal_sakht_darkhast_edari_tejari_lable=tk.Label(frame_up_right_darkhast_edari_tejari,text="سال ساخت",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
sal_sakht_darkhast_edari_tejari_lable.place(x=465, y=80, anchor="e")

sal_sakht_darkhast_edari_tejari_entry=tk.Entry(frame_up_right_darkhast_edari_tejari,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
sal_sakht_darkhast_edari_tejari_entry.place(x=18, y=70, width=350, height=25)

metraj_melk_darkhast_edari_tejari_lable=tk.Label(frame_up_right_darkhast_edari_tejari,text="متراژ",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
metraj_melk_darkhast_edari_tejari_lable.place(x=465, y=120, anchor="e")

metraj_melk_darkhast_edari_tejari_entry=tk.Entry(frame_up_right_darkhast_edari_tejari,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
metraj_melk_darkhast_edari_tejari_entry.place(x=18, y=110, width=350, height=25)

tabaghe_darkhast_edari_tejari_lable=tk.Label(frame_up_right_darkhast_edari_tejari,text="طبقه",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
tabaghe_darkhast_edari_tejari_lable.place(x=465, y=160, anchor="e")

tabaghe_darkhast_edari_tejari_entry=tk.Entry(frame_up_right_darkhast_edari_tejari,bg="#ffffff", fg="#000000",font=("Shabnam", 10),)
tabaghe_darkhast_edari_tejari_entry.place(x=18, y=150, width=350, height=25)

vahed_darkhast_edari_tejari_lable=tk.Label(frame_up_right_darkhast_edari_tejari,text="واحد",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
vahed_darkhast_edari_tejari_lable.place(x=465, y=200, anchor="e")

vahed_darkhast_edari_tejari_entry=tk.Entry(frame_up_right_darkhast_edari_tejari,bg="#ffffff", fg="#000000",font=("Shabnam", 10),)
vahed_darkhast_edari_tejari_entry.place(x=18, y=190, width=350, height=25)
#----------------------------فریم بالا سمت چپ----------------------------------
photo_lbl2_darkhast_edari_tejari = tk.Label(frame_up_left_darkhast_edari_tejari, text="[تصویر ملک]", bg="#ffffff",width=79, height=15,relief="solid")
photo_lbl2_darkhast_edari_tejari.place(x=40, y=10)

add_img_btn_darkhast_edari_tejari = tk.Button(frame_up_left_darkhast_edari_tejari, text="افزودن تصویر", bg="#00BFFF", fg="black",command=open_file,height=2,width=13)
add_img_btn_darkhast_edari_tejari.place(x=240, y=250)
#---------------------------فریم وسط سمت راست----------------------------------------
gheimat_kol_darkhast_edari_tejari_lable=tk.Label(frame_midde_right_darkhast_edari_tejari,text= "قیمت کل",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
gheimat_kol_darkhast_edari_tejari_lable.place(x=465, y=20, anchor="e")

gheimat_kol_darkhast_edari_tejari_entry=tk.Entry(frame_midde_right_darkhast_edari_tejari,bg="#ffffff", fg="#000000",font=("Shabnam", 10),)
gheimat_kol_darkhast_edari_tejari_entry.place(x=18, y=20, width=350, height=25)

mablagh_ejareh_darkhast_edari_tejari_lable=tk.Label(frame_midde_right_darkhast_edari_tejari,text="مبلغ اجاره",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
mablagh_ejareh_darkhast_edari_tejari_lable.place_forget()

mablagh_ejareh_darkhast_edari_tejari_entry=tk.Entry(frame_midde_right_darkhast_edari_tejari,bg="#FFFFFF", fg="#000000",font=("Shabnam", 10),)
mablagh_ejareh_darkhast_edari_tejari_entry.place_forget()

mablagh_vadie_darkhast_edari_tejari_lable=tk.Label(frame_midde_right_darkhast_edari_tejari,text="مبلغ ودیعه",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
mablagh_vadie_darkhast_edari_tejari_lable.place_forget()

mablagh_vadie_darkhast_edari_tejari_entry=tk.Entry(frame_midde_right_darkhast_edari_tejari,bg="#FFFFFF", fg="#000000",font=("Shabnam", 10),)
mablagh_vadie_darkhast_edari_tejari_entry.place_forget()

addrres_darkhast_edari_tejari_lable=tk.Label(frame_midde_right_darkhast_edari_tejari,text="آدرس",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
addrres_darkhast_edari_tejari_lable.place(x=465, y=110, anchor="e")

addrres_darkhast_edari_tejari_entry=tk.Text(frame_midde_right_darkhast_edari_tejari,bg="#ffffff", fg="#000000",font=("Shabnam", 10),)
addrres_darkhast_edari_tejari_entry.place(x=18, y=90, width=350, height=50)

#---------------------------فریم سمت چپ وسط-----------------------------------
name_moshtari_darkhast_edari_tejari_lable=tk.Label(frame_midde_left_darkhast_edari_tejari,text="نام مشتری",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
name_moshtari_darkhast_edari_tejari_lable.place(x=600, y=30,anchor="e")

name_moshtari_darkhast_edari_tejari_entry=tk.Entry(frame_midde_left_darkhast_edari_tejari,bg="#FFFFFF", fg="#000000",font=("Shabnam", 10),)
name_moshtari_darkhast_edari_tejari_entry.place(x=30, y=20, width=350, height=25)

shomareh_moshtari_darkhast_edari_tejari_lable=tk.Label(frame_midde_left_darkhast_edari_tejari,text="شماره مشتری",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
shomareh_moshtari_darkhast_edari_tejari_lable.place(x=600, y=80,anchor="e")

shomareh_moshtari_darkhast_edari_tejari_entry=tk.Entry(frame_midde_left_darkhast_edari_tejari,bg="#FFFFFF", fg="#000000",font=("Shabnam", 10))
shomareh_moshtari_darkhast_edari_tejari_entry.place(x=30, y=70, width=350, height=25)
#-------------------------فریم پایین----------------------------------
parking_darkhast_edari_tejari_var=tk.IntVar(value=0)
anbari_darkhast_edari_tejari_var=tk.IntVar(value=0)
asansor_darkhast_edari_tejari_var=tk.IntVar(value=0)

parking_check_btn_darkhast_edari_tejari=tk.Checkbutton(frame_down_darkhast_edari_tejari,variable=parking_darkhast_edari_tejari_var,image=parking_pic,background="#052340")
parking_check_btn_darkhast_edari_tejari.place(x=1050, y=10)

parking_ch_btn_darkhast_edari_tejari_label=tk.Label(frame_down_darkhast_edari_tejari,text="پارکینگ", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
parking_ch_btn_darkhast_edari_tejari_label.place(x=1055,y=60)

asansor_check_btn_darkhast_edari_tejari=tk.Checkbutton(frame_down_darkhast_edari_tejari,variable=asansor_darkhast_edari_tejari_var,image=elvator_pic,background="#052340")
asansor_check_btn_darkhast_edari_tejari.place(x=950, y=10)

asansor_ch_btn_darkhast_edari_tejari_label=tk.Label(frame_down_darkhast_edari_tejari,text="آسانسور", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
asansor_ch_btn_darkhast_edari_tejari_label.place(x=955,y=60)

anbari_check_btn_darkhast_edari_tejari=tk.Checkbutton(frame_down_darkhast_edari_tejari,variable=anbari_darkhast_edari_tejari_var,image=warehouse_pic,background="#052340")
anbari_check_btn_darkhast_edari_tejari.place(x=850, y=10)

anbari_checkbuton_darkhast_edari_tejari_label=tk.Label(frame_down_darkhast_edari_tejari,text="انباری", bg="#052340", fg="#ffffff", font=("Shabnam", 9), width=7)
anbari_checkbuton_darkhast_edari_tejari_label.place(x=855,y=60)

aab_va_gaz_emkanat_darkhast_edari_tejari=tk.Label(frame_down_darkhast_edari_tejari,text="وضعیت آب و گاز",background="#052340",fg="#ffffff",font=("Shabnam",11),width=17)
aab_va_gaz_emkanat_darkhast_edari_tejari.place(x=650, y=15)

aab_va_gaz_combo_emkanat_darkhast_edari_tejari=ttk.Combobox(frame_down_darkhast_edari_tejari)
aab_va_gaz_combo_emkanat_darkhast_edari_tejari["values"] = ("فقط گاز دارد","فقط آب دارد","آب و گاز دارد")
aab_va_gaz_combo_emkanat_darkhast_edari_tejari["state"]=["readonly"]
aab_va_gaz_combo_emkanat_darkhast_edari_tejari.place(x=475, y=15)

sarmayesh_emkanat_darkhast_edari_tejari=tk.Label(frame_down_darkhast_edari_tejari,text="سیستم سرمایش",background="#052340",fg="#ffffff",font=("Shabnam",11),width=17)
sarmayesh_emkanat_darkhast_edari_tejari.place(x=350, y=45)

sarmayesh_combo_emkanat_darkhast_edari_tejari=ttk.Combobox(frame_down_darkhast_edari_tejari)
sarmayesh_combo_emkanat_darkhast_edari_tejari["values"] = (" کولر گازی"," کولرآبی","پنکه سقفی","ندارد")
sarmayesh_combo_emkanat_darkhast_edari_tejari["state"]=["readonly"]
sarmayesh_combo_emkanat_darkhast_edari_tejari.place(x=150,y=45)

garmayesh_emkanat_darkhast_edari_tejari=tk.Label(frame_down_darkhast_edari_tejari,text="سیستم گرمایش",background="#052340",fg="#ffffff",font=("Shabnam",11),width=17)
garmayesh_emkanat_darkhast_edari_tejari.place(x=350, y=15)

garmayesh_combo_emkanat_darkhast_edari_tejari=ttk.Combobox(frame_down_darkhast_edari_tejari)
garmayesh_combo_emkanat_darkhast_edari_tejari["values"] = (" شوفاژ"," بخاری","ندارد")
garmayesh_combo_emkanat_darkhast_edari_tejari["state"]=["readonly"]
garmayesh_combo_emkanat_darkhast_edari_tejari.place(x=150, y=15)

back_to_home_darkhast_edari_tejari=tk.Button(darkhast_edari_tejari_window,text="بازگشت",bg="#052340", fg="#ffffff",width=10,height=1,command=back_home_darkhast_edari_tejari)
back_to_home_darkhast_edari_tejari.place(x=400,y=30)

zakhire_darkhast_edari_tejari=tk.Button(darkhast_edari_tejari_window,text="ذخیره",bg="#00BFFF", fg="#ffffff",width=10,height=1,command=sabt_darkhast_edari_tejari)
zakhire_darkhast_edari_tejari.place(x=300,y=30)

delete_btn_darkhast_edari_tejari=tk.Button(darkhast_edari_tejari_window,text="حذف",command=delete_darkhast_edari_tejari,bg="#8B0000",fg="#ffffff",height=1,width=10 )
delete_btn_darkhast_edari_tejari.place_forget()

edit_btn_darkhast_edari_tejari=tk.Button(darkhast_edari_tejari_window,text="ثبت ویرایش",command=update_darkhast_edari_tejari,bg="#00BFFF", fg="#ffffff",width=10,height=1,)
edit_btn_darkhast_edari_tejari.place_forget()

darkhast_edari_tejari_window.protocol("WM_DELETE_WINDOW", lambda: None)
darkhast_edari_tejari_window.resizable(False, False)
#endregion
#--------------------پنجره درخواست باغ/زمین-----------------------
#region
darkhast_bagh_zamin_window = tk.Toplevel(root)
darkhast_bagh_zamin_window.title("درخواست باغ و زمین")
darkhast_bagh_zamin_window.geometry("1200x700")
darkhast_bagh_zamin_window.configure(bg="#052340")
darkhast_bagh_zamin_window.withdraw()
#---------------------کادر درخواست باغ و زمین---------------------#
frame_up_right_darkhast_bagh_zamin= tk.Frame(darkhast_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=520,height=300)
frame_up_right_darkhast_bagh_zamin.configure(bg="#052340")
frame_up_right_darkhast_bagh_zamin.place(x=670,y=75)


frame_up_left_darkhast_bagh_zamin= tk.Frame(darkhast_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=300)
frame_up_left_darkhast_bagh_zamin.configure(bg="#052340")
frame_up_left_darkhast_bagh_zamin.place(x=10,y=75)

frame_midde_right_darkhast_bagh_zamin= tk.Frame(darkhast_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=520,height=150)
frame_midde_right_darkhast_bagh_zamin.configure(bg="#052340")
frame_midde_right_darkhast_bagh_zamin.place(x=670,y=390)

frame_midde_left_darkhast_bagh_zamin= tk.Frame(darkhast_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=150)
frame_midde_left_darkhast_bagh_zamin.configure(bg="#052340")
frame_midde_left_darkhast_bagh_zamin.place(x=10,y=390)

frame_down_darkhast_bagh= tk.Frame(darkhast_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=1180,height=130)
frame_down_darkhast_bagh.configure(bg="#052340")
frame_down_darkhast_bagh.place(x=10,y=555)

frame_down_darkhast_zamin=tk.Frame(darkhast_bagh_zamin_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=1180,height=130)
frame_down_darkhast_zamin.configure(bg="#052340")
frame_down_darkhast_zamin.place_forget()

title_label_darkhast_bagh_zamin_up1 = tk.Label(darkhast_bagh_zamin_window,text="درخواست  باغ و زمین",bg="#052340",fg="#00BFFF",font=("Shabnam", 16))
title_label_darkhast_bagh_zamin_up1.place(x=600, y=10)

title_label_darkhast_bagh_zamin_up2 = tk.Label(darkhast_bagh_zamin_window,text="ثبت اطلاعات درخواست باغ و زمین",bg="#052340",fg="#ffffff",font=("Shabnam", 11))
title_label_darkhast_bagh_zamin_up2.place(x=585, y=40)

label_up_right_darkhast_bagh_zamin=tk.Label(darkhast_bagh_zamin_window,text="اطلاعات ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_right_darkhast_bagh_zamin.place(x=760,y=60)

label_up_left_darkhast_bagh_zamin=tk.Label(darkhast_bagh_zamin_window,text="تصویر ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_left_darkhast_bagh_zamin.place(x=100,y=60)

label_midde_right_darkhast_bagh_zamin=tk.Label(darkhast_bagh_zamin_window,text="اطلاعات معامله",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_right_darkhast_bagh_zamin.place(x=800,y=375)

label_midde_left_darkhast_bagh_zamin=tk.Label(darkhast_bagh_zamin_window,text="اطلاعات مالک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_left_darkhast_bagh_zamin.place(x=100,y=375)

label_down_darkhast_bagh_zamin=tk.Label(darkhast_bagh_zamin_window,text="امکانات",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_down_darkhast_bagh_zamin.place(x=100,y=545)
#-------------------------------فریم بالا راست-----------------------------
melk_type_darkhast_bagh_zamin_lable=tk.Label(frame_up_right_darkhast_bagh_zamin,text="نوع ملک",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=10)
melk_type_darkhast_bagh_zamin_lable.place(x=490,y=40, anchor="e")

melk_type_darkhast_bagh_zamin_entry=ttk.Combobox(frame_up_right_darkhast_bagh_zamin,font=("Shabnam", 10),justify="center",state="readonly")
melk_type_darkhast_bagh_zamin_entry["values"]=("درخواست خرید باغ زمین","درخواست اجاره باغ زمین")
melk_type_darkhast_bagh_zamin_entry.set("درخواست خرید باغ زمین")
melk_type_darkhast_bagh_zamin_entry.configure(justify="center")
melk_type_darkhast_bagh_zamin_entry.bind("<<ComboboxSelected>>",sabt_darkhast_bagh_zamin)
melk_type_darkhast_bagh_zamin_entry.place(x=28, y=30, width=350, height=25)

metraj_zamin_darkhast_bagh_zamin_lable=tk.Label(frame_up_right_darkhast_bagh_zamin,text="متراژ",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=10)
metraj_zamin_darkhast_bagh_zamin_lable.place(x=490, y=80, anchor="e")

metraj_zamin_darkhast_bagh_zamin_entry=tk.Entry(frame_up_right_darkhast_bagh_zamin,bg="#ffffff", fg="#000000",font=("Shabnam", 10),textvariable="متر مربع")
metraj_zamin_darkhast_bagh_zamin_entry.place(x=28, y=70, width=350, height=25)

bagh_type_darkhast_bagh_zamin_lable=tk.Label(frame_up_right_darkhast_bagh_zamin,text="کاربری زمین",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=10)
bagh_type_darkhast_bagh_zamin_lable.place(x=490, y=120, anchor="e")

bagh_type_darkhast_bagh_zamin_combo=ttk.Combobox(frame_up_right_darkhast_bagh_zamin,state="readonly")
bagh_type_darkhast_bagh_zamin_combo["values"]=("باغ","زمین کشاورزی")
bagh_type_darkhast_bagh_zamin_combo.set("باغ")
bagh_type_darkhast_bagh_zamin_combo.configure(justify="center")
bagh_type_darkhast_bagh_zamin_combo.place(x=28, y=110, width=350, height=25)
bagh_type_darkhast_bagh_zamin_combo.bind("<<ComboboxSelected>>",change_bagh_zamin_darkhast_bagh)

time_ejareh_bagh_darkhast_zamin_lable=tk.Label(frame_up_right_darkhast_bagh_zamin,text="مدت اجاره",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
time_ejareh_bagh_darkhast_zamin_lable.place_forget()

bagh_time_darkhast_combo=ttk.Combobox(frame_up_right_darkhast_bagh_zamin,state="readonly")
bagh_time_darkhast_combo["values"]=("بلندمدت","کوتاه مدت","فصلی","سالانه")
bagh_time_darkhast_combo.configure(justify="center")
bagh_time_darkhast_combo.set("فصلی")
bagh_time_darkhast_combo.place_forget()
#-------------------------------------------فریم بالا سمت چپ----------------------
photo_darkhast_bagh_zamin_lable= tk.Label(frame_up_left_darkhast_bagh_zamin, text="[تصویر ملک]", bg="#ffffff", width=79, height=15)
photo_darkhast_bagh_zamin_lable.place(x=40, y=10)

add_img_btn_darkhast_bagh_zamin = tk.Button(frame_up_left_darkhast_bagh_zamin, text="افزودن تصویر", bg="#00BFFF", fg="#000000",command=open_file,height=2,width=13)
add_img_btn_darkhast_bagh_zamin.place(x=240, y=250)
#---------------------------------------فریم وسط سمت راست--------------------------
gheimat_kol_bagh_zamin_darkhast_lable=tk.Label(frame_midde_right_darkhast_bagh_zamin,text='قیمت کل',bg="#052340",fg="#ffffff",font=("Shabnam",12),width=10)
gheimat_kol_bagh_zamin_darkhast_lable.place(x=490, y=20, anchor="e")

gheimat_kol_bagh_zamin_darkhast_entry=tk.Entry(frame_midde_right_darkhast_bagh_zamin,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
gheimat_kol_bagh_zamin_darkhast_entry.place(x=28, y=13, width=350, height=25)

gheimat_har_matr_bagh_zamin_darkhast_lable=tk.Label(frame_midde_right_darkhast_bagh_zamin,text='قیمت هر متر',bg="#052340",fg="#ffffff",font=("Shabnam",12),width=10)
gheimat_har_matr_bagh_zamin_darkhast_lable.place(x=490, y=65, anchor="e")

gheimat_har_metr_bagh_zamin_darkhast_entry=tk.Entry(frame_midde_right_darkhast_bagh_zamin,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
gheimat_har_metr_bagh_zamin_darkhast_entry.place(x=28, y=55, width=350, height=25)

gheimat_ejareh_bagh_darkhast_zamin_lable=tk.Label(frame_midde_right_darkhast_bagh_zamin, text="ودیعه", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
gheimat_ejareh_bagh_darkhast_zamin_lable.place_forget()

gheimat_ejareh_bagh_darkhast_zamin_entry=tk.Entry(frame_midde_right_darkhast_bagh_zamin, bg="#FFFFFF", fg="#000000", font=("Shabnam", 10))
gheimat_ejareh_bagh_darkhast_zamin_entry.place_forget()

mablagh_ejareh_mahaneh_darkhast_lable=tk.Label(frame_midde_right_darkhast_bagh_zamin,text='اجاره ماهانه',bg="#052340",fg="#ffffff",font=("Shabnam",12),width=10)
mablagh_ejareh_mahaneh_darkhast_lable.place_forget()

mablagh_ejareh_mahaneh_darkhast_entry=tk.Entry(frame_midde_right_darkhast_bagh_zamin,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
mablagh_ejareh_mahaneh_darkhast_entry.place_forget()

bagh_loctaion_darkhast_bagh_zamin_lable=tk.Label(frame_midde_right_darkhast_bagh_zamin,text="منطقه و آدرس ",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=10)
bagh_loctaion_darkhast_bagh_zamin_lable.place(x=490, y=110, anchor="e")

bagh_loctaion_darkhast_bagh_zamin_entry=tk.Text(frame_midde_right_darkhast_bagh_zamin,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
bagh_loctaion_darkhast_bagh_zamin_entry.place(x=28, y=95, width=350, height=40)
#----------------------------فریم وسط سمت چپ-------------------------------
name_moshtari_darkhast_bagh_lable=tk.Label(frame_midde_left_darkhast_bagh_zamin,text="نام مشتری",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
name_moshtari_darkhast_bagh_lable.place(x=600, y=30, anchor="e")

name_moshtari_darkhast_bagh_entry=tk.Entry(frame_midde_left_darkhast_bagh_zamin,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
name_moshtari_darkhast_bagh_entry.place(x=30, y=20, width=350, height=25)

shomareh_moshtari_darkhast_bagh_lable=tk.Label(frame_midde_left_darkhast_bagh_zamin,text="شماره مشتری",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=9)
shomareh_moshtari_darkhast_bagh_lable.place(x=600, y=80, anchor="e")

shomareh_moshtari_darkhast_bagh_entry=tk.Entry(frame_midde_left_darkhast_bagh_zamin,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
shomareh_moshtari_darkhast_bagh_entry.place(x=30, y=70, width=350, height=25)
#--------------------------فریم پایین-----------------------------
abyari_darkhast_bagh_zamin_lable=tk.Label(frame_down_darkhast_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="نوع آبیاری")
abyari_darkhast_bagh_zamin_lable.place(x=1080, y=20)

abyari_darkhast_bagh_zamin_combo=ttk.Combobox(frame_down_darkhast_bagh,width=12)
abyari_darkhast_bagh_zamin_combo["values"]=("سطحی","بارانی","قطره ای","تحت فشار")
abyari_darkhast_bagh_zamin_combo["state"]=["readonly"]
abyari_darkhast_bagh_zamin_combo.configure(justify="center")
abyari_darkhast_bagh_zamin_combo.set("سطحی")
abyari_darkhast_bagh_zamin_combo.place(x=970, y=20)

metraj_derakht_darkhast_bagh_zamin_lable=tk.Label(frame_down_darkhast_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="متراژ درخت کاری")
metraj_derakht_darkhast_bagh_zamin_lable.place(x=1080, y=60)

metraj_derakht_darkhast_bagh_zamin_entry=tk.Entry(frame_down_darkhast_bagh,width=10,bg="#ffffff",fg="#000000")
metraj_derakht_darkhast_bagh_zamin_entry.place(x=1000, y=60)

tedad_derakht_darkhast_bagh_zamin_lable=tk.Label(frame_down_darkhast_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="تعداد درخت")
tedad_derakht_darkhast_bagh_zamin_lable.place(x=1080, y=100)

tedad_derakht_darkhast_bagh_zamin_entry=tk.Entry(frame_down_darkhast_bagh,width=10,bg="#ffffff",fg="#000000")
tedad_derakht_darkhast_bagh_zamin_entry.place(x=1000, y=100)

type_tree_darkhast_bagh_zamin_lable=tk.Label(frame_down_darkhast_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="نوع درخت")
type_tree_darkhast_bagh_zamin_lable.place(x=870, y=20)

type_tree_darkhast_bagh_zamin_combo=ttk.Combobox(frame_down_darkhast_bagh,width=12)
type_tree_darkhast_bagh_zamin_combo["values"]=(" ","پسته","بادام","گردو","شلیل","هلو","سیب","انگور","انجیر","زردالو","گیلاس","آلبالو")
type_tree_darkhast_bagh_zamin_combo.set("گردو")
type_tree_darkhast_bagh_zamin_combo.configure(justify="center")
type_tree_darkhast_bagh_zamin_combo["state"]=["readonly"]
type_tree_darkhast_bagh_zamin_combo.place(x=770, y=20)

type_tree_darkhast_btn=tk.Button(frame_down_darkhast_bagh,text="افزودن درخت",command=add_tree3,bg="#00BFFF",font=("Shabnam",9),width=10)
type_tree_darkhast_btn.place(x=870, y=60)

label_natige_darkhast_bagh_zamin=tk.Text(frame_down_darkhast_bagh,wrap="word",width=11,height=4,font=("Shabnam",9))
label_natige_darkhast_bagh_zamin.place(x=770, y=50)

chah_darkhast_bagh_zamin_var=tk.IntVar(value=0)
chah_darkhast_bagh_zamin=tk.Checkbutton(frame_down_darkhast_bagh,variable=chah_darkhast_bagh_zamin_var,text="چاه",background="#052340",fg="#00BFFF",font=("Shabnam",9))
chah_darkhast_bagh_zamin.place(x=685, y=20)

estakhr_darkhast_bagh_zamin_var=tk.IntVar(value=0)
estakhr_darkhast_bagh_zamin=tk.Checkbutton(frame_down_darkhast_bagh,variable=estakhr_darkhast_bagh_zamin_var,text="استخر",background="#052340",fg="#00BFFF",font=("Shabnam",9))
estakhr_darkhast_bagh_zamin.place(x=600, y=20)

divar_darkhast_bagh_zamin_var=tk.IntVar(value=0)
divar_darkhast_bagh_zamin=tk.Checkbutton(frame_down_darkhast_bagh,variable=divar_darkhast_bagh_zamin_var,text="دیوار کشی",background="#052340",fg="#00BFFF",font=("Shabnam",9))
divar_darkhast_bagh_zamin.place(x=685, y=80)

bargh_keshi_darkhast_bagh_zamin_var=tk.IntVar(value=0)
bargh_keshi_darkhast_bagh_zamin=tk.Checkbutton(frame_down_darkhast_bagh,variable=bargh_keshi_darkhast_bagh_zamin_var,text="برق کشی",background="#052340",fg="#00BFFF",font=("Shabnam",9))
bargh_keshi_darkhast_bagh_zamin.place(x=600, y=50)

gaz_keshi_darkhast_bagh_zamin_var=tk.IntVar(value=0)
gaz_keshi_darkhast_bagh_zamin=tk.Checkbutton(frame_down_darkhast_bagh,variable=bargh_keshi_darkhast_bagh_zamin_var,text="گاز کشی",background="#052340",fg="#00BFFF",font=("Shabnam",9))
gaz_keshi_darkhast_bagh_zamin.place(x=685, y=50)

var0_darkhast_bagh_zamin=tk.IntVar(value=0)#چک باتن پیش فرض تیک نخورده باشه

otagh_check_btn_darkhast_bagh_zamin=tk.Checkbutton(frame_down_darkhast_bagh,variable=var0_darkhast_bagh_zamin,image=warehouse_pic,background="#052340",text="ساختمان",command=home_true_false3)
otagh_check_btn_darkhast_bagh_zamin.place(x=600, y=70)

otagh_check_btn_darkhast_bagh_zamin_label=tk.Label(frame_down_darkhast_bagh,text="ویلا", bg="#052340", fg="#ffffff", font=("Shabnam", 7), width=7)
otagh_check_btn_darkhast_bagh_zamin_label.place(x=620,y=110)


type_vila_darkhast_bagh_zamin=tk.Label(frame_down_darkhast_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="نوع سازه")
type_vila_darkhast_bagh_zamin.place(x=490, y=20)

type_vila_darkhast_bagh_zamin_combo=ttk.Combobox(frame_down_darkhast_bagh,width=12)
type_vila_darkhast_bagh_zamin_combo["values"]=("آجری","بلوکی","کانکس","چوبی")
type_vila_darkhast_bagh_zamin_combo.config(state="disabled")
type_vila_darkhast_bagh_zamin_combo.configure(justify="center")
type_vila_darkhast_bagh_zamin_combo.set("آجری")
type_vila_darkhast_bagh_zamin_combo.place(x=380, y=20)

toilet_darkhast_bagh_zamin_lable=tk.Label(frame_down_darkhast_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="سرویس بهداشتی")
toilet_darkhast_bagh_zamin_lable.place(x=490, y=60)

toilet_darkhast_bagh_zamin_combo=ttk.Combobox(frame_down_darkhast_bagh,state="disabled",width=12)
toilet_darkhast_bagh_zamin_combo["values"]=("ندارد","فرنگی","ایرانی","هردو")
toilet_darkhast_bagh_zamin_combo.set("ندارد")
toilet_darkhast_bagh_zamin_combo.configure(justify="center")
toilet_darkhast_bagh_zamin_combo.place(x=380, y=60)

hamam_darkhast_bagh_zamin=tk.Label(frame_down_darkhast_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="حمام")
hamam_darkhast_bagh_zamin.place(x=490, y=95)

hamam_darkhast_bagh_zamin_combo=ttk.Combobox(frame_down_darkhast_bagh,state="disabled",width=12)
hamam_darkhast_bagh_zamin_combo["values"]=("ندارد","دارد")
hamam_darkhast_bagh_zamin_combo.set("ندارد")
hamam_darkhast_bagh_zamin_combo.configure(justify="center")
hamam_darkhast_bagh_zamin_combo.place(x=380, y=100)

sanad_darkhast_bagh_zamin_lable=tk.Label(frame_down_darkhast_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="سند")
sanad_darkhast_bagh_zamin_lable.place(x=300, y=60)

sanad_darkhast_bagh_zamin_combo=ttk.Combobox(frame_down_darkhast_bagh,state="disabled",width=12)
sanad_darkhast_bagh_zamin_combo["values"]=("ندارد","تک برگ","قولنامه ای","مشاع")
sanad_darkhast_bagh_zamin_combo.set("ندارد")
sanad_darkhast_bagh_zamin_combo.configure(justify="center")
sanad_darkhast_bagh_zamin_combo.place(x=180,y=60)

option_darkhast_bagh_zamin=tk.Label(frame_down_darkhast_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="امکانات تفریحی")
option_darkhast_bagh_zamin.place(x=310, y=20)

option_darkhast_bagh_zamin_combo=ttk.Combobox(frame_down_darkhast_bagh,state="disabled",width=12)
option_darkhast_bagh_zamin_combo["values"]=("استخر","جکوزی","باربیکیو")
option_darkhast_bagh_zamin_combo.set("استخر")
option_darkhast_bagh_zamin_combo.configure(justify="center")
option_darkhast_bagh_zamin_combo.place(x=180, y=20)

add_option_button_darkhast_bagh_zamin=tk.Button(frame_down_darkhast_bagh,text="افزودن امکانات",command=add_option3,bg="#00BFFF",font=("Shabnam",9),width=10)
add_option_button_darkhast_bagh_zamin.place(x=85, y=20)

lable_natige_add_darkhast_bagh_zamin=tk.Label(frame_down_darkhast_bagh,text="",width=10,bg="#ffffff",fg="#000000")
lable_natige_add_darkhast_bagh_zamin.place(x=5, y=20)

metraj_vila_darkhast_bagh_zamin=tk.Label(frame_down_darkhast_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="متراژ سازه")
metraj_vila_darkhast_bagh_zamin.place(x=290, y=100)

metraj_vila_darkhast_bagh_zamin_entry=tk.Entry(frame_down_darkhast_bagh,width=10,bg="#ffffff",fg="#000000",state="disabled")
metraj_vila_darkhast_bagh_zamin_entry.place(x=230, y=100)

sal_sakht_vila_darkhast_bagh_zamin_lable=tk.Label(frame_down_darkhast_bagh,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=13,text="سال ساخت")
sal_sakht_vila_darkhast_bagh_zamin_lable.place(x=80, y=60)

sal_sakht_vila_darkhast_bagh_zamin_entry=tk.Entry(frame_down_darkhast_bagh,width=10,bg="#ffffff",fg="#000000",state="disabled")
sal_sakht_vila_darkhast_bagh_zamin_entry.place(x=10, y=60)

mojavez_sakht_darkhast_bagh_zamin_var=tk.IntVar(value=0)
mojavez_sakht_check_btn_darkhast_bagh_zamin=tk.Checkbutton(frame_down_darkhast_bagh,variable=mojavez_sakht_darkhast_bagh_zamin_var,text="مجوز ساختن",background="#052340",fg="#00BFFF",font=("Shabnam",9),state="disabled")
mojavez_sakht_check_btn_darkhast_bagh_zamin.place(x=10, y=95)

mohavate_sazi_darkhast_bagh_zamin_var=tk.IntVar(value=0)
mohavate_sazi_check_btn_darkhast_bagh_zamin=tk.Checkbutton(frame_down_darkhast_bagh,variable=mohavate_sazi_darkhast_bagh_zamin_var,text="محوطه سازی",background="#052340",fg="#00BFFF",font=("Shabnam",9),state="disabled")
mohavate_sazi_check_btn_darkhast_bagh_zamin.place(x=120, y=95)
#-------------------------تعویض کاربری به زمین در قسمت درخواست باغ/زمین-------------
karbari_darkhast_bagh_zamin_lable=tk.Label(frame_down_darkhast_zamin,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="نوع کاربری")
karbari_darkhast_bagh_zamin_lable.place(x=1000,y=20)

karbari_darkhast_bagh_zamin_combo=ttk.Combobox(frame_down_darkhast_zamin)
karbari_darkhast_bagh_zamin_combo["values"]=(" ","زراعی","باغی","گلخانه ای","دامداری ","مرغداری","دامداری و مرغداری","آیش")           
karbari_darkhast_bagh_zamin_combo["state"]=["readonly"]                  
karbari_darkhast_bagh_zamin_combo.set(" ")
karbari_darkhast_bagh_zamin_combo.configure(justify="center")
karbari_darkhast_bagh_zamin_combo.place(x=800, y=20)

khak_darkhast_bagh_zamin=tk.Label(frame_down_darkhast_zamin,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="نوع خاک")
khak_darkhast_bagh_zamin.place(x=1000, y=60)

khak_darkhast_bagh_zamin_combo=ttk.Combobox(frame_down_darkhast_zamin)
khak_darkhast_bagh_zamin_combo["values"]=(" ","رسی","شنی","لومی","رسی_شنی","شنی_لومی","رسی_لومی")        
khak_darkhast_bagh_zamin_combo["state"]=["readonly"]                     
khak_darkhast_bagh_zamin_combo.set(" ")
khak_darkhast_bagh_zamin_combo.configure(justify="center")
khak_darkhast_bagh_zamin_combo.place(x=800, y=60)

ab_darkhast_bagh_zamin=tk.Label(frame_down_darkhast_zamin,bg="#052340",fg="#ffffff",font=("Shabnam",9),width=10,text="منبع آب")
ab_darkhast_bagh_zamin.place(x=1000, y=100)

ab_darkhast_bagh_zamin_combo=ttk.Combobox(frame_down_darkhast_zamin)
ab_darkhast_bagh_zamin_combo["values"]=(" ","چاه","قنات","رودخانه","کانال آبیاری","چشمه","آب لوله کشی کشاورزی","تانکر","استخر")  
ab_darkhast_bagh_zamin_combo["state"]=["readonly"]
ab_darkhast_bagh_zamin_combo.set(" ")
ab_darkhast_bagh_zamin_combo.configure(justify="center")
ab_darkhast_bagh_zamin_combo.place(x=800, y=100)

security_zamin_darkhast_bagh_zamin_var=tk.IntVar(value=0)
security_zamin_darkhast_bagh_zamin=tk.Checkbutton(frame_down_darkhast_zamin,variable=security_zamin_darkhast_bagh_zamin_var,text="اتاق نگهبان",background="#052340",fg="#00BFFF",font=("Shabnam",9))
security_zamin_darkhast_bagh_zamin.place(x=300, y=20)

bargh_kesi_zamin_darkhast_bagh_zamin_var=tk.IntVar(value=0)
bargh_kesi_zamin_darkhast_bagh_zamin=tk.Checkbutton(frame_down_darkhast_zamin,variable=bargh_kesi_zamin_darkhast_bagh_zamin_var,text="برق تک فاز",background="#052340",fg="#00BFFF",font=("Shabnam",9))
bargh_kesi_zamin_darkhast_bagh_zamin.place(x=500, y=20)

bargh_zamin_darkhast_bagh_zamin2_var=tk.IntVar(value=0)
bargh_keshi_zamin_darkhast_bagh_zamin2=tk.Checkbutton(frame_down_darkhast_zamin,variable=bargh_zamin_darkhast_bagh_zamin2_var,text="برق سه فاز",background="#052340",fg="#00BFFF",font=("Shabnam",9))
bargh_keshi_zamin_darkhast_bagh_zamin2.place(x=400, y=20)

anbar_zamin_darkhast_bagh_zamin_var=tk.IntVar(value=0)
anbar_zamin_darkhast_bagh_zamin=tk.Checkbutton(frame_down_darkhast_zamin,variable=anbar_zamin_darkhast_bagh_zamin_var,text="انبار/سوله",background="#052340",fg="#00BFFF",font=("Shabnam",9))
anbar_zamin_darkhast_bagh_zamin.place(x=500, y=60)

fans_zamin_darkhast_bagh_zamin_var=tk.IntVar(value=0)
fans_zamin_darkhast_bagh_zamin=tk.Checkbutton(frame_down_darkhast_zamin,variable=fans_zamin_darkhast_bagh_zamin_var,text="فنس/دیوار",background="#052340",fg="#00BFFF",font=("Shabnam",9))
fans_zamin_darkhast_bagh_zamin.place(x=400, y=60)

mojavez_chah_zamin_darkhast_bagh_zamin_var=tk.IntVar(value=0)
mojavez_chah_zamin_darkhast_bagh_zamin=tk.Checkbutton(frame_down_darkhast_zamin,variable=mojavez_chah_zamin_darkhast_bagh_zamin_var,text="اجازه حفر چاه",background="#052340",fg="#00BFFF",font=("Shabnam",9))
mojavez_chah_zamin_darkhast_bagh_zamin.place(x=300, y=60)

back_to_home_darkhast_bagh_zamin=tk.Button(darkhast_bagh_zamin_window,text="بازگشت",bg="#052340", fg="#ffffff",width=10,height=1,command=back_home_darkhast_bagh)
back_to_home_darkhast_bagh_zamin.place(x=300,y=30)

zakhire_darkhast_bagh_zamin=tk.Button(darkhast_bagh_zamin_window,text="ذخیره",bg="#00BFFF", fg="#ffffff",width=10,height=1,command=sabt_darkhast_bagh_zamin)
zakhire_darkhast_bagh_zamin.place(x=200,y=30)

darkhast_bagh_zamin_window.protocol("WM_DELETE_WINDOW", lambda: None)
darkhast_bagh_zamin_window.resizable(False, False)
#endregion
#-------------------پنجره درخواست کارگاه------------------------
#region
darkhast_karghah_window= tk.Toplevel(root)
darkhast_karghah_window.title("درخواست کارگاه")
darkhast_karghah_window.geometry("1200x700")
darkhast_karghah_window.configure(bg="#052340")
darkhast_karghah_window.withdraw()

frame_up_right_darkhast_karghah = tk.Frame(darkhast_karghah_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=510,height=300)
frame_up_right_darkhast_karghah.configure(bg="#052340")
frame_up_right_darkhast_karghah.place(x=670,y=90)

frame_up_left_darkhast_karghah= tk.Frame(darkhast_karghah_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=300)
frame_up_left_darkhast_karghah.configure(bg="#052340")
frame_up_left_darkhast_karghah.place(x=10,y=90)

frame_midde_right_darkhast_karghah= tk.Frame(darkhast_karghah_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=510,height=150)
frame_midde_right_darkhast_karghah.configure(bg="#052340")
frame_midde_right_darkhast_karghah.place(x=670,y=410)

frame_midde_left_darkhast_karghah= tk.Frame(darkhast_karghah_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=640,height=150)
frame_midde_left_darkhast_karghah.configure(bg="#052340")
frame_midde_left_darkhast_karghah.place(x=10,y=410)

frame_down_darkhast_karghah= tk.Frame(darkhast_karghah_window,bd=0,highlightthickness=1,highlightbackground="#00BFFF",width=1170,height=85)
frame_down_darkhast_karghah.configure(bg="#052340")
frame_down_darkhast_karghah.place(x=10,y=580)

title_label_up1_darkhast_karghah= tk.Label(darkhast_karghah_window,text="درخواست کارگاه",bg="#052340",fg="#00BFFF",font=("Shabnam", 16))
title_label_up1_darkhast_karghah.place(x=570, y=17)   

title_label_up2_darkhast_karghah = tk.Label(darkhast_karghah_window,text="ثبت اطلاعات درخواست کارگاه",bg="#052340",fg="#ffffff",font=("Shabnam", 11))
title_label_up2_darkhast_karghah.place(x=555, y=45)

label_up_right_darkhast_karghah=tk.Label(darkhast_karghah_window,text="اطلاعات ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_right_darkhast_karghah.place(x=800,y=73)

label_up_left_darkhast_karghah=tk.Label(darkhast_karghah_window,text="تصویر ملک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_up_left_darkhast_karghah.place(x=100,y=73)

label_midde_right_darkhast_karghah=tk.Label(darkhast_karghah_window,text="اطلاعات معامله",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_right_darkhast_karghah.place(x=800,y=395)

label_midde_left_darkhast_karghah=tk.Label(darkhast_karghah_window,text="اطلاعات مالک",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_midde_left_darkhast_karghah.place(x=100,y=395)

label_down_darkhast_karghah=tk.Label(darkhast_karghah_window,text="امکانات",bg="#052340",fg="#00BFFF",font=("Shabnam", 11))
label_down_darkhast_karghah.place(x=100,y=565)

frame_darkhast_kargah=tk.Frame((darkhast_karghah_window),bd=0,highlightthickness=0)
frame_darkhast_kargah.pack(side="left", fill="y", padx=6, pady=15)
#-------------------------------فریم بالا سمت راست----------------------------
sal_sakht_darkhast_kargah_lable = tk.Label(frame_up_right_darkhast_karghah, text="سال ساخت", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
sal_sakht_darkhast_kargah_lable.place(x=465, y=80, anchor="e")

sal_sakht_darkhast_kargah_entry = tk.Entry(frame_up_right_darkhast_karghah, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
sal_sakht_darkhast_kargah_entry.place(x=18, y=70, width=350, height=25)

karbari_darkhast_kargah = tk.Label(frame_up_right_darkhast_karghah, text="کاربری زمین", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
karbari_darkhast_kargah.place(x=465, y=40, anchor="e")

combo_darkhast_kargah=ttk.Combobox(frame_up_right_darkhast_karghah)
combo_darkhast_kargah["values"] = ("درخواست اجاره کارگاه","درخواست خرید کارگاه")
combo_darkhast_kargah["state"]=["readonly"]
combo_darkhast_kargah.set("درخواست خرید کارگاه")
combo_darkhast_kargah.configure(justify="center")
combo_darkhast_kargah.bind("<<ComboboxSelected>>",sabt_darkhast_kargah)
combo_darkhast_kargah.place(x=18, y=30, width=350, height=25)

metraj_darkhast_kargah = tk.Label(frame_up_right_darkhast_karghah, text="متراژ", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
metraj_darkhast_kargah.place(x=465, y=120,anchor="e")

metraj_darkhast_kargah_entry = tk.Entry(frame_up_right_darkhast_karghah, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
metraj_darkhast_kargah_entry.place(x=18, y=110, width=350, height=25)

#--------------------فریم چپ بالا---------------------------
photo_lbl2_darkhast_kargah = tk.Label(frame_up_left_darkhast_karghah, text="[تصویر ملک]", bg="#FFFFFF", width=79, height=15,relief="solid")
photo_lbl2_darkhast_kargah.place(x=40, y=10)

add_img_btn_darkhast_kargah = tk.Button(frame_up_left_darkhast_karghah, text="افزودن تصویر", bg="#00BFFF", fg="#ffffff",command=open_file, height=2,width=13)
add_img_btn_darkhast_kargah.place(x=240, y=250)

#--------------------------فریم راست وسط---------------------
loctaion_darkhast_kargah = tk.Label(frame_midde_right_darkhast_karghah, text="آدرس", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=12)
loctaion_darkhast_kargah.place(x=465, y=80, anchor="e")

loctaion_darkhast_kargah_entry = tk.Text(frame_midde_right_darkhast_karghah, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
loctaion_darkhast_kargah_entry.place(x=18, y=70, width=350, height=50)

ejareh_mahaneh_darkhast_kargah_lable=tk.Label(frame_midde_right_darkhast_karghah,text="اجاره ماهانه",bg="#052340",fg="#ffffff",font=("Shabnam", 12),width=9)
ejareh_mahaneh_darkhast_kargah_lable.place_forget()

ejareh_mahaneh_darkhast_kargah_entry=tk.Entry(frame_midde_right_darkhast_karghah,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
ejareh_mahaneh_darkhast_kargah_entry.place_forget()

gheimat_kol_darkhast_kargah_lable=tk.Label(frame_midde_right_darkhast_karghah,text="قیمت کل",bg="#052340",fg="#ffffff",font=("Shabnam", 12),width=9)
gheimat_kol_darkhast_kargah_lable.place(x=465, y=30, anchor="e")

gheimat_kol_darkhast_kargah_entry=tk.Entry(frame_midde_right_darkhast_karghah,bg="#ffffff", fg="#000000",font=("Shabnam", 10))
gheimat_kol_darkhast_kargah_entry.place(x=18, y=20, width=350, height=25)

mablagh_pish_darkhast_kargah_lable = tk.Label(frame_midde_right_darkhast_karghah, text="مبلغ پیش", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
mablagh_pish_darkhast_kargah_lable.place_forget()

mablagh_pish_darkhast_kargah_entry = tk.Entry(frame_midde_right_darkhast_karghah, bg="#ffffff", fg="#ffffff", font=("Shabnam", 10))
mablagh_pish_darkhast_kargah_entry.place_forget()
#------------------------------------فریم چپ وسط-----------------------
name_moshtari_darkhast_kargah_lable = tk.Label(frame_midde_left_darkhast_karghah,text="نام مالک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
name_moshtari_darkhast_kargah_lable.place(x=600, y=30,anchor="e")

name_moshtari_darkhast_kargah_entry = tk.Entry(frame_midde_left_darkhast_karghah, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
name_moshtari_darkhast_kargah_entry.place(x=30, y=20, width=350, height=25)

shomareh_moshtari_darkhast_kargah_lable = tk.Label(frame_midde_left_darkhast_karghah, text="شماره مالک", bg="#052340", fg="#ffffff", font=("Shabnam", 12), width=9)
shomareh_moshtari_darkhast_kargah_lable.place(x=600, y=80,anchor="e")

shomareh_moshtari_darkhast_kargah_entry = tk.Entry(frame_midde_left_darkhast_karghah, bg="#ffffff", fg="#000000", font=("Shabnam", 10))
shomareh_moshtari_darkhast_kargah_entry.place(x=30, y=70, width=350, height=25)

#---------------------------------فریم پایین--------------------------------
sarmayesh_darkhast_kargah = tk.Label(frame_down_darkhast_karghah, text="سیستم سرمایش", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
sarmayesh_darkhast_kargah.place(x=485, y=15)
sarmayesh_darkhast_kargah_combo = ttk.Combobox(frame_down_darkhast_karghah)
sarmayesh_darkhast_kargah_combo["values"] = ("ندارد", "پنکه سقفی", "کولر ابی", "کولر گازی ", "ابی/گازی")
sarmayesh_darkhast_kargah_combo["state"] = "readonly"
sarmayesh_darkhast_kargah_combo.configure(justify="center")
sarmayesh_darkhast_kargah_combo.place(x=325, y=15)

garmayesh_type_darkhast_kargah = tk.Label(frame_down_darkhast_karghah, text="سیستم گرمایش", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
garmayesh_type_darkhast_kargah.place(x=485, y=45)
garmayesh_type_darkhast_kargah_combo = ttk.Combobox(frame_down_darkhast_karghah)
garmayesh_type_darkhast_kargah_combo["values"] = ("ندارد", "بخاری", " شوفاژ", "گرمایش از کف ")
garmayesh_type_darkhast_kargah_combo["state"] = "readonly"
garmayesh_type_darkhast_kargah_combo.configure(justify="center")
garmayesh_type_darkhast_kargah_combo.place(x=325, y=45)

toilet_darkhast_kargah = tk.Label(frame_down_darkhast_karghah, text="سرویس بهداشتی", bg="#052340", fg="#ffffff", font=("Shabnam", 11))
toilet_darkhast_kargah.place(x=190, y=45)
toilet_darkhast_kargah_combo = ttk.Combobox(frame_down_darkhast_karghah)
toilet_darkhast_kargah_combo["values"] = ("ایرانی", "فرنگی", "هردو")
toilet_darkhast_kargah_combo["state"] = "readonly"
toilet_darkhast_kargah_combo.configure(justify="center")
toilet_darkhast_kargah_combo.place(x=20, y=45)

vaziat_bagh_darkhast_kargah=tk.Label(frame_down_darkhast_karghah,bg="#052340",fg="#ffffff",font=("Shabnam", 11),width=15,text="وضعیت برق")
vaziat_bagh_darkhast_kargah.place(x=190, y=15)

vaziat_bargh_darkhast_kargah_combo=ttk.Combobox(frame_down_darkhast_karghah)
vaziat_bargh_darkhast_kargah_combo["values"]=("","برق شهری","سه فاز","تک فاز")
vaziat_bargh_darkhast_kargah_combo.set("")
vaziat_bargh_darkhast_kargah_combo["state"]=["readonly"]
vaziat_bargh_darkhast_kargah_combo.place(x=20, y=15)

vaziat_ab_darkhast_kargah=tk.Label(frame_down_darkhast_karghah,bg="#052340",fg="#ffffff",width=13,text=" وضعیت آب",font=("Shabnam", 9))
vaziat_ab_darkhast_kargah.place(x=1064, y=13)

vaziat_ab_darkhast_kargah_combo=ttk.Combobox(frame_down_darkhast_karghah,width=24)
vaziat_ab_darkhast_kargah_combo["values"]=(""," آب  لوله کشی (بدون فشار) " ," آب لوله کشی (همراه موتور فشار) ","دارای منبع(همراه موتور فشار)","دارای منبع(بدون فشار)")
vaziat_ab_darkhast_kargah_combo.set("")
vaziat_ab_darkhast_kargah_combo["state"]=["readonly"]
vaziat_ab_darkhast_kargah_combo.place(x=910, y=13)

abzar_darkhast_kargah=tk.Label(frame_down_darkhast_karghah,bg="#052340",fg="#ffffff",font=("Shabnam", 11),width=15,text=" ابزار صنعتی ")
abzar_darkhast_kargah.place(x=745, y=45)

abzar_darkhast_kargah_combo=ttk.Combobox(frame_down_darkhast_karghah,width=18)
abzar_darkhast_kargah_combo["values"]=("","(کارگاه خالی) بدون دستگاه ","دارای دستگاه")
abzar_darkhast_kargah_combo.set("")
abzar_darkhast_kargah_combo["state"]=["readonly"]
abzar_darkhast_kargah_combo.place(x=620, y=45)

hamam_darkhast_kargah=tk.Label(frame_down_darkhast_karghah,bg="#052340",fg="#ffffff",font=("Shabnam", 11),width=13,text="حمام")
hamam_darkhast_kargah.place(x=1050, y=45)

hamam_darkhast_kargah_combo=ttk.Combobox(frame_down_darkhast_karghah,width=24)
hamam_darkhast_kargah_combo["values"]=("","ندارد","دارد")
hamam_darkhast_kargah_combo.set("")
hamam_darkhast_kargah_combo["state"]=["readonly"]
hamam_darkhast_kargah_combo.place(x=910, y=45)

otagh_darkhast_kargah=tk.Label(frame_down_darkhast_karghah,bg="#052340",fg="#ffffff",font=("Shabnam", 11),width=17,text="اتاق رختکن")
otagh_darkhast_kargah.place(x=740, y=15)

otagh_darkhast_kargah_combo=ttk.Combobox(frame_down_darkhast_karghah,width=18)
otagh_darkhast_kargah_combo["values"]=("","ندارد","دارد")
otagh_darkhast_kargah_combo.set("")
otagh_darkhast_kargah_combo["state"]=["readonly"]
otagh_darkhast_kargah_combo.place(x=620, y=15)
back_to_home_darkhast_kargah=tk.Button(darkhast_karghah_window,text="بازگشت",bg="#052340", fg="#ffffff",width=10,height=1,command=back_home_darkhast_kargah)
back_to_home_darkhast_kargah.place(x=300,y=30)

zakhire_darkhast_kargah=tk.Button(darkhast_karghah_window,text="ذخیره",bg="#00BFFF", fg="#ffffff",width=10,height=1,command=sabt_darkhast_kargah)
zakhire_darkhast_kargah.place(x=200,y=30
                              )

darkhast_karghah_window.protocol("WM_DELETE_WINDOW", lambda: None)
darkhast_karghah_window.resizable(False, False)

#endregion
#=============================== پنجره گزارشات ===============================
#--------------------------پنجره گزارش های مسکونی------------------------------
#region
gozaresh_maskoni = tk.Toplevel(root)
gozaresh_maskoni.title("گزارش مسکونی")
gozaresh_maskoni.geometry("600x380")
gozaresh_maskoni.withdraw()

bg_image = image_gozaresh_maskoni
bg_image = image_gozaresh_maskoni.resize((600, 380))
bg_photo = ImageTk.PhotoImage(bg_image)

# لیبل پس‌زمینه
bg_label = tk.Label(gozaresh_maskoni, image=bg_photo)
bg_label.image = bg_photo
bg_label.place(x=0, y=0, relwidth=1, relheight=1)

frame_gozaresh_maskoni= tk.Frame(gozaresh_maskoni,bd=0,highlightthickness=0)
frame_gozaresh_maskoni.pack(side="left", fill="y", padx=6, pady=15)

title_lbl_gozaresh_maskoni = tk.Label(gozaresh_maskoni,text="گزارش مسکونی",bg="#052340",fg="#00BFFF",font=("Shabnam", 15))
title_lbl_gozaresh_maskoni.place(x=40, y=25)

noe_gozaresh_maskoni=tk.Label(gozaresh_maskoni,text=" نوع گزارش ",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=10)
noe_gozaresh_maskoni.place(x=470, y=76)

error_label_maskoni=tk.Label(gozaresh_maskoni,text='',font=("Shabnam",10),fg="#E91414",bg="#FFFFFF")
error_label_maskoni.place(x=190,y=230,width=260,height=25)

gozaresh_file_combo_maskoni=ttk.Combobox(gozaresh_maskoni)
gozaresh_file_combo_maskoni["values"] = ("گزارش فایل اجاره","گزارش فایل فروش","گزارش فایل درخواست اجاره","گزارش فایل درخواست خرید")
gozaresh_file_combo_maskoni["state"]=["readonly"]
gozaresh_file_combo_maskoni.config(width=40)
gozaresh_file_combo_maskoni.configure(justify="center")
gozaresh_file_combo_maskoni.place(x=190, y=80)

save_gozaresh_maskoni = tk.Button(gozaresh_maskoni, text="تایید", command=excel_gozaresh_maskoni, bg="#00BFFF", fg="#000000", width=10, height=1)
save_gozaresh_maskoni.place(x=95, y=320)

back_to_home_gozaresh_maskoni = tk.Button(gozaresh_maskoni, text="بازگشت", command=back_home_gozaresh_maskoni, bg="#00BFFF", fg="#000000", width=10, height=1)
back_to_home_gozaresh_maskoni.place(x=215, y=320)

gozaresh_maskoni.protocol("WM_DELETE_WINDOW", lambda: None)
gozaresh_maskoni.resizable(False, False)
#endregion
#------------------------------پنجره گزارش اداری/تجاری------------------------------
#region
gozaresh_edari_tejari = tk.Toplevel(root)
gozaresh_edari_tejari.title("گزارش اداری / تجاری")
gozaresh_edari_tejari.geometry("600x380")
gozaresh_edari_tejari.withdraw()

bg_image = image_gozaresh_edari_tejari
bg_image = image_gozaresh_edari_tejari.resize((600, 380))
bg_photo = ImageTk.PhotoImage(bg_image)

# لیبل پس‌زمینه
bg_label = tk.Label(gozaresh_edari_tejari, image=bg_photo)
bg_label.image = bg_photo  # خیلی مهم: جلوگیری از پاک شدن عکس
bg_label.place(x=0, y=0, relwidth=1, relheight=1)

frame_gozaresh_edari_tejari= tk.Frame(gozaresh_edari_tejari,bd=0,highlightthickness=0)
frame_gozaresh_edari_tejari.pack(side="left", fill="y", padx=6, pady=15)

title_lbl_gozaresh_edari_tejari = tk.Label(gozaresh_edari_tejari,text="گزارش اداری/تجاری",bg="#052340",fg="#00BFFF",font=("Shabnam", 15))
title_lbl_gozaresh_edari_tejari.place(x=40, y=25)

noe_gozaresh_edari_tejari=tk.Label(gozaresh_edari_tejari,text=" نوع گزارش ",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=10)
noe_gozaresh_edari_tejari.place(x=470, y=76)

error_label_edari_tejari=tk.Label(gozaresh_edari_tejari,text='',font=("Shabnam",10),fg="#E91414",bg="#FFFFFF")
error_label_edari_tejari.place(x=190,y=230,width=260,height=25)

gozaresh_file_combo_edari_tejari=ttk.Combobox(gozaresh_edari_tejari)
gozaresh_file_combo_edari_tejari["values"] = ("گزارش فایل اجاره","گزارش فایل فروش","گزارش فایل درخواست اجاره","گزارش فایل درخواست خرید")
gozaresh_file_combo_edari_tejari["state"]=["readonly"]
gozaresh_file_combo_edari_tejari.config(width=40)
gozaresh_file_combo_edari_tejari.configure(justify="center")
gozaresh_file_combo_edari_tejari.place(x=190, y=80)

save_gozaresh_edari_tejari = tk.Button(gozaresh_edari_tejari,text="تایید", command=excel_gozaresh_edari_tejari, bg="#00BFFF", fg="#000000", width=10, height=1)
save_gozaresh_edari_tejari.place(x=95, y=320)

back_to_home_gozaresh_edari_tejari = tk.Button(gozaresh_edari_tejari, text="بازگشت", command=back_home_gozaresh_edari_tejari, bg="#00BFFF", fg="#000000", width=10, height=1)
back_to_home_gozaresh_edari_tejari.place(x=215, y=320)

gozaresh_edari_tejari.protocol("WM_DELETE_WINDOW", lambda: None)
gozaresh_edari_tejari.resizable(False, False)
#endregion
#---------------------------------پنجره گزارش های باغ/زمین-------------------------
#region
gozaresh_bagh_zamin = tk.Toplevel(root)
gozaresh_bagh_zamin.title("گزارش باغ/زمین")
gozaresh_bagh_zamin.geometry("600x380")
gozaresh_bagh_zamin.withdraw()

bg_image = image_gozaresh_bagh_zamin
bg_image = image_gozaresh_bagh_zamin.resize((600, 380))
bg_photo = ImageTk.PhotoImage(bg_image)

# لیبل پس‌زمینه
bg_label = tk.Label(gozaresh_bagh_zamin, image=bg_photo)
bg_label.image = bg_photo  # خیلی مهم: جلوگیری از پاک شدن عکس
bg_label.place(x=0, y=0, relwidth=1, relheight=1)

frame_gozaresh_bagh_zamin= tk.Frame(gozaresh_bagh_zamin,bd=0,highlightthickness=0)
frame_gozaresh_bagh_zamin.pack(side="left", fill="y", padx=6, pady=15)

title_lbl_gozaresh_bagh_zamin = tk.Label(gozaresh_bagh_zamin,text="گزارش باغ/زمین",bg="#052340",fg="#00BFFF",font=("Shabnam", 15))
title_lbl_gozaresh_bagh_zamin.place(x=40, y=25)

noe_gozaresh_bagh_zamin=tk.Label(gozaresh_bagh_zamin,text=" نوع گزارش ",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=10)
noe_gozaresh_bagh_zamin.place(x=470, y=76)

error_label_bagh_zamin=tk.Label(gozaresh_bagh_zamin,text='',font=("Shabnam",10),fg="#E91414",bg="#FFFFFF")
error_label_bagh_zamin.place(x=60,y=230,width=390,height=25)

gozaresh_file_combo_bagh_zamin=ttk.Combobox(gozaresh_bagh_zamin)
gozaresh_file_combo_bagh_zamin["values"] = ("گزارش فایل اجاره باغ","گزارش فایل فروش باغ","گزارش فایل درخواست اجاره باغ","گزارش فایل درخواست خرید باغ",
                                            "گزارش فایل اجاره زمین","گزارش فایل فروش زمین","گزارش فایل درخواست اجاره زمین","گزارش فایل درخواست خرید زمین")
gozaresh_file_combo_bagh_zamin["state"]=["readonly"]
gozaresh_file_combo_bagh_zamin.config(width=60)
gozaresh_file_combo_bagh_zamin.configure(justify="center")
gozaresh_file_combo_bagh_zamin.place(x=60 ,y=80)

save_gozaresh_bagh_zamin = tk.Button(gozaresh_bagh_zamin, text="تایید", command=excel_gozaresh_bagh_zamin, bg="#00BFFF", fg="#000000", width=10, height=1)
save_gozaresh_bagh_zamin.place(x=95, y=320)

back_to_home_gozaresh_bagh_zamin = tk.Button(gozaresh_bagh_zamin, text="بازگشت", command=back_home_gozaresh_bagh_zamin, bg="#00BFFF", fg="#000000", width=10, height=1)
back_to_home_gozaresh_bagh_zamin.place(x=215, y=320)

gozaresh_bagh_zamin.protocol("WM_DELETE_WINDOW", lambda: None)
gozaresh_bagh_zamin.resizable(False, False)
#endregion
#---------------------------------پنجره گزارش های کارگاه---------------------------
#region
gozaresh_kargah = tk.Toplevel(root)
gozaresh_kargah.title("گزارش  کارگاه")
gozaresh_kargah.geometry("600x380")
gozaresh_kargah.withdraw()

bg_image = image_gozaresh_kargah
bg_image = image_gozaresh_kargah.resize((600, 380))
bg_photo = ImageTk.PhotoImage(bg_image)

# لیبل پس‌زمینه
bg_label = tk.Label(gozaresh_kargah, image=bg_photo)
bg_label.image = bg_photo  # خیلی مهم: جلوگیری از پاک شدن عکس
bg_label.place(x=0, y=0, relwidth=1, relheight=1)

frame_gozaresh_kargah= tk.Frame(gozaresh_kargah,bd=0,highlightthickness=0)
frame_gozaresh_kargah.pack(side="left", fill="y", padx=6, pady=15)

title_lbl_gozaresh_kargah = tk.Label(gozaresh_kargah,text="گزارش کارگاه",bg="#052340",fg="#00BFFF",font=("Shabnam", 15))
title_lbl_gozaresh_kargah.place(x=40, y=25)

type_gozaresh_kargah=tk.Label(gozaresh_kargah,text=" نوع گزارش ",bg="#052340",fg="#ffffff",font=("Shabnam",12),width=10)
type_gozaresh_kargah.place(x=470, y=76)

error_label_kargah=tk.Label(gozaresh_kargah,text='',font=("Shabnam",10),fg="#E91414",bg="#FFFFFF")
error_label_kargah.place(x=190,y=230,width=260,height=25)

gozaresh_file_combo_kargah=ttk.Combobox(gozaresh_kargah)
gozaresh_file_combo_kargah["values"] = ("گزارش فایل اجاره","گزارش فایل فروش","گزارش فایل درخواست اجاره","گزارش فایل درخواست خرید")
gozaresh_file_combo_kargah["state"]=["readonly"]
gozaresh_file_combo_kargah.config(width=40)
gozaresh_file_combo_kargah.configure(justify="center")
gozaresh_file_combo_kargah.place(x=190, y=80)

save_gozaresh_edari_kargah = tk.Button(gozaresh_kargah, text="تایید", command=excel_gozaresh_kargah, bg="#00BFFF", fg="#000000", width=10, height=1)
save_gozaresh_edari_kargah.place(x=95, y=320)

back_to_home_gozaresh_kargah = tk.Button(gozaresh_kargah, text="بازگشت", command=back_home_gozaresh_kargah, bg="#00BFFF", fg="#000000", width=10, height=1)
back_to_home_gozaresh_kargah.place(x=215, y=320)

gozaresh_kargah.protocol("WM_DELETE_WINDOW", lambda: None)
gozaresh_kargah.resizable(False, False)

#endregion
#-------------------------------------پنجره قرارداد ------------------------------
#region
gharardad_window = tk.Toplevel(root)
gharardad_window.title("قراردادها")
gharardad_window.geometry("600x650")
gharardad_window.configure(bg="#052340")
gharardad_window.withdraw()


titr_window=tk.Label(gharardad_window,text="قراردادها",fg="#00BFFF",bg="#052340",font=("Shabnam",15))
titr_window.place(x=250,y=5)

titr_aval_page=tk.Label(gharardad_window,text="در این بخش میتوانید فایل قرارداد مورد نظر خود را ایجاد و دریافت کنید",fg="#FFFFFF",bg="#052340",font=("Shabnam",8))
titr_aval_page.place(x=130,y=30)

main_frame=tk.Frame(gharardad_window,highlightbackground="#00BFFF",
highlightthickness=1,width=520,height=473)
main_frame.configure(bg="#052340")
main_frame.place(x=40,y=50)

text_label1=tk.Label(main_frame,text="ایجاد قراردادها",font=("Shabnam",11),bg="#052340",fg="#00BFFF",bd=1)
text_label1.place(x=240,y=15)

khat_up_1=tk.Frame(main_frame, bg="#3A6EA5",height=1)
khat_up_1.place(x=40,y=20,width=130)

khat_up_2=tk.Frame(main_frame, bg="#3A6EA5",height=1)
khat_up_2.place(x=340,y=20,width=130)

ghararadad_image=tk.Label(main_frame,bg="#052340",fg="#00BFFF",image=image_gharardad)
ghararadad_image.place(x=180,y=3)

melk_image=tk.Label(main_frame,bg="#052340",fg="#00BFFF",image=image_melk)
melk_image.place(x=440,y=20)

type_gharardad_image=tk.Label(main_frame,bg="#052340",fg="#00BFFF",image=image_type_gharardad)
type_gharardad_image.place(x=440,y=100)

shakhs_aval_image=tk.Label(main_frame,bg="#052340",fg="#00BFFF",image=image_person)
shakhs_aval_image.place(x=440,y=170)

shakhs_dovom_image=tk.Label(main_frame,bg="#052340",fg="#00BFFF",image=image_gharardad)
shakhs_dovom_image.place(x=440,y=240)

tozihat_image=tk.Label(main_frame,bg="#052340",fg="#00BFFF",image=image_massage)
tozihat_image.place(x=440,y=320)

type_melk_gharardad=tk.Label(main_frame,text="نوع ملک",bg="#052340",fg="#FFFFFF",font=("Shabnam",8))
type_melk_gharardad.place(x=350,y=50)
type_melk_gharardad_combo=ttk.Combobox(main_frame)
type_melk_gharardad_combo["values"]=("مسکونی","اداری_تجاری","باغ","زمین","کارگاه")
type_melk_gharardad_combo["state"]=["readonly"]
type_melk_gharardad_combo.config(width=40)
type_melk_gharardad_combo.configure(justify="center")
type_melk_gharardad_combo.set("")
type_melk_gharardad_combo.place(x=25,y=50)

type_gharardad=tk.Label(main_frame,text="نوع قرارداد",bg="#052340",fg="#FFFFFF",font=("Shabnam",8))
type_gharardad.place(x=350,y=115)
type_gharardad_combo=ttk.Combobox(main_frame)
type_gharardad_combo["values"]=("خرید و فروش ","اجاره","مشارکت")
type_gharardad_combo["state"]=["readonly"]
type_gharardad_combo.config(width=40)
type_gharardad_combo.configure(justify="center")
type_gharardad_combo.set("")
type_gharardad_combo.place(x=25,y=105)


name_shakhs_aval_gharardad=tk.Label(main_frame,text="نام طرف اول قرارداد",bg="#052340",fg="#FFFFFF",font=("Shabnam",8))
name_shakhs_aval_gharardad.place(x=305,y=180)
name_shakhs_aval_gharardad_entry=tk.Entry(main_frame,bg="#ffffff",font=("Shabnam",10),fg='black')
name_shakhs_aval_gharardad_entry.place(x=25,y=180,width=260)

name_shakhs_dovom_gharardad=tk.Label(main_frame,text="نام طرف دوم قرارداد",bg="#052340",fg="#FFFFFF",font=("Shabnam",8))
name_shakhs_dovom_gharardad.place(x=305,y=250)
name_shakhs_dovom_gharardad_entry=tk.Entry(main_frame,bg="#ffffff",font=("Shabnam",10),fg='black')
name_shakhs_dovom_gharardad_entry.place(x=25,y=250,width=260)

tozih_gharardad=tk.Label(main_frame,text="توضیحات(اختیاری)",bg="#052340",fg="#FFFFFF",font=("Shabnam",8))
tozih_gharardad.place(x=310,y=315)
tozih_gharardad_entry=tk.Text(main_frame,bg="#ffffff",font=("Shabnam",10),fg='black')
tozih_gharardad_entry.place(x=25,y=315,width=260,height=45)

khat_vasat=tk.Frame(main_frame, bg="#3A6EA5",height=1)
khat_vasat.place(x=25,y=370,width=450)

word_btn=tk.Button(main_frame,bg="#0096D6",text="دریافت فایل ورد قرارداد",fg="#ffffff",image=image_word,compound="left",command=creat_word_gharardad)
word_btn.place(x=140,y=385,width=200)
label_titr_vasat1 = tk.Label(main_frame,text="با کلیک بر دکمه بالا، فایل قرارداد مورد نظر در قالب",fg="white",bg="#052340",font=("Shabnam", 8))
label_titr_vasat1.place(x=230, y=440)

label_titr_vasat2 = tk.Label(main_frame,text="Word",fg="#00B7EB",bg="#052340",font=("Shabnam", 8, "bold"))
label_titr_vasat2.place(x=200, y=440)

label_titr_vasat3= tk.Label(main_frame,text="در دسکتاپ ذخیره میشود",fg="white",bg="#052340",font=("Shabnam", 8))
label_titr_vasat3.place(x=80, y=440)

seconde_frame=tk.Frame(gharardad_window,highlightbackground="#00BFFF",
highlightthickness=1,width=520,height=80)
seconde_frame.configure(bg="#052340")
seconde_frame.place(x=40,y=530)

label_titr_rahgiri= tk.Label(seconde_frame,text="اخرین کد رهگیری تولید شده",fg="#ffffff",bg="#052340",font=("Shabnam", 8))
label_titr_rahgiri.place(x=195, y=2)
code_frame=tk.Frame(seconde_frame,highlightbackground="#00BFFF",highlightthickness=1,width=200,height=35)
code_frame.configure(bg="#052340")
code_frame.place(x=160,y=24)

code_label= tk.Label(code_frame,text=" ",fg="#ffffff",bg="#052340",font=("Shabnam", 8))
code_label.place(x=160,y=24)


label_titr_akhar= tk.Label(seconde_frame,text="این کد برای پیگیری قرارداد در بخش گزارش استفاده میشود",fg="#ffffff",bg="#052340",font=("Shabnam", 7))
label_titr_akhar.place(x=167, y=60)


back_to_main=tk.Button(gharardad_window,text="بازگشت",fg="#ffffff",bg="#052340",command=back_main_ghararadad)
back_to_main.place(x=230,y=620,width=140)

gharardad_window.protocol("WM_DELETE_WINDOW", lambda: None)
gharardad_window.resizable(False, False)
#endregion
#############################################################################
# ----------------------اجرای برنامه-------------------
#region
root.protocol("WM_DELETE_WINDOW",close_window)
root.mainloop()
#endregion